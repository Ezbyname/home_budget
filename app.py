import os
import csv
import io
import re
import sqlite3
import json
import random
import smtplib
import hashlib
import secrets
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import threading
import time as _time
from datetime import datetime, date, timedelta
from functools import wraps
from flask import Flask, request, jsonify, send_from_directory, send_file, session, redirect
import xlrd
import openpyxl

import sys

# When running as a PyInstaller exe, use the exe's directory for data files
if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
    STATIC_DIR = os.path.join(sys._MEIPASS, 'static')
else:
    BASE_DIR = os.path.dirname(__file__)
    STATIC_DIR = os.path.join(BASE_DIR, 'static')

APP_VERSION = '1.0.1000040'

# ---- Smart Tips Configuration ----
TIP_CONFIG = {
    # Scoring weights (sum ~1.0)
    'w_severity': 0.30,
    'w_impact': 0.25,
    'w_urgency': 0.20,
    'w_confidence': 0.10,
    'w_recency': 0.10,
    'w_action_bonus': 0.05,
    # Severity base scores
    'sev_important': 1.0,
    'sev_watch': 0.6,
    'sev_opportunity': 0.3,
    # Impact normalization cap (NIS)
    'impact_cap': 5000,
    # Dedup limits
    'max_per_category': 3,
    'max_total_tips': 8,
    # Subsumption pairs: (winner keeps, loser removed)
    'subsumption_rules': [
        ('negative_balance', 'balance_decline'),
        ('installment_avoid', 'installment_burden'),
        ('subscriptions_cost', 'subscriptions_growth'),
        ('true_savings_rate', 'low_savings_rate'),
        ('income_drop', 'income_gap'),
    ],
    # Coordination: trajectory overlap penalty
    'overlap_penalty': 0.15,
    # Positive reinforcement boost
    'boost_positive_traj': 0.20,
    'boost_mixed_traj': 0.10,
    'positive_tip_ids': {'good_savings_rate', 'passive_income_good', 'installment_freed', 'installment_ending', 'income_stable'},
    # Generator thresholds
    'food_monthly_threshold': 2000,
    'restaurant_ratio_threshold': 0.40,
    'entertainment_monthly_threshold': 1000,
    'cash_total_threshold': 500,
    'bit_total_threshold': 500,
    'savings_low_threshold': 0.10,
    'savings_good_threshold': 0.20,
    'emergency_buffer_months': 2,
    'installment_burden_threshold': 0.25,
    'passive_income_low_threshold': 0.10,
    'passive_income_good_threshold': 0.25,
    'deficit_months_threshold': 2,
    'fixed_ratio_threshold': 0.70,
    'category_spike_threshold': 0.25,
    'latte_min_vendors': 5,
    'latte_max_amount': 50,
    'installment_ending_days': 60,
    'balance_decline_months': 3,
    'balance_decline_pct': 0.05,
    'income_gap_threshold': 2,
    'national_comparison_min_cats': 2,
    'national_comparison_pp_above': 5,
    # New category thresholds
    'subscriptions_monthly_threshold': 200,
    'dining_out_monthly_threshold': 800,
    # Salary tips
    'vacation_days_threshold': 15,
    'true_savings_low_threshold': 15,
    # Income stability
    'income_cv_stable_threshold': 0.05,
    'income_stable_min_months': 3,
    'bonus_reliance_threshold': 0.15,
    'income_drop_threshold': -10,
    'bonus_share_contradicts_stable': 0.20,
    # Analytics
    'analytics_prune_days': 180,
}

app = Flask(__name__, static_folder=STATIC_DIR)
# Store user data in a stable folder that survives exe rebuilds
DATA_DIR = os.path.join(os.path.expanduser('~'), '.budget_tracker_data')
os.makedirs(DATA_DIR, exist_ok=True)
app.config['UPLOAD_FOLDER'] = os.path.join(DATA_DIR, 'uploads')
DB_PATH = os.path.join(DATA_DIR, 'budget.db')

# Session config — store secret key in user's home dir so it survives exe rebuilds
SECRET_FILE = os.path.join(os.path.expanduser('~'), '.budget_tracker_secret_key')
if os.path.exists(SECRET_FILE):
    with open(SECRET_FILE, 'r') as f:
        app.secret_key = f.read().strip()
else:
    app.secret_key = secrets.token_hex(32)
    with open(SECRET_FILE, 'w') as f:
        f.write(app.secret_key)
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=30)

# SMTP config file (created on first email setup)
SMTP_CONFIG_PATH = os.path.join(DATA_DIR, 'smtp_config.json')

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# External secrets for admin credentials
SECRETS_PATH = os.path.join(os.path.expanduser('~'), '.budget_tracker_secrets.json')
ADMIN_SECRETS = {}
if os.path.exists(SECRETS_PATH):
    with open(SECRETS_PATH, 'r') as f:
        ADMIN_SECRETS = json.load(f)

# Hebrew category mapping from the XLS structure
CATEGORY_MAP = {
    'דיור ואחזקת בית': 'housing',
    'מזון': 'food',
    'הורים': 'parents',
    'ילדים': 'children',
    'רכב': 'vehicle',
    'תקשורת': 'communication',
    'מוצרי טיפוח ובריאות': 'health_beauty',
    'ריפוי': 'medical',
    'ביטוחים': 'insurance',
    'בילוי,פנאי ובידור': 'entertainment',
    'אישי': 'personal',
    'חיסכון והתחייבויות': 'savings',
    'תשלומים שונים': 'misc',
    'ביגוד ואופנה': 'clothing',
    'מנויים': 'subscriptions',
    'חינוך ולימודים': 'education',
    'אוכל בחוץ': 'dining_out',
    'מתנות': 'gifts',
}

DEFAULT_CATEGORIES = [
    ('housing', 'דיור ואחזקת בית', '#4e79a7'),
    ('food', 'מזון', '#f28e2b'),
    ('children', 'ילדים', '#e15759'),
    ('vehicle', 'רכב', '#76b7b2'),
    ('communication', 'תקשורת', '#59a14f'),
    ('health_beauty', 'טיפוח ובריאות', '#edc948'),
    ('medical', 'ריפוי', '#b07aa1'),
    ('insurance', 'ביטוחים', '#ff9da7'),
    ('entertainment', 'בילוי ופנאי', '#9c755f'),
    ('personal', 'אישי', '#bab0ac'),
    ('savings', 'חיסכון והתחייבויות', '#4dc9f6'),
    ('misc', 'שונות', '#a5a5a5'),
    ('parents', 'הורים', '#d4a373'),
    ('clothing', 'ביגוד ואופנה', '#e377c2'),
    ('subscriptions', 'מנויים', '#17becf'),
    ('education', 'חינוך ולימודים', '#bcbd22'),
    ('dining_out', 'אוכל בחוץ', '#ff6b6b'),
    ('gifts', 'מתנות', '#c084fc'),
]


def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    return conn


def apply_category_rule(conn, description, category_id, frequency='random', user_id=None):
    """Check if user has a saved category/frequency rule for this description."""
    if description:
        rule = conn.execute(
            "SELECT category_id FROM category_rules WHERE description=?", (description,)
        ).fetchone()
        if rule:
            category_id = rule['category_id']
        # Also check if existing expenses with same description have a non-random frequency
        if user_id is not None:
            freq_row = conn.execute(
                "SELECT frequency FROM expenses WHERE description=? AND frequency != 'random' AND user_id=? LIMIT 1",
                (description, user_id)
            ).fetchone()
        else:
            freq_row = conn.execute(
                "SELECT frequency FROM expenses WHERE description=? AND frequency != 'random' LIMIT 1",
                (description,)
            ).fetchone()
        if freq_row:
            frequency = freq_row['frequency']
    return category_id, frequency


def init_db():
    conn = get_db()
    conn.executescript('''
        CREATE TABLE IF NOT EXISTS categories (
            id TEXT PRIMARY KEY,
            name_he TEXT NOT NULL,
            color TEXT NOT NULL DEFAULT '#888888',
            sort_order INTEGER DEFAULT 0
        );

        CREATE TABLE IF NOT EXISTS expenses (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT NOT NULL,
            category_id TEXT NOT NULL,
            subcategory TEXT,
            description TEXT,
            amount REAL NOT NULL,
            source TEXT DEFAULT 'manual',
            frequency TEXT DEFAULT 'random',
            card TEXT DEFAULT '',
            user_id INTEGER NOT NULL DEFAULT 0,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (category_id) REFERENCES categories(id)
        );

        CREATE TABLE IF NOT EXISTS income (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT NOT NULL,
            person TEXT NOT NULL,
            source TEXT NOT NULL,
            amount REAL NOT NULL,
            description TEXT,
            is_recurring INTEGER DEFAULT 0,
            user_id INTEGER NOT NULL DEFAULT 0,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS budget_plans (
            id INTEGER PRIMARY KEY,
            name TEXT NOT NULL DEFAULT 'תקציב 1',
            description TEXT NOT NULL DEFAULT '',
            user_id INTEGER NOT NULL DEFAULT 0
        );

        CREATE TABLE IF NOT EXISTS budget (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            category_id TEXT NOT NULL,
            month TEXT NOT NULL,
            planned_amount REAL NOT NULL DEFAULT 0,
            plan_id INTEGER NOT NULL DEFAULT 1,
            user_id INTEGER NOT NULL DEFAULT 0,
            UNIQUE(category_id, month, plan_id, user_id),
            FOREIGN KEY (category_id) REFERENCES categories(id),
            FOREIGN KEY (plan_id) REFERENCES budget_plans(id)
        );

        CREATE INDEX IF NOT EXISTS idx_expenses_date ON expenses(date);
        CREATE INDEX IF NOT EXISTS idx_expenses_category ON expenses(category_id);
        CREATE INDEX IF NOT EXISTS idx_income_date ON income(date);

        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            email TEXT DEFAULT '',
            phone TEXT DEFAULT '',
            verified INTEGER DEFAULT 0,
            verification_method TEXT DEFAULT '',
            otp_code TEXT DEFAULT '',
            otp_expires TEXT DEFAULT '',
            is_admin INTEGER DEFAULT 0,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS financial_products (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            type TEXT NOT NULL,
            subtype TEXT NOT NULL DEFAULT '',
            company TEXT NOT NULL DEFAULT '',
            name TEXT NOT NULL DEFAULT '',
            policy_number TEXT DEFAULT '',
            monthly_cost REAL DEFAULT 0,
            coverage_amount REAL DEFAULT 0,
            balance REAL DEFAULT 0,
            balance_date TEXT DEFAULT '',
            employee_pct REAL DEFAULT 0,
            employer_pct REAL DEFAULT 0,
            return_rate REAL DEFAULT 0,
            start_date TEXT DEFAULT '',
            renewal_date TEXT DEFAULT '',
            notes TEXT DEFAULT '',
            expense_pattern TEXT DEFAULT '',
            status TEXT DEFAULT 'active',
            user_id INTEGER NOT NULL DEFAULT 0,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS reminders (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL DEFAULT 'תזכורת ייבוא',
            method TEXT NOT NULL DEFAULT 'email',
            destination TEXT NOT NULL DEFAULT '',
            frequency TEXT NOT NULL DEFAULT 'monthly',
            day_of_month INTEGER DEFAULT 1,
            day_of_week INTEGER DEFAULT 0,
            hour INTEGER DEFAULT 9,
            minute INTEGER DEFAULT 0,
            message TEXT DEFAULT '',
            enabled INTEGER DEFAULT 1,
            last_sent TEXT DEFAULT '',
            user_id INTEGER NOT NULL DEFAULT 0,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS category_rules (
            description TEXT PRIMARY KEY,
            category_id TEXT NOT NULL,
            FOREIGN KEY (category_id) REFERENCES categories(id)
        );

        CREATE TABLE IF NOT EXISTS installments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            description TEXT NOT NULL,
            store TEXT DEFAULT '',
            total_amount REAL NOT NULL,
            total_payments INTEGER NOT NULL,
            payments_made INTEGER NOT NULL DEFAULT 0,
            monthly_payment REAL NOT NULL,
            start_date TEXT NOT NULL,
            card TEXT DEFAULT '',
            notes TEXT DEFAULT '',
            user_id INTEGER NOT NULL DEFAULT 0,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS chat_aliases (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_typed TEXT NOT NULL,
            actual_match TEXT NOT NULL,
            times_used INTEGER DEFAULT 1,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        );
        CREATE INDEX IF NOT EXISTS idx_chat_aliases_typed ON chat_aliases(user_typed);

        CREATE TABLE IF NOT EXISTS chat_feedback (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            query TEXT NOT NULL,
            rating INTEGER NOT NULL CHECK(rating BETWEEN 1 AND 5),
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS bank_balances (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL DEFAULT 0,
            account_name TEXT NOT NULL DEFAULT 'main',
            month TEXT NOT NULL,
            closing_balance REAL NOT NULL,
            last_transaction_date TEXT NOT NULL,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(user_id, account_name, month)
        );
        CREATE INDEX IF NOT EXISTS idx_bank_balances_user ON bank_balances(user_id, month);

        CREATE TABLE IF NOT EXISTS excluded_months (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            month TEXT NOT NULL,
            reason TEXT DEFAULT '',
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(user_id, month)
        );

        CREATE TABLE IF NOT EXISTS insurance_suggestions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            source_type TEXT NOT NULL DEFAULT '',
            source_expense_id INTEGER,
            transaction_date TEXT NOT NULL,
            merchant_name TEXT NOT NULL DEFAULT '',
            raw_description TEXT NOT NULL DEFAULT '',
            amount REAL NOT NULL DEFAULT 0,
            insurer_name_guess TEXT DEFAULT '',
            insurance_type_guess TEXT DEFAULT '',
            confidence_score REAL DEFAULT 0,
            detection_reasons TEXT DEFAULT '[]',
            status TEXT DEFAULT 'suggested',
            linked_product_id INTEGER,
            currency TEXT DEFAULT 'ILS',
            normalized_merchant TEXT DEFAULT '',
            dedupe_key TEXT DEFAULT '',
            suggested_market TEXT DEFAULT '',
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(user_id, source_expense_id)
        );
        CREATE INDEX IF NOT EXISTS idx_insurance_suggestions_user ON insurance_suggestions(user_id, status);

        CREATE TABLE IF NOT EXISTS insurance_blacklist (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            merchant_pattern TEXT NOT NULL,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(user_id, merchant_pattern)
        );

        CREATE TABLE IF NOT EXISTS insurance_ignore_rules (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            rule_type TEXT NOT NULL DEFAULT 'merchant',
            rule_value TEXT NOT NULL,
            reason TEXT DEFAULT '',
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(user_id, rule_type, rule_value)
        );

        CREATE TABLE IF NOT EXISTS insurance_overlap_alerts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            policy_a_id INTEGER NOT NULL,
            policy_b_id INTEGER NOT NULL,
            overlap_score INTEGER DEFAULT 0,
            alert_level TEXT DEFAULT 'info',
            overlap_type TEXT DEFAULT '',
            reasons_json TEXT DEFAULT '[]',
            estimated_duplicate_cost_monthly REAL DEFAULT 0,
            status TEXT DEFAULT 'open',
            user_note TEXT DEFAULT '',
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(user_id, policy_a_id, policy_b_id)
        );
        CREATE INDEX IF NOT EXISTS idx_overlap_alerts_user ON insurance_overlap_alerts(user_id, status);

        CREATE TABLE IF NOT EXISTS assets (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            asset_type TEXT NOT NULL DEFAULT 'custom',
            name TEXT NOT NULL DEFAULT '',
            current_value REAL DEFAULT 0,
            currency TEXT DEFAULT 'ILS',
            address TEXT DEFAULT '',
            mortgage_balance REAL DEFAULT 0,
            rent_income_monthly REAL DEFAULT 0,
            property_expenses_monthly REAL DEFAULT 0,
            platform_name TEXT DEFAULT '',
            dividend_income_monthly REAL DEFAULT 0,
            ticker_summary TEXT DEFAULT '',
            institution_name TEXT DEFAULT '',
            interest_rate REAL DEFAULT 0,
            interest_income_monthly REAL DEFAULT 0,
            notes TEXT DEFAULT '',
            status TEXT DEFAULT 'active',
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP
        );
        CREATE INDEX IF NOT EXISTS idx_assets_user ON assets(user_id, status);

        CREATE TABLE IF NOT EXISTS liabilities (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            liability_type TEXT NOT NULL DEFAULT 'custom',
            name TEXT NOT NULL DEFAULT '',
            current_balance REAL DEFAULT 0,
            currency TEXT DEFAULT 'ILS',
            monthly_payment REAL DEFAULT 0,
            interest_rate REAL DEFAULT 0,
            linked_asset_id INTEGER DEFAULT NULL,
            notes TEXT DEFAULT '',
            status TEXT DEFAULT 'active',
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP
        );
        CREATE INDEX IF NOT EXISTS idx_liabilities_user ON liabilities(user_id, status);

        CREATE TABLE IF NOT EXISTS net_worth_snapshots (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            month TEXT NOT NULL,
            total_assets REAL DEFAULT 0,
            total_liabilities REAL DEFAULT 0,
            net_worth REAL DEFAULT 0,
            assets_breakdown TEXT DEFAULT '{}',
            passive_income REAL DEFAULT 0,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(user_id, month)
        );
        CREATE INDEX IF NOT EXISTS idx_net_worth_user ON net_worth_snapshots(user_id);

        CREATE TABLE IF NOT EXISTS transaction_links (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            transaction_type TEXT NOT NULL,
            transaction_id INTEGER NOT NULL,
            asset_id INTEGER DEFAULT NULL,
            liability_id INTEGER DEFAULT NULL,
            status TEXT NOT NULL DEFAULT 'suggested',
            confidence REAL DEFAULT 0,
            source TEXT NOT NULL DEFAULT 'auto',
            reasons_json TEXT DEFAULT '[]',
            link_version INTEGER DEFAULT 1,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP,
            CHECK(asset_id IS NOT NULL OR liability_id IS NOT NULL)
        );
        CREATE INDEX IF NOT EXISTS idx_tl_user ON transaction_links(user_id, status);
        CREATE INDEX IF NOT EXISTS idx_tl_txn ON transaction_links(transaction_type, transaction_id);
        CREATE INDEX IF NOT EXISTS idx_tl_asset ON transaction_links(asset_id);
        CREATE INDEX IF NOT EXISTS idx_tl_liability ON transaction_links(liability_id);

        CREATE TABLE IF NOT EXISTS link_rules (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            rule_type TEXT NOT NULL,
            description_pattern TEXT NOT NULL,
            target_type TEXT NOT NULL,
            target_id INTEGER NOT NULL,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        );
        CREATE INDEX IF NOT EXISTS idx_lr_user ON link_rules(user_id, rule_type);

        CREATE TABLE IF NOT EXISTS installment_suggestions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            vendor_normalized TEXT NOT NULL DEFAULT '',
            vendor_display TEXT NOT NULL DEFAULT '',
            card TEXT DEFAULT '',
            monthly_amount REAL NOT NULL DEFAULT 0,
            estimated_total_payments INTEGER DEFAULT 0,
            estimated_payments_made INTEGER DEFAULT 0,
            estimated_start_date TEXT DEFAULT '',
            estimated_total_amount REAL DEFAULT 0,
            confidence_score REAL DEFAULT 0,
            confidence_level TEXT DEFAULT 'moderate',
            detection_reasons TEXT DEFAULT '[]',
            expense_ids TEXT DEFAULT '[]',
            months_seen TEXT DEFAULT '[]',
            group_id TEXT NOT NULL DEFAULT '',
            status TEXT NOT NULL DEFAULT 'suggested',
            linked_installment_id INTEGER DEFAULT NULL,
            dedupe_key TEXT NOT NULL DEFAULT '',
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(user_id, dedupe_key)
        );
        CREATE INDEX IF NOT EXISTS idx_isug_user ON installment_suggestions(user_id, status);
        CREATE INDEX IF NOT EXISTS idx_isug_dedupe ON installment_suggestions(user_id, dedupe_key);

        CREATE TABLE IF NOT EXISTS installment_transaction_links (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            installment_id INTEGER NOT NULL,
            expense_id INTEGER NOT NULL,
            status TEXT NOT NULL DEFAULT 'confirmed',
            confidence REAL DEFAULT 1.0,
            match_reasons_json TEXT DEFAULT '[]',
            payment_number INTEGER DEFAULT NULL,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(user_id, installment_id, expense_id)
        );
        CREATE INDEX IF NOT EXISTS idx_itl_inst ON installment_transaction_links(installment_id);
        CREATE INDEX IF NOT EXISTS idx_itl_exp ON installment_transaction_links(expense_id);

        CREATE TABLE IF NOT EXISTS installment_ignore_rules (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            rule_type TEXT NOT NULL DEFAULT 'never_suggest',
            rule_value TEXT NOT NULL,
            reason TEXT DEFAULT '',
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(user_id, rule_type, rule_value)
        );
        CREATE INDEX IF NOT EXISTS idx_iir_user ON installment_ignore_rules(user_id, rule_type);

        CREATE TABLE IF NOT EXISTS salary_statements (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL DEFAULT 0,
            person TEXT NOT NULL DEFAULT 'husband',
            month TEXT NOT NULL,
            company_name TEXT DEFAULT '',
            gross_salary REAL DEFAULT 0,
            net_salary REAL DEFAULT 0,
            income_tax REAL DEFAULT 0,
            social_security REAL DEFAULT 0,
            health_insurance REAL DEFAULT 0,
            pension_employee REAL DEFAULT 0,
            pension_employer REAL DEFAULT 0,
            education_fund_employee REAL DEFAULT 0,
            education_fund_employer REAL DEFAULT 0,
            severance_employer REAL DEFAULT 0,
            other_deductions REAL DEFAULT 0,
            bonus_amount REAL DEFAULT 0,
            vacation_days REAL DEFAULT 0,
            sick_days REAL DEFAULT 0,
            extraction_confidence REAL DEFAULT 0,
            raw_text TEXT DEFAULT '',
            source_filename TEXT DEFAULT '',
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(user_id, person, month)
        );
        CREATE INDEX IF NOT EXISTS idx_salary_stmts_user ON salary_statements(user_id, month);
    ''')

    # Migrate budget table: add plan_id column and fix unique constraint
    cols = [r[1] for r in conn.execute("PRAGMA table_info(budget)").fetchall()]
    if 'plan_id' not in cols:
        conn.execute("ALTER TABLE budget RENAME TO budget_old")
        conn.execute("""
            CREATE TABLE budget (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                category_id TEXT NOT NULL,
                month TEXT NOT NULL,
                planned_amount REAL NOT NULL DEFAULT 0,
                plan_id INTEGER NOT NULL DEFAULT 1,
                UNIQUE(category_id, month, plan_id),
                FOREIGN KEY (category_id) REFERENCES categories(id),
                FOREIGN KEY (plan_id) REFERENCES budget_plans(id)
            )
        """)
        conn.execute("INSERT INTO budget (id, category_id, month, planned_amount, plan_id) SELECT id, category_id, month, planned_amount, 1 FROM budget_old")
        conn.execute("DROP TABLE budget_old")
        conn.commit()

    # Migrate budget_plans: add description column if missing
    bp_cols = [r[1] for r in conn.execute("PRAGMA table_info(budget_plans)").fetchall()]
    if 'description' not in bp_cols:
        conn.execute("ALTER TABLE budget_plans ADD COLUMN description TEXT NOT NULL DEFAULT ''")
        conn.commit()

    # Migration: add is_admin column to users
    user_cols = [r[1] for r in conn.execute("PRAGMA table_info(users)").fetchall()]
    if 'is_admin' not in user_cols:
        conn.execute("ALTER TABLE users ADD COLUMN is_admin INTEGER DEFAULT 0")
        conn.commit()

    # Migration: add user_id to data tables for per-user isolation
    exp_cols = [r[1] for r in conn.execute("PRAGMA table_info(expenses)").fetchall()]
    if 'user_id' not in exp_cols:
        first_user = conn.execute("SELECT id FROM users ORDER BY id LIMIT 1").fetchone()
        default_uid = first_user['id'] if first_user else 1
        for tbl in ['expenses', 'income', 'budget_plans', 'installments', 'reminders', 'financial_products']:
            conn.execute(f"ALTER TABLE {tbl} ADD COLUMN user_id INTEGER NOT NULL DEFAULT {default_uid}")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_expenses_user ON expenses(user_id)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_income_user ON income(user_id)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_installments_user ON installments(user_id)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_reminders_user ON reminders(user_id)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_financial_products_user ON financial_products(user_id)")
        conn.commit()

    # Migration: add user_id to budget table (requires recreate for UNIQUE constraint)
    bud_cols = [r[1] for r in conn.execute("PRAGMA table_info(budget)").fetchall()]
    if 'user_id' not in bud_cols:
        first_user = conn.execute("SELECT id FROM users ORDER BY id LIMIT 1").fetchone()
        default_uid = first_user['id'] if first_user else 1
        conn.execute("ALTER TABLE budget RENAME TO budget_old_uid")
        conn.execute(f"""
            CREATE TABLE budget (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                category_id TEXT NOT NULL,
                month TEXT NOT NULL,
                planned_amount REAL NOT NULL DEFAULT 0,
                plan_id INTEGER NOT NULL DEFAULT 1,
                user_id INTEGER NOT NULL DEFAULT {default_uid},
                UNIQUE(category_id, month, plan_id, user_id),
                FOREIGN KEY (category_id) REFERENCES categories(id),
                FOREIGN KEY (plan_id) REFERENCES budget_plans(id)
            )
        """)
        conn.execute(f"""
            INSERT INTO budget (id, category_id, month, planned_amount, plan_id, user_id)
            SELECT id, category_id, month, planned_amount, plan_id, {default_uid} FROM budget_old_uid
        """)
        conn.execute("DROP TABLE budget_old_uid")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_budget_user ON budget(user_id)")
        conn.commit()

    # Migration: add is_unusual flag to expenses
    exp_cols2 = [r[1] for r in conn.execute("PRAGMA table_info(expenses)").fetchall()]
    if 'is_unusual' not in exp_cols2:
        conn.execute("ALTER TABLE expenses ADD COLUMN is_unusual INTEGER DEFAULT 0")
        conn.commit()

    # Migration: add new columns to insurance_suggestions for dual-market support
    ins_cols = [r[1] for r in conn.execute("PRAGMA table_info(insurance_suggestions)").fetchall()]
    for col, default in [('currency', "'ILS'"), ('normalized_merchant', "''"), ('dedupe_key', "''"), ('suggested_market', "''")]:
        if col not in ins_cols:
            conn.execute(f"ALTER TABLE insurance_suggestions ADD COLUMN {col} TEXT DEFAULT {default}")
    if ins_cols and 'currency' not in ins_cols:
        conn.commit()

    # Migration: add overlap-related columns to financial_products
    fp_cols = [r[1] for r in conn.execute("PRAGMA table_info(financial_products)").fetchall()]
    for col, default in [('insured_person', "''"), ('insured_object', "''"), ('coverage_tags', "'[]'")]:
        if col not in fp_cols:
            conn.execute(f"ALTER TABLE financial_products ADD COLUMN {col} TEXT DEFAULT {default}")
    if fp_cols and 'insured_person' not in fp_cols:
        conn.commit()

    # Migration: add link_keywords to assets and liabilities
    asset_cols2 = [r[1] for r in conn.execute("PRAGMA table_info(assets)").fetchall()]
    if 'link_keywords' not in asset_cols2:
        conn.execute("ALTER TABLE assets ADD COLUMN link_keywords TEXT DEFAULT '[]'")
        conn.commit()
    liab_cols2 = [r[1] for r in conn.execute("PRAGMA table_info(liabilities)").fetchall()]
    if 'link_keywords' not in liab_cols2:
        conn.execute("ALTER TABLE liabilities ADD COLUMN link_keywords TEXT DEFAULT '[]'")
        conn.commit()

    # Migration: add detection columns to installments table
    inst_cols2 = [r[1] for r in conn.execute("PRAGMA table_info(installments)").fetchall()]
    for col, default in [('status', "'active'"), ('source', "'manual'"), ('vendor_normalized', "''"),
                         ('updated_at', "CURRENT_TIMESTAMP"), ('last_matched_date', "''")]:
        if col not in inst_cols2:
            conn.execute(f"ALTER TABLE installments ADD COLUMN {col} TEXT DEFAULT {default}")
    if 'status' not in inst_cols2:
        conn.commit()

    # Tip interaction analytics
    conn.execute("""CREATE TABLE IF NOT EXISTS tip_events (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id TEXT NOT NULL,
        tip_id TEXT NOT NULL,
        event_type TEXT NOT NULL,
        action_target TEXT,
        month TEXT,
        created_at TEXT DEFAULT (datetime('now'))
    )""")

    # Insert default categories — add missing ones for existing DBs too
    existing_cats = {r[0] for r in conn.execute("SELECT id FROM categories").fetchall()}
    for i, (cat_id, name_he, color) in enumerate(DEFAULT_CATEGORIES):
        if cat_id not in existing_cats:
            conn.execute(
                "INSERT INTO categories (id, name_he, color, sort_order) VALUES (?, ?, ?, ?)",
                (cat_id, name_he, color, i)
            )
    conn.commit()
    conn.close()


init_db()


# ============================================================
# Authentication
# ============================================================

def hash_password(password):
    salt = secrets.token_hex(16)
    h = hashlib.pbkdf2_hmac('sha256', password.encode(), salt.encode(), 100000)
    return salt + ':' + h.hex()


def verify_password(password, stored):
    salt, h = stored.split(':')
    check = hashlib.pbkdf2_hmac('sha256', password.encode(), salt.encode(), 100000)
    return check.hex() == h


def generate_otp():
    return str(random.randint(100000, 999999))


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({'error': 'Not authenticated'}), 401
        return f(*args, **kwargs)
    return decorated


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({'error': 'Not authenticated'}), 401
        conn = get_db()
        user = conn.execute("SELECT is_admin FROM users WHERE id=?", (session['user_id'],)).fetchone()
        conn.close()
        if not user or not user['is_admin']:
            return jsonify({'error': 'Admin access required'}), 403
        return f(*args, **kwargs)
    return decorated


def get_uid():
    """Return current user's id from session."""
    return session['user_id']


def ensure_user_budget_plan(conn, user_id):
    """Ensure user has a default budget plan."""
    existing = conn.execute(
        "SELECT id FROM budget_plans WHERE user_id=?", (user_id,)
    ).fetchone()
    if not existing:
        conn.execute(
            "INSERT INTO budget_plans (name, description, user_id) VALUES ('תקציב 1', '', ?)",
            (user_id,)
        )
        conn.commit()


def has_any_users():
    conn = get_db()
    count = conn.execute("SELECT COUNT(*) FROM users").fetchone()[0]
    conn.close()
    return count > 0


def ensure_admin_user():
    """Auto-create admin user from external secrets file."""
    if not ADMIN_SECRETS.get('ADMIN_EMAIL') or not ADMIN_SECRETS.get('ADMIN_PASSWORD'):
        return
    admin_user = ADMIN_SECRETS.get('ADMIN_USERNAME', 'admin')
    conn = get_db()
    admin_exists = conn.execute("SELECT id FROM users WHERE username=?", (admin_user,)).fetchone()
    if not admin_exists:
        pw_hash = hash_password(ADMIN_SECRETS['ADMIN_PASSWORD'])
        conn.execute(
            "INSERT INTO users (username, password_hash, email, verified, is_admin) VALUES (?, ?, ?, 1, 1)",
            (admin_user, pw_hash, ADMIN_SECRETS['ADMIN_EMAIL'])
        )
        conn.commit()
    else:
        conn.execute("UPDATE users SET is_admin=1 WHERE username=?", (admin_user,))
        conn.commit()
    conn.close()


ensure_admin_user()


def send_email_otp(to_email, otp_code):
    """Send OTP via email. Returns True on success."""
    if not os.path.exists(SMTP_CONFIG_PATH):
        return False, 'SMTP not configured'
    with open(SMTP_CONFIG_PATH, 'r') as f:
        cfg = json.load(f)
    try:
        msg = MIMEMultipart()
        msg['From'] = cfg['from_email']
        msg['To'] = to_email
        msg['Subject'] = 'Family Budget Tracker - Verification Code'
        body = f"""
        <div dir="rtl" style="font-family:Arial;text-align:center;padding:20px">
            <h2>קוד אימות</h2>
            <p>הקוד שלך להרשמה למערכת ניהול תקציב:</p>
            <div style="font-size:36px;font-weight:bold;letter-spacing:8px;
                        background:#f0f5ff;padding:20px;border-radius:12px;
                        color:#2563eb;margin:20px 0">{otp_code}</div>
            <p style="color:#666">הקוד תקף ל-10 דקות</p>
        </div>
        """
        msg.attach(MIMEText(body, 'html'))
        server = smtplib.SMTP(cfg['smtp_server'], cfg.get('smtp_port', 587))
        server.starttls()
        server.login(cfg['smtp_user'], cfg['smtp_password'])
        server.send_message(msg)
        server.quit()
        return True, 'OK'
    except Exception as e:
        return False, str(e)


def send_sms_otp(phone, otp_code):
    """Send OTP via SMS (Twilio). Returns True on success."""
    if not os.path.exists(SMTP_CONFIG_PATH):
        return False, 'SMS not configured'
    with open(SMTP_CONFIG_PATH, 'r') as f:
        cfg = json.load(f)
    if not cfg.get('twilio_sid'):
        return False, 'Twilio not configured'
    try:
        import http.client
        import urllib.parse
        conn_http = http.client.HTTPSConnection("api.twilio.com")
        data = urllib.parse.urlencode({
            'To': phone,
            'From': cfg['twilio_from'],
            'Body': f'Family Budget Tracker - Your code is: {otp_code}'
        })
        auth = (cfg['twilio_sid'] + ':' + cfg['twilio_token']).encode()
        import base64
        auth_header = 'Basic ' + base64.b64encode(auth).decode()
        conn_http.request('POST',
            f'/2010-04-01/Accounts/{cfg["twilio_sid"]}/Messages.json',
            body=data,
            headers={
                'Authorization': auth_header,
                'Content-Type': 'application/x-www-form-urlencoded'
            })
        resp = conn_http.getresponse()
        if resp.status in (200, 201):
            return True, 'OK'
        return False, resp.read().decode()
    except Exception as e:
        return False, str(e)


# --- Auth Routes ---

@app.route('/auth')
def auth_page():
    """Serve the login/signup page."""
    return send_from_directory('static', 'auth.html')


@app.route('/api/auth/status', methods=['GET'])
def auth_status():
    """Check if user is logged in and if any users exist."""
    has_users = has_any_users()
    logged_in = 'user_id' in session
    username = session.get('username', '')
    is_admin = False
    if logged_in:
        conn = get_db()
        user = conn.execute("SELECT is_admin FROM users WHERE id=?", (session['user_id'],)).fetchone()
        conn.close()
        is_admin = bool(user and user['is_admin'])
    return jsonify({
        'has_users': has_users,
        'logged_in': logged_in,
        'username': username,
        'is_admin': is_admin,
    })


@app.route('/api/auth/signup', methods=['POST'])
def auth_signup():
    data = request.json
    username = data.get('username', '').strip()
    password = data.get('password', '')
    email = data.get('email', '').strip()
    phone = data.get('phone', '').strip()
    method = data.get('verification_method', 'email')  # 'email' or 'sms'

    if not username or not password:
        return jsonify({'error': 'Username and password required'}), 400
    if len(password) < 6:
        return jsonify({'error': 'Password must be at least 6 characters'}), 400
    if method == 'email' and not email:
        return jsonify({'error': 'Email required for email verification'}), 400
    if method == 'sms' and not phone:
        return jsonify({'error': 'Phone required for SMS verification'}), 400

    conn = get_db()
    existing = conn.execute("SELECT id FROM users WHERE username=?", (username,)).fetchone()
    if existing:
        conn.close()
        return jsonify({'error': 'Username already exists'}), 400

    otp = generate_otp()
    otp_expires = (datetime.now() + timedelta(minutes=10)).strftime('%Y-%m-%d %H:%M:%S')

    # Try to send OTP
    if method == 'email':
        success, msg = send_email_otp(email, otp)
    else:
        success, msg = send_sms_otp(phone, otp)

    if not success:
        # If sending fails, allow skip verification (for local/offline use)
        # Store user as unverified but let them set up verification later
        pass

    pw_hash = hash_password(password)
    conn.execute("""
        INSERT INTO users (username, password_hash, email, phone, verification_method, otp_code, otp_expires)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, (username, pw_hash, email, phone, method, otp, otp_expires))
    conn.commit()
    user_id = conn.execute("SELECT id FROM users WHERE username=?", (username,)).fetchone()[0]
    conn.close()

    if not success:
        # Auto-verify if we can't send OTP (offline mode)
        conn = get_db()
        conn.execute("UPDATE users SET verified=1 WHERE id=?", (user_id,))
        conn.commit()
        conn.close()
        session.permanent = True
        session['user_id'] = user_id
        session['username'] = username
        return jsonify({
            'status': 'ok',
            'auto_verified': True,
            'message': f'Verification skipped ({msg}). Account created.',
        })

    return jsonify({
        'status': 'verification_sent',
        'method': method,
        'user_id': user_id,
        'destination': email if method == 'email' else phone[-4:],
    })


@app.route('/api/auth/verify', methods=['POST'])
def auth_verify():
    data = request.json
    user_id = data.get('user_id')
    code = data.get('code', '').strip()

    conn = get_db()
    user = conn.execute("SELECT * FROM users WHERE id=?", (user_id,)).fetchone()
    if not user:
        conn.close()
        return jsonify({'error': 'User not found'}), 404

    if user['otp_code'] != code:
        conn.close()
        return jsonify({'error': 'Invalid code'}), 400

    if user['otp_expires'] and datetime.now().strftime('%Y-%m-%d %H:%M:%S') > user['otp_expires']:
        conn.close()
        return jsonify({'error': 'Code expired. Please request a new one.'}), 400

    conn.execute("UPDATE users SET verified=1, otp_code='', otp_expires='' WHERE id=?", (user_id,))
    conn.commit()

    session.permanent = True
    session['user_id'] = user_id
    session['username'] = user['username']
    conn.close()
    return jsonify({'status': 'ok', 'username': user['username']})


@app.route('/api/auth/resend', methods=['POST'])
def auth_resend():
    data = request.json
    user_id = data.get('user_id')

    conn = get_db()
    user = conn.execute("SELECT * FROM users WHERE id=?", (user_id,)).fetchone()
    if not user:
        conn.close()
        return jsonify({'error': 'User not found'}), 404

    otp = generate_otp()
    otp_expires = (datetime.now() + timedelta(minutes=10)).strftime('%Y-%m-%d %H:%M:%S')
    conn.execute("UPDATE users SET otp_code=?, otp_expires=? WHERE id=?", (otp, otp_expires, user_id))
    conn.commit()

    if user['verification_method'] == 'email':
        success, msg = send_email_otp(user['email'], otp)
    else:
        success, msg = send_sms_otp(user['phone'], otp)

    conn.close()
    if success:
        return jsonify({'status': 'ok'})
    return jsonify({'error': f'Failed to send: {msg}'}), 500


@app.route('/api/auth/login', methods=['POST'])
def auth_login():
    data = request.json
    identifier = data.get('username', '').strip()
    password = data.get('password', '')

    conn = get_db()
    # Allow login by username OR email
    user = conn.execute("SELECT * FROM users WHERE username=? OR email=?", (identifier, identifier)).fetchone()
    conn.close()

    if not user or not verify_password(password, user['password_hash']):
        return jsonify({'error': 'Invalid username or password'}), 401

    if not user['verified']:
        return jsonify({'error': 'Account not verified', 'needs_verification': True, 'user_id': user['id']}), 403

    session.permanent = True
    session['user_id'] = user['id']
    session['username'] = user['username']
    return jsonify({'status': 'ok', 'username': user['username']})


@app.route('/api/auth/logout', methods=['POST'])
def auth_logout():
    session.clear()
    return jsonify({'status': 'ok'})


@app.route('/api/reset-data', methods=['POST'])
@login_required
def reset_user_data():
    """Delete all data for the current user but keep the account."""
    conn = get_db()
    uid = get_uid()
    for tbl in ['expenses', 'income', 'budget', 'budget_plans', 'installments',
                'reminders', 'financial_products', 'bank_balances', 'excluded_months',
                'insurance_suggestions', 'insurance_blacklist', 'insurance_ignore_rules',
                'installment_suggestions', 'installment_transaction_links', 'installment_ignore_rules']:
        conn.execute(f"DELETE FROM {tbl} WHERE user_id=?", (uid,))
    conn.commit()
    conn.close()
    return jsonify({'status': 'ok'})


@app.route('/api/auth/smtp-config', methods=['GET'])
def get_smtp_config():
    if os.path.exists(SMTP_CONFIG_PATH):
        with open(SMTP_CONFIG_PATH, 'r') as f:
            cfg = json.load(f)
        # Mask sensitive fields
        return jsonify({
            'configured': True,
            'smtp_server': cfg.get('smtp_server', ''),
            'from_email': cfg.get('from_email', ''),
            'has_twilio': bool(cfg.get('twilio_sid')),
        })
    return jsonify({'configured': False})


@app.route('/api/auth/smtp-config', methods=['POST'])
def set_smtp_config():
    data = request.json
    cfg = {
        'smtp_server': data.get('smtp_server', 'smtp.gmail.com'),
        'smtp_port': int(data.get('smtp_port', 587)),
        'smtp_user': data.get('smtp_user', ''),
        'smtp_password': data.get('smtp_password', ''),
        'from_email': data.get('from_email', ''),
        'twilio_sid': data.get('twilio_sid', ''),
        'twilio_token': data.get('twilio_token', ''),
        'twilio_from': data.get('twilio_from', ''),
    }
    with open(SMTP_CONFIG_PATH, 'w') as f:
        json.dump(cfg, f)
    return jsonify({'status': 'ok'})


# ============================================================
# Reminders
# ============================================================

REMINDER_MESSAGES = {
    'bank': 'תזכורת: הגיע הזמן להוריד את דוח התנועות מהבנק ולייבא אותו למערכת ניהול התקציב.',
    'visa': 'תזכורת: הגיע הזמן להוריד את פירוט כרטיס האשראי ולייבא אותו למערכת ניהול התקציב.',
    'general': 'תזכורת: הגיע הזמן לעדכן את נתוני ההוצאות וההכנסות במערכת ניהול התקציב.',
}


def send_reminder(reminder):
    """Send a reminder via email or SMS. Returns (success, msg)."""
    dest = reminder['destination']
    text = reminder['message'] or REMINDER_MESSAGES.get('general', '')
    name = reminder['name']

    if reminder['method'] == 'email':
        if not os.path.exists(SMTP_CONFIG_PATH):
            return False, 'SMTP not configured'
        with open(SMTP_CONFIG_PATH, 'r') as f:
            cfg = json.load(f)
        try:
            msg = MIMEMultipart()
            msg['From'] = cfg['from_email']
            msg['To'] = dest
            msg['Subject'] = f'Family Budget Tracker - {name}'
            body = f"""
            <div dir="rtl" style="font-family:Arial;text-align:center;padding:20px">
                <h2>⏰ {name}</h2>
                <p style="font-size:16px;line-height:1.8">{text}</p>
                <div style="margin-top:20px;padding:15px;background:#f0f5ff;border-radius:12px">
                    <p style="color:#64748b;font-size:14px">
                        💡 היכנס לאתר הבנק או חברת האשראי, הורד את הדוח, ולחץ על "ייבוא" באפליקציה.
                    </p>
                </div>
            </div>
            """
            msg.attach(MIMEText(body, 'html'))
            server = smtplib.SMTP(cfg['smtp_server'], cfg.get('smtp_port', 587))
            server.starttls()
            server.login(cfg['smtp_user'], cfg['smtp_password'])
            server.sendmail(cfg['from_email'], dest, msg.as_string())
            server.quit()
            return True, 'Email sent'
        except Exception as e:
            return False, str(e)
    elif reminder['method'] == 'sms':
        return send_sms_otp(dest, text)
    return False, 'Unknown method'


def reminder_scheduler():
    """Background thread: check every 60s if any reminder should fire."""
    while True:
        _time.sleep(60)
        try:
            now = datetime.now()
            conn = get_db()
            reminders = conn.execute("SELECT * FROM reminders WHERE enabled=1").fetchall()
            for r in reminders:
                should_send = False
                last = r['last_sent']
                today_str = now.strftime('%Y-%m-%d')

                # Skip if already sent today
                if last and last[:10] == today_str:
                    continue

                # Check time
                if now.hour < r['hour'] or (now.hour == r['hour'] and now.minute < r['minute']):
                    continue

                freq = r['frequency']
                if freq == 'daily':
                    should_send = True
                elif freq == 'weekly':
                    if now.weekday() == (r['day_of_week'] or 0):
                        should_send = True
                elif freq == 'monthly':
                    if now.day == (r['day_of_month'] or 1):
                        should_send = True
                elif freq == 'biweekly':
                    # Send on 1st and 15th
                    if now.day in (1, 15):
                        should_send = True

                if should_send:
                    success, msg = send_reminder(dict(r))
                    if success:
                        conn.execute(
                            "UPDATE reminders SET last_sent=? WHERE id=?",
                            (now.strftime('%Y-%m-%d %H:%M:%S'), r['id'])
                        )
                        conn.commit()
            conn.close()
        except Exception:
            pass


# Start scheduler thread
_scheduler_thread = threading.Thread(target=reminder_scheduler, daemon=True)
_scheduler_thread.start()


@app.route('/api/reminders', methods=['GET'])
@login_required
def get_reminders():
    conn = get_db()
    rows = conn.execute("SELECT * FROM reminders WHERE user_id=? ORDER BY created_at DESC", (get_uid(),)).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route('/api/reminders', methods=['POST'])
@login_required
def add_reminder():
    data = request.json
    conn = get_db()
    conn.execute("""
        INSERT INTO reminders (name, method, destination, frequency, day_of_month, day_of_week, hour, minute, message, enabled, user_id)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        data.get('name', 'תזכורת ייבוא'),
        data.get('method', 'email'),
        data.get('destination', ''),
        data.get('frequency', 'monthly'),
        data.get('day_of_month', 1),
        data.get('day_of_week', 0),
        data.get('hour', 9),
        data.get('minute', 0),
        data.get('message', ''),
        1,
        get_uid()
    ))
    conn.commit()
    conn.close()
    return jsonify({'status': 'ok'})


@app.route('/api/reminders/<int:rid>', methods=['PUT'])
@login_required
def update_reminder(rid):
    data = request.json
    conn = get_db()
    conn.execute("""
        UPDATE reminders SET name=?, method=?, destination=?, frequency=?,
        day_of_month=?, day_of_week=?, hour=?, minute=?, message=?, enabled=?
        WHERE id=? AND user_id=?
    """, (
        data.get('name', ''),
        data.get('method', 'email'),
        data.get('destination', ''),
        data.get('frequency', 'monthly'),
        data.get('day_of_month', 1),
        data.get('day_of_week', 0),
        data.get('hour', 9),
        data.get('minute', 0),
        data.get('message', ''),
        data.get('enabled', 1),
        rid,
        get_uid()
    ))
    conn.commit()
    conn.close()
    return jsonify({'status': 'ok'})


@app.route('/api/reminders/<int:rid>', methods=['DELETE'])
@login_required
def delete_reminder(rid):
    conn = get_db()
    conn.execute("DELETE FROM reminders WHERE id=? AND user_id=?", (rid, get_uid()))
    conn.commit()
    conn.close()
    return jsonify({'status': 'ok'})


@app.route('/api/reminders/<int:rid>/test', methods=['POST'])
@login_required
def test_reminder(rid):
    """Send a test reminder immediately."""
    conn = get_db()
    r = conn.execute("SELECT * FROM reminders WHERE id=? AND user_id=?", (rid, get_uid())).fetchone()
    conn.close()
    if not r:
        return jsonify({'error': 'Reminder not found'}), 404
    success, msg = send_reminder(dict(r))
    if success:
        return jsonify({'status': 'ok', 'message': msg})
    return jsonify({'error': msg}), 500


# --- Static files ---
@app.route('/')
def index():
    resp = send_from_directory('static', 'index.html')
    resp.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    return resp


@app.route('/static/<path:path>')
def static_files(path):
    return send_from_directory('static', path)


# --- Categories API ---
@app.route('/api/categories', methods=['GET'])
@login_required
def get_categories():
    conn = get_db()
    rows = conn.execute("SELECT * FROM categories ORDER BY sort_order").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route('/api/categories', methods=['POST'])
@login_required
def add_category():
    data = request.json
    conn = get_db()
    conn.execute(
        "INSERT OR REPLACE INTO categories (id, name_he, color, sort_order) VALUES (?, ?, ?, ?)",
        (data['id'], data['name_he'], data.get('color', '#888888'), data.get('sort_order', 99))
    )
    conn.commit()
    conn.close()
    return jsonify({'status': 'ok'})


# --- Standing Orders API ---
@app.route('/api/standing-orders', methods=['GET'])
@login_required
def get_standing_orders():
    """Return latest occurrence of each monthly expense (grouped by description)."""
    conn = get_db()
    rows = conn.execute("""
        SELECT e.description, e.category_id, c.name_he as category_name,
               c.color as category_color, e.card, e.amount, MAX(e.date) as last_date
        FROM expenses e
        JOIN categories c ON e.category_id = c.id
        WHERE e.frequency = 'monthly' AND e.user_id = ?
        GROUP BY e.description
        ORDER BY e.amount DESC
    """, (get_uid(),)).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


# --- Expenses API ---
@app.route('/api/expenses', methods=['GET'])
@login_required
def get_expenses():
    conn = get_db()
    month = request.args.get('month')  # format: YYYY-MM
    from_date = request.args.get('from_date')  # YYYY-MM-DD
    to_date = request.args.get('to_date')  # YYYY-MM-DD
    cat_id = request.args.get('category_id')
    subcat = request.args.get('subcategory')
    cards = request.args.getlist('card')  # multi-select source/card filter

    sql = """SELECT e.*, c.name_he as category_name, c.color as category_color
             FROM expenses e JOIN categories c ON e.category_id = c.id WHERE e.user_id = ?"""
    params = [get_uid()]

    if from_date and to_date:
        sql += " AND e.date >= ? AND e.date <= ?"
        params += [from_date, to_date]
    elif month:
        sql += " AND e.date LIKE ?"
        params.append(month + '%')

    if cat_id:
        sql += " AND e.category_id = ?"
        params.append(cat_id)

    if subcat:
        sql += " AND (e.subcategory LIKE ? OR e.description LIKE ?)"
        params += [f'%{subcat}%', f'%{subcat}%']

    if cards:
        placeholders = ','.join('?' * len(cards))
        sql += f" AND e.card IN ({placeholders})"
        params += cards

    sql += " ORDER BY e.date DESC"
    if not from_date and not to_date and not month:
        sql += " LIMIT 500"

    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route('/api/expenses', methods=['POST'])
@login_required
def add_expense():
    data = request.json
    conn = get_db()
    conn.execute(
        "INSERT INTO expenses (date, category_id, subcategory, description, amount, source, frequency, user_id) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
        (data['date'], data['category_id'], data.get('subcategory', ''),
         data.get('description', ''), data['amount'], data.get('source', 'manual'),
         data.get('frequency', 'random'), get_uid())
    )
    conn.commit()
    conn.close()
    return jsonify({'status': 'ok'})


@app.route('/api/expenses/<int:expense_id>', methods=['PUT'])
@login_required
def update_expense(expense_id):
    data = request.json
    conn = get_db()
    fields = []
    values = []
    for col in ('date', 'category_id', 'subcategory', 'description', 'amount', 'frequency', 'card', 'is_unusual'):
        if col in data:
            fields.append(f"{col}=?")
            values.append(data[col])
    if not fields:
        conn.close()
        return jsonify({'error': 'No fields to update'}), 400
    values.append(expense_id)
    values.append(get_uid())
    conn.execute(f"UPDATE expenses SET {','.join(fields)} WHERE id=? AND user_id=?", values)

    propagated = 0
    # If category changed, update ALL expenses with the same description and save a rule
    if 'category_id' in data:
        exp = conn.execute("SELECT description FROM expenses WHERE id=? AND user_id=?", (expense_id, get_uid())).fetchone()
        if exp and exp['description']:
            desc = exp['description']
            cur = conn.execute(
                "UPDATE expenses SET category_id=? WHERE description=? AND user_id=? AND id!=?",
                (data['category_id'], desc, get_uid(), expense_id)
            )
            propagated = cur.rowcount
            conn.execute(
                "INSERT OR REPLACE INTO category_rules (description, category_id) VALUES (?, ?)",
                (desc, data['category_id'])
            )

    # If frequency changed, update ALL expenses with the same description
    if 'frequency' in data:
        exp = conn.execute("SELECT description FROM expenses WHERE id=? AND user_id=?", (expense_id, get_uid())).fetchone()
        if exp and exp['description']:
            conn.execute(
                "UPDATE expenses SET frequency=? WHERE description=? AND user_id=?",
                (data['frequency'], exp['description'], get_uid())
            )

    conn.commit()
    conn.close()
    return jsonify({'status': 'ok', 'propagated': propagated})


@app.route('/api/expenses/<int:expense_id>', methods=['DELETE'])
@login_required
def delete_expense(expense_id):
    conn = get_db()
    conn.execute("DELETE FROM expenses WHERE id = ? AND user_id = ?", (expense_id, get_uid()))
    conn.commit()
    conn.close()
    return jsonify({'status': 'ok'})


@app.route('/api/expenses/<int:eid>/unusual', methods=['POST'])
@login_required
def toggle_unusual(eid):
    """Toggle the is_unusual flag on an expense (reversible override)."""
    conn = get_db()
    uid = get_uid()
    row = conn.execute("SELECT is_unusual FROM expenses WHERE id=? AND user_id=?", (eid, uid)).fetchone()
    if not row:
        conn.close()
        return jsonify({'error': 'Not found'}), 404
    new_val = 0 if row['is_unusual'] else 1
    conn.execute("UPDATE expenses SET is_unusual=? WHERE id=? AND user_id=?", (new_val, eid, uid))
    conn.commit()
    conn.close()
    return jsonify({'is_unusual': new_val})


# --- Income API ---
@app.route('/api/income', methods=['GET'])
@login_required
def get_income():
    conn = get_db()
    month = request.args.get('month')
    if month:
        rows = conn.execute(
            "SELECT * FROM income WHERE user_id=? AND date LIKE ? ORDER BY date DESC",
            (get_uid(), month + '%')
        ).fetchall()
    else:
        rows = conn.execute("SELECT * FROM income WHERE user_id=? ORDER BY date DESC LIMIT 200", (get_uid(),)).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route('/api/income', methods=['POST'])
@login_required
def add_income():
    data = request.json
    conn = get_db()
    conn.execute(
        "INSERT INTO income (date, person, source, amount, description, is_recurring, user_id) VALUES (?, ?, ?, ?, ?, ?, ?)",
        (data['date'], data['person'], data['source'], data['amount'],
         data.get('description', ''), data.get('is_recurring', 0), get_uid())
    )
    conn.commit()
    conn.close()
    return jsonify({'status': 'ok'})


@app.route('/api/income/<int:income_id>', methods=['DELETE'])
@login_required
def delete_income(income_id):
    conn = get_db()
    conn.execute("DELETE FROM income WHERE id = ? AND user_id = ?", (income_id, get_uid()))
    conn.commit()
    conn.close()
    return jsonify({'status': 'ok'})


@app.route('/api/income/<int:iid>', methods=['PUT'])
@login_required
def update_income(iid):
    """Edit an income entry — enables manual salary marking via source='salary'."""
    data = request.json
    conn = get_db()
    fields, values = [], []
    for col in ('date', 'person', 'source', 'amount', 'description', 'is_recurring'):
        if col in data:
            fields.append(f"{col}=?")
            values.append(data[col])
    if not fields:
        conn.close()
        return jsonify({'error': 'No fields to update'}), 400
    values.extend([iid, get_uid()])
    conn.execute(f"UPDATE income SET {','.join(fields)} WHERE id=? AND user_id=?", values)
    conn.commit()
    conn.close()
    return jsonify({'status': 'ok'})


# --- Budget API ---
@app.route('/api/available-months', methods=['GET'])
@login_required
def get_available_months():
    """Return all months that have expense data, sorted ascending."""
    conn = get_db()
    rows = conn.execute(
        "SELECT DISTINCT substr(date, 1, 7) as month FROM expenses WHERE user_id=? ORDER BY month",
        (get_uid(),)
    ).fetchall()
    conn.close()
    return jsonify([r['month'] for r in rows])


@app.route('/api/category-averages', methods=['GET'])
@login_required
def get_category_averages():
    """Return monthly average expense per category. Optional ?from=YYYY-MM filter."""
    conn = get_db()
    from_month = request.args.get('from')
    uid = get_uid()
    if from_month and from_month != 'all':
        rows = conn.execute("""
            SELECT e.category_id,
                   SUM(e.amount) as total,
                   COUNT(DISTINCT substr(e.date, 1, 7)) as months
            FROM expenses e
            WHERE e.user_id=? AND substr(e.date, 1, 7) >= ?
            GROUP BY e.category_id
        """, (uid, from_month)).fetchall()
    else:
        rows = conn.execute("""
            SELECT e.category_id,
                   SUM(e.amount) as total,
                   COUNT(DISTINCT substr(e.date, 1, 7)) as months
            FROM expenses e
            WHERE e.user_id=?
            GROUP BY e.category_id
        """, (uid,)).fetchall()
    conn.close()
    result = {}
    for r in rows:
        months = max(r['months'], 1)
        result[r['category_id']] = round(r['total'] / months, 0)
    return jsonify(result)


@app.route('/api/budget', methods=['GET'])
@login_required
def get_budget():
    conn = get_db()
    month = request.args.get('month')
    plan_id = request.args.get('plan', '1')
    uid = get_uid()
    ensure_user_budget_plan(conn, uid)
    if month:
        rows = conn.execute(
            """SELECT b.*, c.name_he, c.color FROM budget b
               JOIN categories c ON b.category_id = c.id
               WHERE b.month = ? AND b.plan_id = ? AND b.user_id = ? ORDER BY c.sort_order""",
            (month, plan_id, uid)
        ).fetchall()
    else:
        rows = conn.execute(
            """SELECT b.*, c.name_he, c.color FROM budget b
               JOIN categories c ON b.category_id = c.id
               WHERE b.plan_id = ? AND b.user_id = ?
               ORDER BY b.month DESC, c.sort_order""",
            (plan_id, uid)
        ).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route('/api/budget', methods=['POST'])
@login_required
def set_budget():
    data = request.json
    plan_id = data.get('plan_id', 1)
    conn = get_db()
    conn.execute(
        """INSERT INTO budget (category_id, month, planned_amount, plan_id, user_id)
           VALUES (?, ?, ?, ?, ?)
           ON CONFLICT(category_id, month, plan_id, user_id)
           DO UPDATE SET planned_amount = excluded.planned_amount""",
        (data['category_id'], data['month'], data['planned_amount'], plan_id, get_uid())
    )
    conn.commit()
    conn.close()
    return jsonify({'status': 'ok'})


@app.route('/api/budget-plans', methods=['GET'])
@login_required
def get_budget_plans():
    conn = get_db()
    ensure_user_budget_plan(conn, get_uid())
    rows = conn.execute("SELECT * FROM budget_plans WHERE user_id=? ORDER BY id", (get_uid(),)).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route('/api/budget-plans', methods=['POST'])
@login_required
def save_budget_plan():
    data = request.json
    conn = get_db()
    uid = get_uid()
    count = conn.execute("SELECT COUNT(*) FROM budget_plans WHERE user_id=?", (uid,)).fetchone()[0]
    desc = data.get('description', '')
    if data.get('id'):
        conn.execute("UPDATE budget_plans SET name = ?, description = ? WHERE id = ? AND user_id = ?",
                     (data['name'], desc, data['id'], uid))
    elif count < 3:
        conn.execute("INSERT INTO budget_plans (name, description, user_id) VALUES (?, ?, ?)",
                     (data['name'], desc, uid))
    else:
        conn.close()
        return jsonify({'error': 'max 3 plans'}), 400
    conn.commit()
    conn.close()
    return jsonify({'status': 'ok'})


@app.route('/api/budget-plans/<int:plan_id>', methods=['DELETE'])
@login_required
def delete_budget_plan(plan_id):
    if plan_id == 1:
        return jsonify({'error': 'cannot delete default plan'}), 400
    conn = get_db()
    uid = get_uid()
    conn.execute("DELETE FROM budget WHERE plan_id = ? AND user_id = ?", (plan_id, uid))
    conn.execute("DELETE FROM budget_plans WHERE id = ? AND user_id = ?", (plan_id, uid))
    conn.commit()
    conn.close()
    return jsonify({'status': 'ok'})


# --- Dashboard summary ---
@app.route('/api/summary', methods=['GET'])
@login_required
def get_summary():
    conn = get_db()
    month = request.args.get('month', date.today().strftime('%Y-%m'))
    uid = get_uid()

    # Expenses by category
    cat_expenses = conn.execute(
        """SELECT c.id, c.name_he, c.color, COALESCE(SUM(e.amount), 0) as total
           FROM categories c LEFT JOIN expenses e ON c.id = e.category_id AND e.date LIKE ? AND e.user_id = ?
           GROUP BY c.id ORDER BY total DESC""",
        (month + '%', uid)
    ).fetchall()

    # Daily expenses for the month
    daily = conn.execute(
        """SELECT date, SUM(amount) as total FROM expenses
           WHERE user_id = ? AND date LIKE ? GROUP BY date ORDER BY date""",
        (uid, month + '%')
    ).fetchall()

    # Total income for the month
    income_total = conn.execute(
        "SELECT COALESCE(SUM(amount), 0) as total FROM income WHERE user_id = ? AND date LIKE ?",
        (uid, month + '%')
    ).fetchone()['total']

    # Income by person
    income_by_person = conn.execute(
        """SELECT person, SUM(amount) as total FROM income
           WHERE user_id = ? AND date LIKE ? GROUP BY person""",
        (uid, month + '%')
    ).fetchall()

    # Budget vs actual
    plan_id = request.args.get('plan', '1')
    budget_vs_actual = conn.execute(
        """SELECT c.id, c.name_he, c.color,
                  COALESCE(b.planned_amount, 0) as planned,
                  COALESCE(SUM(e.amount), 0) as actual
           FROM categories c
           LEFT JOIN budget b ON c.id = b.category_id AND b.month = ? AND b.plan_id = ? AND b.user_id = ?
           LEFT JOIN expenses e ON c.id = e.category_id AND e.date LIKE ? AND e.user_id = ?
           GROUP BY c.id
           HAVING planned > 0 OR actual > 0
           ORDER BY c.sort_order""",
        (month, plan_id, uid, month + '%', uid)
    ).fetchall()

    # Monthly trend (last 6 months) — normalized insight, excludes unusual + excluded months
    excl_months = get_excluded_month_set(conn, uid)
    monthly_trend_raw = conn.execute(
        """SELECT substr(date, 1, 7) as month, SUM(amount) as total
           FROM expenses WHERE user_id = ? AND is_unusual = 0 GROUP BY substr(date, 1, 7)
           ORDER BY month DESC""",
        (uid,)
    ).fetchall()
    monthly_trend = [r for r in monthly_trend_raw if r['month'] not in excl_months][:6]

    # Expenses by card
    by_card = conn.execute(
        """SELECT CASE WHEN card = '' THEN 'אחר' ELSE card END as card_name,
                  SUM(amount) as total, COUNT(*) as count
           FROM expenses WHERE user_id = ? AND date LIKE ?
           GROUP BY card_name ORDER BY total DESC""",
        (uid, month + '%')
    ).fetchall()

    # Expenses by frequency
    by_frequency = conn.execute(
        """SELECT frequency, SUM(amount) as total, COUNT(*) as count
           FROM expenses WHERE user_id = ? AND date LIKE ?
           GROUP BY frequency""",
        (uid, month + '%')
    ).fetchall()

    expense_total = sum(r['total'] for r in cat_expenses)

    # Bank closing balance (separate metric from cashflow)
    bank_bal_row = conn.execute(
        "SELECT closing_balance, last_transaction_date FROM bank_balances WHERE user_id=? AND account_name='main' AND month=?",
        (uid, month)
    ).fetchone()
    bank_balance = bank_bal_row['closing_balance'] if bank_bal_row else None
    bank_balance_date = bank_bal_row['last_transaction_date'] if bank_bal_row else None

    # Bank balance trend (last 6 months, only months with actual data)
    bank_balance_trend = conn.execute(
        "SELECT month, closing_balance FROM bank_balances WHERE user_id=? AND account_name='main' ORDER BY month DESC LIMIT 6",
        (uid,)
    ).fetchall()

    # Trust signals: last import timestamp, salary day, upcoming fixed total
    last_import_row = conn.execute(
        "SELECT created_at FROM bank_balances WHERE user_id=? AND account_name='main' ORDER BY created_at DESC LIMIT 1",
        (uid,)
    ).fetchone()
    last_bank_import = last_import_row['created_at'] if last_import_row else None

    # Quick salary detection (typical day from income records)
    salary_rows = conn.execute(
        "SELECT date FROM income WHERE user_id=? AND source='salary' ORDER BY date DESC LIMIT 6",
        (uid,)
    ).fetchall()
    salary_day = None
    if len(salary_rows) >= 2:
        sal_days = []
        for r in salary_rows:
            parts = r['date'].split('-')
            if len(parts) == 3:
                sal_days.append(int(parts[2]))
        if sal_days:
            sorted_d = sorted(sal_days)
            salary_day = sorted_d[len(sorted_d) // 2]

    # Upcoming fixed expenses (monthly standing orders total)
    fixed_total = conn.execute(
        """SELECT COALESCE(SUM(amount), 0) FROM (
            SELECT amount FROM expenses WHERE user_id=? AND frequency='monthly'
            GROUP BY description)""",
        (uid,)
    ).fetchone()[0]

    conn.close()
    return jsonify({
        'month': month,
        'expense_total': expense_total,
        'income_total': income_total,
        'balance': income_total - expense_total,
        'bank_balance': bank_balance,
        'bank_balance_date': bank_balance_date,
        'bank_balance_trend': [dict(r) for r in reversed(list(bank_balance_trend))],
        'by_category': [dict(r) for r in cat_expenses if r['total'] > 0],
        'daily': [dict(r) for r in daily],
        'income_by_person': [dict(r) for r in income_by_person],
        'budget_vs_actual': [dict(r) for r in budget_vs_actual],
        'monthly_trend': [dict(r) for r in reversed(list(monthly_trend))],
        'by_card': [dict(r) for r in by_card],
        'by_frequency': [dict(r) for r in by_frequency],
        'last_bank_import': last_bank_import,
        'salary_day': salary_day,
        'fixed_monthly_total': round(fixed_total, 0),
    })


# --- Safe to Spend ---
@app.route('/api/safe-to-spend', methods=['GET'])
@login_required
def safe_to_spend():
    """Available after safety buffer: bank balance minus a conservative buffer.
    Simple and honest — the balance already reflects past debits, so we don't
    subtract expenses again (that would double-count)."""
    conn = get_db()
    uid = get_uid()
    month = request.args.get('month', date.today().strftime('%Y-%m'))

    # Current bank balance
    bal = conn.execute(
        "SELECT closing_balance, last_transaction_date FROM bank_balances WHERE user_id=? AND account_name='main' AND month=?",
        (uid, month)
    ).fetchone()
    if not bal:
        conn.close()
        return jsonify({'available': False, 'reason': 'no_balance_data'})
    bank_balance = bal['closing_balance']
    balance_date = bal['last_transaction_date']

    # Safety buffer: fixed 500 NIS (or equivalent baseline)
    buffer = 500

    # Upcoming installment commitments (active plans, remaining this month)
    inst_rows = conn.execute(
        "SELECT monthly_payment, total_payments, payments_made, status FROM installments WHERE user_id=?",
        (uid,)
    ).fetchall()
    installment_monthly = sum(
        r['monthly_payment'] for r in inst_rows
        if max(r['total_payments'] - r['payments_made'], 0) > 0 and dict(r).get('status') != 'completed'
    )

    available = bank_balance - buffer - installment_monthly

    # Detect salary timing for smarter daily budget
    today = date.today()
    days_left = None
    daily_available = None
    days_until_salary = None
    salary_detected = False

    if month == today.strftime('%Y-%m'):
        import calendar
        days_left = calendar.monthrange(today.year, today.month)[1] - today.day

        # Check for salary pattern to use as budget horizon
        salary_rows = conn.execute(
            "SELECT date FROM income WHERE user_id=? AND source='salary' ORDER BY date",
            (uid,)
        ).fetchall()
        if len(salary_rows) >= 2:
            sal_days = []
            for r in salary_rows:
                parts = r['date'].split('-')
                if len(parts) == 3:
                    sal_days.append(int(parts[2]))
            if sal_days:
                sorted_d = sorted(sal_days)
                typical_day = sorted_d[len(sorted_d) // 2]
                salary_detected = True
                if today.day <= typical_day:
                    days_until_salary = typical_day - today.day
                else:
                    if today.month == 12:
                        next_m_days = calendar.monthrange(today.year + 1, 1)[1]
                    else:
                        next_m_days = calendar.monthrange(today.year, today.month + 1)[1]
                    days_until_salary = (calendar.monthrange(today.year, today.month)[1] - today.day) + min(typical_day, next_m_days)

        # Use days until salary if detected, otherwise days until end of month
        horizon = days_until_salary if days_until_salary is not None else days_left
        daily_available = round(available / max(horizon, 1), 0) if available > 0 else 0

    # "Why" inputs for frontend explainability
    why_inputs = {
        'balance': round(bank_balance, 0),
        'buffer': round(buffer, 0),
        'installments': round(installment_monthly, 0),
        'result': round(available, 0),
    }
    if daily_available is not None:
        horizon = days_until_salary if days_until_salary is not None else days_left
        why_inputs['daily_result'] = round(daily_available, 0)
        why_inputs['horizon'] = horizon

    conn.close()
    return jsonify({
        'available': True,
        'bank_balance': round(bank_balance, 0),
        'buffer': round(buffer, 0),
        'installment_monthly': round(installment_monthly, 0),
        'safe_to_spend': round(available, 0),
        'daily_budget': daily_available,
        'days_remaining': days_left,
        'days_until_salary': days_until_salary,
        'salary_detected': salary_detected,
        'balance_date': balance_date,
        'why_inputs': why_inputs,
    })


# --- Salary Detection ---
@app.route('/api/salary-info', methods=['GET'])
@login_required
def salary_info():
    """Detect salary pattern from income records: typical day, amount, person."""
    conn = get_db()
    uid = get_uid()

    # Find all salary-tagged income entries (source='salary')
    rows = conn.execute(
        "SELECT date, amount, person FROM income WHERE user_id=? AND source='salary' ORDER BY date",
        (uid,)
    ).fetchall()

    if len(rows) < 2:
        conn.close()
        return jsonify({'detected': False, 'reason': 'not_enough_data'})

    # Group by person to detect multiple salaries (husband/wife)
    from collections import defaultdict
    by_person = defaultdict(list)
    for r in rows:
        by_person[r['person']].append({'date': r['date'], 'amount': r['amount']})

    salaries = []
    for person, entries in by_person.items():
        if len(entries) < 2:
            continue
        days = []
        amounts = []
        for e in entries:
            parts = e['date'].split('-')
            if len(parts) == 3:
                days.append(int(parts[2]))
                amounts.append(e['amount'])

        if not days:
            continue

        # Typical day: median (robust to occasional shifts)
        sorted_days = sorted(days)
        typical_day = sorted_days[len(sorted_days) // 2]
        avg_amount = round(sum(amounts) / len(amounts), 0)
        latest_amount = amounts[-1]

        # Days until next salary
        today = date.today()
        if today.day <= typical_day:
            next_salary = today.replace(day=typical_day)
        else:
            # Next month
            if today.month == 12:
                next_salary = today.replace(year=today.year + 1, month=1, day=typical_day)
            else:
                import calendar
                next_month = today.month + 1
                max_day = calendar.monthrange(today.year, next_month)[1]
                next_salary = today.replace(month=next_month, day=min(typical_day, max_day))
        days_until = (next_salary - today).days

        salaries.append({
            'person': person,
            'typical_day': typical_day,
            'avg_amount': avg_amount,
            'latest_amount': round(latest_amount, 0),
            'occurrences': len(entries),
            'days_until_next': days_until,
            'next_date': next_salary.strftime('%Y-%m-%d'),
        })

    conn.close()

    if not salaries:
        return jsonify({'detected': False, 'reason': 'not_enough_data'})

    # Sort by amount descending (primary salary first)
    salaries.sort(key=lambda s: s['avg_amount'], reverse=True)
    total_monthly = sum(s['avg_amount'] for s in salaries)
    nearest_salary_days = min(s['days_until_next'] for s in salaries)

    return jsonify({
        'detected': True,
        'salaries': salaries,
        'total_monthly_salary': total_monthly,
        'days_until_nearest': nearest_salary_days,
    })


# --- Salary Statements (Payslip Import) ---

@app.route('/api/salary-statements', methods=['GET'])
@login_required
def get_salary_statements():
    """List all parsed payslips for the current user."""
    conn = get_db()
    uid = get_uid()
    rows = conn.execute(
        "SELECT id, person, month, company_name, gross_salary, net_salary, "
        "income_tax, social_security, health_insurance, "
        "pension_employee, pension_employer, "
        "education_fund_employee, education_fund_employer, "
        "severance_employer, other_deductions, bonus_amount, "
        "vacation_days, sick_days, extraction_confidence, "
        "source_filename, created_at "
        "FROM salary_statements WHERE user_id=? ORDER BY month DESC",
        (uid,)
    ).fetchall()
    conn.close()
    result = []
    for r in rows:
        d = dict(r)
        employer_total = (d.get('pension_employer') or 0) + (d.get('education_fund_employer') or 0) + (d.get('severance_employer') or 0)
        d['employer_contributions_total'] = round(employer_total, 2)
        d['true_compensation'] = round((d.get('gross_salary') or 0) + employer_total, 2)
        result.append(d)
    return jsonify(result)


@app.route('/api/salary-statements', methods=['POST'])
@login_required
def save_salary_statement():
    """Save a user-reviewed payslip to the database."""
    data = request.json
    if not data or not data.get('month'):
        return jsonify({'error': 'Month is required'}), 400
    conn = get_db()
    uid = get_uid()
    conn.execute("""
        INSERT OR REPLACE INTO salary_statements
        (user_id, person, month, company_name, gross_salary, net_salary,
         income_tax, social_security, health_insurance,
         pension_employee, pension_employer,
         education_fund_employee, education_fund_employer,
         severance_employer, other_deductions, bonus_amount,
         vacation_days, sick_days, extraction_confidence,
         raw_text, source_filename)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (uid,
          data.get('person', 'husband'),
          data['month'],
          data.get('company_name', ''),
          data.get('gross_salary', 0),
          data.get('net_salary', 0),
          data.get('income_tax', 0),
          data.get('social_security', 0),
          data.get('health_insurance', 0),
          data.get('pension_employee', 0),
          data.get('pension_employer', 0),
          data.get('education_fund_employee', 0),
          data.get('education_fund_employer', 0),
          data.get('severance_employer', 0),
          data.get('other_deductions', 0),
          data.get('bonus_amount', 0),
          data.get('vacation_days', 0),
          data.get('sick_days', 0),
          data.get('extraction_confidence', 0),
          data.get('raw_text', ''),
          data.get('source_filename', '')))
    conn.commit()
    # Compute return values
    employer_total = (data.get('pension_employer', 0) +
                      data.get('education_fund_employer', 0) +
                      data.get('severance_employer', 0))
    true_comp = data.get('gross_salary', 0) + employer_total
    conn.close()
    return jsonify({
        'status': 'ok',
        'true_compensation': round(true_comp, 2),
        'employer_contributions': round(employer_total, 2),
    })


@app.route('/api/salary-statements/summary', methods=['GET'])
@login_required
def salary_statements_summary():
    """True Compensation summary with baseline normalization (excludes bonus months)."""
    conn = get_db()
    uid = get_uid()
    rows = conn.execute(
        "SELECT * FROM salary_statements WHERE user_id=? ORDER BY month DESC",
        (uid,)
    ).fetchall()
    conn.close()
    if not rows:
        return jsonify({'has_data': False})

    statements = [dict(r) for r in rows]
    latest = statements[0]
    n = len(statements)

    # Baseline: exclude bonus months for stable averages
    baseline = [s for s in statements if (s.get('bonus_amount') or 0) == 0] or statements
    bn = len(baseline)

    avg_gross = sum((s.get('gross_salary') or 0) for s in baseline) / bn
    avg_net = sum((s.get('net_salary') or 0) for s in baseline) / bn
    avg_employer = sum(
        (s.get('pension_employer') or 0) + (s.get('education_fund_employer') or 0) + (s.get('severance_employer') or 0)
        for s in baseline
    ) / bn
    avg_deductions = avg_gross - avg_net
    avg_true_comp = avg_gross + avg_employer

    _lpe = (latest.get('pension_employer') or 0)
    _lee = (latest.get('education_fund_employer') or 0)
    _lse = (latest.get('severance_employer') or 0)
    latest_employer = {
        'pension': _lpe,
        'education_fund': _lee,
        'severance': _lse,
        'total': round(_lpe + _lee + _lse, 2),
    }
    _lit = (latest.get('income_tax') or 0)
    _lss = (latest.get('social_security') or 0)
    _lhi = (latest.get('health_insurance') or 0)
    _lpe2 = (latest.get('pension_employee') or 0)
    _lefe = (latest.get('education_fund_employee') or 0)
    _lod = (latest.get('other_deductions') or 0)
    latest_deductions = {
        'income_tax': _lit,
        'social_security': _lss,
        'health_insurance': _lhi,
        'pension_employee': _lpe2,
        'education_fund_employee': _lefe,
        'other': _lod,
        'total': round(_lit + _lss + _lhi + _lpe2 + _lefe + _lod, 2),
    }

    # ── Income Risk Score ──
    grosses = [(s.get('gross_salary') or 0) for s in statements]
    mean_g = sum(grosses) / len(grosses) if grosses else 0
    cv = 0
    if mean_g > 0:
        variance = sum((g - mean_g) ** 2 for g in grosses) / len(grosses)
        cv = (variance ** 0.5) / mean_g

    total_gross_all = sum(grosses)
    total_bonus_all = sum((s.get('bonus_amount') or 0) for s in statements)
    bonus_share = total_bonus_all / total_gross_all if total_gross_all > 0 else 0

    baseline_gross = sum((s.get('gross_salary') or 0) for s in baseline) / bn
    latest_gross_val = (latest.get('gross_salary') or 0)
    latest_vs_baseline = ((latest_gross_val - baseline_gross) / baseline_gross * 100) if baseline_gross > 0 else 0

    # Component scores (each 0.0–1.0, then weighted)
    cv_score = min(cv / 0.20, 1.0)
    bonus_sc = min(bonus_share / 0.40, 1.0)
    drop_sc = min(max(-latest_vs_baseline, 0) / 30.0, 1.0)
    risk_score = round(cv_score * 0.4 + bonus_sc * 0.3 + drop_sc * 0.3, 2)

    if risk_score < 0.20:
        risk_level = 'low'
    elif risk_score < 0.50:
        risk_level = 'medium'
    else:
        risk_level = 'high'

    # Status key: pick the strongest signal for the dashboard card text
    if latest_vs_baseline <= -10 and n >= 2:
        status_key = 'income_status_drop'
        status_params = {'pct': f'{abs(latest_vs_baseline):.0f}'}
    elif bonus_share >= 0.15:
        status_key = 'income_status_bonus_reliance'
        status_params = {'pct': f'{bonus_share * 100:.0f}'}
    elif cv < 0.05 and n >= 3 and bonus_share < 0.20:
        status_key = 'income_status_stable'
        status_params = {'months': str(n)}
    else:
        status_key = 'income_status_neutral'
        status_params = {}

    return jsonify({
        'has_data': True,
        'months_count': n,
        'baseline_months': bn,
        'latest_month': latest['month'],
        'company_name': latest['company_name'],
        'person': latest['person'],
        'latest_gross': (latest.get('gross_salary') or 0),
        'latest_net': (latest.get('net_salary') or 0),
        'latest_employer': latest_employer,
        'latest_deductions': latest_deductions,
        'latest_true_compensation': round((latest.get('gross_salary') or 0) + latest_employer['total'], 2),
        'latest_bonus': latest.get('bonus_amount', 0),
        'latest_vacation_days': latest.get('vacation_days', 0),
        'latest_sick_days': latest.get('sick_days', 0),
        'avg_gross': round(avg_gross, 0),
        'avg_net': round(avg_net, 0),
        'avg_employer_contributions': round(avg_employer, 0),
        'avg_true_compensation': round(avg_true_comp, 0),
        'avg_deductions': round(avg_deductions, 0),
        'income_risk': {
            'level': risk_level,
            'score': risk_score,
            'status_key': status_key,
            'status_params': status_params,
            'cv': round(cv, 4),
            'bonus_share': round(bonus_share, 4),
            'latest_vs_baseline': round(latest_vs_baseline, 1),
        },
    })


@app.route('/api/salary-statements/<int:sid>', methods=['DELETE'])
@login_required
def delete_salary_statement(sid):
    conn = get_db()
    uid = get_uid()
    cursor = conn.execute("DELETE FROM salary_statements WHERE id=? AND user_id=?", (sid, uid))
    deleted = cursor.rowcount
    conn.commit()
    conn.close()
    return jsonify({'status': 'ok', 'deleted': deleted})


# --- Monthly Financial Summary ---
@app.route('/api/monthly-summary', methods=['GET'])
@login_required
def monthly_financial_summary():
    """Concise end-of-month financial snapshot."""
    conn = get_db()
    uid = get_uid()
    month = request.args.get('month', date.today().strftime('%Y-%m'))

    income = conn.execute(
        "SELECT COALESCE(SUM(amount),0) FROM income WHERE user_id=? AND date LIKE ?",
        (uid, month + '%')
    ).fetchone()[0]
    expenses = conn.execute(
        "SELECT COALESCE(SUM(amount),0) FROM expenses WHERE user_id=? AND date LIKE ?",
        (uid, month + '%')
    ).fetchone()[0]
    net_cashflow = income - expenses

    # Bank balance this month
    bal = conn.execute(
        "SELECT closing_balance FROM bank_balances WHERE user_id=? AND account_name='main' AND month=?",
        (uid, month)
    ).fetchone()
    bank_balance = bal['closing_balance'] if bal else None

    # Previous month balance
    y, m = int(month[:4]), int(month[5:7])
    pm = f"{y if m > 1 else y - 1}-{m - 1 if m > 1 else 12:02d}"
    prev_bal = conn.execute(
        "SELECT closing_balance FROM bank_balances WHERE user_id=? AND account_name='main' AND month=?",
        (uid, pm)
    ).fetchone()
    prev_bank_balance = prev_bal['closing_balance'] if prev_bal else None
    balance_change = round(bank_balance - prev_bank_balance, 0) if bank_balance is not None and prev_bank_balance is not None else None

    # Emergency buffer: months of expenses covered (use whichever is higher: total avg or fixed avg)
    # Normalized insight — excludes unusual transactions and excluded months
    excl_months = get_excluded_month_set(conn, uid)
    total_by_month = conn.execute(
        """SELECT substr(date,1,7) as m, SUM(amount) as monthly_total
           FROM expenses WHERE user_id=? AND is_unusual=0 GROUP BY m""",
        (uid,)
    ).fetchall()
    total_vals = [r['monthly_total'] for r in total_by_month if r['m'] not in excl_months]
    avg_total_expenses = sum(total_vals) / len(total_vals) if total_vals else 0
    fixed_by_month = conn.execute(
        """SELECT substr(date,1,7) as m, SUM(amount) as monthly_total
           FROM expenses WHERE user_id=? AND frequency='monthly' AND is_unusual=0 GROUP BY m""",
        (uid,)
    ).fetchall()
    fixed_vals = [r['monthly_total'] for r in fixed_by_month if r['m'] not in excl_months]
    avg_fixed_expenses = sum(fixed_vals) / len(fixed_vals) if fixed_vals else 0
    essential_monthly = max(avg_total_expenses, avg_fixed_expenses)
    emergency_months = round(bank_balance / essential_monthly, 1) if bank_balance and bank_balance > 0 and essential_monthly > 0 else None

    # "Why" inputs for frontend explainability
    why_inputs = {
        'cashflow': {
            'income': round(income, 0),
            'expenses': round(expenses, 0),
            'result': round(net_cashflow, 0),
        },
    }
    if emergency_months is not None:
        why_inputs['emergency_buffer'] = {
            'balance': round(bank_balance, 0),
            'essential_monthly': round(essential_monthly, 0),
            'result': emergency_months,
        }
    if balance_change is not None:
        why_inputs['balance_change'] = {
            'current': round(bank_balance, 0),
            'previous': round(prev_bank_balance, 0),
            'change': round(balance_change, 0),
        }

    conn.close()
    return jsonify({
        'month': month,
        'income': round(income, 0),
        'expenses': round(expenses, 0),
        'net_cashflow': round(net_cashflow, 0),
        'bank_balance': bank_balance,
        'prev_bank_balance': prev_bank_balance,
        'balance_change': balance_change,
        'emergency_buffer_months': emergency_months,
        'why_inputs': why_inputs,
    })


# --- Excluded Months (user override) ---
@app.route('/api/excluded-months', methods=['GET'])
@login_required
def get_excluded_months():
    conn = get_db()
    rows = conn.execute("SELECT id, month, reason FROM excluded_months WHERE user_id=? ORDER BY month",
                        (get_uid(),)).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route('/api/excluded-months', methods=['POST'])
@login_required
def add_excluded_month():
    data = request.json
    month = data.get('month', '').strip()
    if not month:
        return jsonify({'error': 'Month required'}), 400
    conn = get_db()
    conn.execute("INSERT OR IGNORE INTO excluded_months (user_id, month, reason) VALUES (?, ?, ?)",
                 (get_uid(), month, data.get('reason', '')))
    conn.commit()
    conn.close()
    return jsonify({'status': 'ok'})


@app.route('/api/excluded-months/<int:mid>', methods=['DELETE'])
@login_required
def delete_excluded_month(mid):
    conn = get_db()
    conn.execute("DELETE FROM excluded_months WHERE id=? AND user_id=?", (mid, get_uid()))
    conn.commit()
    conn.close()
    return jsonify({'status': 'ok'})


def get_excluded_month_set(conn, uid):
    """Return set of 'YYYY-MM' months the user excluded from trend analysis."""
    rows = conn.execute("SELECT month FROM excluded_months WHERE user_id=?", (uid,)).fetchall()
    return {r['month'] for r in rows}


# --- Bank Balances Management ---
@app.route('/api/bank-balances', methods=['GET'])
@login_required
def list_bank_balances():
    conn = get_db()
    rows = conn.execute(
        "SELECT id, account_name, month, closing_balance, last_transaction_date FROM bank_balances WHERE user_id=? ORDER BY month DESC",
        (get_uid(),)
    ).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route('/api/bank-balances/<int:bid>', methods=['PUT'])
@login_required
def update_bank_balance(bid):
    """Edit bank balance record — enables account name correction."""
    data = request.json
    conn = get_db()
    fields, values = [], []
    for col in ('account_name',):
        if col in data and data[col].strip():
            fields.append(f"{col}=?")
            values.append(data[col].strip())
    if not fields:
        conn.close()
        return jsonify({'error': 'No fields to update'}), 400
    values.extend([bid, get_uid()])
    try:
        conn.execute(f"UPDATE bank_balances SET {','.join(fields)} WHERE id=? AND user_id=?", values)
        conn.commit()
    except Exception:
        conn.close()
        return jsonify({'error': 'Account name conflict'}), 409
    conn.close()
    return jsonify({'status': 'ok'})


# --- File Import ---
@app.route('/api/import', methods=['POST'])
@login_required
def import_file():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400

    file = request.files['file']
    if not file.filename:
        return jsonify({'error': 'No file selected'}), 400

    filepath = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
    file.save(filepath)

    try:
        uid = get_uid()
        if file.filename.endswith('.xls'):
            result = parse_budget_xls(filepath, uid)
        elif file.filename.endswith('.xlsx'):
            # Auto-detect: insurance portfolio vs visa report
            if _is_insurance_portfolio_xlsx(filepath):
                result = parse_insurance_portfolio_xlsx(filepath, uid)
            else:
                result = parse_visa_xlsx(filepath, uid)
        elif file.filename.endswith('.csv'):
            result = parse_bank_csv(filepath, uid)
        elif file.filename.lower().endswith('.pdf'):
            result = parse_payslip_pdf(filepath, uid)
        else:
            return jsonify({'error': 'Unsupported file format. Use .xls, .xlsx, .csv or .pdf'}), 400
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 400


def parse_budget_xls(filepath, user_id=None):
    """Parse the Hebrew budget XLS format and import expenses."""
    wb = xlrd.open_workbook(filepath)
    sheet = wb.sheets()[0]
    conn = get_db()

    imported = 0
    current_category = None
    category_id = None

    # Try to extract month from cell B1 (Excel date serial)
    month_str = None
    try:
        date_val = sheet.cell_value(1, 1)
        if date_val and isinstance(date_val, float):
            dt = xlrd.xldate_as_datetime(date_val, wb.datemode)
            month_str = dt.strftime('%Y-%m')
    except Exception:
        pass

    if not month_str:
        month_str = date.today().strftime('%Y-%m')

    # Category row markers (col 0 values that indicate category headers)
    category_markers = {
        'דיור': 'housing',
        'מזון': 'food',
        'הורים': 'parents',
        'ילדים': 'children',
        'רכב': 'vehicle',
        'תקשורת': 'communication',
        'טיפוח': 'health_beauty',
        'ריפוי': 'medical',
        'ביטוחים': 'insurance',
        'בילוי': 'entertainment',
        'אישי': 'personal',
        'חיסכון': 'savings',
        'תשלומים': 'misc',
        'נוספים': 'misc',
    }

    for row in range(2, sheet.nrows):
        col0 = str(sheet.cell_value(row, 0)).strip()
        col1 = str(sheet.cell_value(row, 1)).strip()

        # Check if this is a category header row
        if col0 and not col0.startswith('סה'):
            for marker, cat_id in category_markers.items():
                if marker in col0:
                    current_category = col0
                    category_id = cat_id
                    break

        # Skip header/summary rows
        if col0.startswith('סה') or col0 == 'סוג ההוצאה' or not category_id:
            continue

        # Extract amounts from weekly columns (actual columns: 3, 6, 9, 12)
        subcategory = col1 if col1 else current_category
        week_cols = [(3, 1), (6, 8), (9, 16), (12, 24)]  # (col_index, day_start)

        for amount_col, day_start in week_cols:
            try:
                amount = float(sheet.cell_value(row, amount_col))
                if amount > 0:
                    # Use middle of the week as the date
                    day = min(day_start + 3, 28)
                    expense_date = f"{month_str}-{day:02d}"
                    # Skip duplicates
                    dup = conn.execute(
                        "SELECT COUNT(*) FROM expenses WHERE date=? AND category_id=? AND amount=? AND source='xls_import' AND user_id=?",
                        (expense_date, category_id, amount, user_id)
                    ).fetchone()[0]
                    if not dup:
                        conn.execute(
                            """INSERT INTO expenses (date, category_id, subcategory, description, amount, source, user_id)
                               VALUES (?, ?, ?, ?, ?, ?, ?)""",
                            (expense_date, category_id, subcategory, f'Imported from XLS', amount, 'xls_import', user_id)
                        )
                        imported += 1
            except (ValueError, TypeError, IndexError):
                continue

    conn.commit()
    conn.close()
    return {'status': 'ok', 'imported': imported, 'month': month_str}


# Visa category (ענף) to our category mapping
VISA_CATEGORY_MAP = {
    'מזון ומשקאות': 'food',
    'מזון מהיר': 'food',
    'מסעדות': 'food',
    'סופרמרקט': 'food',
    'אופנה': 'personal',
    'הלבשה': 'personal',
    'פנאי בילוי': 'entertainment',
    'פנאי ובילוי': 'entertainment',
    'אנרגיה': 'vehicle',
    'דלק': 'vehicle',
    'רכב ותחבורה': 'vehicle',
    'חינוך': 'children',
    'תקשורת ומחשבים': 'communication',
    'תקשורת': 'communication',
    'בריאות': 'medical',
    'קוסמטיקה': 'health_beauty',
    'ביטוח': 'insurance',
    'ביטוח ופיננסים': 'insurance',
    'ממשלה ומוניציפלי': 'housing',
    'מוסדות': 'housing',
    'שירותים': 'misc',
    'שונות': 'misc',
    'ציוד ומשרד': 'personal',
    'ריהוט ובית': 'housing',
    'ריהוט': 'housing',
    'בית וגן': 'housing',
    'ספורט': 'entertainment',
    'נסיעות ותיירות': 'entertainment',
    'תיירות': 'entertainment',
    'מלונאות ואירוח': 'entertainment',
    'טיפוח ויופי': 'health_beauty',
    'רפואה ובריאות': 'medical',
    'ילדים': 'children',
}


def parse_visa_xlsx(filepath, user_id=None):
    """Parse Visa credit card XLSX export and import expenses."""
    import re
    wb = openpyxl.load_workbook(filepath)
    sheet = wb.active
    conn = get_db()

    imported = 0
    skipped = 0

    # Detect card number from row 1 or filename
    card_label = ''
    row1 = str(sheet.cell(1, 1).value or '')
    card_match = re.search(r'(\d{4})\s*$', row1)
    if card_match:
        card_label = 'ויזה ' + card_match.group(1)
    else:
        fname_match = re.search(r'(\d{4})', os.path.basename(filepath))
        if fname_match:
            card_label = 'ויזה ' + fname_match.group(1)

    # Find the header row (contains 'תאריך')
    data_start = None
    for row in range(1, min(sheet.max_row + 1, 10)):
        val = str(sheet.cell(row, 1).value or '')
        if 'תאריך' in val:
            data_start = row + 1
            break

    # If no header found, data likely starts at row 5 or 6
    if data_start is None:
        for row in range(4, 8):
            val = sheet.cell(row, 1).value
            if val and hasattr(val, 'strftime'):
                data_start = row
                break

    if data_start is None:
        data_start = 6

    for row in range(data_start, sheet.max_row + 1):
        date_val = sheet.cell(row, 1).value
        business = sheet.cell(row, 2).value
        amount_val = sheet.cell(row, 4).value or sheet.cell(row, 3).value  # prefer charge amount
        visa_category = str(sheet.cell(row, 6).value or '').strip()

        if not date_val or not amount_val:
            continue

        # Parse date
        if hasattr(date_val, 'strftime'):
            expense_date = date_val.strftime('%Y-%m-%d')
        else:
            continue

        # Parse amount
        try:
            amount = float(str(amount_val).replace(',', ''))
            if amount <= 0:
                continue
        except (ValueError, TypeError):
            continue

        # Map visa category to our category
        category_id = 'misc'
        for visa_key, cat_id in VISA_CATEGORY_MAP.items():
            if visa_key in visa_category:
                category_id = cat_id
                break

        description = str(business or '').strip()
        subcategory = visa_category

        # If still misc, try to reclassify by description
        if category_id == 'misc' and description:
            for pattern, cat_id, subcat in VISA_DESCRIPTION_MAP:
                if pattern in description:
                    category_id = cat_id
                    subcategory = subcat
                    break

        # Apply user's saved category/frequency rules
        category_id, freq = apply_category_rule(conn, description, category_id, user_id=user_id)

        # Skip duplicates (same date, description, amount, user)
        dup = conn.execute(
            "SELECT COUNT(*) FROM expenses WHERE date=? AND description=? AND amount=? AND user_id=?",
            (expense_date, description, amount, user_id)
        ).fetchone()[0]
        if dup:
            skipped += 1
            continue

        conn.execute(
            """INSERT INTO expenses (date, category_id, subcategory, description, amount, source, card, frequency, user_id)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (expense_date, category_id, subcategory, description, amount, 'visa_import', card_label, freq, user_id)
        )
        imported += 1

    conn.commit()
    # Auto-link imported transactions to assets/liabilities
    link_result = run_linking_engine(user_id)
    inst_match = run_installment_matching(user_id)
    conn.close()
    return {'status': 'ok', 'imported': imported, 'skipped_duplicates': skipped,
            'auto_linked': link_result.get('auto_linked', 0),
            'link_suggestions': link_result.get('suggested', 0),
            'inst_matched': inst_match.get('auto_matched', 0), 'source': 'visa'}


# ---- Bank CSV Import ----

# Patterns to SKIP — credit card summary charges in bank statement
# (individual transactions come from the card's own XLSX/statement)
BANK_SKIP_PATTERNS = [
    # Israeli cards
    'חיוב לכרטיס ויזה',
    'דיינרס קלו',
    'ישראכרט חיוב',
    'ישראכרט גביה',
    'זיכוי הנחות מפתח מכרטיס',
    'החזרת חיוב דיסקונט למשכנתאות',  # mortgage charge-back (paired with charge)
    # US/international cards
    'VISA PAYMENT',
    'MASTERCARD PAYMENT',
    'AMEX PAYMENT',
    'AMERICAN EXPRESS PAYMENT',
    'DISCOVER PAYMENT',
    'CREDIT CARD PAYMENT',
    'CARD PAYMENT',
    'PAYMENT - THANK YOU',
]

# Income patterns: (pattern_in_description, person, source, is_recurring)
BANK_INCOME_PATTERNS = [
    ('בנק פועלים משכורת', 'husband', 'salary', 1),
    ('בנק לאומי משכורת', 'wife', 'salary', 1),
    ('ביטוח לאומי - ילדים', 'family', 'child_allowance', 1),
    ('בטוח לאומי', 'family', 'child_allowance', 0),
    ('מס-הכנסה ה', 'family', 'other', 0),
    ('העברה מגל נעמי', 'wife', 'other', 0),
    ('העברה מליגת כדורסל', 'husband', 'other', 0),
    ('העברה מר.ס.ג.ר', 'family', 'other', 0),
    ('קבלת תשלום על יתרת זכות', 'family', 'other', 0),
    ('החזרת חיוב ישראכרט', 'family', 'other', 0),
    ('החזרת חיוב הראל', 'family', 'other', 0),
    ('מור גמל ופ', 'husband', 'other', 0),  # positive = withdrawal from provident
]

# Expense patterns: (pattern, category_id, subcategory, frequency)
# frequency: 'monthly', 'bimonthly', 'random'
BANK_EXPENSE_PATTERNS = [
    ('דסק-משכנתא', 'housing', 'משכנתא', 'monthly'),
    ('לועד מקומי', 'housing', 'ועד מקומי', 'monthly'),
    ('לגל נעמי', 'housing', 'העברה קבועה', 'monthly'),
    ('מ.א. חוף ה', 'housing', 'ארנונה', 'bimonthly'),
    ('הראל פנסיה', 'savings', 'פנסיה הראל', 'monthly'),
    ('הראל בטוח', 'insurance', 'ביטוח הראל', 'monthly'),
    ('הראלהלואה', 'savings', 'הלוואה הראל', 'monthly'),
    ('כלל חיים/ב', 'insurance', 'ביטוח חיים כלל', 'monthly'),
    ('חיובי הלוו', 'savings', 'הלוואה', 'monthly'),
    # דיינרס/ישראכרט charges are now in BANK_SKIP_PATTERNS (skipped)
    ('השתלמות אג', 'savings', 'קרן השתלמות', 'monthly'),
    ('מור גמל ופ חיוב', 'savings', 'גמל מור', 'monthly'),
    ('ריבית על משיכת יתר', 'misc', 'ריבית מינוס', 'monthly'),
    ('תשלום מס במקור', 'misc', 'מס במקור', 'monthly'),
    ('כספומט', 'misc', 'משיכת מזומן', 'random'),
    ('מש\' מכספומט', 'misc', 'משיכת מזומן', 'random'),
    ('משיכה מכספומט', 'misc', 'משיכת מזומן', 'random'),
    ('משיכת שיק', 'misc', 'שיק', 'random'),
    ('עמלת פנקסי', 'misc', 'עמלת בנק', 'random'),
    ('עמלת החזר', 'misc', 'עמלת בנק', 'random'),
    ('העברה לגל', 'housing', 'העברה משפחתית', 'random'),
    ('הע. ל', 'misc', 'העברה', 'random'),
    ('העברה ל', 'misc', 'העברה', 'random'),
    ('דמי כרטיס', 'misc', 'עמלת בנק', 'monthly'),
]

# Extra description-based classification for Visa "שונות" (misc) entries
VISA_DESCRIPTION_MAP = [
    # Subscriptions (monthly digital services)
    ('NETFLIX', 'subscriptions', 'נטפליקס'),
    ('Netflix', 'subscriptions', 'נטפליקס'),
    ('SPOTIFY', 'subscriptions', 'ספוטיפיי'),
    ('Spotify', 'subscriptions', 'ספוטיפיי'),
    ('APPLE.COM', 'subscriptions', 'אפל'),
    ('GOOGLE *', 'subscriptions', 'גוגל'),
    ('DISNEY+', 'subscriptions', 'דיסני+'),
    ('YOUTUBE', 'subscriptions', 'יוטיוב'),
    ('HBO', 'subscriptions', 'HBO'),
    ('AMAZON PRIME', 'subscriptions', 'אמזון פריים'),
    ('YES ', 'subscriptions', 'yes'),
    ('HOT ', 'subscriptions', 'הוט'),
    ('PARTNER', 'communication', 'פרטנר'),
    # Dining out
    ('מסעד', 'dining_out', 'מסעדה'),
    ('מזון מהיר', 'dining_out', 'מזון מהיר'),
    ('פיצ', 'dining_out', 'פיצה'),
    ('קפה', 'dining_out', 'בית קפה'),
    ('WOLT', 'dining_out', 'וולט'),
    ('JAPANIKA', 'dining_out', 'מסעדה'),
    ('מקדונלד', 'dining_out', 'מזון מהיר'),
    ('בורגר', 'dining_out', 'מזון מהיר'),
    ('DOMINOS', 'dining_out', 'פיצה'),
    # Clothing
    ('H&M', 'clothing', 'ביגוד'),
    ('ZARA', 'clothing', 'ביגוד'),
    ('FOX', 'clothing', 'ביגוד'),
    ('SHEIN', 'clothing', 'ביגוד'),
    ('GOLF', 'clothing', 'ביגוד'),
    ('CASTRO', 'clothing', 'ביגוד'),
    ('רנואר', 'clothing', 'ביגוד'),
    ('נעלי', 'clothing', 'נעליים'),
    ('NIKE', 'clothing', 'ביגוד ספורט'),
    ('ADIDAS', 'clothing', 'ביגוד ספורט'),
    # Education
    ('אוניברסיט', 'education', 'לימודים'),
    ('מכללת', 'education', 'לימודים'),
    ('קורס', 'education', 'קורס'),
    ('UDEMY', 'education', 'קורס אונליין'),
    ('COURSERA', 'education', 'קורס אונליין'),
    ('סטימצקי', 'education', 'ספרים'),
    # Gifts
    ('ספלנדו', 'gifts', 'מתנות'),
    ('מתנות', 'gifts', 'מתנות'),
    ('פרחי', 'gifts', 'פרחים'),
    # Vehicle
    ('חניון', 'vehicle', 'חניה'),
    ('חניה', 'vehicle', 'חניה'),
    ('חצב חניון', 'vehicle', 'חניה'),
    ('מנהרות', 'vehicle', 'מנהרות'),
    ('פנגו', 'vehicle', 'חניה'),
    ('מוביט', 'vehicle', 'תחבורה'),
    # Housing
    ('מי חוף הכרמל', 'housing', 'מים'),
    ('מועצה אזורית', 'housing', 'מועצה אזורית'),
    ('המועצה האזורית', 'housing', 'מועצה אזורית'),
    ('משקי רם', 'housing', 'משקי רם'),
    # Misc / transfers
    ('BIT', 'misc', 'העברות BIT'),
    ('PAYBOX', 'misc', 'העברות PAYBOX'),
    ('העברה ב BIT', 'misc', 'העברות BIT'),
    ('דמי כרטיס', 'misc', 'עמלת בנק'),
    ('PAYPAL', 'personal', 'PayPal'),
    ('AMAZON', 'personal', 'אמזון'),
    # Children
    ('צומת ספרים', 'children', 'ספרים'),
    ('סיגטס', 'children', 'ילדים'),
    # Entertainment
    ('כפר נופש', 'entertainment', 'חופשה'),
    ('אקספו', 'entertainment', 'אירועים'),
]


def parse_bank_csv(filepath, user_id=None):
    """Parse Israeli bank CSV (עובר ושב) and import expenses + income."""
    # Try different encodings
    content = None
    for enc in ('utf-8-sig', 'utf-8', 'windows-1255', 'iso-8859-8'):
        try:
            with open(filepath, 'r', encoding=enc) as f:
                content = f.read()
            break
        except (UnicodeDecodeError, UnicodeError):
            continue

    if not content:
        return {'error': 'Could not decode CSV file'}

    reader = csv.reader(io.StringIO(content))
    header = next(reader)  # skip header row

    conn = get_db()
    imported_expenses = 0
    imported_income = 0
    skipped_visa = 0
    skipped_dup = 0
    skipped_other = 0
    month_balances = {}  # {month_str: (balance_float, date_str)}

    for row in reader:
        if len(row) < 5:
            continue

        date_str = row[0].strip().strip('"')
        description = row[2].strip().strip('"')
        amount_str = row[3].strip().strip('"').replace(',', '')

        if not date_str or not amount_str:
            continue

        # Parse date (DD/MM/YY format)
        try:
            parts = date_str.split('/')
            day, month_num, year = int(parts[0]), int(parts[1]), int(parts[2])
            if year < 100:
                year += 2000
            expense_date = f"{year}-{month_num:02d}-{day:02d}"
        except (ValueError, IndexError):
            continue

        # Parse amount
        try:
            amount = float(amount_str)
        except ValueError:
            continue

        # Track balance from column 4 (running account balance)
        balance_str = row[4].strip().strip('"').replace(',', '') if len(row) > 4 else ''
        if balance_str:
            try:
                row_balance = float(balance_str)
                row_month = f"{year}-{month_num:02d}"
                # Keep the balance from the LATEST date in each month
                if row_month not in month_balances or expense_date >= month_balances[row_month][1]:
                    month_balances[row_month] = (row_balance, expense_date)
            except ValueError:
                pass

        # Skip zero amounts
        if amount == 0:
            continue

        # Always skip visa/card charge lines — they are summary totals,
        # the individual transactions come from the visa XLSX import
        should_skip = False
        for pattern in BANK_SKIP_PATTERNS:
            if pattern in description:
                should_skip = True
                skipped_visa += 1
                break

        if should_skip:
            continue

        # POSITIVE amounts = income
        if amount > 0:
            person = 'family'
            source = 'other'
            is_recurring = 0

            for pattern, p, s, rec in BANK_INCOME_PATTERNS:
                if pattern in description:
                    person = p
                    source = s
                    is_recurring = rec
                    break

            # Skip duplicates (same date, description, amount, user)
            dup = conn.execute(
                "SELECT COUNT(*) FROM income WHERE date=? AND description=? AND amount=? AND user_id=?",
                (expense_date, description, amount, user_id)
            ).fetchone()[0]
            if dup:
                skipped_dup += 1
                continue

            conn.execute(
                """INSERT INTO income (date, person, source, amount, description, is_recurring, user_id)
                   VALUES (?, ?, ?, ?, ?, ?, ?)""",
                (expense_date, person, source, amount, description, is_recurring, user_id)
            )
            imported_income += 1

        # NEGATIVE amounts = expenses
        else:
            abs_amount = abs(amount)
            category_id = 'misc'
            subcategory = ''
            frequency = 'random'

            matched = False
            for pattern, cat, subcat, freq in BANK_EXPENSE_PATTERNS:
                if pattern in description:
                    category_id = cat
                    subcategory = subcat
                    frequency = freq
                    matched = True
                    break

            if not matched:
                subcategory = description
                skipped_other += 1

            # Apply user's saved category/frequency rules
            category_id, frequency = apply_category_rule(conn, description, category_id, frequency, user_id=user_id)

            # Skip duplicates (same date, description, amount, user)
            dup = conn.execute(
                "SELECT COUNT(*) FROM expenses WHERE date=? AND description=? AND amount=? AND user_id=?",
                (expense_date, description, abs_amount, user_id)
            ).fetchone()[0]
            if dup:
                skipped_dup += 1
                continue

            conn.execute(
                """INSERT INTO expenses (date, category_id, subcategory, description, amount, source, frequency, card, user_id)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (expense_date, category_id, subcategory, description, abs_amount, 'bank_csv', frequency, 'בנק דיסקונט', user_id)
            )
            imported_expenses += 1

    # Store monthly closing balances
    balances_saved = 0
    for month_str, (balance, last_date) in month_balances.items():
        conn.execute("""INSERT OR REPLACE INTO bank_balances
            (user_id, account_name, month, closing_balance, last_transaction_date)
            VALUES (?, 'main', ?, ?, ?)""", (user_id, month_str, balance, last_date))
        balances_saved += 1

    conn.commit()
    # Auto-link imported transactions to assets/liabilities
    link_result = run_linking_engine(user_id)
    inst_match = run_installment_matching(user_id)
    conn.close()
    return {
        'status': 'ok',
        'imported_expenses': imported_expenses,
        'imported_income': imported_income,
        'skipped_visa': skipped_visa,
        'skipped_duplicates': skipped_dup,
        'balances_saved': balances_saved,
        'auto_linked': link_result.get('auto_linked', 0),
        'link_suggestions': link_result.get('suggested', 0),
        'inst_matched': inst_match.get('auto_matched', 0),
        'source': 'bank_csv'
    }


# ---- Payslip PDF Import ----

def _extract_payslip_month(text):
    """Extract month string 'YYYY-MM' from payslip text, or None if not found."""
    # Pattern: MM/YYYY
    m = re.search(r'(\d{1,2})/(\d{4})', text)
    if m:
        mm, yyyy = int(m.group(1)), int(m.group(2))
        if 1 <= mm <= 12 and 2000 <= yyyy <= 2099:
            return f"{yyyy}-{mm:02d}"
    # Pattern: YYYY-MM
    m = re.search(r'(20\d{2})-(0[1-9]|1[0-2])', text)
    if m:
        return f"{m.group(1)}-{m.group(2)}"
    return None


def _extract_company_name(text):
    """Extract company name from payslip (usually near top)."""
    m = re.search(r'([\u0590-\u05FF\s]+(?:בע"מ|בע״מ))', text)
    if m:
        return m.group(1).strip()
    m = re.search(r'([\w\s]+(?:Ltd\.?|LTD\.?))', text)
    if m:
        return m.group(1).strip()[:60]
    return ''


_PAYSLIP_KEY_FIELDS = ['gross_salary', 'net_salary', 'income_tax', 'social_security',
                        'pension_employer', 'education_fund_employer']


def _extract_payslip_fields(text):
    """Extract numeric salary fields from Hebrew payslip text using regex."""
    result = {}
    patterns = {
        'gross_salary':     [r'(?:שכר\s*ברוטו|ברוטו\s*כולל|סה["\u05F4]?כ.*?ברוטו)\s*[:\-]?\s*([\d,]+\.?\d*)'],
        'net_salary':       [r'(?:שכר\s*נטו|נטו\s*לתשלום|סה["\u05F4]?כ.*?נטו)\s*[:\-]?\s*([\d,]+\.?\d*)'],
        'income_tax':       [r'(?:מס\s*הכנסה)\s*[:\-]?\s*([\d,]+\.?\d*)'],
        'social_security':  [r'(?:ביטוח\s*לאומי)\s*[:\-]?\s*([\d,]+\.?\d*)'],
        'health_insurance': [r'(?:ביטוח\s*בריאות|מס\s*בריאות)\s*[:\-]?\s*([\d,]+\.?\d*)'],
        'pension_employee': [r'(?:פנסי[הת]\s*עובד|הפרשת?\s*עובד.*?פנסי)\s*[:\-]?\s*([\d,]+\.?\d*)'],
        'pension_employer': [r'(?:פנסי[הת]\s*מעביד|הפרשת?\s*מעביד.*?פנסי)\s*[:\-]?\s*([\d,]+\.?\d*)'],
        'education_fund_employee': [r'(?:(?:קרן\s*)?השתלמות\s*עובד|הפרשת?\s*עובד.*?השתלמות)\s*[:\-]?\s*([\d,]+\.?\d*)'],
        'education_fund_employer': [r'(?:(?:קרן\s*)?השתלמות\s*מעביד|הפרשת?\s*מעביד.*?השתלמות)\s*[:\-]?\s*([\d,]+\.?\d*)'],
        'severance_employer': [r'(?:פיצויי?\s*(?:פיטורי[םן])?|הפרשת?\s*פיצויי?ם?)\s*[:\-]?\s*([\d,]+\.?\d*)'],
        'vacation_days':    [r'(?:ימי\s*חופש[הה]?|יתרת?\s*חופש[הה]?)\s*[:\-]?\s*([\d]+\.?\d*)'],
        'sick_days':        [r'(?:ימי\s*מחלה|יתרת?\s*מחלה)\s*[:\-]?\s*([\d]+\.?\d*)'],
    }
    for field, field_patterns in patterns.items():
        for pat in field_patterns:
            m = re.search(pat, text)
            if m:
                val_str = m.group(1).replace(',', '')
                try:
                    result[field] = float(val_str)
                except ValueError:
                    pass
                break
    return result


def _detect_bonus(text):
    """Detect bonus amounts from payslip text."""
    m = re.search(r'(?:בונוס|פרמי[הה]|מענק|תשלום\s*חד[- ]?פעמי)\s*[:\-]?\s*([\d,]+\.?\d*)', text)
    if m:
        try:
            return float(m.group(1).replace(',', ''))
        except ValueError:
            pass
    return 0


def _compute_extraction_confidence(fields):
    """0.0-1.0 based on how many key fields were found."""
    found = sum(1 for f in _PAYSLIP_KEY_FIELDS if fields.get(f, 0) > 0)
    return round(found / len(_PAYSLIP_KEY_FIELDS), 2)


def _detect_person_from_payslip(conn, uid):
    """Try to match payslip to an existing salary person, else default 'husband'."""
    existing = conn.execute(
        "SELECT DISTINCT person FROM salary_statements WHERE user_id=?", (uid,)
    ).fetchall()
    if len(existing) == 1:
        return existing[0]['person']
    inc_persons = conn.execute(
        "SELECT DISTINCT person FROM income WHERE user_id=? AND source='salary'", (uid,)
    ).fetchall()
    if len(inc_persons) == 1:
        return inc_persons[0]['person']
    return 'husband'


def parse_payslip_pdf(filepath, user_id=None):
    """Parse an Israeli payslip PDF and return extracted fields as preview (no DB write).
    Returns dict with source='payslip_pdf_preview' for frontend review modal."""
    import fitz  # lazy import — only loaded when PDF is actually uploaded

    filename = os.path.basename(filepath)

    # Password strategy: try first part of filename before '_', then no password
    password = None
    if '_' in filename:
        password = filename.split('_')[0]

    doc = None
    try:
        doc = fitz.open(filepath)
        if doc.is_encrypted:
            decrypted = False
            if password:
                decrypted = doc.authenticate(password)
            if not decrypted:
                decrypted = doc.authenticate('')
            if not decrypted:
                doc.close()
                raise ValueError('Could not decrypt PDF. Expected password = first part of filename before underscore.')
    except ValueError:
        raise
    except Exception as e:
        if doc:
            doc.close()
        raise ValueError(f'PDF open failed: {str(e)}')

    try:
        full_text = ''
        for page in doc:
            full_text += page.get_text() + '\n'
    finally:
        doc.close()

    month_str = _extract_payslip_month(full_text)  # None if not found
    company = _extract_company_name(full_text)
    fields = _extract_payslip_fields(full_text)
    fields['bonus_amount'] = _detect_bonus(full_text)
    confidence = _compute_extraction_confidence(fields)
    missing = [f for f in _PAYSLIP_KEY_FIELDS if f not in fields]

    conn = get_db()
    uid = user_id or get_uid()
    person = _detect_person_from_payslip(conn, uid)
    conn.close()

    employer_total = (fields.get('pension_employer', 0) +
                      fields.get('education_fund_employer', 0) +
                      fields.get('severance_employer', 0))
    true_comp = fields.get('gross_salary', 0) + employer_total

    return {
        'source': 'payslip_pdf_preview',
        'month': month_str,                       # None if not detected
        'person': person,
        'company_name': company,
        'gross_salary': fields.get('gross_salary', 0),
        'net_salary': fields.get('net_salary', 0),
        'income_tax': fields.get('income_tax', 0),
        'social_security': fields.get('social_security', 0),
        'health_insurance': fields.get('health_insurance', 0),
        'pension_employee': fields.get('pension_employee', 0),
        'pension_employer': fields.get('pension_employer', 0),
        'education_fund_employee': fields.get('education_fund_employee', 0),
        'education_fund_employer': fields.get('education_fund_employer', 0),
        'severance_employer': fields.get('severance_employer', 0),
        'other_deductions': fields.get('other_deductions', 0),
        'bonus_amount': fields.get('bonus_amount', 0),
        'vacation_days': fields.get('vacation_days', 0),
        'sick_days': fields.get('sick_days', 0),
        'extraction_confidence': confidence,
        'missing_fields': missing,
        'employer_contributions': round(employer_total, 2),
        'true_compensation': round(true_comp, 2),
        'raw_text': full_text,
        'source_filename': filename,
    }


# ---- Insurance Portfolio Import (Har HaBituach / Masbiron) ----

# Map Hebrew branch names to insurance subtypes
INSURANCE_BRANCH_MAP = {
    'ביטוח בריאות': 'health',
    'ביטוח סיעודי': 'health',
    'כתב שירות': 'health',
    'שרותי בריאות נוספים': 'health',
    'ביטוח רכב': 'car',
    'ביטוח דירה': 'home',
    'ביטוח חיים': 'life',
    'אבדן כושר עבודה': 'disability',
    'ביטוח משכנתא': 'home',
    'ביטוח נסיעות': 'general',
    'ביטוח עסק': 'general',
    'ביטוח תאונות אישיות': 'general',
}


def _is_insurance_portfolio_xlsx(filepath):
    """Quick check if an XLSX file is an insurance portfolio from Har HaBituach / Masbiron."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, read_only=True)
        ws = wb.active
        # Check sheet name
        if 'תיק ביטוחי' in (ws.title or ''):
            wb.close()
            return True
        # Check first 5 rows for insurance portfolio markers
        for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
            text = ' '.join(str(v) for v in row if v)
            if 'הר הביטוח' in text or 'מסביר ביטוח' in text or 'התיק הביטוחי' in text or 'כיסויי השב' in text:
                wb.close()
                return True
        wb.close()
    except Exception:
        pass
    return False


def parse_insurance_portfolio_xlsx(filepath, user_id):
    """Parse insurance portfolio XLSX from Har HaBituach / Masbiron portals.
    Handles two formats:
    - HitResults: health fund supplementary (שב"ן) + health insurance policies
    - HbResults: general insurance (car, apartment, etc.)
    Returns dict with imported count and details."""
    import openpyxl
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    conn = get_db()
    imported = 0
    skipped = 0
    products = []

    # Read all rows as plain values
    rows = []
    for r in range(1, ws.max_row + 1):
        row = []
        for c in range(1, ws.max_column + 1):
            row.append(ws.cell(row=r, column=c).value)
        rows.append(row)

    # Detect file type by scanning header rows
    sections = []  # list of (header_row_idx, col_map)

    for i, row in enumerate(rows):
        # Find header rows by looking for known column names
        row_strs = [str(v).strip() if v else '' for v in row]
        row_text = ' '.join(row_strs)

        if 'ענף ראשי' in row_text and 'פרמיה' in row_text:
            # Full insurance format (HbResults or HitResults section 2)
            col_map = {}
            for j, val in enumerate(row_strs):
                if val == 'ענף ראשי':
                    col_map['main_branch'] = j
                elif val == 'ענף (משני)':
                    col_map['sub_branch'] = j
                elif val == 'סוג מוצר':
                    col_map['product_type'] = j
                elif val == 'חברה':
                    col_map['company'] = j
                elif val.startswith('תקופת'):
                    col_map['period'] = j
                elif val.startswith('פרמיה'):
                    col_map['premium'] = j
                elif val == 'סוג פרמיה':
                    col_map['premium_type'] = j
                elif val == 'מספר פוליסה':
                    col_map['policy_number'] = j
                elif val.startswith('פרטים'):
                    col_map['details'] = j
            sections.append((i, col_map, 'full'))

        elif 'ענף (משני)' in row_text and 'פרמיה' in row_text and 'ענף ראשי' not in row_text:
            # Health fund supplementary format (HitResults section 1 - שב"ן)
            col_map = {}
            for j, val in enumerate(row_strs):
                if val == 'ענף (משני)':
                    col_map['sub_branch'] = j
                elif val == 'סוג מוצר':
                    col_map['product_type'] = j
                elif val == 'חברה':
                    col_map['company'] = j
                elif val.startswith('תקופת'):
                    col_map['period'] = j
                elif val.startswith('פרמיה'):
                    col_map['premium'] = j
                elif val == 'סוג פרמיה':
                    col_map['premium_type'] = j
            sections.append((i, col_map, 'shaban'))

    if not sections:
        conn.close()
        return {'error': 'unrecognized_format', 'imported': 0}

    # Get existing policies to avoid duplicates
    existing_policies = set()
    for r in conn.execute("SELECT policy_number FROM financial_products WHERE user_id=? AND type='insurance' AND policy_number != ''",
                          (user_id,)).fetchall():
        existing_policies.add(str(r[0]).strip())

    # First pass: collect all rows, then aggregate by policy number
    raw_entries = []

    for sec_idx, (header_idx, col_map, sec_type) in enumerate(sections):
        # Determine end of section: next header or end of file
        next_header = None
        if sec_idx + 1 < len(sections):
            # Look for empty rows or next section header before next section
            next_header = sections[sec_idx + 1][0]

        for i in range(header_idx + 1, len(rows)):
            if next_header and i >= next_header - 2:
                break

            row = rows[i]
            if not row or not any(row):
                continue

            row_strs = [str(v).strip() if v else '' for v in row]

            # Skip section header rows (like "תחום - כללי")
            if row_strs[0] == 'None' or row_strs[0] == '':
                if any('תחום' in s for s in row_strs):
                    continue
                continue

            # Skip rows that don't start with an ID number
            id_val = row_strs[0]
            if not id_val.replace('-', '').isdigit():
                continue

            # Extract fields based on section type
            if sec_type == 'full':
                main_branch = row_strs[col_map.get('main_branch', 1)] if 'main_branch' in col_map else ''
                sub_branch = row_strs[col_map.get('sub_branch', 2)] if 'sub_branch' in col_map else ''
                product_type = row_strs[col_map.get('product_type', 3)] if 'product_type' in col_map else ''
                company = row_strs[col_map.get('company', 4)] if 'company' in col_map else ''
                period = row_strs[col_map.get('period', 5)] if 'period' in col_map else ''
                premium_raw = row[col_map.get('premium', 6)] if 'premium' in col_map else 0
                premium_type = row_strs[col_map.get('premium_type', 7)] if 'premium_type' in col_map else ''
                policy_num = row_strs[col_map.get('policy_number', 8)] if 'policy_number' in col_map else ''
            else:  # shaban
                main_branch = 'שרותי בריאות נוספים'
                sub_branch = row_strs[col_map.get('sub_branch', 1)] if 'sub_branch' in col_map else ''
                product_type = row_strs[col_map.get('product_type', 2)] if 'product_type' in col_map else ''
                company = row_strs[col_map.get('company', 3)] if 'company' in col_map else ''
                period = row_strs[col_map.get('period', 4)] if 'period' in col_map else ''
                premium_raw = row[col_map.get('premium', 5)] if 'premium' in col_map else 0
                premium_type = row_strs[col_map.get('premium_type', 6)] if 'premium_type' in col_map else ''
                policy_num = ''

            # Parse premium amount
            try:
                premium = float(premium_raw) if premium_raw else 0
            except (ValueError, TypeError):
                try:
                    premium = float(str(premium_raw).replace(',', ''))
                except (ValueError, TypeError):
                    premium = 0

            # Skip zero-cost items
            if premium == 0:
                skipped += 1
                continue

            # Convert yearly to monthly
            if 'שנתית' in premium_type:
                monthly_cost = round(premium / 12, 2)
            else:
                monthly_cost = round(premium, 2)

            # Map to insurance subtype
            subtype = 'general'
            for branch_key, branch_subtype in INSURANCE_BRANCH_MAP.items():
                if branch_key in main_branch:
                    subtype = branch_subtype
                    break

            # Build product name from sub_branch
            name = sub_branch.strip().replace('\n', ' ') if sub_branch else main_branch
            company_clean = company.replace('בע"מ', '').replace('חברה לביטוח', '').strip()

            # Parse renewal date from period
            renewal_date = ''
            start_date = ''
            if period and period != 'מתחדש' and '-' in period:
                parts = period.split('-')
                if len(parts) == 2:
                    try:
                        end_str = parts[1].strip()
                        # Try MM/DD/YYYY or DD/MM/YYYY format
                        for fmt in ('%m/%d/%Y', '%d/%m/%Y'):
                            try:
                                end_dt = datetime.strptime(end_str, fmt)
                                renewal_date = end_dt.strftime('%Y-%m-%d')
                                break
                            except ValueError:
                                continue
                        start_str = parts[0].strip()
                        for fmt in ('%m/%d/%Y', '%d/%m/%Y'):
                            try:
                                start_dt = datetime.strptime(start_str, fmt)
                                start_date = start_dt.strftime('%Y-%m-%d')
                                break
                            except ValueError:
                                continue
                    except Exception:
                        pass

            raw_entries.append({
                'main_branch': main_branch,
                'sub_branch': sub_branch,
                'product_type': product_type,
                'company': company,
                'company_clean': company_clean,
                'period': period,
                'premium': premium,
                'premium_type': premium_type,
                'monthly_cost': monthly_cost,
                'policy_num': policy_num,
                'subtype': subtype,
                'name': name,
                'renewal_date': renewal_date,
                'start_date': start_date,
                'notes_parts': [product_type] if product_type and product_type not in ('None', '') else [],
                'yearly_note': f'פרמיה שנתית: ₪{premium:,.0f}' if 'שנתית' in premium_type else '',
            })

    # Aggregate entries by policy number (multiple coverages under same policy)
    aggregated = {}
    no_policy = []
    for entry in raw_entries:
        pn = entry['policy_num']
        if not pn:
            no_policy.append(entry)
            continue
        if pn in aggregated:
            agg = aggregated[pn]
            agg['monthly_cost'] += entry['monthly_cost']
            agg['coverage_names'].append(entry['name'])
            if entry['notes_parts']:
                agg['all_notes'].extend(entry['notes_parts'])
            if entry['yearly_note']:
                agg['yearly_premium'] += entry['premium']
        else:
            aggregated[pn] = {
                **entry,
                'coverage_names': [entry['name']],
                'all_notes': list(entry['notes_parts']),
                'yearly_premium': entry['premium'] if entry['yearly_note'] else 0,
            }

    # Process aggregated entries + no-policy entries
    all_entries = list(aggregated.values()) + no_policy
    for entry in all_entries:
        policy_num = entry['policy_num']
        monthly_cost = round(entry['monthly_cost'], 2)

        if monthly_cost == 0:
            skipped += 1
            continue

        if policy_num and policy_num in existing_policies:
            skipped += 1
            continue

        # Build name: if multiple coverages, combine them
        coverage_names = entry.get('coverage_names', [entry['name']])
        if len(coverage_names) > 1:
            name = f"{entry['main_branch']} ({len(coverage_names)} כיסויים)"
        else:
            name = coverage_names[0]

        # Build notes
        notes_parts = entry.get('all_notes', entry.get('notes_parts', []))
        if len(coverage_names) > 1:
            notes_parts = [n.strip().replace('\n', ' ') for n in coverage_names]
        yearly_prem = entry.get('yearly_premium', 0)
        if yearly_prem:
            notes_parts.append(f'פרמיה שנתית: ₪{yearly_prem:,.0f}')
        notes = ' | '.join([n for n in notes_parts if n and n not in ('None', '')])

        company_clean = entry['company_clean']
        expense_pattern = company_clean.split()[0] if company_clean else ''

        conn.execute("""
            INSERT INTO financial_products (type, subtype, company, name, monthly_cost, policy_number,
                renewal_date, start_date, notes, expense_pattern, status, user_id)
            VALUES ('insurance', ?, ?, ?, ?, ?, ?, ?, ?, ?, 'active', ?)
        """, (entry['subtype'], company_clean, name, monthly_cost, policy_num,
              entry['renewal_date'], entry['start_date'], notes, expense_pattern, user_id))

        if policy_num:
            existing_policies.add(policy_num)

        imported += 1
        products.append({
            'name': name,
            'company': company_clean,
            'subtype': entry['subtype'],
            'monthly_cost': monthly_cost,
            'policy_number': policy_num,
        })

    conn.commit()
    conn.close()
    return {
        'imported': imported,
        'skipped': skipped,
        'products': products,
        'source': 'insurance_portfolio'
    }


@app.route('/api/import-insurance', methods=['POST'])
@login_required
def import_insurance_portfolio():
    """Import insurance portfolio XLSX from Har HaBituach / Masbiron."""
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400

    file = request.files['file']
    if not file.filename:
        return jsonify({'error': 'No file selected'}), 400

    if not file.filename.endswith('.xlsx'):
        return jsonify({'error': 'Only XLSX files are supported'}), 400

    filepath = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
    file.save(filepath)

    try:
        uid = get_uid()
        result = parse_insurance_portfolio_xlsx(filepath, uid)
        if result.get('error') == 'unrecognized_format':
            return jsonify({'error': 'unrecognized_format'}), 400
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 400


# ---- Recurring Transaction Detection ----

@app.route('/api/detect-recurring', methods=['POST'])
@login_required
def detect_recurring():
    """Detect transactions that look recurring (same business, similar amount, 2+ months)."""
    conn = get_db()
    uid = get_uid()
    # Find descriptions that appear in 2+ distinct months with consistent amounts
    patterns = conn.execute("""
        SELECT description,
               COUNT(DISTINCT substr(date,1,7)) as months_count,
               ROUND(AVG(amount),2) as avg_amount,
               MIN(amount) as min_amount,
               MAX(amount) as max_amount,
               GROUP_CONCAT(DISTINCT substr(date,1,7)) as months,
               GROUP_CONCAT(id) as expense_ids
        FROM expenses
        WHERE user_id = ? AND frequency IN ('random', 'once')
          AND description != '' AND source IN ('bank_csv', 'visa_import')
        GROUP BY description
        HAVING months_count >= 2
           AND (max_amount - min_amount) <= avg_amount * 0.15
        ORDER BY months_count DESC, avg_amount DESC
    """, (uid,)).fetchall()

    result = []
    for p in patterns:
        months_list = sorted(p['months'].split(','))
        result.append({
            'description': p['description'],
            'avg_amount': p['avg_amount'],
            'months_count': p['months_count'],
            'months': months_list,
            'expense_ids': [int(x) for x in p['expense_ids'].split(',')],
        })
    conn.close()
    return jsonify(result)


@app.route('/api/expenses/set-recurring', methods=['POST'])
@login_required
def set_recurring():
    """Bulk-update frequency for a set of expenses matching a recurring pattern."""
    data = request.json
    description = data.get('description', '')
    frequency = data.get('frequency', 'monthly')  # monthly or bimonthly
    end_date = data.get('end_date')  # null = ongoing
    if not description:
        return jsonify({'error': 'Missing description'}), 400

    conn = get_db()
    uid = get_uid()
    conn.execute(
        "UPDATE expenses SET frequency = ? WHERE description = ? AND user_id = ?",
        (frequency, description, uid)
    )
    conn.commit()
    conn.close()
    return jsonify({'status': 'ok', 'description': description, 'frequency': frequency})


# ---- Smart Budget Tips v2 ----

def _build_tip_context(conn, uid, month):
    """Single-pass data gather for all tip generators."""
    ctx = {'uid': uid, 'month': month}

    # Excluded months + normalized month count
    excl_months = get_excluded_month_set(conn, uid)
    ctx['excl_months'] = excl_months
    months_list = [r[0] for r in conn.execute(
        "SELECT DISTINCT substr(date,1,7) FROM expenses WHERE user_id=? AND is_unusual=0 ORDER BY substr(date,1,7)", (uid,)
    ).fetchall() if r[0] not in excl_months]
    ctx['num_months'] = max(len(months_list), 1)

    # Current month totals
    ctx['exp_total'] = conn.execute(
        "SELECT COALESCE(SUM(amount),0) FROM expenses WHERE user_id=? AND date LIKE ?", (uid, month + '%')
    ).fetchone()[0]
    ctx['inc_total'] = conn.execute(
        "SELECT COALESCE(SUM(amount),0) FROM income WHERE user_id=? AND date LIKE ?", (uid, month + '%')
    ).fetchone()[0]

    # All-time income
    ctx['total_income_all'] = conn.execute(
        "SELECT COALESCE(SUM(amount),0) FROM income WHERE user_id=?", (uid,)
    ).fetchone()[0]

    # Expense aggregates (normalized — excludes unusual)
    ctx['overdraft_total'] = conn.execute(
        "SELECT COALESCE(SUM(amount),0) FROM expenses WHERE user_id=? AND is_unusual=0 AND subcategory='ריבית מינוס'", (uid,)
    ).fetchone()[0]
    ctx['cash_total'] = conn.execute(
        "SELECT COALESCE(SUM(amount),0) FROM expenses WHERE user_id=? AND is_unusual=0 AND subcategory='משיכת מזומן'", (uid,)
    ).fetchone()[0]
    ctx['food_total'] = conn.execute(
        "SELECT COALESCE(SUM(amount),0) FROM expenses WHERE user_id=? AND is_unusual=0 AND category_id='food'", (uid,)
    ).fetchone()[0]
    ctx['dining_out_total'] = conn.execute(
        "SELECT COALESCE(SUM(amount),0) FROM expenses WHERE user_id=? AND is_unusual=0 AND category_id='dining_out'", (uid,)
    ).fetchone()[0]
    # Combined food = groceries + dining for overall food spending tip
    ctx['food_combined_total'] = ctx['food_total'] + ctx['dining_out_total']
    ctx['food_monthly'] = ctx['food_combined_total'] / ctx['num_months']
    ctx['dining_out_monthly'] = ctx['dining_out_total'] / ctx['num_months']
    ctx['dining_out_this_month'] = conn.execute(
        "SELECT COALESCE(SUM(amount),0) FROM expenses WHERE user_id=? AND date LIKE ? AND category_id='dining_out'", (uid, month + '%')
    ).fetchone()[0]
    ctx['dining_out_vs_avg_pct'] = round((ctx['dining_out_this_month'] / ctx['dining_out_monthly'] - 1) * 100) if ctx['dining_out_monthly'] > 0 else 0
    ctx['dining_out_pct_of_income'] = round(ctx['dining_out_monthly'] / ctx['inc_total'] * 100, 1) if ctx['inc_total'] > 0 else 0
    # Restaurant total: dining_out category + legacy food subcategory matches
    ctx['restaurant_total'] = ctx['dining_out_total'] + conn.execute(
        "SELECT COALESCE(SUM(amount),0) FROM expenses WHERE user_id=? AND is_unusual=0 AND category_id='food' AND (subcategory LIKE '%מסעד%' OR subcategory LIKE '%מזון מהיר%')", (uid,)
    ).fetchone()[0]
    ctx['ent_total'] = conn.execute(
        "SELECT COALESCE(SUM(amount),0) FROM expenses WHERE user_id=? AND is_unusual=0 AND category_id='entertainment'", (uid,)
    ).fetchone()[0]
    ctx['ent_monthly'] = ctx['ent_total'] / ctx['num_months']
    ctx['insurance_total'] = conn.execute(
        "SELECT COALESCE(SUM(amount),0) FROM expenses WHERE user_id=? AND is_unusual=0 AND category_id='insurance'", (uid,)
    ).fetchone()[0]
    ctx['insurance_monthly'] = ctx['insurance_total'] / ctx['num_months']
    ctx['savings_total'] = conn.execute(
        "SELECT COALESCE(SUM(amount),0) FROM expenses WHERE user_id=? AND is_unusual=0 AND category_id='savings'", (uid,)
    ).fetchone()[0]
    ctx['savings_rate'] = (ctx['savings_total'] / ctx['total_income_all'] * 100) if ctx['total_income_all'] > 0 else 0

    # New category aggregates
    ctx['subscriptions_total'] = conn.execute(
        "SELECT COALESCE(SUM(amount),0) FROM expenses WHERE user_id=? AND is_unusual=0 AND category_id='subscriptions'", (uid,)
    ).fetchone()[0]
    ctx['subscriptions_monthly'] = ctx['subscriptions_total'] / ctx['num_months']
    ctx['subscriptions_this_month'] = conn.execute(
        "SELECT COALESCE(SUM(amount),0) FROM expenses WHERE user_id=? AND date LIKE ? AND category_id='subscriptions'", (uid, month + '%')
    ).fetchone()[0]
    # Subscriptions: last month, count, new-this-month detection
    _cur = datetime.strptime(month, '%Y-%m')
    _prev_m = (_cur.replace(day=1) - timedelta(days=1)).strftime('%Y-%m')
    ctx['subscriptions_last_month'] = conn.execute(
        "SELECT COALESCE(SUM(amount),0) FROM expenses WHERE user_id=? AND date LIKE ? AND category_id='subscriptions'", (uid, _prev_m + '%')
    ).fetchone()[0]
    # Subscription vendor counts and growth — use normalized descriptions for dedup
    _all_sub_descs = [r[0] for r in conn.execute(
        "SELECT DISTINCT description FROM expenses WHERE user_id=? AND is_unusual=0 AND category_id='subscriptions'", (uid,)
    ).fetchall()]
    ctx['subscriptions_count'] = len({_normalize_subscription_desc(d) for d in _all_sub_descs})
    _sub_descs_this_month = [r[0] for r in conn.execute(
        "SELECT DISTINCT description FROM expenses WHERE user_id=? AND date LIKE ? AND category_id='subscriptions'", (uid, month + '%')
    ).fetchall()]
    ctx['subscriptions_count_this_month'] = len({_normalize_subscription_desc(d) for d in _sub_descs_this_month})
    # Normalized descriptions seen this month but never before
    _prev_descs = {_normalize_subscription_desc(r[0]) for r in conn.execute(
        "SELECT DISTINCT description FROM expenses WHERE user_id=? AND date < ? AND category_id='subscriptions'", (uid, month + '-01')
    ).fetchall()}
    _cur_descs = {_normalize_subscription_desc(r[0]) for r in conn.execute(
        "SELECT DISTINCT description FROM expenses WHERE user_id=? AND date LIKE ? AND category_id='subscriptions'", (uid, month + '%')
    ).fetchall()}
    ctx['subscriptions_new_this_month'] = len(_cur_descs - _prev_descs)
    ctx['subscriptions_pct_of_income'] = round(ctx['subscriptions_monthly'] / ctx['inc_total'] * 100, 1) if ctx['inc_total'] > 0 else 0
    ctx['subscriptions_vs_avg_pct'] = round((ctx['subscriptions_this_month'] / ctx['subscriptions_monthly'] - 1) * 100) if ctx['subscriptions_monthly'] > 0 else 0
    # Subscription concentration: top vendors by total spend
    _sub_vendor_rows = conn.execute(
        "SELECT description, SUM(amount) as total FROM expenses WHERE user_id=? AND is_unusual=0 AND category_id='subscriptions' AND description != '' GROUP BY description ORDER BY total DESC", (uid,)
    ).fetchall()
    # Group by normalized name
    _vendor_totals = {}
    for r in _sub_vendor_rows:
        norm = _normalize_subscription_desc(r[0])
        _vendor_totals[norm] = _vendor_totals.get(norm, 0) + r[1]
    _sorted_vendors = sorted(_vendor_totals.items(), key=lambda x: x[1], reverse=True)
    ctx['subscription_top_vendors'] = _sorted_vendors[:3]  # list of (name, total)
    if ctx['subscriptions_total'] > 0 and len(_sorted_vendors) >= 2:
        top2_total = sum(v[1] for v in _sorted_vendors[:2])
        ctx['subscription_concentration_pct'] = round(top2_total / ctx['subscriptions_total'] * 100)
    else:
        ctx['subscription_concentration_pct'] = 0
    ctx['clothing_total'] = conn.execute(
        "SELECT COALESCE(SUM(amount),0) FROM expenses WHERE user_id=? AND is_unusual=0 AND category_id='clothing'", (uid,)
    ).fetchone()[0]
    ctx['gifts_total'] = conn.execute(
        "SELECT COALESCE(SUM(amount),0) FROM expenses WHERE user_id=? AND is_unusual=0 AND category_id='gifts'", (uid,)
    ).fetchone()[0]
    ctx['education_total'] = conn.execute(
        "SELECT COALESCE(SUM(amount),0) FROM expenses WHERE user_id=? AND is_unusual=0 AND category_id='education'", (uid,)
    ).fetchone()[0]

    ctx['bit_total'] = conn.execute(
        "SELECT COALESCE(SUM(amount),0) FROM expenses WHERE user_id=? AND is_unusual=0 AND (subcategory LIKE '%BIT%' OR subcategory LIKE '%PAYBOX%')", (uid,)
    ).fetchone()[0]
    ctx['diners_total'] = conn.execute(
        "SELECT COALESCE(SUM(amount),0) FROM expenses WHERE user_id=? AND is_unusual=0 AND subcategory LIKE '%דיינרס%'", (uid,)
    ).fetchone()[0]
    ctx['isracard_total'] = conn.execute(
        "SELECT COALESCE(SUM(amount),0) FROM expenses WHERE user_id=? AND is_unusual=0 AND subcategory LIKE '%ישראכרט%'", (uid,)
    ).fetchone()[0]

    # Current month overdraft (for recency)
    ctx['overdraft_this_month'] = conn.execute(
        "SELECT COALESCE(SUM(amount),0) FROM expenses WHERE user_id=? AND date LIKE ? AND subcategory='ריבית מינוס'", (uid, month + '%')
    ).fetchone()[0]
    ctx['cash_this_month'] = conn.execute(
        "SELECT COALESCE(SUM(amount),0) FROM expenses WHERE user_id=? AND date LIKE ? AND subcategory='משיכת מזומן'", (uid, month + '%')
    ).fetchone()[0]

    # Deficit months calculation
    all_months_exp = conn.execute(
        "SELECT substr(date,1,7) as m, SUM(amount) as exp_total FROM expenses WHERE user_id=? AND is_unusual=0 GROUP BY m", (uid,)
    ).fetchall()
    neg_months = 0
    for row in all_months_exp:
        m = row[0]
        if m in excl_months:
            continue
        inc = conn.execute(
            "SELECT COALESCE(SUM(amount),0) FROM income WHERE user_id=? AND substr(date,1,7)=?", (uid, m)
        ).fetchone()[0]
        if inc - row[1] < 0:
            neg_months += 1
    ctx['neg_months'] = neg_months

    # Bank balances
    bal_rows = conn.execute(
        "SELECT month, closing_balance FROM bank_balances WHERE user_id=? AND account_name='main' ORDER BY month DESC LIMIT 4", (uid,)
    ).fetchall()
    ctx['bal_rows'] = bal_rows
    ctx['latest_balance'] = bal_rows[0]['closing_balance'] if bal_rows else 0

    # Emergency buffer calculation
    tip_total_by_m = conn.execute(
        "SELECT substr(date,1,7) as m, SUM(amount) as monthly_total FROM expenses WHERE user_id=? AND is_unusual=0 GROUP BY m", (uid,)
    ).fetchall()
    tip_total_vals = [r['monthly_total'] for r in tip_total_by_m if r['m'] not in excl_months]
    avg_total_exp = sum(tip_total_vals) / len(tip_total_vals) if tip_total_vals else 0
    tip_fixed_by_m = conn.execute(
        "SELECT substr(date,1,7) as m, SUM(amount) as monthly_total FROM expenses WHERE user_id=? AND frequency='monthly' AND is_unusual=0 GROUP BY m", (uid,)
    ).fetchall()
    tip_fixed_vals = [r['monthly_total'] for r in tip_fixed_by_m if r['m'] not in excl_months]
    avg_fixed_exp = sum(tip_fixed_vals) / len(tip_fixed_vals) if tip_fixed_vals else 0
    ctx['essential_monthly_avg'] = max(avg_total_exp, avg_fixed_exp)

    # Installments
    inst_rows = conn.execute(
        "SELECT monthly_payment, total_payments, payments_made, status, start_date, description, store FROM installments WHERE user_id=?", (uid,)
    ).fetchall()
    active_insts = [r for r in inst_rows if max(r['total_payments'] - r['payments_made'], 0) > 0 and dict(r).get('status') != 'completed']
    ctx['active_installments'] = active_insts
    ctx['inst_monthly'] = sum(r['monthly_payment'] for r in active_insts)

    # Installment ending soon + recently freed
    today_d = date.today()
    ending_soon = 0
    recently_freed = []
    for r in inst_rows:
        try:
            start = datetime.strptime(r['start_date'], '%Y-%m-%d').date()
            end_year = start.year + (start.month - 1 + r['total_payments']) // 12
            end_month = (start.month - 1 + r['total_payments']) % 12 + 1
            end_d = date(end_year, end_month, min(start.day, 28))
            if dict(r).get('status') != 'completed' and max(r['total_payments'] - r['payments_made'], 0) > 0:
                if 0 < (end_d - today_d).days <= 60:
                    ending_soon += 1
            if dict(r).get('status') == 'completed' or r['payments_made'] >= r['total_payments']:
                if 0 <= (today_d - end_d).days <= 90:
                    recently_freed.append(r)
        except (ValueError, TypeError):
            pass
    ctx['ending_soon_count'] = ending_soon
    ctx['recently_freed'] = recently_freed

    # Insurance overlap count
    try:
        overlap_rows = conn.execute(
            "SELECT COUNT(*) as cnt, SUM(estimated_duplicate_cost_monthly) as cost FROM insurance_overlap_alerts WHERE user_id=? AND status='open'", (uid,)
        ).fetchone()
        ctx['overlap_count'] = overlap_rows['cnt'] if overlap_rows else 0
        ctx['overlap_monthly_cost'] = (overlap_rows['cost'] or 0) if overlap_rows else 0
    except Exception:
        ctx['overlap_count'] = 0
        ctx['overlap_monthly_cost'] = 0

    # Passive income
    try:
        assets = conn.execute("SELECT * FROM assets WHERE user_id=? AND status='active'", (uid,)).fetchall()
        total_passive = 0
        for a in assets:
            a_dict = dict(a)
            if a_dict.get('asset_type') == 'real_estate' and a_dict.get('generates_income'):
                total_passive += a_dict.get('monthly_income', 0) or 0
            elif a_dict.get('asset_type') in ('stocks', 'bonds', 'funds', 'crypto'):
                val = a_dict.get('current_value', 0) or 0
                rate = a_dict.get('annual_return_pct', 0) or 0
                total_passive += val * rate / 100 / 12
        ctx['passive_income'] = total_passive
    except Exception:
        ctx['passive_income'] = 0

    # Per-category monthly averages + current month by category
    cat_rows = conn.execute(
        "SELECT category_id, substr(date,1,7) as m, SUM(amount) as total FROM expenses WHERE user_id=? AND is_unusual=0 GROUP BY category_id, m", (uid,)
    ).fetchall()
    cat_monthly = {}
    for r in cat_rows:
        if r['m'] in excl_months:
            continue
        cid = r['category_id']
        if cid not in cat_monthly:
            cat_monthly[cid] = []
        cat_monthly[cid].append(r['total'])
    ctx['category_avgs'] = {cid: sum(vals) / len(vals) for cid, vals in cat_monthly.items() if vals}
    ctx['current_month_by_cat'] = {}
    for r in cat_rows:
        if r['m'] == month:
            ctx['current_month_by_cat'][r['category_id']] = r['total']

    # Fixed expenses ratio
    fixed_total = conn.execute(
        "SELECT COALESCE(SUM(amount),0) FROM expenses WHERE user_id=? AND is_unusual=0 AND frequency='monthly'", (uid,)
    ).fetchone()[0]
    ctx['fixed_monthly'] = fixed_total / ctx['num_months']

    # Latte-factor candidates (small recurring)
    latte_rows = conn.execute(
        "SELECT description, AVG(amount) as avg_amt, COUNT(DISTINCT substr(date,1,7)) as months FROM expenses WHERE user_id=? AND amount < 50 AND amount > 5 AND description != '' GROUP BY description HAVING months >= 3",
        (uid,)
    ).fetchall()
    ctx['latte_count'] = len(latte_rows)
    ctx['latte_monthly_total'] = sum(r['avg_amt'] for r in latte_rows) if latte_rows else 0

    # Income gap: months with expenses but no income
    income_months = {r[0] for r in conn.execute(
        "SELECT DISTINCT substr(date,1,7) FROM income WHERE user_id=?", (uid,)
    ).fetchall()}
    expense_months = {r[0] for r in conn.execute(
        "SELECT DISTINCT substr(date,1,7) FROM expenses WHERE user_id=?", (uid,)
    ).fetchall()}
    ctx['income_gap_months'] = len(expense_months - income_months)

    # Salary statements data for tips
    try:
        sal_rows = conn.execute(
            "SELECT * FROM salary_statements WHERE user_id=? ORDER BY month DESC LIMIT 12",
            (uid,)
        ).fetchall()
        if sal_rows:
            sal_dicts = [dict(r) for r in sal_rows]
            baseline = [s for s in sal_dicts if (s.get('bonus_amount') or 0) == 0] or sal_dicts
            ctx['has_salary_data'] = True
            ctx['sal_gross'] = sum((s.get('gross_salary') or 0) for s in baseline) / len(baseline)
            ctx['sal_net'] = sum((s.get('net_salary') or 0) for s in baseline) / len(baseline)
            ctx['sal_employer_total'] = sum(
                (s.get('pension_employer') or 0) + (s.get('education_fund_employer') or 0) + (s.get('severance_employer') or 0)
                for s in baseline
            ) / len(baseline)
            ctx['sal_true_comp'] = ctx['sal_gross'] + ctx['sal_employer_total']
            latest = sal_dicts[0]
            ctx['sal_vacation_days'] = latest.get('vacation_days') or 0
            ctx['sal_bonus'] = latest.get('bonus_amount') or 0
            ctx['sal_months_count'] = len(sal_dicts)
            # Income stability metrics
            grosses = [(s.get('gross_salary') or 0) for s in sal_dicts]
            mean_g = sum(grosses) / len(grosses) if grosses else 0
            if mean_g > 0:
                variance = sum((g - mean_g) ** 2 for g in grosses) / len(grosses)
                ctx['sal_cv'] = (variance ** 0.5) / mean_g
            else:
                ctx['sal_cv'] = 0
            bonus_month_count = sum(1 for s in sal_dicts if (s.get('bonus_amount') or 0) > 0)
            ctx['sal_bonus_prevalence'] = bonus_month_count / len(sal_dicts)
            total_gross_all = sum(grosses)
            total_bonus_all = sum((s.get('bonus_amount') or 0) for s in sal_dicts)
            ctx['sal_bonus_share'] = total_bonus_all / total_gross_all if total_gross_all > 0 else 0
            ctx['sal_latest_gross'] = sal_dicts[0].get('gross_salary') or 0
            ctx['sal_latest_vs_baseline'] = (
                (ctx['sal_latest_gross'] - ctx['sal_gross']) / ctx['sal_gross'] * 100
            ) if ctx['sal_gross'] > 0 else 0
        else:
            ctx['has_salary_data'] = False
    except Exception:
        ctx['has_salary_data'] = False

    return ctx


def _score_tip(tip, overlapping_ids=None):
    """Score a tip using 6 weighted factors for sharp ranking separation."""
    C = TIP_CONFIG
    SEVERITY_W = {'important': C['sev_important'], 'watch': C['sev_watch'], 'opportunity': C['sev_opportunity']}
    impact = min(tip.get('impact_amount', 0) / C['impact_cap'], 1.0)
    severity = SEVERITY_W.get(tip.get('severity', 'watch'), 0.5)
    urgency = tip.get('urgency', 0.5)
    confidence = tip.get('confidence', 0.7)
    recency = tip.get('recency', 0.7)
    action_bonus = 0.1 if tip.get('action') else 0

    score = (
        severity * C['w_severity'] +
        impact * C['w_impact'] +
        urgency * C['w_urgency'] +
        confidence * C['w_confidence'] +
        recency * C['w_recency'] +
        action_bonus * C['w_action_bonus']
    )
    # Coordination penalty: lower score when trajectory already shows this signal
    if overlapping_ids and tip.get('id') in overlapping_ids:
        score -= C['overlap_penalty']
    return round(max(min(score, 1.0), 0.0), 2)


def _deduplicate_tips(tips):
    """Apply subsumption rules and category caps."""
    C = TIP_CONFIG
    tip_ids = {t['id'] for t in tips}

    # Subsumption: more severe tip replaces less severe
    for winner, loser in C['subsumption_rules']:
        if winner in tip_ids and loser in tip_ids:
            tips = [t for t in tips if t['id'] != loser]

    # Category cap
    cat_counts = {}
    result = []
    for t in tips:
        cat = t.get('category', 'other')
        cat_counts[cat] = cat_counts.get(cat, 0) + 1
        if cat_counts[cat] <= C['max_per_category']:
            result.append(t)

    return result[:C['max_total_tips']]


# ---- Tip Generators (each returns dict or None) ----

def _tip_overdraft_interest(ctx):
    if ctx['overdraft_total'] <= 0:
        return None
    monthly_avg = ctx['overdraft_total'] / ctx['num_months']
    yearly_est = monthly_avg * 12
    return {
        'id': 'overdraft_interest', 'category': 'debt', 'severity': 'important',
        'icon': 'bi-exclamation-triangle-fill', 'color': '#dc2626',
        'impact_amount': yearly_est, 'impact_type': 'yearly_cost',
        'urgency': 0.9, 'confidence': 0.95,
        'recency': 1.0 if ctx['overdraft_this_month'] > 0 else 0.6,
        'params': {'total': f"{ctx['overdraft_total']:,.0f}", 'yearly': f'{yearly_est:,.0f}', 'monthly': f'{monthly_avg:,.0f}'},
        'action': 'bank_balances', 'action_label_key': 'tip_action_view_bank',
    }


def _tip_negative_balance(ctx):
    if not ctx['bal_rows'] or ctx['latest_balance'] >= 0:
        return None
    return {
        'id': 'negative_balance', 'category': 'debt', 'severity': 'important',
        'icon': 'bi-exclamation-octagon-fill', 'color': '#dc2626',
        'impact_amount': abs(ctx['latest_balance']) * 0.15, 'impact_type': 'risk',
        'urgency': 1.0, 'confidence': 0.95, 'recency': 1.0,
        'params': {'balance': f"{ctx['latest_balance']:,.0f}"},
        'action': 'bank_balances', 'action_label_key': 'tip_action_view_bank',
    }


def _tip_balance_decline(ctx):
    C = TIP_CONFIG
    if len(ctx['bal_rows']) < C['balance_decline_months']:
        return None
    balances = [r['closing_balance'] for r in reversed(ctx['bal_rows'])]
    if balances[0] <= 0 or balances[-1] >= balances[0]:
        return None
    drop = balances[0] - balances[-1]
    pct_drop = drop / balances[0]
    n = len(balances) - 1
    if pct_drop <= C['balance_decline_pct'] * n:
        return None
    monthly_drop = drop / n
    return {
        'id': 'balance_decline', 'category': 'debt', 'severity': 'watch',
        'icon': 'bi-graph-down-arrow', 'color': '#f59e0b',
        'impact_amount': monthly_drop * 12, 'impact_type': 'risk',
        'urgency': 0.7, 'confidence': 0.8, 'recency': 0.9,
        'params': {'drop': f'{drop:,.0f}', 'months': str(n), 'monthly_drop': f'{monthly_drop:,.0f}'},
        'action': 'bank_balances', 'action_label_key': 'tip_action_view_bank',
    }


def _tip_installment_burden(ctx):
    if ctx['inst_monthly'] <= 0:
        return None
    inc = ctx['inc_total']
    inst_pct = round((ctx['inst_monthly'] / inc) * 100, 0) if inc > 0 else 0
    if inst_pct <= TIP_CONFIG['installment_burden_threshold'] * 100:
        return None
    return {
        'id': 'installment_burden', 'category': 'debt', 'severity': 'watch',
        'icon': 'bi-credit-card-2-front', 'color': '#7c3aed',
        'impact_amount': ctx['inst_monthly'] * 12, 'impact_type': 'monthly_savings',
        'urgency': 0.5, 'confidence': 0.9, 'recency': 0.8,
        'params': {
            'monthly': f"{ctx['inst_monthly']:,.0f}", 'count': str(len(ctx['active_installments'])),
            'pct': str(int(inst_pct)), 'ending_soon': str(ctx['ending_soon_count']),
        },
        'action': 'installments', 'action_label_key': 'tip_action_view_installments',
    }


def _tip_installment_avoid(ctx):
    if ctx['inst_monthly'] <= 0 or ctx['inc_total'] <= 0:
        return None
    inst_pct = round((ctx['inst_monthly'] / ctx['inc_total']) * 100, 0)
    if inst_pct <= TIP_CONFIG['installment_burden_threshold'] * 100 or ctx['exp_total'] <= ctx['inc_total']:
        return None
    deficit = ctx['exp_total'] - ctx['inc_total']
    return {
        'id': 'installment_avoid', 'category': 'debt', 'severity': 'important',
        'icon': 'bi-exclamation-diamond', 'color': '#dc2626',
        'impact_amount': deficit * 12, 'impact_type': 'risk',
        'urgency': 0.8, 'confidence': 0.85, 'recency': 1.0,
        'params': {'pct': str(int(inst_pct)), 'deficit': f'{deficit:,.0f}'},
        'action': 'installments', 'action_label_key': 'tip_action_view_installments',
    }


def _tip_food_spending(ctx):
    if ctx['food_monthly'] <= TIP_CONFIG['food_monthly_threshold']:
        return None
    rest_pct = (ctx['restaurant_total'] / ctx['food_combined_total'] * 100) if ctx['food_combined_total'] else 0
    savings = ctx['food_monthly'] * 0.2
    return {
        'id': 'food_spending', 'category': 'spending', 'severity': 'watch',
        'icon': 'bi-cart4', 'color': '#f28e2b',
        'impact_amount': savings * 12, 'impact_type': 'monthly_savings',
        'urgency': 0.3, 'confidence': 0.9, 'recency': 0.5,
        'params': {'monthly': f"{ctx['food_monthly']:,.0f}", 'rest_pct': f'{rest_pct:.0f}', 'monthly_savings': f'{savings:,.0f}'},
        'action': 'expenses_food', 'action_label_key': 'tip_action_view_food',
    }


def _tip_restaurant_ratio(ctx):
    if ctx['food_combined_total'] <= 0:
        return None
    rest_pct = ctx['restaurant_total'] / ctx['food_combined_total'] * 100
    if rest_pct <= TIP_CONFIG['restaurant_ratio_threshold'] * 100:
        return None
    savings = ctx['restaurant_total'] / ctx['num_months'] * 0.5
    return {
        'id': 'restaurant_ratio', 'category': 'spending', 'severity': 'opportunity',
        'icon': 'bi-cup-hot', 'color': '#e15759',
        'impact_amount': savings * 12, 'impact_type': 'monthly_savings',
        'urgency': 0.2, 'confidence': 0.85, 'recency': 0.5,
        'params': {'rest_pct': f'{rest_pct:.0f}', 'monthly_restaurant': f'{ctx["restaurant_total"] / ctx["num_months"]:,.0f}'},
        'action': 'expenses_food', 'action_label_key': 'tip_action_view_food',
    }


def _tip_entertainment_high(ctx):
    if ctx['ent_monthly'] <= TIP_CONFIG['entertainment_monthly_threshold']:
        return None
    savings = max(ctx['ent_monthly'] - 1200, 0)
    return {
        'id': 'entertainment_high', 'category': 'spending', 'severity': 'watch',
        'icon': 'bi-film', 'color': '#9c755f',
        'impact_amount': savings * 12, 'impact_type': 'monthly_savings',
        'urgency': 0.3, 'confidence': 0.9, 'recency': 0.5,
        'params': {'monthly': f"{ctx['ent_monthly']:,.0f}", 'potential_savings': f'{savings:,.0f}'},
        'action': 'expenses_entertainment', 'action_label_key': 'tip_action_view_entertainment',
    }


def _tip_subscriptions_cost(ctx):
    if ctx['subscriptions_monthly'] <= TIP_CONFIG.get('subscriptions_monthly_threshold', 200):
        return None
    yearly = ctx['subscriptions_monthly'] * 12
    vs_avg = ctx['subscriptions_vs_avg_pct']
    params = {
        'monthly': f"{ctx['subscriptions_monthly']:,.0f}",
        'yearly': f'{yearly:,.0f}',
        'count': str(ctx['subscriptions_count']),
        'count_this_month': str(ctx['subscriptions_count_this_month']),
        'pct_of_income': f"{ctx['subscriptions_pct_of_income']:.1f}",
    }
    # Context comparison: vs average
    if vs_avg > 10:
        params['vs_avg'] = f"+{vs_avg}%"
    elif vs_avg < -10:
        params['vs_avg'] = f"{vs_avg}%"
    else:
        params['vs_avg'] = ''
    # New subscriptions this month
    params['new_count'] = str(ctx['subscriptions_new_this_month'])
    # Concentration insight: when top 2 vendors dominate
    conc = ctx.get('subscription_concentration_pct', 0)
    top_vendors = ctx.get('subscription_top_vendors', [])
    if conc >= 60 and len(top_vendors) >= 2:
        params['concentration_pct'] = str(conc)
        params['top_vendor'] = top_vendors[0][0]
    else:
        params['concentration_pct'] = ''
        params['top_vendor'] = ''
    # Text variant for frontend i18n key selection
    if vs_avg > 10:
        params['_variant'] = 'up'
    elif vs_avg < -10:
        params['_variant'] = 'down'
    else:
        params['_variant'] = ''
    return {
        'id': 'subscriptions_cost', 'category': 'spending', 'severity': 'watch',
        'icon': 'bi-repeat', 'color': '#17becf',
        'impact_amount': yearly, 'impact_type': 'yearly_cost',
        'urgency': 0.6 if ctx['subscriptions_new_this_month'] > 0 else 0.4,
        'confidence': 0.9,
        'recency': 1.0 if ctx['subscriptions_this_month'] > 0 else 0.5,
        'params': params,
        'action': 'expenses_subscriptions', 'action_label_key': 'tip_action_view_subscriptions',
    }


def _tip_subscriptions_growth(ctx):
    new_count = ctx['subscriptions_new_this_month']
    if new_count < 1:
        return None
    return {
        'id': 'subscriptions_growth', 'category': 'spending', 'severity': 'watch' if new_count >= 2 else 'opportunity',
        'icon': 'bi-plus-circle', 'color': '#17becf',
        'impact_amount': ctx['subscriptions_this_month'], 'impact_type': 'monthly_savings',
        'urgency': 0.7 if new_count >= 2 else 0.4, 'confidence': 0.8,
        'recency': 1.0,
        'params': {'new_count': str(new_count), 'this_month': f"{ctx['subscriptions_this_month']:,.0f}"},
        'action': 'expenses_subscriptions', 'action_label_key': 'tip_action_view_subscriptions',
    }


def _tip_dining_out_high(ctx):
    if ctx['dining_out_monthly'] <= TIP_CONFIG.get('dining_out_monthly_threshold', 800):
        return None
    savings = ctx['dining_out_monthly'] * 0.25
    vs_avg = ctx['dining_out_vs_avg_pct']
    params = {
        'monthly': f"{ctx['dining_out_monthly']:,.0f}",
        'potential_savings': f'{savings:,.0f}',
        'pct_of_income': f"{ctx['dining_out_pct_of_income']:.1f}",
    }
    if vs_avg > 10:
        params['vs_avg'] = f"+{vs_avg}%"
        params['_variant'] = 'up'
    elif vs_avg < -10:
        params['vs_avg'] = f"{vs_avg}%"
        params['_variant'] = 'down'
    else:
        params['vs_avg'] = ''
        params['_variant'] = ''
    return {
        'id': 'dining_out_high', 'category': 'spending', 'severity': 'watch',
        'icon': 'bi-cup-straw', 'color': '#ff6b6b',
        'impact_amount': savings * 12, 'impact_type': 'monthly_savings',
        'urgency': 0.5 if vs_avg > 20 else 0.3, 'confidence': 0.85,
        'recency': 1.0 if vs_avg > 10 else 0.7,
        'params': params,
        'action': 'expenses_dining_out', 'action_label_key': 'tip_action_view_dining_out',
    }


def _tip_cash_untracked(ctx):
    if ctx['cash_total'] <= TIP_CONFIG['cash_total_threshold']:
        return None
    monthly_cash = ctx['cash_total'] / ctx['num_months']
    savings = monthly_cash * 0.3
    return {
        'id': 'cash_untracked', 'category': 'spending', 'severity': 'watch',
        'icon': 'bi-cash', 'color': '#f59e0b',
        'impact_amount': savings * 12, 'impact_type': 'yearly_cost',
        'urgency': 0.4, 'confidence': 0.7,
        'recency': 0.9 if ctx['cash_this_month'] > 0 else 0.5,
        'params': {'total': f"{ctx['cash_total']:,.0f}", 'monthly': f'{monthly_cash:,.0f}', 'monthly_savings': f'{savings:,.0f}'},
        'action': 'expenses', 'action_label_key': 'tip_action_view_expenses',
    }


def _tip_bit_untracked(ctx):
    if ctx['bit_total'] <= TIP_CONFIG['bit_total_threshold']:
        return None
    monthly_bit = ctx['bit_total'] / ctx['num_months']
    return {
        'id': 'bit_untracked', 'category': 'spending', 'severity': 'opportunity',
        'icon': 'bi-phone', 'color': '#59a14f',
        'impact_amount': ctx['bit_total'] * 0.15, 'impact_type': 'yearly_cost',
        'urgency': 0.2, 'confidence': 0.6, 'recency': 0.5,
        'params': {'total': f"{ctx['bit_total']:,.0f}", 'monthly': f'{monthly_bit:,.0f}'},
        'action': 'expenses', 'action_label_key': 'tip_action_view_expenses',
    }


def _tip_multi_cards(ctx):
    if ctx['diners_total'] <= 0 or ctx['isracard_total'] <= 0:
        return None
    return {
        'id': 'multi_cards', 'category': 'spending', 'severity': 'opportunity',
        'icon': 'bi-credit-card-2-back', 'color': '#8b5cf6',
        'impact_amount': 480, 'impact_type': 'yearly_cost',
        'urgency': 0.2, 'confidence': 0.95, 'recency': 0.6,
        'params': {'isracard': f"{ctx['isracard_total']:,.0f}", 'diners': f"{ctx['diners_total']:,.0f}"},
        'action': 'expenses', 'action_label_key': 'tip_action_view_expenses',
    }


def _tip_category_spike(ctx):
    C = TIP_CONFIG
    spike_cat = None
    spike_pct = 0
    spike_amount = 0
    spike_avg = 0
    for cat_id, avg in ctx['category_avgs'].items():
        current = ctx['current_month_by_cat'].get(cat_id, 0)
        if avg > 0 and current > avg * (1 + C['category_spike_threshold']):
            pct = ((current - avg) / avg) * 100
            if pct > spike_pct:
                spike_pct = pct
                spike_cat = cat_id
                spike_amount = current
                spike_avg = avg
    if not spike_cat or spike_pct < C['category_spike_threshold'] * 100:
        return None
    return {
        'id': 'category_spike', 'category': 'spending', 'severity': 'watch',
        'icon': 'bi-arrow-up-right-circle', 'color': '#ef4444',
        'impact_amount': spike_amount - spike_avg, 'impact_type': 'monthly_savings',
        'urgency': 0.6, 'confidence': 0.75, 'recency': 1.0,
        'params': {'category': spike_cat, 'current': f'{spike_amount:,.0f}', 'avg': f'{spike_avg:,.0f}', 'pct': f'{spike_pct:.0f}'},
        'action': 'expenses', 'action_label_key': 'tip_action_view_expenses',
    }


def _tip_latte_effect(ctx):
    if ctx['latte_count'] < TIP_CONFIG['latte_min_vendors']:
        return None
    yearly = ctx['latte_monthly_total'] * 12
    return {
        'id': 'latte_effect', 'category': 'spending', 'severity': 'opportunity',
        'icon': 'bi-cup-hot-fill', 'color': '#d97706',
        'impact_amount': yearly, 'impact_type': 'monthly_savings',
        'urgency': 0.2, 'confidence': 0.7, 'recency': 0.5,
        'params': {'count': str(ctx['latte_count']), 'monthly': f"{ctx['latte_monthly_total']:,.0f}", 'yearly': f'{yearly:,.0f}'},
        'action': 'standing_orders', 'action_label_key': 'tip_action_view_standing_orders',
    }


def _tip_low_savings_rate(ctx):
    if ctx['total_income_all'] <= 0:
        return None
    rate = ctx['savings_rate']
    if rate >= TIP_CONFIG['savings_low_threshold'] * 100:
        return None
    return {
        'id': 'low_savings_rate', 'category': 'savings', 'severity': 'watch',
        'icon': 'bi-piggy-bank', 'color': '#4dc9f6',
        'impact_amount': 0, 'impact_type': 'coverage',
        'urgency': 0.5, 'confidence': 0.85, 'recency': 0.6,
        'params': {'rate': f'{rate:.1f}'},
        'action': 'income', 'action_label_key': 'tip_action_view_income',
    }


def _tip_good_savings_rate(ctx):
    if ctx['total_income_all'] <= 0:
        return None
    rate = ctx['savings_rate']
    if rate < TIP_CONFIG['savings_good_threshold'] * 100:
        return None
    return {
        'id': 'good_savings_rate', 'category': 'savings', 'severity': 'opportunity',
        'icon': 'bi-trophy', 'color': '#16a34a',
        'impact_amount': 0, 'impact_type': 'coverage',
        'urgency': 0.1, 'confidence': 0.85, 'recency': 0.6,
        'params': {'rate': f'{rate:.1f}'},
        'action': 'income', 'action_label_key': 'tip_action_view_income',
    }


def _tip_low_emergency_buffer(ctx):
    if not ctx['bal_rows'] or ctx['latest_balance'] <= 0 or ctx['essential_monthly_avg'] <= 0:
        return None
    emergency_months = ctx['latest_balance'] / ctx['essential_monthly_avg']
    if emergency_months >= TIP_CONFIG['emergency_buffer_months']:
        return None
    emergency_weeks = max(1, round(emergency_months * 4.3))
    return {
        'id': 'low_emergency_buffer', 'category': 'savings', 'severity': 'important',
        'icon': 'bi-shield-exclamation', 'color': '#6366f1',
        'impact_amount': 0, 'impact_type': 'risk',
        'urgency': 0.8, 'confidence': 0.9, 'recency': 0.9,
        'params': {
            'emergency_months': f'{emergency_months:.1f}', 'emergency_weeks': str(emergency_weeks),
            'balance': f"{ctx['latest_balance']:,.0f}", 'essential_monthly': f"{ctx['essential_monthly_avg']:,.0f}",
        },
        'action': 'bank_balances', 'action_label_key': 'tip_action_view_bank',
    }


def _tip_insurance_overlaps(ctx):
    if ctx['overlap_count'] <= 0:
        return None
    monthly_cost = ctx['overlap_monthly_cost']
    return {
        'id': 'insurance_overlaps', 'category': 'savings', 'severity': 'watch',
        'icon': 'bi-shield-check', 'color': '#ff9da7',
        'impact_amount': monthly_cost * 12, 'impact_type': 'monthly_savings',
        'urgency': 0.4, 'confidence': 0.7, 'recency': 0.6,
        'params': {'count': str(ctx['overlap_count']), 'monthly_cost': f'{monthly_cost:,.0f}'},
        'action': 'insurance', 'action_label_key': 'tip_action_view_insurance',
    }


def _tip_passive_income_low(ctx):
    if ctx['essential_monthly_avg'] <= 0 or ctx['passive_income'] <= 0:
        return None
    pct = ctx['passive_income'] / ctx['essential_monthly_avg'] * 100
    if pct >= TIP_CONFIG['passive_income_low_threshold'] * 100:
        return None
    return {
        'id': 'passive_income_low', 'category': 'income', 'severity': 'opportunity',
        'icon': 'bi-graph-up', 'color': '#2563eb',
        'impact_amount': 0, 'impact_type': 'coverage',
        'urgency': 0.2, 'confidence': 0.8, 'recency': 0.5,
        'params': {'passive': f"{ctx['passive_income']:,.0f}", 'pct': f'{pct:.0f}'},
        'action': 'net_worth', 'action_label_key': 'tip_action_view_net_worth',
    }


def _tip_passive_income_good(ctx):
    if ctx['essential_monthly_avg'] <= 0 or ctx['passive_income'] <= 0:
        return None
    pct = ctx['passive_income'] / ctx['essential_monthly_avg'] * 100
    if pct < TIP_CONFIG['passive_income_good_threshold'] * 100:
        return None
    return {
        'id': 'passive_income_good', 'category': 'income', 'severity': 'opportunity',
        'icon': 'bi-gem', 'color': '#16a34a',
        'impact_amount': 0, 'impact_type': 'coverage',
        'urgency': 0.1, 'confidence': 0.8, 'recency': 0.5,
        'params': {'passive': f"{ctx['passive_income']:,.0f}", 'pct': f'{pct:.0f}'},
        'action': 'net_worth', 'action_label_key': 'tip_action_view_net_worth',
    }


def _tip_income_gap(ctx):
    if ctx['income_gap_months'] < TIP_CONFIG['income_gap_threshold']:
        return None
    return {
        'id': 'income_gap', 'category': 'income', 'severity': 'watch',
        'icon': 'bi-calendar-x', 'color': '#d97706',
        'impact_amount': 0, 'impact_type': 'risk',
        'urgency': 0.6, 'confidence': 0.6, 'recency': 0.7,
        'params': {'gap_months': str(ctx['income_gap_months'])},
        'action': 'income', 'action_label_key': 'tip_action_view_income',
    }


def _tip_deficit_months(ctx):
    if ctx['neg_months'] < TIP_CONFIG['deficit_months_threshold']:
        return None
    return {
        'id': 'deficit_months', 'category': 'planning', 'severity': 'important',
        'icon': 'bi-graph-down-arrow', 'color': '#dc2626',
        'impact_amount': 0, 'impact_type': 'risk',
        'urgency': 0.7, 'confidence': 0.9, 'recency': 0.8,
        'params': {'neg_months': str(ctx['neg_months']), 'num_months': str(ctx['num_months'])},
        'action': 'analysis', 'action_label_key': 'tip_action_run_analysis',
    }


def _tip_installment_ending(ctx):
    if ctx['ending_soon_count'] <= 0:
        return None
    freed_monthly = 0
    for r in ctx['active_installments']:
        try:
            start = datetime.strptime(r['start_date'], '%Y-%m-%d').date()
            end_year = start.year + (start.month - 1 + r['total_payments']) // 12
            end_month = (start.month - 1 + r['total_payments']) % 12 + 1
            end_d = date(end_year, end_month, min(start.day, 28))
            if 0 < (end_d - date.today()).days <= TIP_CONFIG['installment_ending_days']:
                freed_monthly += r['monthly_payment']
        except (ValueError, TypeError):
            pass
    return {
        'id': 'installment_ending', 'category': 'planning', 'severity': 'opportunity',
        'icon': 'bi-check-circle', 'color': '#16a34a',
        'impact_amount': freed_monthly * 12, 'impact_type': 'monthly_savings',
        'urgency': 0.3, 'confidence': 0.95, 'recency': 0.9,
        'params': {'count': str(ctx['ending_soon_count']), 'freed_monthly': f'{freed_monthly:,.0f}'},
        'action': 'installments', 'action_label_key': 'tip_action_view_installments',
    }


def _tip_installment_freed(ctx):
    if not ctx['recently_freed']:
        return None
    freed_monthly = sum(r['monthly_payment'] for r in ctx['recently_freed'])
    return {
        'id': 'installment_freed', 'category': 'planning', 'severity': 'opportunity',
        'icon': 'bi-unlock', 'color': '#22c55e',
        'impact_amount': freed_monthly * 12, 'impact_type': 'monthly_savings',
        'urgency': 0.1, 'confidence': 0.95, 'recency': 0.8,
        'params': {'count': str(len(ctx['recently_freed'])), 'freed_monthly': f'{freed_monthly:,.0f}'},
        'action': 'installments', 'action_label_key': 'tip_action_view_installments',
    }


def _tip_high_fixed_ratio(ctx):
    if ctx['inc_total'] <= 0:
        return None
    ratio = ctx['fixed_monthly'] / ctx['inc_total'] * 100
    if ratio <= TIP_CONFIG['fixed_ratio_threshold'] * 100:
        return None
    return {
        'id': 'high_fixed_ratio', 'category': 'planning', 'severity': 'watch',
        'icon': 'bi-lock', 'color': '#7c3aed',
        'impact_amount': 0, 'impact_type': 'risk',
        'urgency': 0.5, 'confidence': 0.85, 'recency': 0.7,
        'params': {'ratio': f'{ratio:.0f}', 'fixed': f"{ctx['fixed_monthly']:,.0f}"},
        'action': 'expenses', 'action_label_key': 'tip_action_view_expenses',
    }


def _tip_national_comparison(ctx):
    C = TIP_CONFIG
    if ctx['inc_total'] <= 0:
        return None
    above_count = 0
    above_cats = []
    for cat_id, my_avg in ctx['category_avgs'].items():
        cbs_pct = CBS_AVERAGES.get(cat_id, 0)
        if cbs_pct <= 0 or ctx['inc_total'] <= 0:
            continue
        my_pct = (my_avg / ctx['inc_total']) * 100
        if my_pct > cbs_pct + C['national_comparison_pp_above']:
            above_count += 1
            above_cats.append(cat_id)
    if above_count < C['national_comparison_min_cats']:
        return None
    return {
        'id': 'national_comparison', 'category': 'planning', 'severity': 'opportunity',
        'icon': 'bi-bar-chart-line', 'color': '#6366f1',
        'impact_amount': 0, 'impact_type': 'monthly_savings',
        'urgency': 0.2, 'confidence': 0.6, 'recency': 0.5,
        'params': {'count': str(above_count), 'categories': ', '.join(above_cats[:3])},
        'action': 'analysis', 'action_label_key': 'tip_action_run_analysis',
    }


# ---- Salary-aware tips ----

def _tip_vacation_days_unused(ctx):
    if not ctx.get('has_salary_data') or (ctx.get('sal_vacation_days') or 0) < TIP_CONFIG['vacation_days_threshold']:
        return None
    return {
        'id': 'vacation_days_unused', 'category': 'income', 'severity': 'watch',
        'icon': 'bi-sun', 'color': '#f59e0b',
        'impact_amount': 0, 'impact_type': 'coverage',
        'urgency': 0.4, 'confidence': 0.9, 'recency': 0.7,
        'params': {'days': f'{ctx["sal_vacation_days"]:.0f}'},
        'action': 'income', 'action_label_key': 'tip_action_check_vacation',
    }


def _tip_bonus_detected(ctx):
    if not ctx.get('has_salary_data') or (ctx.get('sal_bonus') or 0) <= 0:
        return None
    C = TIP_CONFIG
    original_bonus = ctx['sal_bonus']
    remaining = original_bonus
    alloc = {}
    # Priority 1: Emergency buffer (if below threshold months)
    ema = ctx.get('essential_monthly_avg') or 0
    buffer_months = (ctx.get('latest_balance') or 0) / ema if ema > 0 else 99
    if buffer_months < C['emergency_buffer_months']:
        gap = ema * C['emergency_buffer_months'] - max(ctx.get('latest_balance') or 0, 0)
        emergency_alloc = round(min(remaining * 0.50, max(gap, 0)))
        if emergency_alloc > 0:
            alloc['emergency'] = emergency_alloc
            remaining -= emergency_alloc
    # Priority 2: Debt paydown (if installments or overdraft)
    if (ctx.get('inst_monthly') or 0) > 0 or (ctx.get('overdraft_total') or 0) > 0:
        debt_alloc = round(min(remaining * 0.40, remaining))
        if debt_alloc > 0:
            alloc['debt'] = debt_alloc
            remaining -= debt_alloc
    # Priority 3: Investment (if savings rate below good threshold)
    if (ctx.get('savings_rate') or 0) < C['savings_good_threshold'] * 100:
        invest_alloc = round(min(remaining * 0.50, remaining))
        if invest_alloc > 0:
            alloc['invest'] = invest_alloc
            remaining -= invest_alloc
    # Remainder: Enjoy
    alloc['enjoy'] = round(max(remaining, 0))
    return {
        'id': 'bonus_detected', 'category': 'income', 'severity': 'opportunity',
        'icon': 'bi-gift', 'color': '#8b5cf6',
        'impact_amount': original_bonus, 'impact_type': 'monthly_savings',
        'urgency': 0.6, 'confidence': 0.95, 'recency': 0.9,
        'params': {
            'amount': f'{original_bonus:,.0f}',
            'emergency': f'{alloc.get("emergency", 0):,.0f}',
            'debt': f'{alloc.get("debt", 0):,.0f}',
            'invest': f'{alloc.get("invest", 0):,.0f}',
            'enjoy': f'{alloc.get("enjoy", 0):,.0f}',
            'buffer_months': f'{buffer_months:.1f}',
        },
        'action': 'income', 'action_label_key': 'tip_action_allocate_bonus',
    }


def _tip_true_savings_rate(ctx):
    if not ctx.get('has_salary_data') or (ctx.get('sal_gross') or 0) <= 0:
        return None
    # Approximate: employer contributions / gross as a savings proxy
    emp = ctx.get('sal_employer_total') or 0
    gross = ctx['sal_gross']
    rate = round(emp / gross * 100, 1)
    if rate >= TIP_CONFIG['true_savings_low_threshold']:
        return None
    return {
        'id': 'true_savings_rate', 'category': 'savings', 'severity': 'watch',
        'icon': 'bi-percent', 'color': '#dc2626',
        'impact_amount': 0, 'impact_type': 'coverage',
        'urgency': 0.5, 'confidence': 0.8, 'recency': 0.6,
        'params': {'rate': str(rate), 'gross': f'{gross:,.0f}'},
        'action': 'income', 'action_label_key': 'tip_action_review_hidden_savings',
    }


def _tip_employer_contributions_value(ctx):
    if not ctx.get('has_salary_data') or (ctx.get('sal_employer_total') or 0) <= 0:
        return None
    monthly = ctx['sal_employer_total']
    annual = monthly * 12
    return {
        'id': 'employer_contributions_value', 'category': 'income', 'severity': 'opportunity',
        'icon': 'bi-building', 'color': '#16a34a',
        'impact_amount': round(monthly), 'impact_type': 'monthly_savings',
        'urgency': 0.2, 'confidence': 0.95, 'recency': 0.5,
        'params': {'monthly': f'{monthly:,.0f}', 'annual': f'{annual:,.0f}'},
        'action': 'income', 'action_label_key': 'tip_action_review_hidden_savings',
    }


# ---- Income stability tips ----

def _tip_income_stable(ctx):
    C = TIP_CONFIG
    if not ctx.get('has_salary_data') or (ctx.get('sal_months_count') or 0) < C['income_stable_min_months']:
        return None
    if (ctx.get('sal_cv') or 0) >= C['income_cv_stable_threshold']:
        return None
    if (ctx.get('sal_bonus_share') or 0) >= C['bonus_share_contradicts_stable']:
        return None
    return {
        'id': 'income_stable', 'category': 'income', 'severity': 'opportunity',
        'icon': 'bi-shield-check', 'color': '#16a34a',
        'impact_amount': 0, 'impact_type': 'coverage',
        'urgency': 0.2, 'confidence': 0.9, 'recency': 0.5,
        'params': {'months': str(ctx['sal_months_count'])},
        'action': 'income', 'action_label_key': 'tip_action_view_salary',
    }


def _tip_bonus_reliance(ctx):
    C = TIP_CONFIG
    if not ctx.get('has_salary_data') or (ctx.get('sal_bonus_share') or 0) < C['bonus_reliance_threshold']:
        return None
    pct = round(ctx['sal_bonus_share'] * 100)
    return {
        'id': 'bonus_reliance', 'category': 'income', 'severity': 'watch',
        'icon': 'bi-exclamation-diamond', 'color': '#d97706',
        'impact_amount': round(ctx.get('sal_bonus') or 0), 'impact_type': 'risk',
        'urgency': 0.5, 'confidence': 0.85, 'recency': 0.7,
        'params': {'pct': str(pct)},
        'action': 'income', 'action_label_key': 'tip_action_view_salary',
    }


def _tip_income_drop(ctx):
    C = TIP_CONFIG
    if not ctx.get('has_salary_data') or (ctx.get('sal_months_count') or 0) < 2:
        return None
    drop = ctx.get('sal_latest_vs_baseline') or 0
    if drop > C['income_drop_threshold']:
        return None
    latest = ctx.get('sal_latest_gross') or 0
    baseline = ctx.get('sal_gross') or 0
    return {
        'id': 'income_drop', 'category': 'income', 'severity': 'important',
        'icon': 'bi-graph-down-arrow', 'color': '#dc2626',
        'impact_amount': round(abs(latest - baseline)), 'impact_type': 'monthly_savings',
        'urgency': 0.8, 'confidence': 0.85, 'recency': 0.9,
        'params': {'drop_pct': str(abs(round(drop))), 'latest': f'{latest:,.0f}', 'baseline': f'{baseline:,.0f}'},
        'action': 'income', 'action_label_key': 'tip_action_view_salary',
    }


# All tip generators in evaluation order
_TIP_GENERATORS = [
    _tip_overdraft_interest, _tip_negative_balance, _tip_balance_decline,
    _tip_installment_burden, _tip_installment_avoid,
    _tip_food_spending, _tip_restaurant_ratio, _tip_entertainment_high,
    _tip_subscriptions_cost, _tip_subscriptions_growth, _tip_dining_out_high,
    _tip_cash_untracked, _tip_bit_untracked, _tip_multi_cards,
    _tip_category_spike, _tip_latte_effect,
    _tip_low_savings_rate, _tip_good_savings_rate, _tip_low_emergency_buffer,
    _tip_insurance_overlaps,
    _tip_passive_income_low, _tip_passive_income_good, _tip_income_gap,
    _tip_deficit_months, _tip_installment_ending, _tip_installment_freed,
    _tip_high_fixed_ratio, _tip_national_comparison,
    _tip_vacation_days_unused, _tip_bonus_detected,
    _tip_true_savings_rate, _tip_employer_contributions_value,
    _tip_income_stable, _tip_bonus_reliance, _tip_income_drop,
]

# Map trajectory signal keys → tip IDs that cover the same ground
_TRAJECTORY_TIP_OVERLAP = {
    'traj_overspending': {'deficit_months', 'high_fixed_ratio', 'income_drop', 'bonus_reliance'},
    'traj_nw_down': {'negative_balance', 'balance_decline', 'income_drop'},
    'traj_high_installments': {'installment_burden', 'installment_avoid'},
    'traj_saving': {'good_savings_rate'},
    'traj_passive_growing': {'passive_income_good'},
    'traj_expenses_rising': {'category_spike', 'food_spending', 'entertainment_high', 'dining_out_high', 'subscriptions_cost'},
    'traj_expenses_falling': {'food_spending', 'entertainment_high', 'dining_out_high'},
    'traj_passive_declining': {'passive_income_low'},
    'traj_income_risk': {'income_drop', 'bonus_reliance'},
    'traj_income_stable': {'income_stable'},
}


def _boost_positive(tips, trajectory_type):
    """When trajectory is positive and no important tips exist, boost positive tips."""
    C = TIP_CONFIG
    has_important = any(t['severity'] == 'important' for t in tips)
    if has_important:
        return tips

    boost = 0.0
    if trajectory_type == 'positive':
        boost = C['boost_positive_traj']
    elif trajectory_type == 'mixed':
        boost = C['boost_mixed_traj']

    if boost > 0:
        for tip in tips:
            if tip['id'] in C['positive_tip_ids']:
                tip['score'] = round(min(tip['score'] + boost, 1.0), 2)
        tips.sort(key=lambda t: t['score'], reverse=True)

    return tips


@app.route('/api/tips', methods=['GET'])
@login_required
def get_tips():
    conn = get_db()
    month = request.args.get('month', date.today().strftime('%Y-%m'))
    trajectory = request.args.get('trajectory', '')  # positive|negative|mixed
    trajectory_signals = request.args.get('signals', '')  # comma-separated signal keys
    uid = get_uid()

    # Build set of tip IDs that overlap with active trajectory signals
    overlapping_ids = set()
    if trajectory_signals:
        for sig in trajectory_signals.split(','):
            overlapping_ids |= _TRAJECTORY_TIP_OVERLAP.get(sig.strip(), set())

    ctx = _build_tip_context(conn, uid, month)

    # Run all generators
    tips = []
    for gen in _TIP_GENERATORS:
        try:
            tip = gen(ctx)
            if tip:
                tip['score'] = _score_tip(tip, overlapping_ids)
                tips.append(tip)
        except Exception:
            pass  # Individual tip failure should not crash the endpoint

    # Sort by score descending, then deduplicate
    tips.sort(key=lambda t: t['score'], reverse=True)
    tips = _deduplicate_tips(tips)

    # Positive reinforcement boost when environment is calm
    if trajectory in ('positive', 'mixed'):
        tips = _boost_positive(tips, trajectory)

    conn.close()
    return jsonify(tips)


# ---- Next Best Action engine ----
# NBA i18n key map: tip_id → nba recommendation key prefix
# Only tips with NBA keys become candidates
_NBA_CANDIDATES = {
    'subscriptions_cost', 'subscriptions_growth', 'dining_out_high',
    'overdraft_interest', 'negative_balance', 'installment_burden',
    'installment_avoid', 'deficit_months', 'low_emergency_buffer',
    'food_spending', 'entertainment_high', 'category_spike',
    'low_savings_rate', 'high_fixed_ratio',
    'vacation_days_unused', 'bonus_detected',
    'true_savings_rate', 'employer_contributions_value',
    'bonus_reliance', 'income_drop',
}


@app.route('/api/next-action', methods=['GET'])
@login_required
def get_next_action():
    """Return the single highest-priority actionable recommendation."""
    conn = get_db()
    month = request.args.get('month', date.today().strftime('%Y-%m'))
    uid = get_uid()
    ctx = _build_tip_context(conn, uid, month)

    # Run generators, score, pick best actionable candidate
    best = None
    for gen in _TIP_GENERATORS:
        try:
            tip = gen(ctx)
            if not tip or tip['id'] not in _NBA_CANDIDATES:
                continue
            if not tip.get('action'):
                continue
            tip['score'] = _score_tip(tip)
            if not best or tip['score'] > best['score']:
                best = tip
        except Exception:
            pass

    conn.close()
    if not best or best['score'] < 0.30:
        return jsonify(None)

    # Build reason chips from params — priority-ordered, max 3 shown
    reasons = []
    p = best.get('params', {})
    if p.get('new_count') and p['new_count'] != '0':
        reasons.append({'key': 'nba_reason_new_services', 'params': {'count': p['new_count']}})
    if p.get('vs_avg'):
        reasons.append({'key': 'nba_reason_vs_avg', 'params': {'vs_avg': p['vs_avg']}})
    if p.get('pct_of_income') and float(p['pct_of_income']) >= 3:
        reasons.append({'key': 'nba_reason_income_pct', 'params': {'pct': p['pct_of_income']}})
    if p.get('concentration_pct'):
        reasons.append({'key': 'nba_reason_concentration', 'params': {
            'pct': p['concentration_pct'], 'vendor': p.get('top_vendor', '')}})
    if p.get('count') and p['count'] != '0':
        reasons.append({'key': 'nba_reason_services', 'params': {'count': p['count']}})
    # Salary-specific reason chips
    if p.get('days'):
        reasons.append({'key': 'nba_reason_vacation_balance', 'params': {'days': p['days']}})
    if p.get('amount') and best['id'] == 'bonus_detected':
        reasons.append({'key': 'nba_reason_bonus_received', 'params': {'amount': p['amount']}})
    if p.get('monthly') and best['id'] == 'employer_contributions_value':
        reasons.append({'key': 'nba_reason_employer_savings', 'params': {'monthly': p['monthly'], 'annual': p['annual']}})
    if p.get('rate') and best['id'] == 'true_savings_rate':
        reasons.append({'key': 'nba_reason_savings_vs_gross', 'params': {'rate': p['rate']}})
    if p.get('pct') and best['id'] == 'bonus_reliance':
        reasons.append({'key': 'nba_reason_bonus_pct', 'params': {'pct': p['pct']}})
    if p.get('drop_pct') and best['id'] == 'income_drop':
        reasons.append({'key': 'nba_reason_income_drop', 'params': {'drop_pct': p['drop_pct']}})
    reasons = reasons[:3]

    return jsonify({
        'tip_id': best['id'],
        'action': best['action'],
        'action_label_key': best.get('action_label_key', ''),
        'icon': best.get('icon', 'bi-lightbulb'),
        'color': best.get('color', '#2563eb'),
        'severity': best.get('severity', 'watch'),
        'score': best['score'],
        'params': p,
        'reasons': reasons,
    })


@app.route('/api/tip-events', methods=['POST'])
@login_required
def post_tip_events():
    data = request.json or {}
    events = data.get('events', [])
    if not events:
        return jsonify({'status': 'ok', 'count': 0})
    uid = get_uid()
    conn = get_db()
    valid_types = {'shown', 'expanded', 'dismissed', 'action',
                   'payslip_upload', 'payslip_preview_shown', 'payslip_field_edited',
                   'payslip_saved', 'payslip_cancelled'}
    count = 0
    for ev in events:
        etype = ev.get('event_type', '')
        tid = ev.get('tip_id', '')
        if etype not in valid_types or not tid:
            continue
        conn.execute(
            "INSERT INTO tip_events (user_id, tip_id, event_type, action_target, month) VALUES (?,?,?,?,?)",
            (uid, tid, etype, ev.get('action_target'), ev.get('month'))
        )
        count += 1
    # Prune old entries
    conn.execute(
        "DELETE FROM tip_events WHERE user_id=? AND created_at < datetime('now', ?)",
        (uid, f"-{TIP_CONFIG['analytics_prune_days']} days")
    )
    conn.commit()
    conn.close()
    return jsonify({'status': 'ok', 'count': count})


@app.route('/api/payslip-analytics', methods=['GET'])
@login_required
def payslip_analytics():
    """Aggregate payslip import flow metrics from tip_events."""
    conn = get_db()
    uid = get_uid()
    rows = conn.execute(
        "SELECT event_type, COUNT(*) as cnt FROM tip_events "
        "WHERE user_id=? AND event_type LIKE 'payslip_%' GROUP BY event_type",
        (uid,)
    ).fetchall()
    stats = {dict(r)['event_type']: dict(r)['cnt'] for r in rows}
    uploads = stats.get('payslip_upload', 0)
    previews = stats.get('payslip_preview_shown', 0)
    saves = stats.get('payslip_saved', 0)
    cancels = stats.get('payslip_cancelled', 0)
    edits = stats.get('payslip_field_edited', 0)
    # Per-field edit breakdown
    field_rows = conn.execute(
        "SELECT action_target, COUNT(*) as cnt FROM tip_events "
        "WHERE user_id=? AND event_type='payslip_field_edited' AND action_target IS NOT NULL "
        "GROUP BY action_target ORDER BY cnt DESC",
        (uid,)
    ).fetchall()
    field_edits_by_field = {dict(r)['action_target']: dict(r)['cnt'] for r in field_rows}
    most_edited = max(field_edits_by_field, key=field_edits_by_field.get) if field_edits_by_field else None
    # Per-company extraction confidence
    company_rows = conn.execute(
        "SELECT company_name, COUNT(*) as cnt, AVG(extraction_confidence) as avg_conf "
        "FROM salary_statements WHERE user_id=? AND company_name != '' "
        "GROUP BY company_name ORDER BY avg_conf ASC",
        (uid,)
    ).fetchall()
    companies = [{'name': dict(r)['company_name'], 'count': dict(r)['cnt'],
                  'avg_confidence': round(dict(r)['avg_conf'] or 0, 2)} for r in company_rows]
    conn.close()
    return jsonify({
        'uploads': uploads, 'previews': previews,
        'saves': saves, 'cancels': cancels, 'field_edits': edits,
        'save_rate': round(saves / previews * 100, 1) if previews > 0 else 0,
        'edit_rate': round(1.0 if edits > 0 and previews > 0 else 0.0, 2),
        'field_edits_by_field': field_edits_by_field,
        'most_edited_field': most_edited,
        'companies': companies,
    })


# ---- Budget Agent - Deep Analysis ----

@app.route('/api/analyze', methods=['GET'])
@login_required
def analyze_budget():
    """Deep analysis engine that processes all expenses and generates conclusions."""
    conn = get_db()
    uid = get_uid()

    # Gather all data (normalized — excludes unusual + excluded months)
    excl_months = get_excluded_month_set(conn, uid)
    months_data_raw = conn.execute("""
        SELECT m,
               exp_total,
               COALESCE((SELECT SUM(amount) FROM income WHERE user_id=? AND substr(date,1,7)=m), 0) as inc_total
        FROM (SELECT substr(date,1,7) as m, SUM(amount) as exp_total
              FROM expenses WHERE user_id=? AND is_unusual=0 GROUP BY m ORDER BY m)
    """, (uid, uid)).fetchall()
    months_data = [r for r in months_data_raw if r[0] not in excl_months]

    categories = conn.execute("""
        SELECT c.name_he, c.id, SUM(e.amount) as total,
               COUNT(DISTINCT substr(e.date,1,7)) as months_active
        FROM expenses e JOIN categories c ON e.category_id=c.id
        WHERE e.user_id=? AND e.is_unusual=0
        GROUP BY c.id ORDER BY total DESC
    """, (uid,)).fetchall()

    num_months = len(months_data)
    total_income = sum(r[2] for r in months_data)
    total_expense = sum(r[1] for r in months_data)
    avg_income = total_income / num_months if num_months else 0
    avg_expense = total_expense / num_months if num_months else 0

    # Fixed vs variable (normalized)
    fixed = conn.execute(
        "SELECT SUM(amount) FROM expenses WHERE user_id=? AND is_unusual=0 AND frequency='monthly'", (uid,)
    ).fetchone()[0] or 0
    fixed_monthly = fixed / num_months if num_months else 0

    variable = conn.execute(
        "SELECT SUM(amount) FROM expenses WHERE user_id=? AND is_unusual=0 AND frequency='random'", (uid,)
    ).fetchone()[0] or 0
    variable_monthly = variable / num_months if num_months else 0

    # Category averages
    cat_analysis = []
    for cat in categories:
        monthly_avg = cat[2] / cat[3] if cat[3] else 0
        pct = cat[2] / total_expense * 100 if total_expense else 0
        cat_analysis.append({
            'name': cat[0],
            'id': cat[1],
            'total': cat[2],
            'monthly_avg': monthly_avg,
            'percentage': pct,
        })

    # Budget health score (0-100)
    score = 50  # base
    savings_rate = 0

    # Positive factors
    savings_exp = conn.execute(
        "SELECT COALESCE(SUM(amount),0) FROM expenses WHERE user_id=? AND is_unusual=0 AND category_id='savings'", (uid,)
    ).fetchone()[0]
    if total_income > 0:
        savings_rate = savings_exp / total_income * 100
        if savings_rate >= 10:
            score += 10
        if savings_rate >= 20:
            score += 10

    # Surplus months
    surplus_months = sum(1 for r in months_data if r[2] > r[1])
    deficit_months = num_months - surplus_months
    score += min(surplus_months * 3, 15)
    score -= min(deficit_months * 3, 15)

    # Overdraft penalty (normalized)
    overdraft = conn.execute(
        "SELECT COALESCE(SUM(amount),0) FROM expenses WHERE user_id=? AND is_unusual=0 AND subcategory='ריבית מינוס'", (uid,)
    ).fetchone()[0]
    if overdraft > 0:
        score -= 10

    # Fixed expenses ratio
    if avg_income > 0:
        fixed_ratio = fixed_monthly / avg_income * 100
        if fixed_ratio < 50:
            score += 5
        elif fixed_ratio > 70:
            score -= 10

    # Bank balance trend factor (max ±10)
    bal_trend = conn.execute(
        "SELECT closing_balance FROM bank_balances WHERE user_id=? AND account_name='main' ORDER BY month DESC LIMIT 3",
        (uid,)
    ).fetchall()
    if len(bal_trend) >= 2:
        recent_bal = bal_trend[0]['closing_balance']
        if recent_bal > 0:
            score += 5
        if recent_bal < 0:
            score -= 10
        # Compare first vs last chronologically (oldest vs newest)
        bal_chrono = [r['closing_balance'] for r in reversed(bal_trend)]
        oldest, newest = bal_chrono[0], bal_chrono[-1]
        # Declining: newest balance lower than oldest
        if oldest > 0 and newest < oldest * 0.90:
            score -= 5
        # Improving: newest balance higher than oldest
        elif newest > oldest * 1.10:
            score += 5

    score = max(0, min(100, score))

    # Recommended budget (50/30/20 rule adapted)
    recommended = {}
    if avg_income > 0:
        recommended = {
            'needs': avg_income * 0.50,  # housing, food, transport, insurance
            'wants': avg_income * 0.30,  # entertainment, personal, dining out
            'savings': avg_income * 0.20,  # savings, investments, emergency fund
            'needs_categories': ['דיור ואחזקת בית', 'מזון', 'רכב', 'ביטוחים', 'ריפוי', 'תקשורת'],
            'wants_categories': ['בילוי ופנאי', 'אישי', 'טיפוח ובריאות', 'ילדים'],
            'savings_categories': ['חיסכון והתחייבויות'],
        }

    # Generate textual analysis
    analysis_sections = []

    # Overall health
    if score >= 70:
        health_text = 'המצב הכלכלי שלכם סביר. יש בסיס טוב לעבוד איתו.'
    elif score >= 50:
        health_text = 'יש מקום לשיפור משמעותי. כמה שינויים קטנים יכולים לעשות הבדל גדול.'
    else:
        health_text = 'המצב דורש תשומת לב מיידית. חשוב להתחיל בשינויים היום.'

    analysis_sections.append({
        'title': 'בריאות תקציבית',
        'icon': 'bi-heart-pulse',
        'text': health_text,
        'score': score,
    })

    # Income vs Expense
    if avg_income > avg_expense:
        ie_text = f'בממוצע אתם מרוויחים {avg_income-avg_expense:,.0f} ש"ח יותר ממה שאתם מוציאים. מצוין!'
    else:
        ie_text = f'בממוצע אתם מוציאים {avg_expense-avg_income:,.0f} ש"ח יותר ממה שאתם מרוויחים. זה חייב להשתנות.'

    analysis_sections.append({
        'title': 'הכנסות מול הוצאות',
        'icon': 'bi-arrow-left-right',
        'text': f'הכנסה ממוצעת: {avg_income:,.0f} ש"ח | הוצאה ממוצעת: {avg_expense:,.0f} ש"ח\n{ie_text}',
    })

    # Fixed vs Variable
    analysis_sections.append({
        'title': 'הוצאות קבועות מול משתנות',
        'icon': 'bi-pie-chart',
        'text': f'הוצאות קבועות: {fixed_monthly:,.0f} ש"ח/חודש ({fixed_monthly/avg_expense*100:.0f}% מההוצאות)\n'
                f'הוצאות משתנות: {variable_monthly:,.0f} ש"ח/חודש ({variable_monthly/avg_expense*100:.0f}% מההוצאות)\n\n'
                + ('ההוצאות הקבועות גבוהות מדי (מעל 60%). נסו לצמצם חיובים חודשיים.' if fixed_monthly/avg_expense > 0.6
                   else 'יחס סביר בין הוצאות קבועות למשתנות.'),
    })

    # Top 3 expense categories
    top3 = cat_analysis[:3]
    top3_text = '\n'.join([f'{i+1}. **{c["name"]}**: {c["monthly_avg"]:,.0f} ש"ח/חודש ({c["percentage"]:.1f}%)' for i, c in enumerate(top3)])
    analysis_sections.append({
        'title': 'הקטגוריות הגדולות',
        'icon': 'bi-bar-chart',
        'text': f'שלוש הקטגוריות הכי יקרות:\n{top3_text}',
    })

    # Investment recommendations
    invest_text = ''
    if avg_income > 0:
        monthly_surplus = avg_income - avg_expense
        if monthly_surplus > 2000:
            invest_text = (
                f'יש לכם עודף ממוצע של {monthly_surplus:,.0f} ש"ח/חודש. הנה מה לעשות איתו:\n\n'
                '1. **חיסכון חירום** (עדיפות ראשונה) - 3 חודשי הוצאות בפיקדון/ממ"ש\n'
                '2. **קרן השתלמות** - אם לא ממקסמים, השלימו. הטבת מס משמעותית\n'
                '3. **קופת גמל להשקעה** - עד 70,000 ש"ח/שנה, הטבת מס במשיכה\n'
                '4. **תיק השקעות** - קרן מחקה מדד (S&P500 / ת"א 125) לטווח ארוך'
            )
        elif monthly_surplus > 0:
            invest_text = (
                f'העודף הממוצע שלכם ({monthly_surplus:,.0f} ש"ח/חודש) קטן מדי להשקעה משמעותית.\n'
                'קודם כל צמצמו הוצאות כדי להגדיל את העודף ל-2,000 ש"ח לפחות.'
            )
        else:
            invest_text = (
                'כרגע אתם בגירעון - אין מקום להשקעות.\n'
                'שלב 1: צמצמו הוצאות והגיעו לאיזון.\n'
                'שלב 2: בנו חיסכון חירום.\n'
                'שלב 3: רק אז תתחילו להשקיע.'
            )
    analysis_sections.append({
        'title': 'המלצות השקעה',
        'icon': 'bi-graph-up-arrow',
        'text': invest_text,
    })

    # Bank balance section (only if data exists)
    if bal_trend:
        latest = bal_trend[0]['closing_balance']
        bal_text = f'יתרת חשבון סוף חודש: {latest:,.0f} ₪.\n'
        if len(bal_trend) >= 2:
            change = bal_trend[0]['closing_balance'] - bal_trend[-1]['closing_balance']
            direction = 'עלתה' if change > 0 else 'ירדה'
            bal_text += f'היתרה {direction} ב-{abs(change):,.0f} ₪ ב-{len(bal_trend)} חודשים אחרונים.\n'
        bal_text += '\nהערה: הנתון משקף את היתרה בעת העסקה האחרונה שיובאה ועלול להשתנות עקב תזמון משכורת, חיובי כרטיס ועוד.'
        analysis_sections.append({
            'title': 'יתרת חשבון',
            'icon': 'bi-bank',
            'text': bal_text,
        })

    conn.close()

    return jsonify({
        'score': score,
        'avg_income': avg_income,
        'avg_expense': avg_expense,
        'savings_rate': savings_rate,
        'num_months': num_months,
        'surplus_months': surplus_months,
        'deficit_months': deficit_months,
        'fixed_monthly': fixed_monthly,
        'variable_monthly': variable_monthly,
        'categories': cat_analysis,
        'recommended': recommended,
        'sections': analysis_sections,
    })


# ---- Export to Excel ----

FREQ_HE = {'monthly': 'חודשי', 'bimonthly': 'דו-חודשי', 'random': 'לא קבוע'}
SOURCE_HE = {'manual': 'ידני', 'visa_import': 'ויזה', 'xls_import': 'XLS', 'bank_csv': 'בנק'}
PERSON_HE = {'husband': 'בעל', 'wife': 'אישה', 'family': 'משפחה', 'other': 'אחר'}
INCOME_SRC_HE = {'salary': 'משכורת', 'bonus': 'בונוס', 'freelance': 'פרילנס',
                 'child_allowance': 'קצבת ילדים', 'rental': 'שכירות', 'other': 'אחר'}


def _style_header(ws, col_count):
    """Apply header styling to first row."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    thin_border = Border(bottom=Side(style='thin', color='CCCCCC'))
    for col in range(1, col_count + 1):
        cell = ws.cell(1, col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border


def _auto_width(ws):
    """Auto-fit column widths."""
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val = str(cell.value or '')
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


def _add_number_format(ws, col_idx, start_row, end_row, fmt='#,##0.00'):
    """Apply number format to a column range."""
    for row in range(start_row, end_row + 1):
        ws.cell(row, col_idx).number_format = fmt


@app.route('/api/export', methods=['GET'])
@login_required
def export_excel():
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import PieChart, BarChart, Reference, LineChart
    from openpyxl.utils import get_column_letter

    month = request.args.get('month', date.today().strftime('%Y-%m'))
    conn = get_db()
    uid = get_uid()

    wb = openpyxl.Workbook()

    # ========== Sheet 1: Summary ==========
    ws_sum = wb.active
    ws_sum.title = 'סיכום'
    ws_sum.sheet_view.rightToLeft = True

    # Get summary data
    income_total = conn.execute(
        "SELECT COALESCE(SUM(amount),0) FROM income WHERE user_id=? AND date LIKE ?", (uid, month+'%')
    ).fetchone()[0]
    expense_total = conn.execute(
        "SELECT COALESCE(SUM(amount),0) FROM expenses WHERE user_id=? AND date LIKE ?", (uid, month+'%')
    ).fetchone()[0]

    title_font = Font(bold=True, size=14, color='1E40AF')
    ws_sum.cell(1, 1, 'דוח תקציב משפחתי').font = title_font
    ws_sum.cell(2, 1, f'חודש: {month}').font = Font(size=12, color='64748B')

    ws_sum.cell(4, 1, 'סה"כ הכנסות').font = Font(bold=True)
    ws_sum.cell(4, 2, income_total).number_format = '#,##0.00'
    ws_sum.cell(5, 1, 'סה"כ הוצאות').font = Font(bold=True)
    ws_sum.cell(5, 2, expense_total).number_format = '#,##0.00'
    ws_sum.cell(6, 1, 'יתרה').font = Font(bold=True, color='16A34A' if income_total >= expense_total else 'DC2626')
    ws_sum.cell(6, 2, income_total - expense_total).number_format = '#,##0.00'

    # By category summary
    ws_sum.cell(8, 1, 'הוצאות לפי קטגוריה').font = Font(bold=True, size=12)
    ws_sum.cell(9, 1, 'קטגוריה')
    ws_sum.cell(9, 2, 'סכום')
    ws_sum.cell(9, 3, 'אחוז')
    _style_header_row(ws_sum, 9, 3)

    cat_rows = conn.execute(
        """SELECT c.name_he, COALESCE(SUM(e.amount),0) as total
           FROM categories c LEFT JOIN expenses e ON c.id=e.category_id AND e.date LIKE ? AND e.user_id=?
           GROUP BY c.id HAVING total>0 ORDER BY total DESC""",
        (month+'%', uid)
    ).fetchall()

    for i, r in enumerate(cat_rows):
        row = 10 + i
        ws_sum.cell(row, 1, r[0])
        ws_sum.cell(row, 2, r[1]).number_format = '#,##0.00'
        ws_sum.cell(row, 3, r[1]/expense_total if expense_total else 0).number_format = '0.0%'

    # By frequency
    freq_start = 10 + len(cat_rows) + 2
    ws_sum.cell(freq_start, 1, 'הוצאות לפי תדירות').font = Font(bold=True, size=12)
    ws_sum.cell(freq_start+1, 1, 'תדירות')
    ws_sum.cell(freq_start+1, 2, 'סכום')
    ws_sum.cell(freq_start+1, 3, 'מספר פריטים')
    _style_header_row(ws_sum, freq_start+1, 3)

    freq_rows = conn.execute(
        """SELECT frequency, SUM(amount), COUNT(*) FROM expenses
           WHERE user_id=? AND date LIKE ? GROUP BY frequency ORDER BY SUM(amount) DESC""",
        (uid, month+'%')
    ).fetchall()
    for i, r in enumerate(freq_rows):
        row = freq_start + 2 + i
        ws_sum.cell(row, 1, FREQ_HE.get(r[0], r[0]))
        ws_sum.cell(row, 2, r[1]).number_format = '#,##0.00'
        ws_sum.cell(row, 3, r[2])

    # By card
    card_start = freq_start + 2 + len(freq_rows) + 2
    ws_sum.cell(card_start, 1, 'הוצאות לפי כרטיס').font = Font(bold=True, size=12)
    ws_sum.cell(card_start+1, 1, 'כרטיס')
    ws_sum.cell(card_start+1, 2, 'סכום')
    ws_sum.cell(card_start+1, 3, 'מספר פריטים')
    _style_header_row(ws_sum, card_start+1, 3)

    card_rows = conn.execute(
        """SELECT CASE WHEN card='' THEN 'אחר' ELSE card END, SUM(amount), COUNT(*)
           FROM expenses WHERE user_id=? AND date LIKE ? GROUP BY card ORDER BY SUM(amount) DESC""",
        (uid, month+'%')
    ).fetchall()
    for i, r in enumerate(card_rows):
        row = card_start + 2 + i
        ws_sum.cell(row, 1, r[0])
        ws_sum.cell(row, 2, r[1]).number_format = '#,##0.00'
        ws_sum.cell(row, 3, r[2])

    # Income by person
    inc_start = card_start + 2 + len(card_rows) + 2
    ws_sum.cell(inc_start, 1, 'הכנסות לפי בן/בת זוג').font = Font(bold=True, size=12)
    ws_sum.cell(inc_start+1, 1, 'בן/בת זוג')
    ws_sum.cell(inc_start+1, 2, 'סכום')
    _style_header_row(ws_sum, inc_start+1, 2)

    inc_person_rows = conn.execute(
        """SELECT person, SUM(amount) FROM income
           WHERE user_id=? AND date LIKE ? GROUP BY person ORDER BY SUM(amount) DESC""",
        (uid, month+'%')
    ).fetchall()
    for i, r in enumerate(inc_person_rows):
        row = inc_start + 2 + i
        ws_sum.cell(row, 1, PERSON_HE.get(r[0], r[0]))
        ws_sum.cell(row, 2, r[1]).number_format = '#,##0.00'

    # Bank balance & cashflow section
    bal_start = inc_start + 2 + len(inc_person_rows) + 2
    bank_bal_row = conn.execute(
        "SELECT closing_balance FROM bank_balances WHERE user_id=? AND account_name='main' AND month=?",
        (uid, month)
    ).fetchone()
    ws_sum.cell(bal_start, 1, 'תזרים ויתרה').font = Font(bold=True, size=12)
    ws_sum.cell(bal_start + 1, 1, 'תזרים מזומנים נטו').font = Font(bold=True)
    ws_sum.cell(bal_start + 1, 2, income_total - expense_total).number_format = '#,##0.00'
    if bank_bal_row:
        ws_sum.cell(bal_start + 2, 1, 'יתרת חשבון סוף חודש').font = Font(bold=True)
        ws_sum.cell(bal_start + 2, 2, bank_bal_row['closing_balance']).number_format = '#,##0.00'

    _auto_width(ws_sum)

    # ========== Sheet 2: All Expenses ==========
    ws_exp = wb.create_sheet('הוצאות')
    ws_exp.sheet_view.rightToLeft = True
    exp_headers = ['תאריך', 'קטגוריה', 'תת-קטגוריה', 'תיאור', 'סכום', 'תדירות', 'כרטיס', 'מקור']
    for i, h in enumerate(exp_headers, 1):
        ws_exp.cell(1, i, h)
    _style_header(ws_exp, len(exp_headers))

    expenses = conn.execute(
        """SELECT e.date, c.name_he, e.subcategory, e.description, e.amount,
                  e.frequency, e.card, e.source
           FROM expenses e JOIN categories c ON e.category_id=c.id
           WHERE e.user_id=? AND e.date LIKE ? ORDER BY e.date DESC""",
        (uid, month+'%')
    ).fetchall()
    for i, r in enumerate(expenses):
        row = i + 2
        ws_exp.cell(row, 1, r[0])
        ws_exp.cell(row, 2, r[1])
        ws_exp.cell(row, 3, r[2] or '')
        ws_exp.cell(row, 4, r[3] or '')
        ws_exp.cell(row, 5, r[4]).number_format = '#,##0.00'
        ws_exp.cell(row, 6, FREQ_HE.get(r[5], r[5] or ''))
        ws_exp.cell(row, 7, r[6] or '')
        ws_exp.cell(row, 8, SOURCE_HE.get(r[7], r[7] or ''))
    _auto_width(ws_exp)

    # ========== Sheet 3: Income ==========
    ws_inc = wb.create_sheet('הכנסות')
    ws_inc.sheet_view.rightToLeft = True
    inc_headers = ['תאריך', 'בן/בת זוג', 'מקור', 'סכום', 'תיאור', 'חוזר']
    for i, h in enumerate(inc_headers, 1):
        ws_inc.cell(1, i, h)
    _style_header(ws_inc, len(inc_headers))

    incomes = conn.execute(
        "SELECT date, person, source, amount, description, is_recurring FROM income WHERE user_id=? AND date LIKE ? ORDER BY date DESC",
        (uid, month+'%')
    ).fetchall()
    for i, r in enumerate(incomes):
        row = i + 2
        ws_inc.cell(row, 1, r[0])
        ws_inc.cell(row, 2, PERSON_HE.get(r[1], r[1]))
        ws_inc.cell(row, 3, INCOME_SRC_HE.get(r[2], r[2]))
        ws_inc.cell(row, 4, r[3]).number_format = '#,##0.00'
        ws_inc.cell(row, 5, r[4] or '')
        ws_inc.cell(row, 6, 'כן' if r[5] else 'לא')
    _auto_width(ws_inc)

    # ========== Sheet 4: Daily Breakdown ==========
    ws_daily = wb.create_sheet('יומי')
    ws_daily.sheet_view.rightToLeft = True
    daily_headers = ['תאריך', 'סה"כ הוצאות', 'מספר עסקאות']
    for i, h in enumerate(daily_headers, 1):
        ws_daily.cell(1, i, h)
    _style_header(ws_daily, len(daily_headers))

    daily_rows = conn.execute(
        """SELECT date, SUM(amount), COUNT(*) FROM expenses
           WHERE user_id=? AND date LIKE ? GROUP BY date ORDER BY date""",
        (uid, month+'%')
    ).fetchall()
    for i, r in enumerate(daily_rows):
        row = i + 2
        ws_daily.cell(row, 1, r[0])
        ws_daily.cell(row, 2, r[1]).number_format = '#,##0.00'
        ws_daily.cell(row, 3, r[2])
    _auto_width(ws_daily)

    # ========== Sheet 5: Budget vs Actual ==========
    ws_bud = wb.create_sheet('תקציב מול ביצוע')
    ws_bud.sheet_view.rightToLeft = True
    bud_headers = ['קטגוריה', 'מתוכנן', 'בפועל', 'הפרש', 'אחוז ניצול']
    for i, h in enumerate(bud_headers, 1):
        ws_bud.cell(1, i, h)
    _style_header(ws_bud, len(bud_headers))

    bud_rows = conn.execute(
        """SELECT c.name_he, COALESCE(b.planned_amount,0), COALESCE(SUM(e.amount),0)
           FROM categories c
           LEFT JOIN budget b ON c.id=b.category_id AND b.month=? AND b.user_id=?
           LEFT JOIN expenses e ON c.id=e.category_id AND e.date LIKE ? AND e.user_id=?
           GROUP BY c.id HAVING COALESCE(b.planned_amount,0)>0 OR COALESCE(SUM(e.amount),0)>0
           ORDER BY c.sort_order""",
        (month, uid, month+'%', uid)
    ).fetchall()

    over_fmt = Font(color='DC2626', bold=True)
    under_fmt = Font(color='16A34A')
    for i, r in enumerate(bud_rows):
        row = i + 2
        planned, actual = r[1], r[2]
        diff = planned - actual
        ws_bud.cell(row, 1, r[0])
        ws_bud.cell(row, 2, planned).number_format = '#,##0.00'
        ws_bud.cell(row, 3, actual).number_format = '#,##0.00'
        c = ws_bud.cell(row, 4, diff)
        c.number_format = '#,##0.00'
        c.font = over_fmt if diff < 0 else under_fmt
        ws_bud.cell(row, 5, actual/planned if planned else 0).number_format = '0.0%'
    _auto_width(ws_bud)

    # ========== Sheet 6: Monthly Trend ==========
    ws_trend = wb.create_sheet('מגמה חודשית')
    ws_trend.sheet_view.rightToLeft = True
    trend_headers = ['חודש', 'הוצאות', 'הכנסות', 'יתרה']
    for i, h in enumerate(trend_headers, 1):
        ws_trend.cell(1, i, h)
    _style_header(ws_trend, len(trend_headers))

    excl_months = get_excluded_month_set(conn, uid)
    trend_rows_raw = conn.execute(
        """SELECT m, exp_total,
                  COALESCE((SELECT SUM(amount) FROM income WHERE user_id=? AND substr(date,1,7)=m), 0),
                  COALESCE((SELECT SUM(amount) FROM income WHERE user_id=? AND substr(date,1,7)=m), 0) - exp_total
           FROM (SELECT substr(date,1,7) as m, SUM(amount) as exp_total
                 FROM expenses WHERE user_id=? AND is_unusual=0 GROUP BY m ORDER BY m)""",
        (uid, uid, uid)
    ).fetchall()
    trend_rows = [r for r in trend_rows_raw if r[0] not in excl_months]
    for i, r in enumerate(trend_rows):
        row = i + 2
        ws_trend.cell(row, 1, r[0])
        ws_trend.cell(row, 2, r[1]).number_format = '#,##0.00'
        ws_trend.cell(row, 3, r[2]).number_format = '#,##0.00'
        c = ws_trend.cell(row, 4, r[3])
        c.number_format = '#,##0.00'
        c.font = Font(color='16A34A' if r[3] >= 0 else 'DC2626', bold=True)
    _auto_width(ws_trend)

    conn.close()

    # Save to BytesIO and send
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'budget_report_{month}.xlsx'
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)


def _style_header_row(ws, row_num, col_count):
    """Apply header styling to a specific row."""
    from openpyxl.styles import Font, PatternFill, Alignment
    header_font = Font(bold=True, color='FFFFFF', size=10)
    header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    for col in range(1, col_count + 1):
        cell = ws.cell(row_num, col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')


# ---- Financial Products (Insurance, Pension, Funds) ----

# ============================================================
# Insurance Suggestion Detection Engine
# ============================================================

# Insurer name patterns: (pattern, insurer_display_name, default_type_guess, market)
INSURANCE_INSURER_PATTERNS = [
    # Israeli insurers
    ('הראל', 'הראל', 'general', 'IL'),
    ('מנורה', 'מנורה מבטחים', 'general', 'IL'),
    ('כלל', 'כלל ביטוח', 'general', 'IL'),
    ('מגדל', 'מגדל', 'general', 'IL'),
    ('הפניקס', 'הפניקס', 'general', 'IL'),
    ('איילון', 'איילון', 'general', 'IL'),
    ('שומרה', 'שומרה', 'general', 'IL'),
    ('ביטוח ישיר', 'ביטוח ישיר', 'car', 'IL'),
    ('שירביט', 'שירביט', 'general', 'IL'),
    ('wesure', 'WeSure', 'general', 'IL'),
    ('וישור', 'WeSure', 'general', 'IL'),
    ('ליברה', 'ליברה', 'general', 'IL'),
    ('libra', 'Libra', 'general', 'IL'),
    # US insurers
    ('state farm', 'State Farm', 'auto', 'US'),
    ('statefarm', 'State Farm', 'auto', 'US'),
    ('geico', 'GEICO', 'auto', 'US'),
    ('progressive', 'Progressive', 'auto', 'US'),
    ('allstate', 'Allstate', 'auto', 'US'),
    ('liberty mutual', 'Liberty Mutual', 'auto', 'US'),
    ('libertymutual', 'Liberty Mutual', 'auto', 'US'),
    ('farmers ins', 'Farmers Insurance', 'auto', 'US'),
    ('nationwide', 'Nationwide', 'general', 'US'),
    ('usaa', 'USAA', 'auto', 'US'),
    ('aetna', 'Aetna', 'health', 'US'),
    ('cigna', 'Cigna', 'health', 'US'),
    ('blue cross', 'Blue Cross Blue Shield', 'health', 'US'),
    ('bluecross', 'Blue Cross Blue Shield', 'health', 'US'),
    ('bcbs', 'Blue Cross Blue Shield', 'health', 'US'),
    ('anthem', 'Anthem', 'health', 'US'),
    ('unitedhealth', 'UnitedHealthcare', 'health', 'US'),
    ('united health', 'UnitedHealthcare', 'health', 'US'),
    ('metlife', 'MetLife', 'life', 'US'),
    ('new york life', 'New York Life', 'life', 'US'),
    ('prudential', 'Prudential', 'life', 'US'),
    ('travelers', 'Travelers', 'homeowners', 'US'),
    ('hartford', 'The Hartford', 'auto', 'US'),
    ('lemonade', 'Lemonade', 'renters', 'US'),
    # International (both markets)
    ('AIG', 'AIG', 'general', 'both'),
    ('aig', 'AIG', 'general', 'both'),
]

# Strong keywords — standalone insurance signal
INSURANCE_STRONG_KEYWORDS = [
    'ביטוח', 'insurance', 'פוליס',
    'premium', 'ins prem', 'ins pmnt',
]

# Weak keywords — only boost confidence when combined with other signals
INSURANCE_WEAK_KEYWORDS = [
    # Hebrew
    ('חיים', 'life'),
    ('בריאות', 'health'),
    ('רכב', 'car'),
    ('דירה', 'home'),
    ('משכנתא', 'home'),  # mortgage insurance
    # English
    ('auto ins', 'auto'),
    ('car ins', 'auto'),
    ('home ins', 'homeowners'),
    ('homeowner', 'homeowners'),
    ('renter', 'renters'),
    ('dental', 'dental'),
    ('vision', 'vision'),
    ('life ins', 'life'),
    ('umbrella', 'umbrella'),
    ('disability', 'disability'),
    ('pet ins', 'pet'),
    ('liability', 'general'),
    ('deductible', 'general'),
    ('coverage', 'general'),
    ('policy', 'general'),
]

# ACH / autopay patterns that boost insurance detection when combined with insurer name
INSURANCE_ACH_PATTERNS = [
    'ach', 'autopay', 'auto pay', 'auto-pay', 'direct debit',
    'recurring pmt', 'recurring payment', 'monthly pmt',
    'הוראת קבע', 'הו"ק',
]

# Visa category hints that indicate insurance
INSURANCE_CATEGORY_HINTS = ['ביטוח', 'ביטוח ופיננסים', 'insurance', 'insurance & finance']

# Confidence weights
CONF_INSURER_NAME = 0.60
CONF_STRONG_KEYWORD = 0.35
CONF_WEAK_KEYWORD = 0.15
CONF_CATEGORY_HINT = 0.30
CONF_RECURRING = 0.15
CONF_ALREADY_CATEGORIZED = 0.20
CONF_ACH_PATTERN = 0.10
CONF_THRESHOLD = 0.40


def _normalize_merchant_name(raw):
    """Normalize a merchant description to a clean comparable name."""
    import re
    name = raw.strip().lower()
    # Remove common ACH/payment prefixes
    for prefix in ['ach ', 'ach-', 'autopay ', 'auto-pay ', 'direct debit ', 'pos ', 'הו"ק ', 'הוראת קבע ']:
        if name.startswith(prefix):
            name = name[len(prefix):]
    # Remove trailing digits/reference numbers
    name = re.sub(r'\s+\d{4,}$', '', name)
    name = re.sub(r'\s+#\d+$', '', name)
    return name.strip()


def _normalize_transaction(exp_row):
    """Convert an expense DB row to a normalized transaction dict."""
    d = dict(exp_row)
    source = d.get('source', '')
    raw_desc = d.get('description', '')
    return {
        'expense_id': d['id'],
        'source_type': source,
        'transaction_date': d.get('date', ''),
        'raw_description': raw_desc,
        'merchant_name': raw_desc,
        'normalized_merchant': _normalize_merchant_name(raw_desc),
        'amount': abs(d.get('amount', 0)),
        'category_id': d.get('category_id', ''),
        'subcategory': d.get('subcategory', ''),
        'card': d.get('card', ''),
        'frequency': d.get('frequency', ''),
    }


def _detect_insurance_signals(txn):
    """Run all detection signals on a normalized transaction.
    Returns (confidence, reasons, insurer_guess, type_guess, market_guess)."""
    text = (txn['raw_description'] + ' ' + txn['subcategory']).lower()
    text_orig = txn['raw_description'] + ' ' + txn['subcategory']
    normalized = txn.get('normalized_merchant', text)
    confidence = 0.0
    reasons = []
    insurer_guess = ''
    type_guess = 'general'
    market_guess = ''

    # Signal 1: Insurer name match
    for pattern, insurer_name, default_type, market in INSURANCE_INSURER_PATTERNS:
        if pattern.lower() in text or pattern in text_orig or pattern.lower() in normalized:
            confidence += CONF_INSURER_NAME
            reasons.append(f'insurer_name_match:{insurer_name}')
            insurer_guess = insurer_name
            type_guess = default_type
            market_guess = market
            break

    # Signal 2: Strong insurance keyword
    for kw in INSURANCE_STRONG_KEYWORDS:
        if kw.lower() in text or kw in text_orig:
            confidence += CONF_STRONG_KEYWORD
            reasons.append(f'strong_keyword:{kw}')
            break

    # Signal 3: Weak keywords (refine type guess)
    for kw, kw_type in INSURANCE_WEAK_KEYWORDS:
        if kw.lower() in text or kw in text_orig:
            confidence += CONF_WEAK_KEYWORD
            reasons.append(f'weak_keyword:{kw}')
            type_guess = kw_type
            break

    # Signal 4: Visa category hint
    if txn['subcategory']:
        sub_lower = txn['subcategory'].lower()
        for hint in INSURANCE_CATEGORY_HINTS:
            if hint.lower() in sub_lower:
                confidence += CONF_CATEGORY_HINT
                reasons.append(f'category_hint:{txn["subcategory"]}')
                break

    # Signal 5: Already categorized as insurance
    if txn['category_id'] == 'insurance':
        confidence += CONF_ALREADY_CATEGORIZED
        reasons.append('already_insurance_category')

    # Signal 7: ACH/autopay pattern (boosts when combined with insurer match)
    if insurer_guess:
        for ach_pat in INSURANCE_ACH_PATTERNS:
            if ach_pat.lower() in text:
                confidence += CONF_ACH_PATTERN
                reasons.append(f'ach_pattern:{ach_pat}')
                break

    # Infer market from text if not set by insurer match
    if not market_guess:
        has_hebrew = any('\u0590' <= c <= '\u05FF' for c in txn['raw_description'])
        market_guess = 'IL' if has_hebrew else 'US'

    return min(confidence, 1.0), reasons, insurer_guess, type_guess, market_guess


def _is_card_settlement(description):
    """Check if a bank row is just a credit card settlement (not a real insurance charge)."""
    for pat in BANK_SKIP_PATTERNS:
        if pat.lower() in description.lower():
            return True
    return False


def scan_insurance_suggestions(conn, uid):
    """Scan all expenses for insurance candidates. Returns list of suggestion dicts."""
    # Get user's blacklisted merchants (legacy table)
    bl_rows = conn.execute("SELECT merchant_pattern FROM insurance_blacklist WHERE user_id=?", (uid,)).fetchall()
    blacklist = {r[0].lower() for r in bl_rows}

    # Get ignore rules (new table)
    ignore_merchants = set()
    ignore_amounts = set()
    try:
        ir_rows = conn.execute("SELECT rule_type, rule_value FROM insurance_ignore_rules WHERE user_id=?", (uid,)).fetchall()
        for r in ir_rows:
            if r['rule_type'] == 'merchant':
                ignore_merchants.add(r['rule_value'].lower())
            elif r['rule_type'] == 'amount':
                try:
                    ignore_amounts.add(float(r['rule_value']))
                except ValueError:
                    pass
    except Exception:
        pass  # table may not exist yet in old DBs

    # Get existing suggestion expense IDs to avoid re-processing
    existing_ids = set()
    rejected_merchants = set()
    for r in conn.execute("SELECT source_expense_id, merchant_name, status FROM insurance_suggestions WHERE user_id=?", (uid,)).fetchall():
        existing_ids.add(r['source_expense_id'])
        if r['status'] in ('rejected', 'confirmed', 'ignored'):
            rejected_merchants.add(r['merchant_name'].lower().strip())

    # Get existing financial product patterns to avoid duplicates
    prod_patterns = set()
    for r in conn.execute("SELECT expense_pattern FROM financial_products WHERE user_id=? AND status='active'", (uid,)).fetchall():
        if r[0]:
            prod_patterns.add(r[0].lower())

    # Fetch all expenses for this user
    expenses = conn.execute("""
        SELECT id, date, category_id, subcategory, description, amount, source, card, frequency
        FROM expenses WHERE user_id=?
        ORDER BY date DESC
    """, (uid,)).fetchall()

    # Build recurring frequency map: merchant -> count of distinct months
    merchant_months = {}
    for exp in expenses:
        desc = (exp['description'] or '').strip()
        month = (exp['date'] or '')[:7]
        if desc and month:
            key = desc.lower()
            if key not in merchant_months:
                merchant_months[key] = set()
            merchant_months[key].add(month)

    candidates = []
    seen_merchants = {}  # for cross-source dedup: key -> best candidate

    for exp in expenses:
        eid = exp['id']
        if eid in existing_ids:
            continue

        txn = _normalize_transaction(exp)

        # Skip card settlement lines
        if txn['source_type'] == 'bank_csv' and _is_card_settlement(txn['raw_description']):
            continue

        # Skip blacklisted merchants (legacy)
        if any(bl in txn['raw_description'].lower() for bl in blacklist):
            continue

        # Skip ignore-ruled merchants
        normalized = txn['normalized_merchant']
        if any(ig in normalized for ig in ignore_merchants) or any(ig in txn['raw_description'].lower() for ig in ignore_merchants):
            continue

        # Skip ignore-ruled amounts
        if txn['amount'] in ignore_amounts:
            continue

        # Skip merchants already processed (confirmed/rejected/ignored)
        if txn['raw_description'].lower().strip() in rejected_merchants:
            continue

        # Skip if already tracked as a financial product
        if any(pat in txn['raw_description'].lower() for pat in prod_patterns):
            continue

        # Run detection
        confidence, reasons, insurer_guess, type_guess, market_guess = _detect_insurance_signals(txn)

        # Signal 6: Recurring pattern boost
        desc_key = txn['raw_description'].lower().strip()
        if desc_key in merchant_months and len(merchant_months[desc_key]) >= 2:
            confidence = min(confidence + CONF_RECURRING, 1.0)
            reasons.append(f'recurring_months:{len(merchant_months[desc_key])}')

        if confidence < CONF_THRESHOLD:
            continue

        # Determine currency from market
        currency = 'USD' if market_guess == 'US' else 'ILS'

        # Build dedup key
        dedup_key = f"{insurer_guess or normalized}|{round(txn['amount'], -1)}"

        candidate = {
            'source_expense_id': eid,
            'source_type': txn['source_type'],
            'transaction_date': txn['transaction_date'],
            'merchant_name': txn['merchant_name'],
            'raw_description': txn['raw_description'],
            'amount': txn['amount'],
            'insurer_name_guess': insurer_guess,
            'insurance_type_guess': type_guess,
            'confidence_score': round(confidence, 2),
            'detection_reasons': reasons,
            'currency': currency,
            'normalized_merchant': normalized,
            'dedupe_key': dedup_key,
            'suggested_market': market_guess,
        }

        # Cross-source dedup: group by (similar merchant, similar amount)
        if dedup_key in seen_merchants:
            existing = seen_merchants[dedup_key]
            # Keep the one with higher confidence or prefer card source over bank
            if candidate['confidence_score'] > existing['confidence_score']:
                seen_merchants[dedup_key] = candidate
            continue
        seen_merchants[dedup_key] = candidate

    candidates = sorted(seen_merchants.values(), key=lambda c: -c['confidence_score'])
    return candidates


@app.route('/api/insurance/scan', methods=['POST'])
@login_required
def insurance_scan():
    """Scan expenses for insurance candidates and store as suggestions."""
    conn = get_db()
    uid = get_uid()

    candidates = scan_insurance_suggestions(conn, uid)

    inserted = 0
    for c in candidates:
        try:
            conn.execute("""
                INSERT OR IGNORE INTO insurance_suggestions
                (user_id, source_type, source_expense_id, transaction_date, merchant_name,
                 raw_description, amount, insurer_name_guess, insurance_type_guess,
                 confidence_score, detection_reasons, status,
                 currency, normalized_merchant, dedupe_key, suggested_market)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'suggested', ?, ?, ?, ?)
            """, (uid, c['source_type'], c['source_expense_id'], c['transaction_date'],
                  c['merchant_name'], c['raw_description'], c['amount'],
                  c['insurer_name_guess'], c['insurance_type_guess'],
                  c['confidence_score'], json.dumps(c['detection_reasons']),
                  c.get('currency', 'ILS'), c.get('normalized_merchant', ''),
                  c.get('dedupe_key', ''), c.get('suggested_market', '')))
            inserted += 1
        except Exception:
            pass

    conn.commit()

    # Return current suggestions
    suggestions = conn.execute("""
        SELECT * FROM insurance_suggestions
        WHERE user_id=? AND status='suggested'
        ORDER BY confidence_score DESC
    """, (uid,)).fetchall()
    conn.close()

    return jsonify({
        'new_found': inserted,
        'suggestions': [dict(r) for r in suggestions],
    })


@app.route('/api/insurance/suggestions', methods=['GET'])
@login_required
def insurance_suggestions_list():
    """List current insurance suggestions."""
    conn = get_db()
    uid = get_uid()
    status = request.args.get('status', 'suggested')
    rows = conn.execute("""
        SELECT * FROM insurance_suggestions
        WHERE user_id=? AND status=?
        ORDER BY confidence_score DESC
    """, (uid, status)).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route('/api/insurance/suggestions/<int:sid>/confirm', methods=['POST'])
@login_required
def insurance_suggestion_confirm(sid):
    """Confirm a suggestion and create or merge into a financial_products record."""
    conn = get_db()
    uid = get_uid()
    data = request.json or {}

    sug = conn.execute("SELECT * FROM insurance_suggestions WHERE id=? AND user_id=?", (sid, uid)).fetchone()
    if not sug:
        conn.close()
        return jsonify({'error': 'Not found'}), 404

    # Allow user overrides
    provider = data.get('provider_name', sug['insurer_name_guess'] or sug['merchant_name'])
    ins_type = data.get('insurance_type', sug['insurance_type_guess'] or 'general')
    amount = data.get('amount', sug['amount'])
    merge_into = data.get('merge_into_product_id')
    market = sug['suggested_market'] if 'suggested_market' in sug.keys() else ''

    if merge_into:
        # Merge: update the existing financial product's expense_pattern to also match this
        existing = conn.execute("SELECT id, expense_pattern FROM financial_products WHERE id=? AND user_id=?",
                                (merge_into, uid)).fetchone()
        if existing:
            old_pattern = existing['expense_pattern'] or ''
            new_pattern = sug['raw_description']
            if new_pattern.lower() not in old_pattern.lower():
                merged = f"{old_pattern}||{new_pattern}" if old_pattern else new_pattern
                conn.execute("UPDATE financial_products SET expense_pattern=? WHERE id=? AND user_id=?",
                             (merged, merge_into, uid))
            product_id = merge_into
        else:
            conn.close()
            return jsonify({'error': 'Target product not found'}), 404
    else:
        # Language-aware product name
        is_hebrew_market = market == 'IL' or any('\u0590' <= c <= '\u05FF' for c in provider)
        product_name = f'ביטוח {provider}' if is_hebrew_market else f'{provider} Insurance'

        # Create financial product
        conn.execute("""
            INSERT INTO financial_products (type, subtype, company, name, monthly_cost, expense_pattern, status, user_id)
            VALUES ('insurance', ?, ?, ?, ?, ?, 'active', ?)
        """, (ins_type, provider, product_name, amount, sug['raw_description'], uid))
        product_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]

    # Update suggestion status
    conn.execute("""
        UPDATE insurance_suggestions SET status='confirmed', linked_product_id=?, updated_at=CURRENT_TIMESTAMP
        WHERE id=? AND user_id=?
    """, (product_id, sid, uid))

    conn.commit()
    conn.close()
    return jsonify({'status': 'ok', 'product_id': product_id, 'merged': bool(merge_into)})


@app.route('/api/insurance/suggestions/<int:sid>/reject', methods=['POST'])
@login_required
def insurance_suggestion_reject(sid):
    """Reject a suggestion (can be re-detected later)."""
    conn = get_db()
    uid = get_uid()
    conn.execute("""
        UPDATE insurance_suggestions SET status='rejected', updated_at=CURRENT_TIMESTAMP
        WHERE id=? AND user_id=?
    """, (sid, uid))
    conn.commit()
    conn.close()
    return jsonify({'status': 'ok'})


@app.route('/api/insurance/suggestions/<int:sid>/ignore', methods=['POST'])
@login_required
def insurance_suggestion_ignore(sid):
    """Ignore suggestion and optionally add an ignore rule."""
    conn = get_db()
    uid = get_uid()
    data = request.json or {}

    sug = conn.execute("SELECT merchant_name, normalized_merchant FROM insurance_suggestions WHERE id=? AND user_id=?", (sid, uid)).fetchone()
    if not sug:
        conn.close()
        return jsonify({'error': 'Not found'}), 404

    conn.execute("""
        UPDATE insurance_suggestions SET status='ignored', updated_at=CURRENT_TIMESTAMP
        WHERE id=? AND user_id=?
    """, (sid, uid))

    # Add ignore rule (new system) and legacy blacklist
    if data.get('blacklist'):
        merchant = sug['merchant_name'].strip()
        normalized = (sug['normalized_merchant'] or merchant).strip()
        reason = data.get('reason', '')
        if merchant:
            # Legacy blacklist for backward compat
            conn.execute("INSERT OR IGNORE INTO insurance_blacklist (user_id, merchant_pattern) VALUES (?, ?)",
                         (uid, merchant))
            # New ignore rule
            rule_value = normalized if normalized else merchant.lower()
            conn.execute("INSERT OR IGNORE INTO insurance_ignore_rules (user_id, rule_type, rule_value, reason) VALUES (?, 'merchant', ?, ?)",
                         (uid, rule_value, reason))

    conn.commit()
    conn.close()
    return jsonify({'status': 'ok'})


@app.route('/api/insurance/ignore-rules', methods=['GET'])
@login_required
def list_ignore_rules():
    """List all insurance ignore rules for the user."""
    conn = get_db()
    uid = get_uid()
    rows = conn.execute("SELECT id, rule_type, rule_value, reason, created_at FROM insurance_ignore_rules WHERE user_id=? ORDER BY created_at DESC",
                        (uid,)).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route('/api/insurance/ignore-rules/<int:rid>', methods=['DELETE'])
@login_required
def delete_ignore_rule(rid):
    """Delete an ignore rule."""
    conn = get_db()
    uid = get_uid()
    conn.execute("DELETE FROM insurance_ignore_rules WHERE id=? AND user_id=?", (rid, uid))
    conn.commit()
    conn.close()
    return jsonify({'status': 'ok'})


@app.route('/api/insurance/merge-candidates', methods=['GET'])
@login_required
def insurance_merge_candidates():
    """Return active insurance products the user could merge a suggestion into."""
    conn = get_db()
    uid = get_uid()
    rows = conn.execute("""
        SELECT id, company, name, subtype, monthly_cost FROM financial_products
        WHERE user_id=? AND type='insurance' AND status='active'
        ORDER BY company
    """, (uid,)).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route('/api/insurance/overlap-scan', methods=['POST'])
@login_required
def insurance_overlap_scan():
    """Scan active insurance products for overlaps and persist alerts."""
    conn = get_db()
    uid = get_uid()
    new_alerts = scan_insurance_overlaps(conn, uid)
    # Return total open count + new findings
    total = conn.execute(
        "SELECT COUNT(*) FROM insurance_overlap_alerts WHERE user_id=? AND status='open'",
        (uid,)
    ).fetchone()[0]
    conn.close()
    return jsonify({'new_found': len(new_alerts), 'total_open': total, 'alerts': new_alerts})


@app.route('/api/insurance/overlap-alerts', methods=['GET'])
@login_required
def insurance_overlap_alerts_list():
    """List overlap alerts with enriched product data."""
    conn = get_db()
    uid = get_uid()
    status_filter = request.args.get('status', 'open')
    rows = conn.execute("""
        SELECT a.*,
            pa.company AS a_company, pa.name AS a_name, pa.subtype AS a_subtype,
            pa.monthly_cost AS a_cost, pa.policy_number AS a_policy,
            pa.insured_person AS a_person, pa.insured_object AS a_object,
            pa.start_date AS a_start, pa.renewal_date AS a_renewal,
            pa.coverage_amount AS a_coverage, pa.notes AS a_notes,
            pa.coverage_tags AS a_tags, pa.expense_pattern AS a_pattern,
            pb.company AS b_company, pb.name AS b_name, pb.subtype AS b_subtype,
            pb.monthly_cost AS b_cost, pb.policy_number AS b_policy,
            pb.insured_person AS b_person, pb.insured_object AS b_object,
            pb.start_date AS b_start, pb.renewal_date AS b_renewal,
            pb.coverage_amount AS b_coverage, pb.notes AS b_notes,
            pb.coverage_tags AS b_tags, pb.expense_pattern AS b_pattern
        FROM insurance_overlap_alerts a
        JOIN financial_products pa ON a.policy_a_id = pa.id
        JOIN financial_products pb ON a.policy_b_id = pb.id
        WHERE a.user_id=? AND a.status=?
        ORDER BY a.overlap_score DESC
    """, (uid, status_filter)).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route('/api/insurance/overlap-alerts/<int:aid>', methods=['PUT'])
@login_required
def insurance_overlap_alert_update(aid):
    """Update an overlap alert's status or note."""
    data = request.json
    conn = get_db()
    uid = get_uid()
    allowed_statuses = ('open', 'dismissed', 'intentional', 'confirmed_duplicate', 'resolved')
    fields, values = [], []
    if 'status' in data and data['status'] in allowed_statuses:
        fields.append('status=?')
        values.append(data['status'])
    if 'user_note' in data:
        fields.append('user_note=?')
        values.append(data['user_note'])
    if not fields:
        conn.close()
        return jsonify({'error': 'No valid fields'}), 400
    fields.append('updated_at=CURRENT_TIMESTAMP')
    values.extend([aid, uid])
    conn.execute(f"UPDATE insurance_overlap_alerts SET {','.join(fields)} WHERE id=? AND user_id=?", values)
    conn.commit()
    conn.close()
    return jsonify({'status': 'ok'})


@app.route('/api/insurance/overlap-summary', methods=['GET'])
@login_required
def insurance_overlap_summary():
    """Quick count of open overlap alerts by level."""
    conn = get_db()
    uid = get_uid()
    rows = conn.execute("""
        SELECT alert_level, COUNT(*) as cnt,
            SUM(estimated_duplicate_cost_monthly) as total_dup_cost
        FROM insurance_overlap_alerts
        WHERE user_id=? AND status='open'
        GROUP BY alert_level
    """, (uid,)).fetchall()
    conn.close()
    summary = {'critical': 0, 'warning': 0, 'info': 0, 'total': 0, 'total_dup_cost': 0}
    for r in rows:
        summary[r['alert_level']] = r['cnt']
        summary['total'] += r['cnt']
        summary['total_dup_cost'] += r['total_dup_cost'] or 0
    return jsonify(summary)


# ============================================================
# Insurance Overlap Detection Engine
# ============================================================

# Coverage tag families — subtypes that share coverage domains
COVERAGE_FAMILIES = {
    'health': {'health', 'dental', 'vision'},
    'life_protection': {'life', 'disability'},
    'vehicle': {'car', 'auto'},
    'property': {'home', 'homeowners', 'renters'},
    'general_protection': {'general', 'umbrella'},
}

# Subtype → canonical category for comparison
SUBTYPE_TO_CATEGORY = {}
for cat, subs in COVERAGE_FAMILIES.items():
    for s in subs:
        SUBTYPE_TO_CATEGORY[s] = cat


def _normalize_target_key(raw, target_type):
    """Normalize a person name, vehicle ID, or property label for comparison.
    Conservative: only strips formatting, not semantic content.
    Future: could add fuzzy matching, transliteration, or vehicle DB lookup."""
    if not raw:
        return ''
    key = raw.lower().strip()
    import re
    # Strip common prefixes/suffixes
    key = re.sub(r'[\s\-_/\\.,]+', ' ', key).strip()
    if target_type == 'person':
        # Remove honorifics and titles (Hebrew + English)
        for title in ['mr', 'mrs', 'ms', 'dr', 'prof', "מר", "גב'", "גברת", "ד\"ר", "פרופ'"]:
            if key.startswith(title + ' '):
                key = key[len(title):].strip()
        # Normalize double spaces
        key = ' '.join(key.split())
    elif target_type == 'vehicle':
        # Strip non-alphanumeric (plate numbers: "12-345-67" → "1234567")
        key = re.sub(r'[^a-z0-9\u0590-\u05ff]', '', key)
        # Strip generic vehicle words for comparison
        for word in ['car', 'auto', 'vehicle', 'רכב', 'מכונית']:
            key = key.replace(word, '').strip()
    elif target_type == 'property':
        # Strip apartment/floor qualifiers for coarser matching
        key = re.sub(r'\b(apt|apartment|floor|דירה|קומה)\s*\d*', '', key).strip()
        key = ' '.join(key.split())
    return key


def _normalize_insurance_product(product):
    """Normalize a financial_products row into a comparable structure."""
    p = dict(product)
    subtype = (p.get('subtype') or '').lower().strip()
    company = (p.get('company') or '').lower().strip()
    name = (p.get('name') or '').lower().strip()
    person = (p.get('insured_person') or '').lower().strip()
    obj = (p.get('insured_object') or '').lower().strip()
    policy = (p.get('policy_number') or '').strip()

    # Derive category from subtype
    category = SUBTYPE_TO_CATEGORY.get(subtype, subtype or 'unknown')

    # Parse coverage_tags
    tags_raw = p.get('coverage_tags', '[]') or '[]'
    try:
        tags = json.loads(tags_raw) if isinstance(tags_raw, str) else tags_raw
    except (json.JSONDecodeError, TypeError):
        tags = []

    # Infer tags from subtype if none set
    if not tags and subtype:
        tags = [subtype]

    # Normalize provider name — strip common suffixes
    provider = company
    for suffix in ['ביטוח', 'insurance', 'ins', 'חיים', 'בריאות']:
        provider = provider.replace(suffix, '').strip()

    # Build target key — what is being insured
    target_type = 'unknown'
    target_key = ''
    if obj:
        if subtype in ('car', 'auto'):
            target_type = 'vehicle'
        elif subtype in ('home', 'homeowners', 'renters'):
            target_type = 'property'
        else:
            target_type = 'object'
        target_key = _normalize_target_key(obj, target_type)
    elif person:
        target_type = 'person'
        target_key = _normalize_target_key(person, 'person')
    elif subtype in ('car', 'auto'):
        target_type = 'vehicle'
    elif subtype in ('home', 'homeowners', 'renters'):
        target_type = 'property'
    elif subtype in ('health', 'dental', 'vision', 'life', 'disability'):
        target_type = 'person'

    return {
        'id': p['id'],
        'category': category,
        'subtype': subtype,
        'coverage_tags': set(t.lower().strip() for t in tags),
        'provider': provider,
        'provider_raw': company,
        'name': name,
        'policy_number': policy,
        'insured_target_type': target_type,
        'insured_target_key': target_key,
        'insured_person': _normalize_target_key(person, 'person'),
        'insured_object': _normalize_target_key(obj, target_type if obj else 'object'),
        'monthly_cost': float(p.get('monthly_cost', 0) or 0),
        'start_date': p.get('start_date', ''),
        'renewal_date': p.get('renewal_date', ''),
        'expense_pattern': (p.get('expense_pattern') or '').lower(),
    }


def _compute_overlap_score(a, b):
    """Compare two normalized products. Return (score, reasons, overlap_type)."""
    score = 0
    reasons = []

    # ── Strong signals ──

    # Same insurance category
    if a['category'] == b['category'] and a['category'] != 'unknown':
        score += 30
        reasons.append({'signal': 'same_category', 'weight': 30,
                        'detail': a['category']})

    # Same insured target (person or object)
    if a['insured_target_key'] and b['insured_target_key']:
        if a['insured_target_key'] == b['insured_target_key']:
            score += 25
            reasons.append({'signal': 'same_target', 'weight': 25,
                            'detail': a['insured_target_key']})
        elif (a['insured_target_type'] == b['insured_target_type'] and
              a['insured_target_type'] != 'unknown'):
            # Same type but different key — minor signal
            score += 5
            reasons.append({'signal': 'same_target_type', 'weight': 5,
                            'detail': a['insured_target_type']})

    # Same provider
    if a['provider'] and b['provider'] and a['provider'] == b['provider']:
        score += 10
        reasons.append({'signal': 'same_provider', 'weight': 10,
                        'detail': a['provider_raw']})

    # Same or similar policy number
    if a['policy_number'] and b['policy_number']:
        if a['policy_number'] == b['policy_number']:
            score += 20
            reasons.append({'signal': 'same_policy_number', 'weight': 20,
                            'detail': a['policy_number']})
        elif (a['policy_number'][:6] == b['policy_number'][:6] and
              len(a['policy_number']) >= 6):
            score += 10
            reasons.append({'signal': 'similar_policy_number', 'weight': 10,
                            'detail': f"{a['policy_number']} / {b['policy_number']}"})

    # Coverage tag overlap
    if a['coverage_tags'] and b['coverage_tags']:
        shared = a['coverage_tags'] & b['coverage_tags']
        if shared:
            tag_score = min(20, len(shared) * 10)
            score += tag_score
            reasons.append({'signal': 'shared_coverage_tags', 'weight': tag_score,
                            'detail': ', '.join(sorted(shared))})

    # ── Medium signals ──

    # Similar monthly premium (within 25%)
    if a['monthly_cost'] > 0 and b['monthly_cost'] > 0:
        ratio = min(a['monthly_cost'], b['monthly_cost']) / max(a['monthly_cost'], b['monthly_cost'])
        if ratio >= 0.75:
            score += 10
            reasons.append({'signal': 'similar_premium', 'weight': 10,
                            'detail': f"{a['monthly_cost']:.0f} / {b['monthly_cost']:.0f}"})

    # Same exact subtype (more specific than category)
    if a['subtype'] and b['subtype'] and a['subtype'] == b['subtype']:
        score += 10
        reasons.append({'signal': 'same_subtype', 'weight': 10,
                        'detail': a['subtype']})

    # Same person when target type is person
    if (a['insured_person'] and b['insured_person'] and
            a['insured_person'] == b['insured_person'] and
            a['insured_target_type'] == 'person'):
        # Don't double-count if already scored as same_target
        if not any(r['signal'] == 'same_target' for r in reasons):
            score += 15
            reasons.append({'signal': 'same_insured_person', 'weight': 15,
                            'detail': a['insured_person']})

    # ── Weak signals ──

    # Similar expense pattern (transaction description match)
    if a['expense_pattern'] and b['expense_pattern']:
        # Check if patterns share significant words
        a_words = set(a['expense_pattern'].split()) - {'ביטוח', 'insurance', 'חיוב', 'payment'}
        b_words = set(b['expense_pattern'].split()) - {'ביטוח', 'insurance', 'חיוב', 'payment'}
        if a_words and b_words:
            common = a_words & b_words
            if len(common) >= 2:
                score += 5
                reasons.append({'signal': 'similar_expense_pattern', 'weight': 5,
                                'detail': ' '.join(sorted(common))})

    # Cap score at 100
    score = min(100, score)

    # Determine overlap type
    if score >= 70 and any(r['signal'] in ('same_policy_number', 'same_target') for r in reasons):
        overlap_type = 'exact_duplicate'
    elif score >= 50 and a['category'] == b['category']:
        overlap_type = 'overlapping_coverage'
    elif (score >= 40 and any(r['signal'] == 'similar_premium' for r in reasons) and
          a['category'] == b['category']):
        overlap_type = 'payment_duplicate'
    else:
        overlap_type = 'possible_overlap'

    return score, reasons, overlap_type


def _determine_alert_level(score):
    """Map overlap score to alert level."""
    if score >= 80:
        return 'critical'
    elif score >= 55:
        return 'warning'
    else:
        return 'info'


def _generate_confidence_explanation(score, reasons):
    """Generate a human-readable confidence label and explanation.
    Returns (confidence_level, explanation_key, explanation_detail).
    The frontend resolves the key to localized text via i18n."""
    if not reasons:
        return 'low', 'overlap_conf_low_generic', ''

    # Sort reasons by weight descending
    top = sorted(reasons, key=lambda r: r['weight'], reverse=True)
    top_signals = [r['signal'] for r in top[:2]]

    # Strong structural match = high confidence
    strong_structural = {'same_target', 'same_policy_number', 'same_insured_person'}
    has_strong = bool(strong_structural & set(top_signals))
    has_category = 'same_category' in [r['signal'] for r in reasons]

    if score >= 80 and has_strong:
        level = 'high'
    elif score >= 55 and (has_strong or has_category):
        level = 'medium'
    else:
        level = 'low'

    # Build explanation detail from the top 2 reasons
    detail_parts = []
    for r in top[:2]:
        if r.get('detail'):
            detail_parts.append(r['detail'])

    return level, f"overlap_conf_{level}", ', '.join(detail_parts)


def _build_recurring_payment_map(conn, uid):
    """Build a map of expense_pattern → {months_seen, avg_amount, day_of_month_mode}
    for recurring insurance-like payments. Used as a supporting signal."""
    try:
        rows = conn.execute("""
            SELECT description, date, amount FROM expenses
            WHERE user_id=? AND amount > 0 AND date != ''
            ORDER BY date
        """, (uid,)).fetchall()
    except Exception:
        return {}

    from collections import defaultdict
    pattern_data = defaultdict(lambda: {'months': set(), 'amounts': [], 'days': []})
    for r in rows:
        desc = (r['description'] or '').lower().strip()
        if not desc:
            continue
        date_str = r['date'] or ''
        month = date_str[:7]  # YYYY-MM
        day = date_str[8:10] if len(date_str) >= 10 else ''
        pattern_data[desc]['months'].add(month)
        pattern_data[desc]['amounts'].append(float(r['amount']))
        if day:
            pattern_data[desc]['days'].append(int(day))

    result = {}
    for desc, data in pattern_data.items():
        if len(data['months']) >= 2:
            avg_amt = sum(data['amounts']) / len(data['amounts'])
            # Mode of day_of_month (most common charge day)
            day_mode = max(set(data['days']), key=data['days'].count) if data['days'] else 0
            result[desc] = {
                'months_seen': len(data['months']),
                'avg_amount': avg_amt,
                'day_of_month': day_mode,
            }
    return result


def scan_insurance_overlaps(conn, uid):
    """Scan all active insurance products for overlaps. Returns list of new alerts.

    Dismissed/intentional alerts are preserved across rescans — the engine checks
    the existing alert status before upserting. A pair key (min_id, max_id) uniquely
    identifies each pair, and any pair with status 'dismissed', 'intentional', or
    'resolved' is skipped entirely, ensuring the user's decision is permanent until
    the alert is manually reopened or the underlying products materially change."""
    # Fetch active insurance products
    rows = conn.execute("""
        SELECT * FROM financial_products
        WHERE user_id=? AND type='insurance' AND status='active'
        ORDER BY id
    """, (uid,)).fetchall()

    if len(rows) < 2:
        return []

    # Normalize all products
    products = [_normalize_insurance_product(r) for r in rows]

    # Build recurring payment map for expense-based confidence boost
    recurring_map = _build_recurring_payment_map(conn, uid)

    # Get existing alerts to avoid re-creating dismissed/intentional ones
    existing = conn.execute("""
        SELECT policy_a_id, policy_b_id, status FROM insurance_overlap_alerts
        WHERE user_id=?
    """, (uid,)).fetchall()
    existing_pairs = {}
    for e in existing:
        key = (min(e['policy_a_id'], e['policy_b_id']),
               max(e['policy_a_id'], e['policy_b_id']))
        existing_pairs[key] = e['status']

    new_alerts = []

    # Pairwise comparison
    for i in range(len(products)):
        for j in range(i + 1, len(products)):
            a, b = products[i], products[j]
            score, reasons, overlap_type = _compute_overlap_score(a, b)

            # Recurring payment boost: if both products match recurring expenses
            # with similar timing/amounts, it strengthens the overlap signal
            if a['expense_pattern'] and b['expense_pattern']:
                a_rec = recurring_map.get(a['expense_pattern'])
                b_rec = recurring_map.get(b['expense_pattern'])
                if a_rec and b_rec:
                    # Both are confirmed recurring charges
                    if a_rec['months_seen'] >= 3 and b_rec['months_seen'] >= 3:
                        score += 5
                        reasons.append({'signal': 'both_recurring', 'weight': 5,
                                        'detail': f"{a_rec['months_seen']}+{b_rec['months_seen']} months"})
                    # Similar charge day (within 3 days) is a weak supporting signal
                    if (a_rec['day_of_month'] and b_rec['day_of_month'] and
                            abs(a_rec['day_of_month'] - b_rec['day_of_month']) <= 3):
                        score += 3
                        reasons.append({'signal': 'similar_charge_day', 'weight': 3,
                                        'detail': f"day ~{a_rec['day_of_month']}"})
                    score = min(100, score)

            # Only alert for meaningful overlaps
            if score < 40:
                continue

            pair_key = (min(a['id'], b['id']), max(a['id'], b['id']))

            # Skip if user already dismissed/marked as intentional
            existing_status = existing_pairs.get(pair_key)
            if existing_status in ('dismissed', 'intentional', 'resolved'):
                continue

            alert_level = _determine_alert_level(score)

            # Estimate duplicate cost (lower of the two premiums)
            dup_cost = min(a['monthly_cost'], b['monthly_cost']) if (
                a['monthly_cost'] > 0 and b['monthly_cost'] > 0 and score >= 60
            ) else 0

            # Generate confidence explanation
            conf_level, conf_key, conf_detail = _generate_confidence_explanation(score, reasons)

            # Bundle reasons + confidence metadata into one JSON blob
            alert_meta = {
                'reasons': reasons,
                'confidence_level': conf_level,
                'confidence_key': conf_key,
                'confidence_detail': conf_detail,
            }

            # Upsert alert
            conn.execute("""
                INSERT INTO insurance_overlap_alerts
                    (user_id, policy_a_id, policy_b_id, overlap_score, alert_level,
                     overlap_type, reasons_json, estimated_duplicate_cost_monthly,
                     status, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'open', CURRENT_TIMESTAMP)
                ON CONFLICT(user_id, policy_a_id, policy_b_id) DO UPDATE SET
                    overlap_score=excluded.overlap_score,
                    alert_level=excluded.alert_level,
                    overlap_type=excluded.overlap_type,
                    reasons_json=excluded.reasons_json,
                    estimated_duplicate_cost_monthly=excluded.estimated_duplicate_cost_monthly,
                    updated_at=CURRENT_TIMESTAMP
            """, (uid, pair_key[0], pair_key[1], score, alert_level,
                  overlap_type, json.dumps(alert_meta, ensure_ascii=False), dup_cost))

            new_alerts.append({
                'policy_a_id': pair_key[0],
                'policy_b_id': pair_key[1],
                'score': score,
                'level': alert_level,
                'type': overlap_type,
                'reasons': reasons,
                'estimated_cost': dup_cost,
                'confidence_level': conf_level,
            })

    # Clean up alerts for products that no longer exist or are deleted
    active_ids = {p['id'] for p in products}
    conn.execute("""
        DELETE FROM insurance_overlap_alerts
        WHERE user_id=? AND (policy_a_id NOT IN ({ids}) OR policy_b_id NOT IN ({ids}))
    """.format(ids=','.join(str(x) for x in active_ids) if active_ids else '0'), (uid,))

    conn.commit()
    return new_alerts


# Auto-detect patterns: (expense_pattern, type, subtype, company, name)
FINANCIAL_DETECT_PATTERNS = [
    # Insurance
    ('הראל בטוח', 'insurance', 'general', 'הראל', 'ביטוח הראל'),
    ('כלל חיים/ב', 'insurance', 'life', 'כלל', 'ביטוח חיים כלל'),
    ('הפניקס חיים', 'insurance', 'life', 'הפניקס', 'ביטוח חיים הפניקס'),
    ('הפניקס חיים וברי', 'insurance', 'health', 'הפניקס', 'ביטוח בריאות הפניקס'),
    ('מגדל ביטוח', 'insurance', 'general', 'מגדל', 'ביטוח מגדל'),
    ('ביטוח ישיר', 'insurance', 'car', 'ביטוח ישיר', 'ביטוח רכב'),
    ('איילון ביטוח', 'insurance', 'general', 'איילון', 'ביטוח איילון'),
    ('שומרה ביטוח', 'insurance', 'general', 'שומרה', 'ביטוח שומרה'),
    ('מנורה ביטוח', 'insurance', 'general', 'מנורה', 'ביטוח מנורה'),
    # Pension
    ('הראל פנסיה', 'pension', 'comprehensive', 'הראל', 'פנסיה הראל'),
    ('מגדל פנסיה', 'pension', 'comprehensive', 'מגדל', 'פנסיה מגדל'),
    ('מנורה פנסיה', 'pension', 'comprehensive', 'מנורה', 'פנסיה מנורה'),
    ('הפניקס פנסיה', 'pension', 'comprehensive', 'הפניקס', 'פנסיה הפניקס'),
    ('כלל פנסיה', 'pension', 'comprehensive', 'כלל', 'פנסיה כלל'),
    ('מיטב פנסיה', 'pension', 'comprehensive', 'מיטב', 'פנסיה מיטב'),
    # Provident / Gemel
    ('מור גמל ופ חיוב', 'fund', 'gemel', 'מור', 'קופת גמל מור'),
    ('מור גמל', 'fund', 'gemel', 'מור', 'קופת גמל מור'),
    ('הראל גמל', 'fund', 'gemel', 'הראל', 'קופת גמל הראל'),
    ('אלטשולר גמל', 'fund', 'gemel', 'אלטשולר שחם', 'קופת גמל אלטשולר'),
    # Hishtalmut (education fund)
    ('השתלמות אג', 'fund', 'hishtalmut', '', 'קרן השתלמות'),
    ('השתלמות חיוב', 'fund', 'hishtalmut', '', 'קרן השתלמות'),
    ('הראל השתלמות', 'fund', 'hishtalmut', 'הראל', 'קרן השתלמות הראל'),
    # Loans linked to savings
    ('הראלהלואה', 'fund', 'loan', 'הראל', 'הלוואה מהראל'),
    ('חיובי הלוו', 'fund', 'loan', '', 'הלוואה'),
]

PRODUCT_TYPE_HE = {
    'insurance': 'ביטוח',
    'pension': 'פנסיה',
    'fund': 'קרן/קופה',
}

PRODUCT_SUBTYPE_HE = {
    'life': 'ביטוח חיים',
    'health': 'ביטוח בריאות',
    'home': 'ביטוח דירה',
    'car': 'ביטוח רכב',
    'general': 'ביטוח כללי',
    'comprehensive': 'פנסיה מקיפה',
    'gemel': 'קופת גמל',
    'hishtalmut': 'קרן השתלמות',
    'investment': 'קופת גמל להשקעה',
    'loan': 'הלוואה',
}


@app.route('/api/financial/detect', methods=['POST'])
@login_required
def financial_detect():
    """Scan expenses and auto-detect financial products."""
    conn = get_db()
    uid = get_uid()

    # Get all unique expense descriptions/subcategories
    expenses = conn.execute("""
        SELECT DISTINCT description, subcategory, AVG(amount) as avg_amt,
               COUNT(DISTINCT substr(date,1,7)) as months
        FROM expenses
        WHERE user_id=? AND category_id IN ('insurance', 'savings')
        GROUP BY description
        HAVING months >= 1
    """, (uid,)).fetchall()

    # Also check bank patterns
    bank_expenses = conn.execute("""
        SELECT DISTINCT description, subcategory, AVG(amount) as avg_amt,
               COUNT(DISTINCT substr(date,1,7)) as months
        FROM expenses
        WHERE user_id=? AND source = 'bank_csv' AND frequency = 'monthly'
        GROUP BY description
        HAVING months >= 1
    """, (uid,)).fetchall()

    all_expenses = list(expenses) + list(bank_expenses)
    seen_patterns = set()
    detected = []

    # Check existing products to avoid duplicates
    existing = conn.execute("SELECT expense_pattern FROM financial_products WHERE user_id=?", (uid,)).fetchall()
    existing_patterns = {r[0] for r in existing if r[0]}

    for exp in all_expenses:
        desc = exp[0] or ''
        subcat = exp[1] or ''
        text = desc + ' ' + subcat

        for pattern, ptype, psubtype, company, name in FINANCIAL_DETECT_PATTERNS:
            if pattern in text and pattern not in seen_patterns and pattern not in existing_patterns:
                seen_patterns.add(pattern)
                detected.append({
                    'type': ptype,
                    'subtype': psubtype,
                    'company': company,
                    'name': name,
                    'expense_pattern': pattern,
                    'monthly_cost': round(exp[2], 0),
                    'months_seen': exp[3],
                    'source_desc': desc,
                })

    conn.close()
    return jsonify({
        'detected': detected,
        'count': len(detected),
        'already_tracked': len(existing_patterns),
    })


@app.route('/api/financial/auto-add', methods=['POST'])
@login_required
def financial_auto_add():
    """Add detected products to the database."""
    data = request.json
    conn = get_db()
    uid = get_uid()
    added = 0

    for item in data.get('items', []):
        # Check if already exists
        existing = conn.execute(
            "SELECT id FROM financial_products WHERE expense_pattern=? AND user_id=?",
            (item['expense_pattern'], uid)
        ).fetchone()
        if existing:
            continue

        conn.execute("""
            INSERT INTO financial_products (type, subtype, company, name, monthly_cost, expense_pattern, status, user_id)
            VALUES (?, ?, ?, ?, ?, ?, 'active', ?)
        """, (item['type'], item['subtype'], item['company'], item['name'],
              item['monthly_cost'], item['expense_pattern'], uid))
        added += 1

    conn.commit()
    conn.close()
    return jsonify({'status': 'ok', 'added': added})


@app.route('/api/financial/products', methods=['GET'])
@login_required
def financial_list():
    """List all financial products, optionally filtered by type."""
    conn = get_db()
    uid = get_uid()
    ptype = request.args.get('type')
    if ptype:
        rows = conn.execute(
            "SELECT * FROM financial_products WHERE user_id=? AND type=? AND status='active' ORDER BY type, company",
            (uid, ptype)
        ).fetchall()
    else:
        rows = conn.execute(
            "SELECT * FROM financial_products WHERE user_id=? AND status='active' ORDER BY type, company",
            (uid,)
        ).fetchall()

    products = []
    for r in rows:
        p = dict(r)
        # Enrich with actual expense data
        if p['expense_pattern']:
            actual = conn.execute("""
                SELECT AVG(amount) as avg_amt, COUNT(DISTINCT substr(date,1,7)) as months,
                       MAX(date) as last_payment
                FROM expenses WHERE user_id=? AND (description LIKE ? OR subcategory LIKE ?)
            """, (uid, '%' + p['expense_pattern'] + '%', '%' + p['expense_pattern'] + '%')).fetchone()
            p['actual_monthly'] = round(actual[0], 0) if actual[0] else 0
            p['months_tracked'] = actual[1] or 0
            p['last_payment'] = actual[2] or ''
        else:
            p['actual_monthly'] = 0
            p['months_tracked'] = 0
            p['last_payment'] = ''
        products.append(p)

    conn.close()
    return jsonify(products)


@app.route('/api/financial/products', methods=['POST'])
@login_required
def financial_add():
    """Add or update a financial product."""
    data = request.json
    conn = get_db()

    uid = get_uid()
    if data.get('id'):
        conn.execute("""
            UPDATE financial_products SET
                type=?, subtype=?, company=?, name=?, policy_number=?,
                monthly_cost=?, coverage_amount=?, balance=?, balance_date=?,
                employee_pct=?, employer_pct=?, return_rate=?,
                start_date=?, renewal_date=?, notes=?, expense_pattern=?,
                insured_person=?, insured_object=?, coverage_tags=?,
                status=?, updated_at=CURRENT_TIMESTAMP
            WHERE id=? AND user_id=?
        """, (data['type'], data.get('subtype', ''), data.get('company', ''),
              data.get('name', ''), data.get('policy_number', ''),
              data.get('monthly_cost', 0), data.get('coverage_amount', 0),
              data.get('balance', 0), data.get('balance_date', ''),
              data.get('employee_pct', 0), data.get('employer_pct', 0),
              data.get('return_rate', 0),
              data.get('start_date', ''), data.get('renewal_date', ''),
              data.get('notes', ''), data.get('expense_pattern', ''),
              data.get('insured_person', ''), data.get('insured_object', ''),
              json.dumps(data.get('coverage_tags', []), ensure_ascii=False) if isinstance(data.get('coverage_tags'), list) else data.get('coverage_tags', '[]'),
              data.get('status', 'active'), data['id'], uid))
    else:
        conn.execute("""
            INSERT INTO financial_products (type, subtype, company, name, policy_number,
                monthly_cost, coverage_amount, balance, balance_date,
                employee_pct, employer_pct, return_rate,
                start_date, renewal_date, notes, expense_pattern,
                insured_person, insured_object, coverage_tags, status, user_id)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (data['type'], data.get('subtype', ''), data.get('company', ''),
              data.get('name', ''), data.get('policy_number', ''),
              data.get('monthly_cost', 0), data.get('coverage_amount', 0),
              data.get('balance', 0), data.get('balance_date', ''),
              data.get('employee_pct', 0), data.get('employer_pct', 0),
              data.get('return_rate', 0),
              data.get('start_date', ''), data.get('renewal_date', ''),
              data.get('notes', ''), data.get('expense_pattern', ''),
              data.get('insured_person', ''), data.get('insured_object', ''),
              json.dumps(data.get('coverage_tags', []), ensure_ascii=False) if isinstance(data.get('coverage_tags'), list) else data.get('coverage_tags', '[]'),
              data.get('status', 'active'), uid))

    conn.commit()
    conn.close()
    return jsonify({'status': 'ok'})


@app.route('/api/financial/products/<int:product_id>', methods=['DELETE'])
@login_required
def financial_delete(product_id):
    conn = get_db()
    conn.execute("UPDATE financial_products SET status='deleted' WHERE id=? AND user_id=?", (product_id, get_uid()))
    conn.commit()
    conn.close()
    return jsonify({'status': 'ok'})


@app.route('/api/financial/summary', methods=['GET'])
@login_required
def financial_summary():
    """Dashboard summary of all financial products."""
    conn = get_db()
    uid = get_uid()

    products = conn.execute(
        "SELECT * FROM financial_products WHERE user_id=? AND status='active'", (uid,)
    ).fetchall()

    insurance_items = [dict(r) for r in products if r['type'] == 'insurance']
    pension_items = [dict(r) for r in products if r['type'] == 'pension']
    fund_items = [dict(r) for r in products if r['type'] == 'fund']

    insurance_monthly = sum(r['monthly_cost'] for r in insurance_items)
    pension_monthly = sum(r['monthly_cost'] for r in pension_items)
    fund_monthly = sum(r['monthly_cost'] for r in fund_items)
    total_monthly = insurance_monthly + pension_monthly + fund_monthly

    total_coverage = sum(r['coverage_amount'] for r in insurance_items)
    total_balance = sum(r['balance'] for r in pension_items + fund_items)

    # Projection: project fund/pension balances 10/20/30 years
    projections = {}
    for years in [10, 20, 30]:
        projected = 0
        for p in pension_items + fund_items:
            bal = p['balance'] or 0
            monthly = p['monthly_cost'] or 0
            rate = (p['return_rate'] or 5) / 100
            # Future value with monthly contributions
            for y in range(years):
                bal = bal * (1 + rate) + monthly * 12
            projected += bal
        projections[years] = round(projected, 0)

    # Coverage check
    coverage_gaps = []
    subtypes_found = {r['subtype'] for r in insurance_items}
    essential = [
        ('life', 'ביטוח חיים'),
        ('health', 'ביטוח בריאות'),
        ('home', 'ביטוח דירה'),
        ('car', 'ביטוח רכב'),
    ]
    for sub_id, sub_name in essential:
        if sub_id not in subtypes_found:
            coverage_gaps.append(sub_name)

    # Renewal alerts (next 60 days)
    upcoming_renewals = []
    today_str = date.today().strftime('%Y-%m-%d')
    for r in products:
        if r['renewal_date'] and r['renewal_date'] >= today_str:
            try:
                rd = datetime.strptime(r['renewal_date'], '%Y-%m-%d').date()
                days_until = (rd - date.today()).days
                if days_until <= 60:
                    upcoming_renewals.append({
                        'name': r['name'],
                        'company': r['company'],
                        'renewal_date': r['renewal_date'],
                        'days_until': days_until,
                    })
            except ValueError:
                pass

    conn.close()
    return jsonify({
        'insurance': {
            'count': len(insurance_items),
            'monthly_cost': round(insurance_monthly, 0),
            'total_coverage': round(total_coverage, 0),
        },
        'pension': {
            'count': len(pension_items),
            'monthly_cost': round(pension_monthly, 0),
            'total_balance': round(sum(r['balance'] for r in pension_items), 0),
        },
        'funds': {
            'count': len(fund_items),
            'monthly_cost': round(fund_monthly, 0),
            'total_balance': round(sum(r['balance'] for r in fund_items), 0),
        },
        'total_monthly': round(total_monthly, 0),
        'total_balance': round(total_balance, 0),
        'projections': projections,
        'coverage_gaps': coverage_gaps,
        'upcoming_renewals': upcoming_renewals,
    })


# ============================================================
# ---- Assets / Net Worth ----
# ============================================================

ASSET_TYPES = ('real_estate', 'stocks', 'cash', 'custom')
LIABILITY_TYPES = ('mortgage', 'investment_loan', 'custom')

ASSET_FIELDS = {
    'common': ['name', 'current_value', 'currency', 'notes', 'status', 'link_keywords'],
    'real_estate': ['address', 'mortgage_balance', 'rent_income_monthly', 'property_expenses_monthly'],
    'stocks': ['platform_name', 'dividend_income_monthly', 'ticker_summary'],
    'cash': ['institution_name', 'interest_rate', 'interest_income_monthly'],
}

LIABILITY_FIELDS = ['name', 'current_balance', 'currency', 'monthly_payment',
                    'interest_rate', 'linked_asset_id', 'notes', 'status', 'link_keywords']


@app.route('/api/assets', methods=['GET'])
@login_required
def list_assets():
    conn = get_db()
    rows = conn.execute(
        "SELECT * FROM assets WHERE user_id=? AND status='active' ORDER BY current_value DESC",
        (get_uid(),)
    ).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route('/api/assets', methods=['POST'])
@login_required
def create_asset():
    data = request.json
    asset_type = data.get('asset_type', 'custom')
    if asset_type not in ASSET_TYPES:
        return jsonify({'error': 'Invalid asset type'}), 400
    allowed = ASSET_FIELDS['common'] + ASSET_FIELDS.get(asset_type, [])
    fields = ['user_id', 'asset_type']
    values = [get_uid(), asset_type]
    for col in allowed:
        if col in data:
            fields.append(col)
            values.append(data[col])
    # Auto-seed link_keywords if not provided
    if 'link_keywords' not in data or not data['link_keywords']:
        seed = _seed_link_keywords(data, 'asset')
        if seed:
            fields.append('link_keywords')
            values.append(json.dumps(seed))
    placeholders = ','.join(['?'] * len(values))
    col_names = ','.join(fields)
    conn = get_db()
    cur = conn.execute(f"INSERT INTO assets ({col_names}) VALUES ({placeholders})", values)
    new_id = cur.lastrowid
    conn.commit()
    conn.close()
    _auto_snapshot(get_uid())
    return jsonify({'status': 'ok', 'id': new_id})


@app.route('/api/assets/<int:aid>', methods=['PUT'])
@login_required
def update_asset(aid):
    data = request.json
    conn = get_db()
    row = conn.execute("SELECT asset_type FROM assets WHERE id=? AND user_id=?", (aid, get_uid())).fetchone()
    if not row:
        conn.close()
        return jsonify({'error': 'Not found'}), 404
    asset_type = row['asset_type']
    allowed = ASSET_FIELDS['common'] + ASSET_FIELDS.get(asset_type, []) + ['asset_type']
    fields, values = [], []
    for col in allowed:
        if col in data:
            fields.append(f"{col}=?")
            values.append(data[col])
    if not fields:
        conn.close()
        return jsonify({'error': 'No fields to update'}), 400
    fields.append("updated_at=CURRENT_TIMESTAMP")
    values.extend([aid, get_uid()])
    conn.execute(f"UPDATE assets SET {','.join(fields)} WHERE id=? AND user_id=?", values)
    conn.commit()
    conn.close()
    _auto_snapshot(get_uid())
    return jsonify({'status': 'ok'})


@app.route('/api/assets/<int:aid>', methods=['DELETE'])
@login_required
def delete_asset(aid):
    conn = get_db()
    conn.execute("UPDATE assets SET status='archived', updated_at=CURRENT_TIMESTAMP WHERE id=? AND user_id=?",
                 (aid, get_uid()))
    conn.commit()
    conn.close()
    _auto_snapshot(get_uid())
    return jsonify({'status': 'ok'})


@app.route('/api/liabilities', methods=['GET'])
@login_required
def list_liabilities():
    conn = get_db()
    rows = conn.execute(
        "SELECT * FROM liabilities WHERE user_id=? AND status='active' ORDER BY current_balance DESC",
        (get_uid(),)
    ).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route('/api/liabilities', methods=['POST'])
@login_required
def create_liability():
    data = request.json
    liability_type = data.get('liability_type', 'custom')
    if liability_type not in LIABILITY_TYPES:
        return jsonify({'error': 'Invalid liability type'}), 400
    fields = ['user_id', 'liability_type']
    values = [get_uid(), liability_type]
    for col in LIABILITY_FIELDS:
        if col in data:
            fields.append(col)
            values.append(data[col])
    # Auto-seed link_keywords if not provided
    if 'link_keywords' not in data or not data['link_keywords']:
        seed = _seed_link_keywords(data, 'liability')
        if seed:
            fields.append('link_keywords')
            values.append(json.dumps(seed))
    placeholders = ','.join(['?'] * len(values))
    col_names = ','.join(fields)
    conn = get_db()
    cur = conn.execute(f"INSERT INTO liabilities ({col_names}) VALUES ({placeholders})", values)
    new_id = cur.lastrowid
    conn.commit()
    conn.close()
    _auto_snapshot(get_uid())
    return jsonify({'status': 'ok', 'id': new_id})


@app.route('/api/liabilities/<int:lid>', methods=['PUT'])
@login_required
def update_liability(lid):
    data = request.json
    conn = get_db()
    row = conn.execute("SELECT id FROM liabilities WHERE id=? AND user_id=?", (lid, get_uid())).fetchone()
    if not row:
        conn.close()
        return jsonify({'error': 'Not found'}), 404
    fields, values = [], []
    for col in LIABILITY_FIELDS + ['liability_type']:
        if col in data:
            fields.append(f"{col}=?")
            values.append(data[col])
    if not fields:
        conn.close()
        return jsonify({'error': 'No fields to update'}), 400
    fields.append("updated_at=CURRENT_TIMESTAMP")
    values.extend([lid, get_uid()])
    conn.execute(f"UPDATE liabilities SET {','.join(fields)} WHERE id=? AND user_id=?", values)
    conn.commit()
    conn.close()
    _auto_snapshot(get_uid())
    return jsonify({'status': 'ok'})


@app.route('/api/liabilities/<int:lid>', methods=['DELETE'])
@login_required
def delete_liability(lid):
    conn = get_db()
    conn.execute("UPDATE liabilities SET status='archived', updated_at=CURRENT_TIMESTAMP WHERE id=? AND user_id=?",
                 (lid, get_uid()))
    conn.commit()
    conn.close()
    _auto_snapshot(get_uid())
    return jsonify({'status': 'ok'})


@app.route('/api/net-worth/summary', methods=['GET'])
@login_required
def net_worth_summary():
    """Calculate net worth, asset breakdown, passive income, and liability totals."""
    conn = get_db()
    uid = get_uid()
    assets = conn.execute("SELECT * FROM assets WHERE user_id=? AND status='active'", (uid,)).fetchall()
    liabs = conn.execute("SELECT * FROM liabilities WHERE user_id=? AND status='active'", (uid,)).fetchall()
    conn.close()

    total_assets = sum(a['current_value'] or 0 for a in assets)
    total_liabilities = sum(l['current_balance'] or 0 for l in liabs)
    net_worth = total_assets - total_liabilities

    # Breakdown by asset type
    by_type = {}
    for a in assets:
        t = a['asset_type'] or 'custom'
        by_type[t] = by_type.get(t, 0) + (a['current_value'] or 0)

    # Passive income
    passive_rent = sum(a['rent_income_monthly'] or 0 for a in assets if a['asset_type'] == 'real_estate')
    passive_dividend = sum(a['dividend_income_monthly'] or 0 for a in assets if a['asset_type'] == 'stocks')
    passive_interest = sum(a['interest_income_monthly'] or 0 for a in assets if a['asset_type'] == 'cash')
    total_passive = passive_rent + passive_dividend + passive_interest

    # Asset-related expenses
    property_expenses = sum(a['property_expenses_monthly'] or 0 for a in assets if a['asset_type'] == 'real_estate')
    liability_payments = sum(l['monthly_payment'] or 0 for l in liabs)
    total_asset_expenses = property_expenses + liability_payments

    # Net passive = income minus asset-related expenses
    net_passive = total_passive - total_asset_expenses

    # Average monthly expenses for passive income coverage ratio
    current_month = datetime.now().strftime('%Y-%m')
    conn2 = get_db()
    avg_row = conn2.execute(
        "SELECT AVG(mt) FROM (SELECT SUM(amount) as mt FROM expenses WHERE user_id=? AND is_unusual=0 GROUP BY substr(date,1,7))",
        (uid,)
    ).fetchone()
    avg_monthly_expenses = round(avg_row[0], 2) if avg_row and avg_row[0] else 0

    # Previous month snapshot for change indicators
    prev_snap = conn2.execute(
        "SELECT net_worth, total_assets, total_liabilities, passive_income, assets_breakdown "
        "FROM net_worth_snapshots WHERE user_id=? AND month < ? ORDER BY month DESC LIMIT 1",
        (uid, current_month)
    ).fetchone()
    prev = dict(prev_snap) if prev_snap else None

    # Actual passive income from linked transactions
    active_st = ('auto_linked', 'confirmed', 'manual')
    actual_rent = 0
    actual_dividend = 0
    actual_interest = 0
    for a in assets:
        at2 = a['asset_type']
        rows2 = conn2.execute("""
            SELECT i.date, i.amount FROM transaction_links tl
            JOIN income i ON tl.transaction_id = i.id AND tl.transaction_type='income'
            WHERE tl.user_id=? AND tl.asset_id=? AND tl.status IN (?,?,?)
        """, (uid, a['id'], *active_st)).fetchall()
        if rows2:
            by_m = {}
            for r2 in rows2:
                m2 = r2['date'][:7]
                by_m[m2] = by_m.get(m2, 0) + abs(r2['amount'])
            avg = sum(by_m.values()) / max(len(by_m), 1)
            if at2 == 'real_estate':
                actual_rent += avg
            elif at2 == 'stocks':
                actual_dividend += avg
            elif at2 == 'cash':
                actual_interest += avg
    pending_suggestions = conn2.execute(
        "SELECT COUNT(*) FROM transaction_links WHERE user_id=? AND status='suggested'",
        (uid,)).fetchone()[0]
    conn2.close()

    return jsonify({
        'total_assets': round(total_assets, 2),
        'total_liabilities': round(total_liabilities, 2),
        'net_worth': round(net_worth, 2),
        'by_type': by_type,
        'passive_income': {
            'rent': round(passive_rent, 2),
            'dividends': round(passive_dividend, 2),
            'interest': round(passive_interest, 2),
            'total': round(total_passive, 2),
        },
        'actual_passive_income': {
            'rent': round(actual_rent, 2),
            'dividends': round(actual_dividend, 2),
            'interest': round(actual_interest, 2),
            'total': round(actual_rent + actual_dividend + actual_interest, 2),
            'has_data': (actual_rent + actual_dividend + actual_interest) > 0,
        },
        'pending_suggestions': pending_suggestions,
        'asset_expenses': {
            'property': round(property_expenses, 2),
            'liability_payments': round(liability_payments, 2),
            'total': round(total_asset_expenses, 2),
        },
        'net_passive': round(net_passive, 2),
        'asset_count': len(assets),
        'liability_count': len(liabs),
        'avg_monthly_expenses': avg_monthly_expenses,
        'previous_month': prev,
    })


@app.route('/api/net-worth/history', methods=['GET'])
@login_required
def net_worth_history():
    conn = get_db()
    rows = conn.execute(
        "SELECT month, total_assets, total_liabilities, net_worth, passive_income, assets_breakdown "
        "FROM net_worth_snapshots WHERE user_id=? ORDER BY month ASC",
        (get_uid(),)
    ).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route('/api/net-worth/snapshot', methods=['POST'])
@login_required
def create_net_worth_snapshot():
    uid = get_uid()
    _auto_snapshot(uid)
    return jsonify({'status': 'ok'})


def _auto_snapshot(uid):
    """Save a net worth snapshot for the current month."""
    import json as _json
    month = datetime.now().strftime('%Y-%m')
    conn = get_db()
    assets = conn.execute("SELECT * FROM assets WHERE user_id=? AND status='active'", (uid,)).fetchall()
    liabs = conn.execute("SELECT * FROM liabilities WHERE user_id=? AND status='active'", (uid,)).fetchall()

    total_assets = sum(a['current_value'] or 0 for a in assets)
    total_liabilities = sum(l['current_balance'] or 0 for l in liabs)
    net_worth = total_assets - total_liabilities

    by_type = {}
    for a in assets:
        t = a['asset_type'] or 'custom'
        by_type[t] = round(by_type.get(t, 0) + (a['current_value'] or 0), 2)

    passive = sum(a['rent_income_monthly'] or 0 for a in assets if a['asset_type'] == 'real_estate')
    passive += sum(a['dividend_income_monthly'] or 0 for a in assets if a['asset_type'] == 'stocks')
    passive += sum(a['interest_income_monthly'] or 0 for a in assets if a['asset_type'] == 'cash')

    conn.execute("""
        INSERT OR REPLACE INTO net_worth_snapshots
        (user_id, month, total_assets, total_liabilities, net_worth, assets_breakdown, passive_income)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, (uid, month, round(total_assets, 2), round(total_liabilities, 2),
          round(net_worth, 2), _json.dumps(by_type), round(passive, 2)))
    conn.commit()
    conn.close()


# ── Transaction-Asset Linking Engine ──────────────────────────────

import unicodedata as _ud

def _normalize_text(text):
    """Normalize for matching: lowercase, strip punctuation, collapse whitespace."""
    if not text:
        return ''
    text = text.lower()
    text = _ud.normalize('NFKD', text)
    text = re.sub(r'[^\w\s\u0590-\u05FF]', ' ', text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text

def _text_matches_keyword(text, keyword):
    """Check if normalized text contains normalized keyword as substring."""
    nt = _normalize_text(text)
    nk = _normalize_text(keyword)
    if not nk or len(nk) < 2:
        return False
    return nk in nt

BUILTIN_KEYWORDS = {
    'real_estate_income': ['שכירות', 'שכ"ד', 'דמי שכירות', 'שכר דירה', 'rent', 'rental', 'lease'],
    'real_estate_expense': ['ארנונה', 'ועד בית', 'תיקון', 'אחזקה', 'property tax', 'hoa', 'maintenance', 'repair'],
    'mortgage': ['משכנתא', 'משכנתה', 'הלוואת דיור', 'mortgage', 'home loan'],
    'stocks_income': ['דיבידנד', 'דיבידנדים', 'רווח הון', 'dividend', 'capital gain', 'distribution'],
    'cash_income': ['ריבית', 'פיקדון', 'תשואה', 'interest', 'deposit', 'yield'],
    'insurance': ['ביטוח', 'פרמיה', 'פוליסה', 'insurance', 'premium', 'policy'],
}

AMOUNT_TOLERANCE = {
    'rent':      {'pct': 0.07, 'score': 0.20},
    'mortgage':  {'pct': 0.10, 'score': 0.20},
    'insurance': {'pct': 0.15, 'score': 0.15},
    'dividend':  {'pct': 0.30, 'score': 0.10},
    'interest':  {'pct': 0.25, 'score': 0.10},
    'other':     {'pct': 1.00, 'score': 0.05},
}

def _amount_score(actual, declared, amt_type):
    """Score based on type-aware amount proximity."""
    if not declared or declared == 0 or not actual:
        return 0
    tol = AMOUNT_TOLERANCE.get(amt_type, AMOUNT_TOLERANCE['other'])
    ratio = abs(actual - declared) / declared
    return tol['score'] if ratio <= tol['pct'] else 0

def _get_declared_amount(target, target_type):
    """Get the declared monthly amount for comparison."""
    if target_type == 'asset':
        at = target.get('asset_type', '')
        if at == 'real_estate':
            return target.get('rent_income_monthly') or 0
        elif at == 'stocks':
            return target.get('dividend_income_monthly') or 0
        elif at == 'cash':
            return target.get('interest_income_monthly') or 0
    elif target_type == 'liability':
        return target.get('monthly_payment') or 0
    return 0

def _get_amount_type(target, target_type):
    """Determine which tolerance band to use."""
    if target_type == 'liability':
        lt = target.get('liability_type', '')
        if lt == 'mortgage':
            return 'mortgage'
        return 'other'
    at = target.get('asset_type', '')
    if at == 'real_estate':
        return 'rent'
    elif at == 'stocks':
        return 'dividend'
    elif at == 'cash':
        return 'interest'
    return 'other'

def _get_builtin_key(target, target_type, txn_type):
    """Get the BUILTIN_KEYWORDS key for this target+transaction combination."""
    if target_type == 'liability':
        lt = target.get('liability_type', '')
        if lt == 'mortgage':
            return 'mortgage'
        return None
    at = target.get('asset_type', '')
    if at == 'real_estate':
        return 'real_estate_income' if txn_type == 'income' else 'real_estate_expense'
    elif at == 'stocks' and txn_type == 'income':
        return 'stocks_income'
    elif at == 'cash' and txn_type == 'income':
        return 'cash_income'
    return None

def _matches_expected_category(txn, txn_type, target, target_type):
    """Check if transaction category/source matches expected type."""
    if txn_type == 'income':
        src = (txn.get('source') or '').lower()
        at = target.get('asset_type', '')
        if at == 'real_estate' and src in ('rental', 'rent', 'שכירות'):
            return True
        if at == 'stocks' and src in ('dividend', 'dividends', 'דיבידנד'):
            return True
        if at == 'cash' and src in ('interest', 'ריבית'):
            return True
    else:
        cat = txn.get('category_id', '')
        sub = (txn.get('subcategory') or '').lower()
        if target_type == 'liability':
            lt = target.get('liability_type', '')
            if lt == 'mortgage' and cat == 'housing' and ('משכנת' in sub or 'mortgage' in sub):
                return True
        else:
            at = target.get('asset_type', '')
            if at == 'real_estate' and cat == 'housing':
                return True
            if cat == 'insurance':
                return True
    return False

def _recurrence_score(conn, uid, table, description, current_date):
    """Score based on monthly repetition of similar transactions."""
    norm_desc = _normalize_text(description)
    if len(norm_desc) < 3:
        return 0, 0
    try:
        six_months_ago = (datetime.strptime(current_date[:10], '%Y-%m-%d')
                          - timedelta(days=180)).strftime('%Y-%m-%d')
    except (ValueError, TypeError):
        return 0, 0
    rows = conn.execute(
        f"SELECT date, description FROM {table} WHERE user_id=? AND date >= ?",
        (uid, six_months_ago)
    ).fetchall()
    matching = [r for r in rows if _normalize_text(r['description'] or '') and
                norm_desc in _normalize_text(r['description'])]
    months = set(r['date'][:7] for r in matching)
    days = [int(r['date'][8:10]) for r in matching if len(r['date']) >= 10]
    recurrence = 0.15 if len(months) >= 2 else 0
    day_score = 0
    if days and len(current_date) >= 10:
        median_day = sorted(days)[len(days) // 2]
        try:
            current_day = int(current_date[8:10])
            if abs(current_day - median_day) <= 3:
                day_score = 0.10
        except (ValueError, IndexError):
            pass
    return recurrence, day_score

def _has_confirmed_pattern(conn, uid, txn_type, description, target_type, target_id):
    """Check if user previously confirmed a link with same description pattern → same target."""
    norm_desc = _normalize_text(description)
    if len(norm_desc) < 3:
        return False
    target_col = 'asset_id' if target_type == 'asset' else 'liability_id'
    table = 'expenses' if txn_type == 'expense' else 'income'
    rows = conn.execute(f"""
        SELECT t.description FROM transaction_links tl
        JOIN {table} t ON tl.transaction_id = t.id AND tl.transaction_type = ?
        WHERE tl.user_id=? AND tl.status IN ('confirmed','manual')
        AND tl.{target_col} = ?
    """, (txn_type, uid, target_id)).fetchall()
    for r in rows:
        if norm_desc in _normalize_text(r['description'] or ''):
            return True
    return False

def _score_transaction(conn, uid, txn_type, txn, target_type, target):
    """Score a single transaction against a single asset or liability.
    Returns (score, reasons_list)."""
    score = 0.0
    reasons = []
    desc = txn.get('description') or ''

    # S1: User keyword match
    keywords = json.loads(target.get('link_keywords') or '[]')
    for kw in keywords:
        if _text_matches_keyword(desc, kw):
            score += 0.35
            reasons.append(f'keyword_match:{kw}')
            break

    # S2: Built-in keyword match
    bk_key = _get_builtin_key(target, target_type, txn_type)
    if bk_key:
        for bk in BUILTIN_KEYWORDS.get(bk_key, []):
            if _text_matches_keyword(desc, bk):
                score += 0.25
                reasons.append(f'builtin_keyword:{bk}')
                break

    # S3: Amount proximity (type-aware)
    declared = _get_declared_amount(target, target_type)
    if declared:
        amt_type = _get_amount_type(target, target_type)
        s3 = _amount_score(abs(txn.get('amount', 0)), declared, amt_type)
        if s3:
            score += s3
            reasons.append(f'amount_match:{amt_type}')

    # S4: Category/source match
    if _matches_expected_category(txn, txn_type, target, target_type):
        score += 0.15
        reasons.append('category_match')

    # S5: Monthly frequency (expenses only)
    if txn_type == 'expense' and txn.get('frequency') == 'monthly':
        score += 0.10
        reasons.append('monthly_frequency')

    # S6 + S7: Recurrence + day-of-month
    table = 'expenses' if txn_type == 'expense' else 'income'
    s6, s7 = _recurrence_score(conn, uid, table, desc, txn.get('date', ''))
    if s6:
        score += s6
        reasons.append('recurrence_pattern')
    if s7:
        score += s7
        reasons.append('day_proximity')

    # S8: Previous confirmation boost
    if _has_confirmed_pattern(conn, uid, txn_type, desc, target_type, target.get('id')):
        score += 0.30
        reasons.append('prior_confirmation')

    return round(score, 2), reasons

def _seed_link_keywords(target, target_type):
    """Auto-generate link_keywords from existing fields if empty."""
    existing = json.loads(target.get('link_keywords') or '[]')
    if existing:
        return existing
    keywords = []
    name = (target.get('name') or '').strip()
    if name and len(name) >= 2:
        keywords.append(name)
    if target_type == 'asset':
        at = target.get('asset_type', '')
        if at == 'real_estate':
            addr = (target.get('address') or '').strip()
            if addr and len(addr) >= 2:
                keywords.append(addr)
        elif at == 'stocks':
            plat = (target.get('platform_name') or '').strip()
            if plat and len(plat) >= 2:
                keywords.append(plat)
        elif at == 'cash':
            inst = (target.get('institution_name') or '').strip()
            if inst and len(inst) >= 2:
                keywords.append(inst)
    return keywords

def run_linking_engine(uid, month=None, conn=None):
    """Run the linking engine for a user. Returns {auto_linked, suggested, skipped}."""
    own_conn = conn is None
    if own_conn:
        conn = get_db()
    result = {'auto_linked': 0, 'suggested': 0, 'skipped': 0}

    # Load active assets and liabilities
    assets = [dict(r) for r in conn.execute(
        "SELECT * FROM assets WHERE user_id=? AND status='active'", (uid,)).fetchall()]
    liabilities = [dict(r) for r in conn.execute(
        "SELECT * FROM liabilities WHERE user_id=? AND status='active'", (uid,)).fetchall()]
    if not assets and not liabilities:
        if own_conn:
            conn.close()
        return result

    # Load link rules
    always_rules = conn.execute(
        "SELECT * FROM link_rules WHERE user_id=? AND rule_type='always_link'", (uid,)).fetchall()
    never_rules = conn.execute(
        "SELECT * FROM link_rules WHERE user_id=? AND rule_type='never_suggest'", (uid,)).fetchall()

    # Build never-suggest lookup: {normalized_pattern: set of (target_type, target_id)}
    never_lookup = {}
    for r in never_rules:
        pat = _normalize_text(r['description_pattern'])
        key = (r['target_type'], r['target_id'])
        never_lookup.setdefault(pat, set()).add(key)

    # Load existing transaction_links for this user
    existing_links = conn.execute(
        "SELECT transaction_type, transaction_id, asset_id, liability_id, status FROM transaction_links WHERE user_id=?",
        (uid,)).fetchall()
    # Build lookup: (txn_type, txn_id) -> list of {target_type, target_id, status}
    link_map = {}
    for el in existing_links:
        key = (el['transaction_type'], el['transaction_id'])
        targets = link_map.setdefault(key, [])
        if el['asset_id']:
            targets.append({'target_type': 'asset', 'target_id': el['asset_id'], 'status': el['status']})
        if el['liability_id']:
            targets.append({'target_type': 'liability', 'target_id': el['liability_id'], 'status': el['status']})

    # Process income and expenses
    for txn_type, table in [('income', 'income'), ('expense', 'expenses')]:
        date_filter = f" AND substr(date,1,7)='{month}'" if month else ""
        # Skip salary income
        skip_clause = " AND source != 'salary'" if txn_type == 'income' else ""
        rows = conn.execute(
            f"SELECT * FROM {table} WHERE user_id=?{date_filter}{skip_clause}",
            (uid,)).fetchall()

        for txn in rows:
            txn_dict = dict(txn)
            txn_key = (txn_type, txn_dict['id'])
            existing = link_map.get(txn_key, [])

            # Precedence 1: skip if already has active link
            active_statuses = ('auto_linked', 'confirmed', 'manual')
            if any(e['status'] in active_statuses for e in existing):
                result['skipped'] += 1
                continue

            # Precedence 2: skip if ignored
            if any(e['status'] == 'ignored' for e in existing):
                result['skipped'] += 1
                continue

            desc = txn_dict.get('description') or ''
            norm_desc = _normalize_text(desc)

            # Precedence 3: check always_link rules
            rule_matched = False
            for rule in always_rules:
                rule_pat = _normalize_text(rule['description_pattern'])
                if rule_pat and rule_pat in norm_desc:
                    target_col = 'asset_id' if rule['target_type'] == 'asset' else 'liability_id'
                    other_col = 'liability_id' if rule['target_type'] == 'asset' else 'asset_id'
                    # Check not already rejected for this specific target
                    rejected = any(e['target_type'] == rule['target_type'] and
                                   e['target_id'] == rule['target_id'] and
                                   e['status'] == 'rejected' for e in existing)
                    if not rejected:
                        conn.execute(f"""
                            INSERT INTO transaction_links
                            (user_id, transaction_type, transaction_id, {target_col}, {other_col},
                             status, confidence, source, reasons_json, link_version)
                            VALUES (?, ?, ?, ?, NULL, 'auto_linked', 1.0, 'rule', ?, 1)
                        """, (uid, txn_type, txn_dict['id'], rule['target_id'],
                              json.dumps(['rule_match:' + rule['description_pattern']])))
                        result['auto_linked'] += 1
                        rule_matched = True
                        break
            if rule_matched:
                continue

            # Build target list: assets for income, liabilities (then assets) for expenses
            targets = []
            if txn_type == 'income':
                for a in assets:
                    targets.append(('asset', a))
            else:
                for l in liabilities:
                    targets.append(('liability', l))
                for a in assets:
                    targets.append(('asset', a))

            best_score = 0
            best_target = None
            best_reasons = []
            best_target_type = None

            for target_type, target in targets:
                target_id = target['id']

                # Precedence 4: never_suggest check
                skip_never = False
                for np, blocked_targets in never_lookup.items():
                    if np and np in norm_desc and (target_type, target_id) in blocked_targets:
                        skip_never = True
                        break
                if skip_never:
                    continue

                # Precedence 5: rejected check for this target
                if any(e['target_type'] == target_type and e['target_id'] == target_id and
                       e['status'] == 'rejected' for e in existing):
                    continue

                # Score
                sc, reasons = _score_transaction(conn, uid, txn_type, txn_dict, target_type, target)
                if sc > best_score:
                    best_score = sc
                    best_target = target
                    best_reasons = reasons
                    best_target_type = target_type

            # Create link based on score
            if best_score >= 0.40 and best_target:
                status = 'auto_linked' if best_score >= 0.60 else 'suggested'
                target_col = 'asset_id' if best_target_type == 'asset' else 'liability_id'
                other_col = 'liability_id' if best_target_type == 'asset' else 'asset_id'
                # Check no duplicate suggestion for same (txn, target) exists
                dup = conn.execute(f"""
                    SELECT 1 FROM transaction_links
                    WHERE user_id=? AND transaction_type=? AND transaction_id=? AND {target_col}=?
                """, (uid, txn_type, txn_dict['id'], best_target['id'])).fetchone()
                if not dup:
                    conn.execute(f"""
                        INSERT INTO transaction_links
                        (user_id, transaction_type, transaction_id, {target_col}, {other_col},
                         status, confidence, source, reasons_json, link_version)
                        VALUES (?, ?, ?, ?, NULL, ?, ?, 'auto', ?, 1)
                    """, (uid, txn_type, txn_dict['id'], best_target['id'],
                          status, best_score, json.dumps(best_reasons)))
                    if status == 'auto_linked':
                        result['auto_linked'] += 1
                        # Auto-ignore other suggestions for this transaction
                        conn.execute("""
                            UPDATE transaction_links SET status='ignored', updated_at=CURRENT_TIMESTAMP
                            WHERE user_id=? AND transaction_type=? AND transaction_id=?
                            AND status='suggested' AND id != last_insert_rowid()
                        """, (uid, txn_type, txn_dict['id']))
                    else:
                        result['suggested'] += 1
                else:
                    result['skipped'] += 1
            else:
                result['skipped'] += 1

    conn.commit()
    if own_conn:
        conn.close()
    return result


# ── Installment Detection Engine ──────────────────────────────────

INSTALLMENT_AMOUNT_TOLERANCE = 0.05  # 5% unified tolerance for grouping and matching

INSTALLMENT_KEYWORDS_HE = ['תשלום', 'תשלומים', 'מתוך', 'קרדיט', 'פריסה']
INSTALLMENT_KEYWORDS_EN = ['installment', 'payment plan', 'split pay', 'financing']
INSTALLMENT_COUNT_RE_HE = re.compile(r'תשלום\s*(\d+)\s*מתוך\s*(\d+)')
INSTALLMENT_COUNT_RE_EN = re.compile(r'(?:payment\s+)?(\d+)\s*(?:of|/)\s*(\d+)', re.IGNORECASE)

SUBSCRIPTION_PATTERNS = [
    'נטפליקס', 'ספוטיפיי', 'אפל', 'גוגל', 'אמזון פריים', 'hot', 'yes',
    'סלקום', 'פלאפון', 'פרטנר', 'בזק', 'מנוי', 'חבר מועדון', 'חדר כושר',
    'ביטוח', 'ועד בית', 'ארנונה', 'חשמל', 'מים', 'גז',
    'netflix', 'spotify', 'apple', 'google', 'amazon prime', 'hulu',
    'disney', 'hbo', 'youtube', 'adobe', 'microsoft 365', 'dropbox',
    'gym', 'fitness', 'membership', 'insurance', 'utility', 'electric',
]

INSTALLMENT_EXCLUDED_CATEGORIES = ['housing', 'insurance', 'savings']
COMMON_INSTALLMENT_COUNTS = [3, 6, 10, 12, 18, 24, 36]
INSTALLMENT_MIN_AMOUNT = 20
INSTALLMENT_MAX_MONTHS = 36


def _normalize_vendor(text):
    """Normalize vendor name: normalize text then strip trailing numbers."""
    n = _normalize_text(text)
    # Strip trailing reference numbers (common in Israeli bank exports)
    n = re.sub(r'\s*\d{4,}$', '', n)
    # Strip common prefixes
    for pfx in ['הו ק ', 'הוראת קבע ', 'ach ', 'pos ', 'autopay ', 'direct debit ']:
        if n.startswith(pfx):
            n = n[len(pfx):]
    return n.strip()


def _normalize_subscription_desc(text):
    """Normalize subscription vendor for dedup: aggressive but deterministic.

    Handles: casing, punctuation, trailing ref numbers, common suffixes
    (BILL, COM, SERVICE, ישראל), payment prefixes, and spacing.
    """
    if not text:
        return ''
    n = _normalize_vendor(text)  # lowercase, strip punctuation, trailing digits, payment prefixes
    # Strip common subscription vendor suffixes
    for sfx in [' com', ' bill', ' billing', ' payment', ' payments', ' pay',
                ' service', ' services', ' subscription', ' sub', ' monthly',
                ' annual', ' ltd', ' inc', ' llc', ' ישראל', ' israel', ' il']:
        if n.endswith(sfx):
            n = n[:-len(sfx)]
    # Strip leading asterisk/star patterns (GOOGLE *SERVICE → google service → google)
    n = re.sub(r'^\w+\s*\*\s*', '', n) if '*' in text.lower() else n
    # After _normalize_text, * is already stripped, so handle the space-separated remnant
    # e.g. "google service" from "GOOGLE *SERVICE" — keep as-is, the suffix strip handled "service"
    # Strip trailing single digits (e.g. "netflix 1" from "NETFLIX 1-MONTH")
    n = re.sub(r'\s+\d{1,2}$', '', n)
    return n.strip()


def _is_subscription_vendor(vendor_norm):
    """Check if vendor matches known subscription patterns."""
    for pat in SUBSCRIPTION_PATTERNS:
        if _normalize_text(pat) in vendor_norm or vendor_norm in _normalize_text(pat):
            return True
    return False


def _extract_installment_count(description):
    """Try to extract (current, total) from description like 'תשלום 3 מתוך 10'."""
    m = INSTALLMENT_COUNT_RE_HE.search(description or '')
    if m:
        return int(m.group(1)), int(m.group(2))
    m = INSTALLMENT_COUNT_RE_EN.search(description or '')
    if m:
        return int(m.group(1)), int(m.group(2))
    return None, None


def _estimate_total_payments(months_seen):
    """Estimate total installment count from observed months."""
    n = len(months_seen)
    if n <= 0:
        return n
    for c in COMMON_INSTALLMENT_COUNTS:
        if c >= n:
            return c
    return n


def _score_installment_group(expenses, vendor_norm, card, always_rules):
    """Score a group of expenses as an installment candidate.
    Returns (score, reasons, estimated_current, estimated_total)."""
    score = 0.0
    reasons = []

    if not expenses:
        return 0, [], 0, 0

    amounts = [e['amount'] for e in expenses]
    dates = sorted([e['date'] for e in expenses])
    months = sorted(set(d[:7] for d in dates))
    descriptions = [e['description'] or '' for e in expenses]
    n = len(months)

    # Check always_installment rules
    for rule in always_rules:
        rule_val = _normalize_text(rule['rule_value'])
        if rule_val and rule_val in vendor_norm:
            score += 0.50
            reasons.append('always_installment_rule')
            break

    # S1: Installment keyword match (+0.30)
    all_descs = ' '.join(descriptions).lower()
    for kw in INSTALLMENT_KEYWORDS_HE + INSTALLMENT_KEYWORDS_EN:
        if kw.lower() in all_descs or _normalize_text(kw) in _normalize_text(all_descs):
            score += 0.30
            reasons.append(f'keyword:{kw}')
            break

    # S2: Explicit count pattern (+0.25)
    est_current, est_total = None, None
    for desc in descriptions:
        cur, tot = _extract_installment_count(desc)
        if cur is not None:
            est_current, est_total = cur, tot
            score += 0.25
            reasons.append(f'explicit_count:{cur}/{tot}')
            break

    # S3: Fixed amount consistency (+0.20)
    avg_amt = sum(amounts) / len(amounts) if amounts else 0
    if avg_amt > 0:
        variance = max(abs(a - avg_amt) / avg_amt for a in amounts)
        if variance <= 0.01:
            score += 0.20
            reasons.append('exact_amount')
        elif variance <= INSTALLMENT_AMOUNT_TOLERANCE:
            score += 0.12
            reasons.append('consistent_amount')

    # S4: Monthly cadence (+0.15)
    if n >= 2:
        consecutive = True
        for i in range(1, len(months)):
            y1, m1 = int(months[i-1][:4]), int(months[i-1][5:7])
            y2, m2 = int(months[i][:4]), int(months[i][5:7])
            diff = (y2 * 12 + m2) - (y1 * 12 + m1)
            if diff > 2:
                consecutive = False
                break
        if consecutive:
            score += 0.15
            reasons.append('monthly_cadence')

    # S5: Known installment count (+0.10)
    if est_total and est_total in COMMON_INSTALLMENT_COUNTS:
        score += 0.10
        reasons.append(f'common_count:{est_total}')
    elif n in COMMON_INSTALLMENT_COUNTS:
        score += 0.05
        reasons.append(f'count_matches_common:{n}')

    # S6: Finite pattern (+0.10) — penalize "no end signal"
    has_end_signal = est_total is not None
    if not has_end_signal:
        # Check if keyword present
        has_keyword = any(r.startswith('keyword:') for r in reasons)
        if n >= 13 and not has_keyword:
            score -= 0.15
            reasons.append('no_end_signal_penalty')
        elif n < 13:
            score += 0.10
            reasons.append('finite_pattern')
    else:
        score += 0.10
        reasons.append('finite_pattern')

    # S7: Day-of-month consistency (+0.08)
    days = [int(d[8:10]) for d in dates]
    if len(days) >= 2:
        med_day = sorted(days)[len(days) // 2]
        if all(abs(d - med_day) <= 3 for d in days):
            score += 0.08
            reasons.append('day_consistency')

    # S8: Same card (+0.05)
    cards = set(e['card'] for e in expenses if e.get('card'))
    if len(cards) == 1:
        score += 0.05
        reasons.append('same_card')

    # S9: NOT subscription (+0.05)
    if not _is_subscription_vendor(vendor_norm):
        score += 0.05
        reasons.append('not_subscription')

    # S10: Amount in typical range (+0.05)
    if INSTALLMENT_MIN_AMOUNT <= avg_amt <= 5000:
        score += 0.05
        reasons.append('typical_amount_range')

    # Determine estimated total
    if est_total:
        final_total = est_total
        final_made = est_current if est_current else n
    else:
        final_total = _estimate_total_payments(months)
        final_made = n

    return round(min(score, 1.0), 2), reasons, final_made, final_total


def scan_installment_suggestions(uid, conn=None):
    """Scan user's expenses for installment patterns. Returns summary dict."""
    own_conn = conn is None
    if own_conn:
        conn = get_db()

    result = {'total_found': 0, 'new_found': 0, 'updated': 0}

    # Load ignore rules
    ignore_rules = conn.execute(
        "SELECT rule_type, rule_value FROM installment_ignore_rules WHERE user_id=? AND rule_type='never_suggest'",
        (uid,)
    ).fetchall()
    ignore_vendors = set(_normalize_text(r['rule_value']) for r in ignore_rules)

    # Load always_installment rules
    always_rules = conn.execute(
        "SELECT rule_type, rule_value FROM installment_ignore_rules WHERE user_id=? AND rule_type='always_installment'",
        (uid,)
    ).fetchall()

    # Load existing suggestions to check dedupe
    existing = {}
    for row in conn.execute(
        "SELECT id, dedupe_key, status, expense_ids FROM installment_suggestions WHERE user_id=?", (uid,)
    ).fetchall():
        existing[row['dedupe_key']] = dict(row)

    # Fetch candidate expenses
    expenses = conn.execute("""
        SELECT id, date, category_id, description, amount, source, frequency, card
        FROM expenses
        WHERE user_id=? AND source IN ('visa_import', 'bank_csv')
        AND frequency IN ('random', 'once')
        AND description != '' AND description IS NOT NULL
        AND amount >= ?
    """, (uid, INSTALLMENT_MIN_AMOUNT)).fetchall()

    # Group by (normalized_vendor, card, amount_bucket)
    groups = {}
    for e in expenses:
        ed = dict(e)
        vendor_n = _normalize_vendor(ed['description'])
        if not vendor_n or len(vendor_n) < 2:
            continue
        # Skip excluded categories
        if ed['category_id'] in INSTALLMENT_EXCLUDED_CATEGORIES:
            continue
        # Skip ignored vendors
        if any(iv in vendor_n or vendor_n in iv for iv in ignore_vendors if iv):
            continue
        # Skip subscription vendors (unless always_installment rule exists)
        is_always = any(_normalize_text(r['rule_value']) in vendor_n for r in always_rules if r['rule_value'])
        if _is_subscription_vendor(vendor_n) and not is_always:
            continue

        amount_bucket = round(ed['amount'] / 10) * 10
        group_key = f"{vendor_n}|{ed.get('card', '')}|{amount_bucket}"
        if group_key not in groups:
            groups[group_key] = {'expenses': [], 'vendor_norm': vendor_n, 'card': ed.get('card', ''), 'amount_bucket': amount_bucket}
        groups[group_key]['expenses'].append(ed)

    # Score each group
    for group_key, grp in groups.items():
        exps = grp['expenses']
        months = sorted(set(e['date'][:7] for e in exps))

        # Need 2+ months
        if len(months) < 2:
            continue

        # FP guard: multiple transactions per month → likely regular shopping, not installments
        # True installments have ~1 charge per month. >1.5 avg suggests repeat purchases.
        tx_per_month = len(exps) / len(months) if months else 0
        has_explicit_kw = any(_extract_installment_count(e['description'])[0] is not None for e in exps)
        if tx_per_month > 1.5 and not has_explicit_kw:
            continue

        # FP guard: skip if same expense appears in both visa_import and bank_csv (duplicate)
        sources = set(e.get('source', '') for e in exps)
        if 'visa_import' in sources and 'bank_csv' in sources:
            # Only keep visa_import entries (bank CSV often has aggregate visa charges)
            exps = [e for e in exps if e.get('source') == 'visa_import']
            if len(set(e['date'][:7] for e in exps)) < 2:
                continue

        # Check amount variance within group
        amounts = [e['amount'] for e in exps]
        avg_amt = sum(amounts) / len(amounts)
        if avg_amt > 0:
            max_var = max(abs(a - avg_amt) / avg_amt for a in amounts)
            if max_var > INSTALLMENT_AMOUNT_TOLERANCE:
                continue

        vendor_n = grp['vendor_norm']
        card = grp['card']

        score, reasons, est_made, est_total = _score_installment_group(
            exps, vendor_n, card, always_rules
        )

        if score < 0.45:
            continue

        result['total_found'] += 1

        # Compute fields
        first_month = months[0]
        monthly_amount = round(avg_amt, 2)
        vendor_display = exps[0]['description'] or vendor_n
        expense_ids = sorted(set(e['id'] for e in exps))
        est_total_amount = round(monthly_amount * est_total, 2) if est_total else 0
        first_date = sorted(e['date'] for e in exps)[0]
        confidence_level = 'strong' if score >= 0.65 else 'moderate'

        # group_id for rescan stability
        group_id = f"{vendor_n}|{card}|{round(monthly_amount)}"
        # dedupe_key includes first_month for uniqueness
        dedupe_key = f"{vendor_n}|{card}|{round(monthly_amount)}|{first_month}"

        ex = existing.get(dedupe_key)
        if ex:
            if ex['status'] in ('confirmed', 'rejected', 'ignored'):
                continue
            # Update existing suggestion
            conn.execute("""
                UPDATE installment_suggestions SET
                    confidence_score=?, confidence_level=?, detection_reasons=?,
                    expense_ids=?, months_seen=?, estimated_payments_made=?,
                    estimated_total_payments=?, estimated_total_amount=?,
                    vendor_display=?, updated_at=CURRENT_TIMESTAMP
                WHERE id=?
            """, (score, confidence_level, json.dumps(reasons),
                  json.dumps(expense_ids), json.dumps(months), est_made,
                  est_total, est_total_amount, vendor_display, ex['id']))
            result['updated'] += 1
        else:
            # Insert new suggestion
            try:
                conn.execute("""
                    INSERT INTO installment_suggestions
                    (user_id, vendor_normalized, vendor_display, card, monthly_amount,
                     estimated_total_payments, estimated_payments_made, estimated_start_date,
                     estimated_total_amount, confidence_score, confidence_level, detection_reasons,
                     expense_ids, months_seen, group_id, status, dedupe_key)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """, (uid, vendor_n, vendor_display, card, monthly_amount,
                      est_total, est_made, first_date, est_total_amount,
                      score, confidence_level, json.dumps(reasons),
                      json.dumps(expense_ids), json.dumps(months),
                      group_id, 'suggested', dedupe_key))
                result['new_found'] += 1
            except Exception:
                pass  # UNIQUE constraint — already exists with slightly different key

    conn.commit()
    if own_conn:
        conn.close()
    return result


def run_installment_matching(uid, conn=None):
    """Match newly imported expenses to active installment plans."""
    own_conn = conn is None
    if own_conn:
        conn = get_db()

    result = {'auto_matched': 0, 'suggested': 0}

    # Load active installment plans with remaining payments
    plans = conn.execute("""
        SELECT * FROM installments
        WHERE user_id=? AND status='active'
        AND (total_payments - payments_made) > 0
    """, (uid,)).fetchall()

    if not plans:
        if own_conn:
            conn.close()
        return result

    # Load recently imported expenses (last 60 days, not already linked to an installment)
    cutoff = (datetime.now() - timedelta(days=60)).strftime('%Y-%m-%d')
    expenses = conn.execute("""
        SELECT id, date, description, amount, card FROM expenses
        WHERE user_id=? AND source IN ('visa_import', 'bank_csv')
        AND date >= ? AND frequency IN ('random', 'once')
        AND id NOT IN (SELECT expense_id FROM installment_transaction_links WHERE user_id=?)
    """, (uid, cutoff, uid)).fetchall()

    for e in expenses:
        ed = dict(e)
        e_vendor = _normalize_vendor(ed['description'] or '')
        if not e_vendor:
            continue

        best_score = 0
        best_plan = None
        best_reasons = []

        for p in plans:
            pd = dict(p)
            p_vendor = pd.get('vendor_normalized', '')
            if not p_vendor:
                p_vendor = _normalize_vendor(pd.get('description', '') or '')

            mscore = 0.0
            mreasons = []

            # M1: Vendor match (+0.35)
            if p_vendor and e_vendor and (p_vendor in e_vendor or e_vendor in p_vendor):
                mscore += 0.35
                mreasons.append('vendor_match')

            # M2: Amount match (+0.25)
            mp = pd.get('monthly_payment', 0) or 0
            if mp > 0:
                ratio = abs(ed['amount'] - mp) / mp
                if ratio <= INSTALLMENT_AMOUNT_TOLERANCE:
                    mscore += 0.25
                    mreasons.append('amount_match')

            # M3: Card match (+0.15)
            if ed.get('card') and pd.get('card') and ed['card'] == pd['card']:
                mscore += 0.15
                mreasons.append('card_match')

            # M4: Expected date window (+0.10)
            if pd.get('start_date') and pd.get('payments_made') is not None:
                try:
                    start = datetime.strptime(pd['start_date'][:10], '%Y-%m-%d')
                    expected_month = start.month + pd['payments_made']
                    expected_year = start.year + (expected_month - 1) // 12
                    expected_month = ((expected_month - 1) % 12) + 1
                    e_date = datetime.strptime(ed['date'][:10], '%Y-%m-%d')
                    if e_date.year == expected_year and e_date.month == expected_month:
                        mscore += 0.10
                        mreasons.append('expected_date')
                except (ValueError, TypeError):
                    pass

            # M5: Prior link pattern (+0.15)
            prior = conn.execute("""
                SELECT e.description FROM installment_transaction_links itl
                JOIN expenses e ON itl.expense_id = e.id
                WHERE itl.installment_id=? AND itl.status IN ('confirmed', 'auto_matched')
                LIMIT 1
            """, (pd['id'],)).fetchone()
            if prior:
                prior_vendor = _normalize_vendor(prior['description'] or '')
                if prior_vendor and (prior_vendor in e_vendor or e_vendor in prior_vendor):
                    mscore += 0.15
                    mreasons.append('prior_link_pattern')

            if mscore > best_score:
                best_score = mscore
                best_plan = pd
                best_reasons = mreasons

        if best_score >= 0.45 and best_plan:
            # Guard: don't exceed total_payments
            if best_plan['payments_made'] >= best_plan['total_payments']:
                continue

            payment_num = best_plan['payments_made'] + 1
            status = 'auto_matched' if best_score >= 0.65 else 'suggested'

            try:
                conn.execute("""
                    INSERT OR IGNORE INTO installment_transaction_links
                    (user_id, installment_id, expense_id, status, confidence, match_reasons_json, payment_number)
                    VALUES (?,?,?,?,?,?,?)
                """, (uid, best_plan['id'], ed['id'], status,
                      round(best_score, 2), json.dumps(best_reasons), payment_num))
            except Exception:
                continue

            if status == 'auto_matched':
                new_paid = best_plan['payments_made'] + 1
                new_status = 'completed' if new_paid >= best_plan['total_payments'] else 'active'
                conn.execute("""
                    UPDATE installments SET payments_made=?, last_matched_date=?, status=?,
                    updated_at=CURRENT_TIMESTAMP WHERE id=?
                """, (new_paid, ed['date'], new_status, best_plan['id']))
                result['auto_matched'] += 1
            else:
                result['suggested'] += 1

    conn.commit()
    if own_conn:
        conn.close()
    return result


# ── Transaction Linking Endpoints ─────────────────────────────────

@app.route('/api/auto-link', methods=['POST'])
@login_required
def auto_link():
    """Run the linking engine on unlinked transactions."""
    data = request.json or {}
    month = data.get('month')
    result = run_linking_engine(get_uid(), month=month)
    return jsonify(result)


@app.route('/api/link-suggestions', methods=['GET'])
@login_required
def get_link_suggestions():
    """Return pending suggestions for user review."""
    conn = get_db()
    uid = get_uid()
    rows = conn.execute("""
        SELECT tl.*, 'income' as _tbl FROM transaction_links tl
        WHERE tl.user_id=? AND tl.status='suggested' AND tl.transaction_type='income'
        UNION ALL
        SELECT tl.*, 'expenses' as _tbl FROM transaction_links tl
        WHERE tl.user_id=? AND tl.status='suggested' AND tl.transaction_type='expense'
        ORDER BY created_at DESC
    """, (uid, uid)).fetchall()
    suggestions = []
    for r in rows:
        rd = dict(r)
        # Fetch transaction details
        table = 'income' if rd['transaction_type'] == 'income' else 'expenses'
        txn = conn.execute(f"SELECT date, description, amount FROM {table} WHERE id=? AND user_id=?",
                           (rd['transaction_id'], uid)).fetchone()
        if not txn:
            continue
        # Fetch target name
        target_name = ''
        if rd['asset_id']:
            a = conn.execute("SELECT name, asset_type FROM assets WHERE id=?", (rd['asset_id'],)).fetchone()
            target_name = a['name'] if a else ''
            rd['target_type'] = a['asset_type'] if a else ''
        elif rd['liability_id']:
            l = conn.execute("SELECT name, liability_type FROM liabilities WHERE id=?", (rd['liability_id'],)).fetchone()
            target_name = l['name'] if l else ''
            rd['target_type'] = l['liability_type'] if l else ''
        suggestions.append({
            'id': rd['id'],
            'transaction_type': rd['transaction_type'],
            'transaction_id': rd['transaction_id'],
            'transaction_date': txn['date'],
            'transaction_desc': txn['description'],
            'transaction_amount': txn['amount'],
            'asset_id': rd['asset_id'],
            'liability_id': rd['liability_id'],
            'target_name': target_name,
            'target_type': rd.get('target_type', ''),
            'confidence': rd['confidence'],
            'reasons': json.loads(rd['reasons_json'] or '[]'),
            'source': rd['source'],
            'status': rd['status'],
        })
    conn.close()
    return jsonify(suggestions)


@app.route('/api/transaction-links', methods=['POST'])
@login_required
def create_transaction_link():
    """Manually link a transaction to an asset or liability."""
    data = request.json
    conn = get_db()
    uid = get_uid()
    txn_type = data.get('transaction_type')
    txn_id = data.get('transaction_id')
    asset_id = data.get('asset_id')
    liability_id = data.get('liability_id')
    if not txn_type or not txn_id or (not asset_id and not liability_id):
        conn.close()
        return jsonify({'error': 'Missing required fields'}), 400
    # Validate transaction belongs to user
    table = 'income' if txn_type == 'income' else 'expenses'
    txn = conn.execute(f"SELECT id FROM {table} WHERE id=? AND user_id=?", (txn_id, uid)).fetchone()
    if not txn:
        conn.close()
        return jsonify({'error': 'Transaction not found'}), 404
    # Remove any existing active link for this transaction
    conn.execute("""
        UPDATE transaction_links SET status='ignored', updated_at=CURRENT_TIMESTAMP
        WHERE user_id=? AND transaction_type=? AND transaction_id=? AND status IN ('auto_linked','confirmed','manual','suggested')
    """, (uid, txn_type, txn_id))
    # Create manual link
    conn.execute("""
        INSERT INTO transaction_links
        (user_id, transaction_type, transaction_id, asset_id, liability_id,
         status, confidence, source, reasons_json, link_version)
        VALUES (?, ?, ?, ?, ?, 'manual', 1.0, 'manual', '["manual_link"]', 1)
    """, (uid, txn_type, txn_id, asset_id, liability_id))
    conn.commit()
    conn.close()
    return jsonify({'status': 'ok'})


@app.route('/api/transaction-links/<int:lid>', methods=['PUT'])
@login_required
def update_transaction_link(lid):
    """Update link status: confirm, reject, ignore. Optionally create a rule."""
    data = request.json
    conn = get_db()
    uid = get_uid()
    link = conn.execute("SELECT * FROM transaction_links WHERE id=? AND user_id=?", (lid, uid)).fetchone()
    if not link:
        conn.close()
        return jsonify({'error': 'Not found'}), 404
    new_status = data.get('status')
    if new_status not in ('confirmed', 'rejected', 'ignored'):
        conn.close()
        return jsonify({'error': 'Invalid status'}), 400
    conn.execute("UPDATE transaction_links SET status=?, updated_at=CURRENT_TIMESTAMP WHERE id=?",
                 (new_status, lid))
    # If confirming, auto-ignore other suggestions for this transaction
    if new_status == 'confirmed':
        conn.execute("""
            UPDATE transaction_links SET status='ignored', updated_at=CURRENT_TIMESTAMP
            WHERE user_id=? AND transaction_type=? AND transaction_id=? AND status='suggested' AND id!=?
        """, (uid, link['transaction_type'], link['transaction_id'], lid))
    # Optionally create a rule
    if data.get('create_rule'):
        table = 'income' if link['transaction_type'] == 'income' else 'expenses'
        txn = conn.execute(f"SELECT description FROM {table} WHERE id=? AND user_id=?",
                           (link['transaction_id'], uid)).fetchone()
        if txn and txn['description']:
            pattern = _normalize_text(txn['description'])
            if pattern and len(pattern) >= 2:
                rule_type = 'always_link' if new_status == 'confirmed' else 'never_suggest'
                target_type = 'asset' if link['asset_id'] else 'liability'
                target_id = link['asset_id'] or link['liability_id']
                # Avoid duplicate rules
                existing_rule = conn.execute("""
                    SELECT id FROM link_rules WHERE user_id=? AND rule_type=? AND description_pattern=? AND target_type=? AND target_id=?
                """, (uid, rule_type, pattern, target_type, target_id)).fetchone()
                if not existing_rule:
                    conn.execute("""
                        INSERT INTO link_rules (user_id, rule_type, description_pattern, target_type, target_id)
                        VALUES (?, ?, ?, ?, ?)
                    """, (uid, rule_type, pattern, target_type, target_id))
                    # If always_link, retroactively link matching unlinked transactions
                    if rule_type == 'always_link':
                        run_linking_engine(uid, conn=conn)
    conn.commit()
    conn.close()
    return jsonify({'status': 'ok'})


@app.route('/api/transaction-links/<int:lid>', methods=['DELETE'])
@login_required
def delete_transaction_link(lid):
    """Remove a link entirely."""
    conn = get_db()
    uid = get_uid()
    conn.execute("DELETE FROM transaction_links WHERE id=? AND user_id=?", (lid, uid))
    conn.commit()
    conn.close()
    return jsonify({'status': 'ok'})


@app.route('/api/link-rules', methods=['GET'])
@login_required
def get_link_rules():
    """List user's link rules."""
    conn = get_db()
    rows = conn.execute("SELECT * FROM link_rules WHERE user_id=? ORDER BY created_at DESC",
                        (get_uid(),)).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route('/api/link-rules', methods=['POST'])
@login_required
def create_link_rule():
    """Create an always/never link rule."""
    data = request.json
    conn = get_db()
    uid = get_uid()
    rule_type = data.get('rule_type')
    if rule_type not in ('always_link', 'never_suggest'):
        conn.close()
        return jsonify({'error': 'Invalid rule_type'}), 400
    pattern = _normalize_text(data.get('description_pattern', ''))
    if not pattern or len(pattern) < 2:
        conn.close()
        return jsonify({'error': 'Pattern too short'}), 400
    target_type = data.get('target_type')
    target_id = data.get('target_id')
    if target_type not in ('asset', 'liability') or not target_id:
        conn.close()
        return jsonify({'error': 'Invalid target'}), 400
    conn.execute("""
        INSERT INTO link_rules (user_id, rule_type, description_pattern, target_type, target_id)
        VALUES (?, ?, ?, ?, ?)
    """, (uid, rule_type, pattern, target_type, target_id))
    conn.commit()
    # Retroactively apply always_link rules
    if rule_type == 'always_link':
        run_linking_engine(uid, conn=conn)
    conn.close()
    return jsonify({'status': 'ok'})


@app.route('/api/link-rules/<int:rid>', methods=['DELETE'])
@login_required
def delete_link_rule(rid):
    """Delete a link rule."""
    conn = get_db()
    conn.execute("DELETE FROM link_rules WHERE id=? AND user_id=?", (rid, get_uid()))
    conn.commit()
    conn.close()
    return jsonify({'status': 'ok'})


@app.route('/api/assets/<int:aid>/linked-transactions', methods=['GET'])
@login_required
def asset_linked_transactions(aid):
    """Get all active links for an asset."""
    conn = get_db()
    uid = get_uid()
    links = conn.execute("""
        SELECT tl.* FROM transaction_links tl
        WHERE tl.user_id=? AND tl.asset_id=? AND tl.status IN ('auto_linked','confirmed','manual')
        ORDER BY tl.created_at DESC
    """, (uid, aid)).fetchall()
    result = []
    for lnk in links:
        ld = dict(lnk)
        table = 'income' if ld['transaction_type'] == 'income' else 'expenses'
        txn = conn.execute(f"SELECT date, description, amount FROM {table} WHERE id=?",
                           (ld['transaction_id'],)).fetchone()
        if txn:
            ld['transaction_date'] = txn['date']
            ld['transaction_desc'] = txn['description']
            ld['transaction_amount'] = txn['amount']
        ld['reasons'] = json.loads(ld.pop('reasons_json', '[]'))
        result.append(ld)
    conn.close()
    return jsonify(result)


@app.route('/api/liabilities/<int:lid>/linked-transactions', methods=['GET'])
@login_required
def liability_linked_transactions(lid):
    """Get all active links for a liability."""
    conn = get_db()
    uid = get_uid()
    links = conn.execute("""
        SELECT tl.* FROM transaction_links tl
        WHERE tl.user_id=? AND tl.liability_id=? AND tl.status IN ('auto_linked','confirmed','manual')
        ORDER BY tl.created_at DESC
    """, (uid, lid)).fetchall()
    result = []
    for lnk in links:
        ld = dict(lnk)
        table = 'income' if ld['transaction_type'] == 'income' else 'expenses'
        txn = conn.execute(f"SELECT date, description, amount FROM {table} WHERE id=?",
                           (ld['transaction_id'],)).fetchone()
        if txn:
            ld['transaction_date'] = txn['date']
            ld['transaction_desc'] = txn['description']
            ld['transaction_amount'] = txn['amount']
        ld['reasons'] = json.loads(ld.pop('reasons_json', '[]'))
        result.append(ld)
    conn.close()
    return jsonify(result)


@app.route('/api/assets/<int:aid>/intelligence', methods=['GET'])
@login_required
def asset_intelligence(aid):
    """Actual vs declared for a single asset."""
    conn = get_db()
    uid = get_uid()
    asset = conn.execute("SELECT * FROM assets WHERE id=? AND user_id=? AND status='active'",
                         (aid, uid)).fetchone()
    if not asset:
        conn.close()
        return jsonify({'error': 'Not found'}), 404
    ad = dict(asset)
    at = ad['asset_type']

    # Declared values
    declared_income = 0
    declared_expenses = 0
    if at == 'real_estate':
        declared_income = ad.get('rent_income_monthly') or 0
        declared_expenses = ad.get('property_expenses_monthly') or 0
    elif at == 'stocks':
        declared_income = ad.get('dividend_income_monthly') or 0
    elif at == 'cash':
        declared_income = ad.get('interest_income_monthly') or 0

    # Actual from linked transactions (active links only)
    active_statuses = ('auto_linked', 'confirmed', 'manual')
    income_links = conn.execute("""
        SELECT i.date, i.amount FROM transaction_links tl
        JOIN income i ON tl.transaction_id = i.id AND tl.transaction_type='income'
        WHERE tl.user_id=? AND tl.asset_id=? AND tl.status IN (?,?,?)
    """, (uid, aid, *active_statuses)).fetchall()
    expense_links = conn.execute("""
        SELECT e.date, e.amount FROM transaction_links tl
        JOIN expenses e ON tl.transaction_id = e.id AND tl.transaction_type='expense'
        WHERE tl.user_id=? AND tl.asset_id=? AND tl.status IN (?,?,?)
    """, (uid, aid, *active_statuses)).fetchall()

    # Also get expenses linked via liabilities that are linked to this asset
    linked_liabs = conn.execute("SELECT id FROM liabilities WHERE linked_asset_id=? AND user_id=? AND status='active'",
                                (aid, uid)).fetchall()
    liability_expense_links = []
    mortgage_monthly = 0
    for ll in linked_liabs:
        le = conn.execute("""
            SELECT e.date, e.amount FROM transaction_links tl
            JOIN expenses e ON tl.transaction_id = e.id AND tl.transaction_type='expense'
            WHERE tl.user_id=? AND tl.liability_id=? AND tl.status IN (?,?,?)
        """, (uid, ll['id'], *active_statuses)).fetchall()
        liability_expense_links.extend(le)
        mp = conn.execute("SELECT monthly_payment FROM liabilities WHERE id=?", (ll['id'],)).fetchone()
        if mp:
            mortgage_monthly += mp['monthly_payment'] or 0

    # Calculate actual averages
    income_by_month = {}
    for r in income_links:
        m = r['date'][:7]
        income_by_month[m] = income_by_month.get(m, 0) + abs(r['amount'])
    expense_by_month = {}
    for r in list(expense_links) + list(liability_expense_links):
        m = r['date'][:7]
        expense_by_month[m] = expense_by_month.get(m, 0) + abs(r['amount'])

    months_with_income = sorted(income_by_month.keys())
    months_with_expenses = sorted(expense_by_month.keys())
    avg_income = round(sum(income_by_month.values()) / max(len(income_by_month), 1), 2) if income_by_month else 0
    avg_expenses = round(sum(expense_by_month.values()) / max(len(expense_by_month), 1), 2) if expense_by_month else 0

    # Missing months detection (last 3 months)
    current_month = date.today().strftime('%Y-%m')
    recent_months = []
    d = date.today().replace(day=1)
    for i in range(3):
        recent_months.append(d.strftime('%Y-%m'))
        d = (d - timedelta(days=1)).replace(day=1)
    missing_income = [m for m in recent_months if m not in income_by_month and declared_income > 0]

    # Variance
    income_var = round(((avg_income - declared_income) / declared_income) * 100, 1) if declared_income else 0
    expense_var = round(((avg_expenses - declared_expenses) / declared_expenses) * 100, 1) if declared_expenses else 0

    # Alerts
    alerts = []
    if missing_income:
        for m in missing_income:
            alerts.append({'type': 'missing_income', 'month': m, 'key': 'nw_alert_no_rent'})
    if declared_income and abs(income_var) > 5:
        alerts.append({'type': 'income_variance', 'pct': income_var, 'key': 'nw_alert_rent_diff'})

    conn.close()
    return jsonify({
        'asset_id': aid,
        'asset_name': ad['name'],
        'asset_type': at,
        'declared': {'income_monthly': declared_income, 'expenses_monthly': declared_expenses},
        'actual': {
            'avg_monthly_income': avg_income,
            'avg_monthly_expenses': avg_expenses,
            'months_with_income': months_with_income,
            'months_with_expenses': months_with_expenses,
            'months_missing_income': missing_income,
            'total_income_txns': len(income_links),
            'total_expense_txns': len(expense_links) + len(liability_expense_links),
        },
        'variance': {'income_pct': income_var, 'expense_pct': expense_var},
        'pnl': {
            'declared_net': round(declared_income - declared_expenses - mortgage_monthly, 2),
            'actual_net': round(avg_income - avg_expenses, 2),
            'mortgage_payment': mortgage_monthly,
        },
        'alerts': alerts,
    })


@app.route('/api/intelligence/actual-vs-declared', methods=['GET'])
@login_required
def actual_vs_declared_summary():
    """Dashboard summary: declared vs actual passive income across all assets."""
    conn = get_db()
    uid = get_uid()
    assets = conn.execute("SELECT * FROM assets WHERE user_id=? AND status='active'", (uid,)).fetchall()

    active_statuses = ('auto_linked', 'confirmed', 'manual')
    by_asset = []
    declared_total = 0
    actual_total = 0

    for a in assets:
        ad = dict(a)
        at = ad['asset_type']
        declared = 0
        income_type = ''
        if at == 'real_estate':
            declared = ad.get('rent_income_monthly') or 0
            income_type = 'rent'
        elif at == 'stocks':
            declared = ad.get('dividend_income_monthly') or 0
            income_type = 'dividend'
        elif at == 'cash':
            declared = ad.get('interest_income_monthly') or 0
            income_type = 'interest'
        if declared == 0:
            continue

        # Actual income from linked transactions
        rows = conn.execute("""
            SELECT i.date, i.amount FROM transaction_links tl
            JOIN income i ON tl.transaction_id = i.id AND tl.transaction_type='income'
            WHERE tl.user_id=? AND tl.asset_id=? AND tl.status IN (?,?,?)
        """, (uid, ad['id'], *active_statuses)).fetchall()
        by_month = {}
        for r in rows:
            m = r['date'][:7]
            by_month[m] = by_month.get(m, 0) + abs(r['amount'])
        actual = round(sum(by_month.values()) / max(len(by_month), 1), 2) if by_month else 0

        declared_total += declared
        actual_total += actual
        by_asset.append({
            'id': ad['id'], 'name': ad['name'], 'type': income_type,
            'declared': declared, 'actual': actual, 'months_data': len(by_month),
        })

    # Count pending suggestions and unlinked passive candidates
    pending = conn.execute(
        "SELECT COUNT(*) FROM transaction_links WHERE user_id=? AND status='suggested'", (uid,)).fetchone()[0]

    conn.close()
    return jsonify({
        'declared_passive_total': round(declared_total, 2),
        'actual_passive_total': round(actual_total, 2),
        'variance_pct': round(((actual_total - declared_total) / declared_total) * 100, 1) if declared_total else 0,
        'by_asset': by_asset,
        'pending_suggestions': pending,
    })


@app.route('/api/trajectory', methods=['GET'])
@login_required
def financial_trajectory():
    """Evaluate user's financial direction — one clear statement.

    Priority order: net worth trend > cash flow > passive income > expense trend.
    Returns ONE primary message with trajectory type and confidence.
    """
    conn = get_db()
    uid = get_uid()
    month = request.args.get('month', date.today().strftime('%Y-%m'))

    # --- Gather signals ---
    signals = []  # list of (weight, direction, key)
    # direction: +1 positive, -1 negative
    confidence = 'low'

    # A. Net worth trend (primary, weight=3)
    current_month = datetime.now().strftime('%Y-%m')
    snap_cur = conn.execute(
        "SELECT net_worth, passive_income FROM net_worth_snapshots WHERE user_id=? AND month=?",
        (uid, current_month)
    ).fetchone()
    snap_prev = conn.execute(
        "SELECT net_worth, passive_income FROM net_worth_snapshots WHERE user_id=? AND month < ? ORDER BY month DESC LIMIT 1",
        (uid, current_month)
    ).fetchone()

    nw_delta = None
    has_nw = snap_cur is not None and snap_prev is not None
    if has_nw:
        nw_delta = snap_cur['net_worth'] - snap_prev['net_worth']
        # Suppress tiny changes (<0.5% of previous)
        prev_nw = abs(snap_prev['net_worth']) if snap_prev['net_worth'] else 1
        if abs(nw_delta) / prev_nw > 0.005:
            direction = 1 if nw_delta > 0 else -1
            signals.append((3, direction, 'traj_nw_up' if direction > 0 else 'traj_nw_down'))
            confidence = 'high'

    # B. Cash flow (weight=2)
    inc_row = conn.execute(
        "SELECT COALESCE(SUM(amount),0) FROM income WHERE user_id=? AND date LIKE ?",
        (uid, month + '%')
    ).fetchone()
    exp_row = conn.execute(
        "SELECT COALESCE(SUM(amount),0) FROM expenses WHERE user_id=? AND date LIKE ?",
        (uid, month + '%')
    ).fetchone()
    income_total = inc_row[0] if inc_row else 0
    expense_total = exp_row[0] if exp_row else 0

    has_cashflow = income_total > 0 or expense_total > 0
    if income_total > 0:
        if expense_total > income_total:
            signals.append((2, -1, 'traj_overspending'))
            if confidence == 'low':
                confidence = 'medium'
        elif expense_total <= income_total * 0.8:
            signals.append((2, 1, 'traj_saving'))
            if confidence == 'low':
                confidence = 'medium'

    # C. Passive income trend (weight=1)
    pi_delta = None
    if has_nw and snap_cur['passive_income'] and snap_prev['passive_income']:
        pi_now = snap_cur['passive_income'] or 0
        pi_prev = snap_prev['passive_income'] or 0
        pi_delta = pi_now - pi_prev
        if pi_prev > 0 and pi_delta > 0:
            signals.append((1, 1, 'traj_passive_growing'))
        elif pi_prev > 0 and pi_delta < 0:
            signals.append((1, -1, 'traj_passive_declining'))

    # D. Expense trend — compare current month vs previous months average
    avg_exp = conn.execute(
        "SELECT AVG(mt) FROM (SELECT SUM(amount) as mt FROM expenses WHERE user_id=? AND is_unusual=0 AND substr(date,1,7) != ? GROUP BY substr(date,1,7))",
        (uid, month)
    ).fetchone()
    avg_monthly = avg_exp[0] if avg_exp and avg_exp[0] else 0
    if avg_monthly > 0 and expense_total > 0:
        exp_change_pct = ((expense_total - avg_monthly) / avg_monthly) * 100
        if exp_change_pct > 15:
            signals.append((1, -1, 'traj_expenses_rising'))
        elif exp_change_pct < -10:
            signals.append((1, 1, 'traj_expenses_falling'))

    # E. High installment burden (weight=1) — flag when >25% of income
    inst_rows = conn.execute(
        "SELECT monthly_payment, total_payments, payments_made, status FROM installments WHERE user_id=?", (uid,)
    ).fetchall()
    inst_monthly = sum(
        r['monthly_payment'] for r in inst_rows
        if max(r['total_payments'] - r['payments_made'], 0) > 0 and dict(r).get('status') != 'completed'
    )
    if income_total > 0 and inst_monthly > 0:
        inst_pct = inst_monthly / income_total
        if inst_pct > 0.25:
            signals.append((1, -1, 'traj_high_installments'))

    # F. Income stability (weight=1) — from salary data
    sal_rows = conn.execute(
        "SELECT gross_salary, bonus_amount FROM salary_statements WHERE user_id=? ORDER BY month DESC",
        (uid,)
    ).fetchall()
    if len(sal_rows) >= 2:
        sal_dicts_f = [dict(r) for r in sal_rows]
        s_grosses = [(s.get('gross_salary') or 0) for s in sal_dicts_f]
        s_mean = sum(s_grosses) / len(s_grosses) if s_grosses else 0
        s_cv = 0
        if s_mean > 0:
            s_var = sum((g - s_mean) ** 2 for g in s_grosses) / len(s_grosses)
            s_cv = (s_var ** 0.5) / s_mean
        s_total_gross = sum(s_grosses)
        s_total_bonus = sum((s.get('bonus_amount') or 0) for s in sal_dicts_f)
        s_bonus_share = s_total_bonus / s_total_gross if s_total_gross > 0 else 0
        # Baseline (exclude bonus months)
        s_baseline = [s for s in sal_dicts_f if (s.get('bonus_amount') or 0) == 0] or sal_dicts_f
        s_base_gross = sum((s.get('gross_salary') or 0) for s in s_baseline) / len(s_baseline)
        s_latest_gross = s_grosses[0] if s_grosses else 0
        s_vs_base = ((s_latest_gross - s_base_gross) / s_base_gross * 100) if s_base_gross > 0 else 0
        # Risk score (same formula as salary summary)
        s_cv_sc = min(s_cv / 0.20, 1.0)
        s_bn_sc = min(s_bonus_share / 0.40, 1.0)
        s_dr_sc = min(max(-s_vs_base, 0) / 30.0, 1.0)
        s_risk = s_cv_sc * 0.4 + s_bn_sc * 0.3 + s_dr_sc * 0.3
        if s_risk >= 0.50:
            signals.append((1, -1, 'traj_income_risk'))
        elif s_risk < 0.20 and len(sal_rows) >= 3:
            signals.append((1, 1, 'traj_income_stable'))

    conn.close()

    # --- No signals → don't show ---
    if not signals and not has_cashflow:
        return jsonify({'show': False})

    # --- Pick ONE primary message ---
    # Sort by weight desc; within same weight, negative first (warn user)
    signals.sort(key=lambda s: (-s[0], s[1]))

    # Weighted score determines trajectory type
    score = sum(w * d for w, d, _ in signals)
    if score > 0:
        traj_type = 'positive'
    elif score < 0:
        traj_type = 'negative'
    else:
        traj_type = 'mixed'

    # Primary = highest weight signal
    primary_key = signals[0][2] if signals else None

    # For mixed: build compound message if top two signals conflict
    if traj_type == 'mixed' and len(signals) >= 2:
        pos_sig = next((s for s in signals if s[1] > 0), None)
        neg_sig = next((s for s in signals if s[1] < 0), None)
        if pos_sig and neg_sig:
            primary_key = 'traj_mixed'

    # If no signals at all but we have cashflow, use a neutral fallback
    if not signals:
        traj_type = 'positive' if income_total > expense_total else 'negative' if expense_total > income_total else 'mixed'
        primary_key = 'traj_saving' if traj_type == 'positive' else 'traj_overspending' if traj_type == 'negative' else None
        if not primary_key:
            return jsonify({'show': False})

    # Build reasons list (secondary signals)
    reasons = [s[2] for s in signals[1:] if s[2] != primary_key]

    # Emotional headline based on trajectory + strength
    abs_score = abs(score)
    if traj_type == 'positive':
        headline_key = 'traj_h_strong_pos' if abs_score >= 5 else 'traj_h_pos'
    elif traj_type == 'negative':
        headline_key = 'traj_h_strong_neg' if abs_score >= 5 else 'traj_h_neg'
    else:
        headline_key = 'traj_h_mixed'

    # Data freshness — most recent update across assets + bank imports
    conn2 = get_db()
    asset_fresh = conn2.execute(
        "SELECT MAX(updated_at) FROM assets WHERE user_id=? AND status='active'", (uid,)
    ).fetchone()
    bank_fresh = conn2.execute(
        "SELECT MAX(created_at) FROM bank_balances WHERE user_id=?", (uid,)
    ).fetchone()
    conn2.close()
    latest_asset = asset_fresh[0] if asset_fresh and asset_fresh[0] else None
    latest_bank = bank_fresh[0] if bank_fresh and bank_fresh[0] else None
    # Pick the most recent of the two
    data_date = max(filter(None, [latest_asset, latest_bank]), default=None)

    return jsonify({
        'show': True,
        'trajectory_type': traj_type,
        'headline_key': headline_key,
        'message_key': primary_key,
        'confidence': confidence,
        'reasons': reasons[:2],  # max 2 supporting reasons
        'data_date': data_date,
    })


# ── Installment Insights ────────────────────────────────────────

@app.route('/api/installment-insights', methods=['GET'])
@login_required
def installment_insights():
    """Actionable installment intelligence: commitments, ending-soon, recently completed, next-month forecast."""
    conn = get_db()
    uid = get_uid()
    today = date.today()

    rows = conn.execute("SELECT * FROM installments WHERE user_id=?", (uid,)).fetchall()
    plans = []
    for r in rows:
        d = dict(r)
        d['payments_remaining'] = max(d['total_payments'] - d['payments_made'], 0)
        d['remaining_amount'] = round(d['monthly_payment'] * d['payments_remaining'], 2)
        is_active = d['payments_remaining'] > 0 and d.get('status') != 'completed'
        d['is_active'] = is_active

        # Estimate end date from start_date + total_payments months
        try:
            start = datetime.strptime(d['start_date'], '%Y-%m-%d').date()
            end_month_offset = d['total_payments']
            end_year = start.year + (start.month - 1 + end_month_offset) // 12
            end_month = (start.month - 1 + end_month_offset) % 12 + 1
            import calendar
            end_day = min(start.day, calendar.monthrange(end_year, end_month)[1])
            d['estimated_end_date'] = date(end_year, end_month, end_day).isoformat()
            d['days_until_end'] = (date(end_year, end_month, end_day) - today).days if is_active else 0
            d['months_remaining'] = d['payments_remaining']
        except (ValueError, TypeError):
            d['estimated_end_date'] = ''
            d['days_until_end'] = 0
            d['months_remaining'] = d['payments_remaining']

        plans.append(d)

    active = [p for p in plans if p['is_active']]
    completed = [p for p in plans if not p['is_active']]

    total_monthly = sum(p['monthly_payment'] for p in active)
    total_remaining = sum(p['remaining_amount'] for p in active)

    # Ending soon: active plans finishing within 60 days
    ending_soon = [p for p in active if 0 < p.get('days_until_end', 999) <= 60]
    ending_soon.sort(key=lambda p: p.get('days_until_end', 999))

    # Recently completed: plans that ended in the last 30 days
    recently_completed = []
    for p in completed:
        end_str = p.get('estimated_end_date', '')
        if end_str:
            try:
                end_d = datetime.strptime(end_str, '%Y-%m-%d').date()
                if 0 <= (today - end_d).days <= 30:
                    recently_completed.append(p)
            except (ValueError, TypeError):
                pass

    # Next-month commitment drop: sum of monthly_payment for plans ending within 30 days
    ending_next_month = [p for p in active if 0 < p.get('days_until_end', 999) <= 30]
    next_month_drop = sum(p['monthly_payment'] for p in ending_next_month)

    # Burden ratio: installment commitment vs income
    inc_total = conn.execute(
        "SELECT COALESCE(SUM(amount),0) FROM income WHERE user_id=? AND date LIKE ?",
        (uid, today.strftime('%Y-%m') + '%')
    ).fetchone()[0]
    if inc_total <= 0:
        # Fallback: average income across all months
        avg_inc = conn.execute(
            "SELECT AVG(mt) FROM (SELECT SUM(amount) as mt FROM income WHERE user_id=? GROUP BY substr(date,1,7))",
            (uid,)
        ).fetchone()
        inc_total = avg_inc[0] if avg_inc and avg_inc[0] else 0

    burden_pct = round((total_monthly / inc_total) * 100, 1) if inc_total > 0 else 0

    # Trend analysis: is the user's installment burden improving?
    recently_freed = sum(p['monthly_payment'] for p in recently_completed)
    soon_freed = sum(p['monthly_payment'] for p in ending_soon)
    # Check for recently added plans (created in last 30 days)
    recently_added = [p for p in active if p.get('created_at', '')]
    new_plans_30d = 0
    new_plans_monthly = 0
    for p in recently_added:
        try:
            created = datetime.strptime(p['created_at'][:10], '%Y-%m-%d').date()
            if (today - created).days <= 30:
                new_plans_30d += 1
                new_plans_monthly += p['monthly_payment']
        except (ValueError, TypeError):
            pass

    # Determine trend direction
    if recently_freed > 0 and new_plans_monthly <= recently_freed:
        trend_direction = 'improving'
        if recently_freed > 0:
            trend_key = 'inst_trend_freed'
            trend_params = {'amount': round(recently_freed, 2), 'count': len(recently_completed)}
        else:
            trend_key = 'inst_trend_ending'
            trend_params = {'amount': round(soon_freed, 2), 'days': ending_soon[0].get('days_until_end', 0) if ending_soon else 0}
    elif soon_freed > 0 and new_plans_monthly <= soon_freed:
        trend_direction = 'improving'
        trend_key = 'inst_trend_ending'
        trend_params = {'amount': round(soon_freed, 2), 'days': ending_soon[0].get('days_until_end', 0) if ending_soon else 0}
    elif new_plans_monthly > 0 and new_plans_monthly > recently_freed + soon_freed:
        trend_direction = 'increasing'
        trend_key = 'inst_trend_increasing'
        trend_params = {'amount': round(new_plans_monthly, 2), 'count': new_plans_30d}
    elif len(active) > 0 and burden_pct <= 25:
        trend_direction = 'stable'
        trend_key = 'inst_trend_healthy'
        trend_params = {'pct': burden_pct}
    else:
        trend_direction = 'stable'
        trend_key = ''
        trend_params = {}

    conn.close()
    return jsonify({
        'total_monthly_commitment': round(total_monthly, 2),
        'total_remaining': round(total_remaining, 2),
        'active_count': len(active),
        'burden_pct': burden_pct,
        'ending_soon': [{
            'id': p['id'], 'description': p['description'], 'store': p.get('store', ''),
            'monthly_payment': p['monthly_payment'], 'payments_remaining': p['payments_remaining'],
            'days_until_end': p.get('days_until_end', 0), 'estimated_end_date': p.get('estimated_end_date', ''),
        } for p in ending_soon],
        'recently_completed': [{
            'id': p['id'], 'description': p['description'], 'store': p.get('store', ''),
            'monthly_payment': p['monthly_payment'], 'total_payments': p['total_payments'],
            'estimated_end_date': p.get('estimated_end_date', ''),
        } for p in recently_completed],
        'next_month_drop': round(next_month_drop, 2),
        'trend': {
            'direction': trend_direction,
            'message_key': trend_key,
            'params': trend_params,
        },
        'plans_summary': [{
            'id': p['id'], 'description': p['description'], 'monthly_payment': p['monthly_payment'],
            'months_remaining': p.get('months_remaining', 0), 'is_active': p['is_active'],
        } for p in plans],
    })


# ---- Insights API Endpoints ----

# Israeli CBS average household spending percentages (2024 data approximation)
CBS_AVERAGES = {
    'housing': 25.2,
    'food': 14.0,
    'vehicle': 13.5,
    'children': 7.5,
    'communication': 4.2,
    'health_beauty': 3.8,
    'medical': 5.1,
    'insurance': 5.5,
    'entertainment': 4.5,
    'personal': 3.0,
    'savings': 5.0,
    'misc': 2.0,
    'clothing': 3.2,
    'subscriptions': 2.5,
    'education': 3.8,
    'dining_out': 3.5,
    'gifts': 1.8,
}


@app.route('/api/insights/heatmap', methods=['GET'])
@login_required
def insights_heatmap():
    """Calendar heatmap: daily spending intensity for the month."""
    conn = get_db()
    uid = get_uid()
    month = request.args.get('month', date.today().strftime('%Y-%m'))
    year, mon = map(int, month.split('-'))

    import calendar
    days_in_month = calendar.monthrange(year, mon)[1]
    first_weekday = calendar.monthrange(year, mon)[0]  # 0=Mon 6=Sun
    # Convert to Sunday-start (Israeli week): Sun=0
    first_weekday = (first_weekday + 1) % 7

    daily = conn.execute(
        "SELECT date, SUM(amount) as total FROM expenses WHERE user_id=? AND date LIKE ? GROUP BY date ORDER BY date",
        (uid, month + '%')
    ).fetchall()

    daily_map = {r[0]: r[1] for r in daily}
    amounts = [r[1] for r in daily] if daily else [0]
    max_amount = max(amounts) if amounts else 1

    days = []
    for d in range(1, days_in_month + 1):
        ds = f"{month}-{d:02d}"
        amt = daily_map.get(ds, 0)
        intensity = min(amt / max_amount, 1.0) if max_amount > 0 else 0
        days.append({'date': ds, 'day': d, 'amount': amt, 'intensity': round(intensity, 3)})

    conn.close()
    return jsonify({
        'month': month,
        'days_in_month': days_in_month,
        'first_weekday': first_weekday,
        'days': days,
        'max_amount': max_amount,
    })


@app.route('/api/insights/burnrate', methods=['GET'])
@login_required
def insights_burnrate():
    """Burn rate gauge: projected month-end spending based on current pace."""
    conn = get_db()
    uid = get_uid()
    month = request.args.get('month', date.today().strftime('%Y-%m'))
    year, mon = map(int, month.split('-'))

    import calendar
    days_in_month = calendar.monthrange(year, mon)[1]

    today = date.today()
    if today.year == year and today.month == mon:
        day_of_month = today.day
    else:
        day_of_month = days_in_month

    spent = conn.execute(
        "SELECT COALESCE(SUM(amount),0) FROM expenses WHERE user_id=? AND date LIKE ?", (uid, month + '%')
    ).fetchone()[0]

    income = conn.execute(
        "SELECT COALESCE(SUM(amount),0) FROM income WHERE user_id=? AND date LIKE ?", (uid, month + '%')
    ).fetchone()[0]

    daily_avg = spent / day_of_month if day_of_month > 0 else 0
    projected = daily_avg * days_in_month
    pct_of_income = (projected / income * 100) if income > 0 else 0

    # Previous month average for comparison (normalized — excludes unusual + excluded months)
    excl_months = get_excluded_month_set(conn, uid)
    prev_month_rows = conn.execute("""
        SELECT substr(date,1,7) as m, SUM(amount) as total FROM expenses
        WHERE user_id=? AND is_unusual=0 AND substr(date,1,7) != ? GROUP BY m
    """, (uid, month)).fetchall()
    prev_vals = [r['total'] for r in prev_month_rows if r['m'] not in excl_months]
    prev_months = sum(prev_vals) / len(prev_vals) if prev_vals else 0

    # Bank balance context
    bank_bal = conn.execute(
        "SELECT closing_balance FROM bank_balances WHERE user_id=? AND account_name='main' AND month=?",
        (uid, month)
    ).fetchone()
    bank_balance = bank_bal['closing_balance'] if bank_bal else None

    # Previous month bank balance for change
    prev_dt = date(year, mon, 1) - timedelta(days=1)
    prev_month_str = prev_dt.strftime('%Y-%m')
    prev_bank = conn.execute(
        "SELECT closing_balance FROM bank_balances WHERE user_id=? AND account_name='main' AND month=?",
        (uid, prev_month_str)
    ).fetchone()
    bank_balance_change = round(bank_balance - prev_bank['closing_balance'], 0) if bank_balance is not None and prev_bank else None

    conn.close()
    return jsonify({
        'month': month,
        'day_of_month': day_of_month,
        'days_in_month': days_in_month,
        'spent_so_far': spent,
        'daily_average': round(daily_avg, 0),
        'projected_total': round(projected, 0),
        'income': income,
        'pct_of_income': round(pct_of_income, 1),
        'prev_month_avg': round(prev_months, 0),
        'on_track': projected <= income if income > 0 else projected <= prev_months,
        'bank_balance': bank_balance,
        'bank_balance_change': bank_balance_change,
    })


@app.route('/api/insights/latte', methods=['GET'])
@login_required
def insights_latte():
    """Latte factor: show what small recurring costs become over time if invested."""
    conn = get_db()
    uid = get_uid()
    # Find small recurring expenses (< 200 NIS each, appearing 3+ months)
    candidates = conn.execute("""
        SELECT description, AVG(amount) as avg_amt, COUNT(DISTINCT substr(date,1,7)) as months,
               category_id, subcategory
        FROM expenses
        WHERE user_id=? AND amount < 200 AND amount > 5 AND description != ''
        GROUP BY description
        HAVING months >= 2
        ORDER BY avg_amt * months DESC
        LIMIT 15
    """, (uid,)).fetchall()

    items = []
    for r in candidates:
        monthly = r[1]
        yearly = monthly * 12
        # Compound interest projections at 7% annual
        y5 = sum(yearly * (1.07 ** i) for i in range(5))
        y10 = sum(yearly * (1.07 ** i) for i in range(10))
        y20 = sum(yearly * (1.07 ** i) for i in range(20))
        items.append({
            'description': r[0],
            'monthly_avg': round(r[1], 0),
            'months_seen': r[2],
            'category': r[3],
            'yearly': round(yearly, 0),
            'invested_5y': round(y5, 0),
            'invested_10y': round(y10, 0),
            'invested_20y': round(y20, 0),
        })

    # Also allow custom calculation via query params
    custom = None
    custom_amount = request.args.get('amount')
    custom_freq = request.args.get('freq', 'monthly')
    if custom_amount:
        amt = float(custom_amount)
        if custom_freq == 'daily':
            monthly = amt * 30
        elif custom_freq == 'weekly':
            monthly = amt * 4.33
        else:
            monthly = amt
        yearly = monthly * 12
        y5 = sum(yearly * (1.07 ** i) for i in range(5))
        y10 = sum(yearly * (1.07 ** i) for i in range(10))
        y20 = sum(yearly * (1.07 ** i) for i in range(20))
        custom = {
            'monthly': round(monthly, 0),
            'yearly': round(yearly, 0),
            'invested_5y': round(y5, 0),
            'invested_10y': round(y10, 0),
            'invested_20y': round(y20, 0),
        }

    conn.close()
    return jsonify({'items': items, 'custom': custom})


@app.route('/api/insights/anomalies', methods=['GET'])
@login_required
def insights_anomalies():
    """Detect spending anomalies: categories significantly above their average."""
    conn = get_db()
    uid = get_uid()
    month = request.args.get('month', date.today().strftime('%Y-%m'))

    # Get current month totals per category
    current = conn.execute("""
        SELECT c.id, c.name_he, c.color, COALESCE(SUM(e.amount),0) as total
        FROM categories c LEFT JOIN expenses e ON c.id=e.category_id AND e.date LIKE ? AND e.user_id=?
        GROUP BY c.id HAVING total > 0
    """, (month + '%', uid)).fetchall()

    anomalies = []
    for r in current:
        cat_id, name, color, cur_total = r[0], r[1], r[2], r[3]
        # Get average of other months
        hist = conn.execute("""
            SELECT AVG(monthly_total) FROM (
                SELECT SUM(amount) as monthly_total FROM expenses
                WHERE user_id=? AND category_id=? AND substr(date,1,7) != ?
                GROUP BY substr(date,1,7)
            )
        """, (uid, cat_id, month)).fetchone()[0]

        if hist and hist > 0:
            pct_change = ((cur_total - hist) / hist) * 100
            if abs(pct_change) > 25:
                # Find top contributors to the change
                top_items = conn.execute("""
                    SELECT description, SUM(amount) as total FROM expenses
                    WHERE user_id=? AND category_id=? AND date LIKE ? AND description != ''
                    GROUP BY description ORDER BY total DESC LIMIT 3
                """, (uid, cat_id, month + '%')).fetchall()

                anomalies.append({
                    'category': name,
                    'category_id': cat_id,
                    'color': color,
                    'current': round(cur_total, 0),
                    'average': round(hist, 0),
                    'pct_change': round(pct_change, 1),
                    'direction': 'up' if pct_change > 0 else 'down',
                    'top_items': [{'desc': t[0], 'amount': round(t[1], 0)} for t in top_items],
                })

    anomalies.sort(key=lambda a: abs(a['pct_change']), reverse=True)
    conn.close()
    return jsonify(anomalies)


@app.route('/api/insights/recurring', methods=['GET'])
@login_required
def insights_recurring():
    """Auto-detect recurring expenses by pattern analysis."""
    conn = get_db()
    uid = get_uid()

    # Find descriptions that appear in 3+ different months
    recurring = conn.execute("""
        SELECT description, category_id, AVG(amount) as avg_amt,
               MIN(amount) as min_amt, MAX(amount) as max_amt,
               COUNT(*) as count, COUNT(DISTINCT substr(date,1,7)) as months,
               GROUP_CONCAT(DISTINCT substr(date,1,7)) as month_list
        FROM expenses
        WHERE user_id=? AND description != '' AND description IS NOT NULL
        GROUP BY description
        HAVING months >= 3
        ORDER BY avg_amt DESC
    """, (uid,)).fetchall()

    # Also find by subcategory for those without description
    recurring_sub = conn.execute("""
        SELECT subcategory, category_id, AVG(amount) as avg_amt,
               MIN(amount) as min_amt, MAX(amount) as max_amt,
               COUNT(*) as count, COUNT(DISTINCT substr(date,1,7)) as months,
               GROUP_CONCAT(DISTINCT substr(date,1,7)) as month_list
        FROM expenses
        WHERE user_id=? AND (description = '' OR description IS NULL) AND subcategory != ''
        GROUP BY subcategory, category_id
        HAVING months >= 3
        ORDER BY avg_amt DESC
    """, (uid,)).fetchall()

    # Get category names
    cat_names = {r[0]: r[1] for r in conn.execute("SELECT id, name_he FROM categories").fetchall()}

    items = []
    seen = set()
    for r in list(recurring) + list(recurring_sub):
        name = r[0]
        if name in seen:
            continue
        seen.add(name)
        months_list = r[7].split(',') if r[7] else []
        # Determine frequency pattern
        if len(months_list) >= 2:
            sorted_months = sorted(months_list)
            gaps = []
            for i in range(1, len(sorted_months)):
                y1, m1 = map(int, sorted_months[i-1].split('-'))
                y2, m2 = map(int, sorted_months[i].split('-'))
                gap = (y2 - y1) * 12 + (m2 - m1)
                gaps.append(gap)
            avg_gap = sum(gaps) / len(gaps) if gaps else 1
            if avg_gap <= 1.2:
                pattern = 'monthly'
            elif avg_gap <= 2.5:
                pattern = 'bimonthly'
            else:
                pattern = 'irregular'
        else:
            pattern = 'unknown'

        items.append({
            'name': name,
            'category': cat_names.get(r[1], r[1]),
            'category_id': r[1],
            'avg_amount': round(r[2], 0),
            'min_amount': round(r[3], 0),
            'max_amount': round(r[4], 0),
            'count': r[5],
            'months': r[6],
            'pattern': pattern,
            'yearly_cost': round(r[2] * 12 if pattern == 'monthly' else r[2] * 6 if pattern == 'bimonthly' else r[2] * r[6], 0),
        })

    total_monthly = sum(i['avg_amount'] for i in items if i['pattern'] == 'monthly')
    total_yearly = sum(i['yearly_cost'] for i in items)

    conn.close()
    return jsonify({
        'items': items,
        'total_monthly': round(total_monthly, 0),
        'total_yearly': round(total_yearly, 0),
        'count': len(items),
    })


@app.route('/api/insights/whatif', methods=['GET'])
@login_required
def insights_whatif():
    """What-if simulator: provide baseline data for interactive sliders."""
    conn = get_db()
    uid = get_uid()

    months_count = conn.execute(
        "SELECT COUNT(DISTINCT substr(date,1,7)) FROM expenses WHERE user_id=?", (uid,)
    ).fetchone()[0] or 1

    cats = conn.execute("""
        SELECT c.id, c.name_he, c.color, COALESCE(SUM(e.amount),0) as total
        FROM categories c LEFT JOIN expenses e ON c.id=e.category_id AND e.user_id=?
        GROUP BY c.id HAVING total > 0
        ORDER BY total DESC
    """, (uid,)).fetchall()

    income_monthly = conn.execute(
        "SELECT COALESCE(SUM(amount),0) FROM income WHERE user_id=?", (uid,)
    ).fetchone()[0] / months_count

    categories_data = []
    for r in cats:
        monthly_avg = r[3] / months_count
        categories_data.append({
            'id': r[0],
            'name': r[1],
            'color': r[2],
            'monthly_avg': round(monthly_avg, 0),
            'total': round(r[3], 0),
        })

    total_monthly = sum(c['monthly_avg'] for c in categories_data)
    conn.close()
    return jsonify({
        'categories': categories_data,
        'total_monthly_expense': round(total_monthly, 0),
        'monthly_income': round(income_monthly, 0),
        'months_analyzed': months_count,
    })


@app.route('/api/insights/weekly-pulse', methods=['GET'])
@login_required
def insights_weekly_pulse():
    """Average spending by day of the week."""
    conn = get_db()
    uid = get_uid()

    # SQLite strftime('%w') = 0(Sun) to 6(Sat)
    raw = conn.execute(
        "SELECT CAST(strftime('%w', date) AS INTEGER) as dow,"
        " SUM(amount) as total, COUNT(*) as count,"
        " COUNT(DISTINCT date) as days_count"
        " FROM expenses WHERE user_id=? GROUP BY dow ORDER BY dow",
        (uid,)
    ).fetchall()

    day_names = ['ראשון', 'שני', 'שלישי', 'רביעי', 'חמישי', 'שישי', 'שבת']
    days = []
    amounts = []
    for r in raw:
        avg_per_day = r[1] / r[3] if r[3] > 0 else 0
        days.append({
            'dow': r[0],
            'name': day_names[r[0]],
            'total': round(r[1], 0),
            'avg_per_day': round(avg_per_day, 0),
            'transaction_count': r[2],
            'days_count': r[3],
        })
        amounts.append(avg_per_day)

    # Find peak and low days
    if amounts:
        max_idx = amounts.index(max(amounts))
        min_idx = amounts.index(min(amounts))
        peak_day = days[max_idx]['name'] if days else ''
        low_day = days[min_idx]['name'] if days else ''
    else:
        peak_day = low_day = ''

    conn.close()
    return jsonify({
        'days': days,
        'peak_day': peak_day,
        'low_day': low_day,
    })


@app.route('/api/insights/projection', methods=['GET'])
@login_required
def insights_projection():
    """12-month forward projection based on trends."""
    conn = get_db()
    uid = get_uid()

    # Get monthly data (normalized — excludes unusual + excluded months for regression)
    excl_months = get_excluded_month_set(conn, uid)
    monthly_raw = conn.execute("""
        SELECT substr(date,1,7) as m, SUM(amount) as total
        FROM expenses WHERE user_id=? AND is_unusual=0 GROUP BY m ORDER BY m
    """, (uid,)).fetchall()
    monthly = [r for r in monthly_raw if r[0] not in excl_months]

    income_monthly_raw = conn.execute("""
        SELECT substr(date,1,7) as m, SUM(amount) as total
        FROM income WHERE user_id=? GROUP BY m ORDER BY m
    """, (uid,)).fetchall()
    income_monthly = [r for r in income_monthly_raw if r[0] not in excl_months]

    inc_map = {r[0]: r[1] for r in income_monthly}

    # Bank balance history
    bank_balances = conn.execute(
        "SELECT month, closing_balance FROM bank_balances WHERE user_id=? AND account_name='main' ORDER BY month",
        (uid,)
    ).fetchall()
    bal_map = {r[0]: r[1] for r in bank_balances}

    if len(monthly) < 2:
        conn.close()
        return jsonify({'error': 'Not enough data for projection', 'history': [], 'projection': []})

    # Calculate trend using simple linear regression
    exp_values = [r[1] for r in monthly]
    n = len(exp_values)
    x_mean = (n - 1) / 2
    y_mean = sum(exp_values) / n
    numerator = sum((i - x_mean) * (v - y_mean) for i, v in enumerate(exp_values))
    denominator = sum((i - x_mean) ** 2 for i in range(n))
    slope = numerator / denominator if denominator else 0
    intercept = y_mean - slope * x_mean

    # Average income
    inc_values = list(inc_map.values())
    avg_income = sum(inc_values) / len(inc_values) if inc_values else 0

    # Bank balance projection: only if 3+ months and low volatility
    bal_values = [r[1] for r in bank_balances]
    bank_projection_suppressed = True
    b_slope = 0
    b_intercept = 0
    bn = len(bal_values)
    if bn >= 3:
        b_mean = sum(bal_values) / bn
        b_x_mean = (bn - 1) / 2
        b_num = sum((i - b_x_mean) * (v - b_mean) for i, v in enumerate(bal_values))
        b_den = sum((i - b_x_mean) ** 2 for i in range(bn))
        b_slope = b_num / b_den if b_den else 0
        b_intercept = b_mean - b_slope * b_x_mean
        # Check volatility: stddev < 30% of mean
        if b_mean != 0:
            variance = sum((v - b_mean) ** 2 for v in bal_values) / bn
            stddev = variance ** 0.5
            if stddev / abs(b_mean) < 0.30:
                bank_projection_suppressed = False

    # Build history
    history = []
    for r in monthly:
        inc = inc_map.get(r[0], 0)
        history.append({
            'month': r[0],
            'expenses': round(r[1], 0),
            'income': round(inc, 0),
            'balance': round(inc - r[1], 0),
            'bank_balance': round(bal_map[r[0]], 0) if r[0] in bal_map else None,
        })

    # Project 12 months forward
    last_month = monthly[-1][0]
    ly, lm = map(int, last_month.split('-'))
    projection = []
    cumulative_savings = 0
    for i in range(1, 13):
        pm = lm + i
        py = ly
        while pm > 12:
            pm -= 12
            py += 1
        proj_month = f"{py}-{pm:02d}"
        proj_expense = max(intercept + slope * (n - 1 + i), 0)
        proj_balance = avg_income - proj_expense
        cumulative_savings += proj_balance
        proj_bank = round(b_intercept + b_slope * (bn - 1 + i), 0) if not bank_projection_suppressed else None
        projection.append({
            'month': proj_month,
            'expenses': round(proj_expense, 0),
            'income': round(avg_income, 0),
            'balance': round(proj_balance, 0),
            'cumulative': round(cumulative_savings, 0),
            'bank_balance': proj_bank,
        })

    # "Why" inputs for frontend explainability
    trend_dir = 'up' if slope > 50 else 'down' if slope < -50 else 'stable'
    why_inputs = {
        'months_of_data': n,
        'trend_direction': trend_dir,
        'monthly_trend': round(abs(slope), 0),
        'avg_income': round(avg_income, 0),
    }
    if not bank_projection_suppressed:
        why_inputs['bank_months_of_data'] = bn
        why_inputs['bank_monthly_trend'] = round(b_slope, 0)

    conn.close()
    return jsonify({
        'history': history,
        'projection': projection,
        'trend_direction': trend_dir,
        'monthly_trend': round(slope, 0),
        'avg_income': round(avg_income, 0),
        'projected_yearly_savings': round(sum(p['balance'] for p in projection), 0),
        'bank_projection_suppressed': bank_projection_suppressed,
        'note': 'Estimates based on recent trends — actual results may vary',
        'why_inputs': why_inputs,
    })


@app.route('/api/insights/comparison', methods=['GET'])
@login_required
def insights_comparison():
    """Compare spending to Israeli CBS household averages."""
    conn = get_db()
    uid = get_uid()

    # Normalized — excludes unusual + excluded months
    excl_months = get_excluded_month_set(conn, uid)
    months_rows = conn.execute(
        "SELECT DISTINCT substr(date,1,7) as m FROM expenses WHERE user_id=? AND is_unusual=0", (uid,)
    ).fetchall()
    counted_months = [r['m'] for r in months_rows if r['m'] not in excl_months]
    months_count = max(len(counted_months), 1)

    total_expense = conn.execute(
        "SELECT COALESCE(SUM(amount),0) FROM expenses WHERE user_id=? AND is_unusual=0", (uid,)
    ).fetchone()[0]
    monthly_total = total_expense / months_count

    cats = conn.execute("""
        SELECT c.id, c.name_he, c.color, COALESCE(SUM(e.amount),0) as total
        FROM categories c LEFT JOIN expenses e ON c.id=e.category_id AND e.user_id=? AND e.is_unusual=0
        GROUP BY c.id HAVING total > 0
        ORDER BY total DESC
    """, (uid,)).fetchall()

    comparisons = []
    for r in cats:
        cat_id, name, color, total = r[0], r[1], r[2], r[3]
        my_pct = (total / total_expense * 100) if total_expense > 0 else 0
        cbs_pct = CBS_AVERAGES.get(cat_id, 0)
        diff = my_pct - cbs_pct
        monthly = total / months_count

        comparisons.append({
            'category': name,
            'category_id': cat_id,
            'color': color,
            'my_pct': round(my_pct, 1),
            'cbs_pct': cbs_pct,
            'diff': round(diff, 1),
            'monthly_avg': round(monthly, 0),
            'status': 'high' if diff > 5 else 'low' if diff < -5 else 'normal',
        })

    comparisons.sort(key=lambda c: abs(c['diff']), reverse=True)
    conn.close()
    return jsonify({
        'comparisons': comparisons,
        'monthly_total': round(monthly_total, 0),
        'months_analyzed': months_count,
    })


@app.route('/api/insights/achievements', methods=['GET'])
@login_required
def insights_achievements():
    """Budget streaks and gamification achievements."""
    conn = get_db()
    uid = get_uid()

    achievements = []

    # Get monthly data (normalized — excludes unusual + excluded months)
    excl_months = get_excluded_month_set(conn, uid)
    monthly_raw = conn.execute("""
        SELECT substr(date,1,7) as m, SUM(amount) as total
        FROM expenses WHERE user_id=? AND is_unusual=0 GROUP BY m ORDER BY m
    """, (uid,)).fetchall()
    monthly = [r for r in monthly_raw if r[0] not in excl_months]

    # Income by month
    income_data_raw = conn.execute("""
        SELECT substr(date,1,7) as m, SUM(amount) as total
        FROM income WHERE user_id=? GROUP BY m ORDER BY m
    """, (uid,)).fetchall()
    income_data = [r for r in income_data_raw if r[0] not in excl_months]
    inc_map = {r[0]: r[1] for r in income_data}

    # 1. Surplus streak
    surplus_streak = 0
    max_surplus_streak = 0
    for r in monthly:
        inc = inc_map.get(r[0], 0)
        if inc > r[1]:
            surplus_streak += 1
            max_surplus_streak = max(max_surplus_streak, surplus_streak)
        else:
            surplus_streak = 0

    if max_surplus_streak >= 2:
        achievements.append({
            'id': 'surplus_streak',
            'icon': 'bi-trophy',
            'color': '#f59e0b',
            'title': f'רצף חיובי - {max_surplus_streak} חודשים ברצף בעודף!',
            'description': f'הצלחתם לסיים {max_surplus_streak} חודשים ברצף עם יותר הכנסות מהוצאות.',
            'current': surplus_streak,
            'best': max_surplus_streak,
            'type': 'streak',
            'unlocked': True,
        })

    # 2. Best savings month
    best_savings = None
    best_savings_month = ''
    for r in monthly:
        inc = inc_map.get(r[0], 0)
        savings = inc - r[1]
        if best_savings is None or savings > best_savings:
            best_savings = savings
            best_savings_month = r[0]

    if best_savings and best_savings > 0:
        achievements.append({
            'id': 'best_savings',
            'icon': 'bi-star',
            'color': '#16a34a',
            'title': f'שיא חיסכון - {best_savings:,.0f} ש"ח!',
            'description': f'החודש הכי טוב שלכם היה {best_savings_month} עם חיסכון של {best_savings:,.0f} ש"ח.',
            'value': best_savings,
            'month': best_savings_month,
            'type': 'record',
            'unlocked': True,
        })

    # 3. Lowest expense month
    if monthly:
        lowest = min(monthly, key=lambda r: r[1])
        achievements.append({
            'id': 'lowest_expense',
            'icon': 'bi-arrow-down-circle',
            'color': '#2563eb',
            'title': f'חודש חסכוני - {lowest[1]:,.0f} ש"ח בלבד!',
            'description': f'ב-{lowest[0]} הוצאתם הכי מעט.',
            'value': lowest[1],
            'month': lowest[0],
            'type': 'record',
            'unlocked': True,
        })

    # 4. No overdraft streak
    overdraft_months = set()
    od_data = conn.execute("""
        SELECT DISTINCT substr(date,1,7) FROM expenses WHERE user_id=? AND subcategory='ריבית מינוס'
    """, (uid,)).fetchall()
    overdraft_months = {r[0] for r in od_data}

    all_months = [r[0] for r in monthly]
    no_od_streak = 0
    max_no_od = 0
    for m in all_months:
        if m not in overdraft_months:
            no_od_streak += 1
            max_no_od = max(max_no_od, no_od_streak)
        else:
            no_od_streak = 0

    if max_no_od >= 2:
        achievements.append({
            'id': 'no_overdraft',
            'icon': 'bi-shield-check',
            'color': '#8b5cf6',
            'title': f'ללא מינוס - {max_no_od} חודשים!',
            'description': f'{max_no_od} חודשים בלי ריבית מינוס. כל שקל ריבית שנחסך = כסף בכיס!',
            'current': no_od_streak,
            'best': max_no_od,
            'type': 'streak',
            'unlocked': True,
        })

    # 5. Category improvement: check if any category went down vs 3-month average
    if len(all_months) >= 4:
        latest = all_months[-1]
        prev3 = all_months[-4:-1]
        cats = conn.execute("SELECT id, name_he FROM categories").fetchall()
        for cat_id, cat_name in cats:
            cur = conn.execute(
                "SELECT COALESCE(SUM(amount),0) FROM expenses WHERE user_id=? AND category_id=? AND substr(date,1,7)=?",
                (uid, cat_id, latest)
            ).fetchone()[0]
            prev_avg_val = 0
            for pm in prev3:
                prev_avg_val += conn.execute(
                    "SELECT COALESCE(SUM(amount),0) FROM expenses WHERE user_id=? AND category_id=? AND substr(date,1,7)=?",
                    (uid, cat_id, pm)
                ).fetchone()[0]
            prev_avg_val /= 3
            if prev_avg_val > 500 and cur < prev_avg_val * 0.7:
                saved = prev_avg_val - cur
                achievements.append({
                    'id': f'improved_{cat_id}',
                    'icon': 'bi-graph-down-arrow',
                    'color': '#16a34a',
                    'title': f'שיפור ב{cat_name} - חסכתם {saved:,.0f} ש"ח!',
                    'description': f'הוצאתם ב-{cat_name} ירדו מממוצע {prev_avg_val:,.0f} ל-{cur:,.0f} ש"ח.',
                    'value': saved,
                    'type': 'improvement',
                    'unlocked': True,
                })

    # 6. Locked achievements (goals to work toward)
    if max_surplus_streak < 6:
        achievements.append({
            'id': 'surplus_6',
            'icon': 'bi-lock',
            'color': '#94a3b8',
            'title': 'אתגר: 6 חודשים ברצף בעודף',
            'description': f'עוד {6 - max_surplus_streak} חודשים בעודף כדי לפתוח את ההישג הזה!',
            'progress': max_surplus_streak,
            'target': 6,
            'type': 'challenge',
            'unlocked': False,
        })

    total_income = sum(inc_map.values())
    total_expense = sum(r[1] for r in monthly)
    savings_rate = ((total_income - total_expense) / total_income * 100) if total_income > 0 else 0

    if savings_rate < 20:
        achievements.append({
            'id': 'savings_20',
            'icon': 'bi-lock',
            'color': '#94a3b8',
            'title': 'אתגר: 20% חיסכון',
            'description': f'שיעור החיסכון שלכם {savings_rate:.1f}%. הגיעו ל-20% כדי לפתוח!',
            'progress': round(savings_rate, 1),
            'target': 20,
            'type': 'challenge',
            'unlocked': False,
        })

    conn.close()
    return jsonify(achievements)


# ============================================================
# Installment Payments
# ============================================================

@app.route('/api/cards', methods=['GET'])
@login_required
def cards_list():
    conn = get_db()
    rows = conn.execute("SELECT DISTINCT card FROM expenses WHERE user_id=? AND card != '' ORDER BY card", (get_uid(),)).fetchall()
    conn.close()
    return jsonify([r['card'] for r in rows])


@app.route('/api/installments', methods=['GET'])
@login_required
def installments_list():
    uid = get_uid()
    conn = get_db()
    rows = conn.execute("SELECT * FROM installments WHERE user_id=? ORDER BY start_date DESC", (uid,)).fetchall()
    result = []
    for r in rows:
        d = dict(r)
        d['payments_remaining'] = max(d['total_payments'] - d['payments_made'], 0)
        d['remaining_amount'] = round(d['monthly_payment'] * d['payments_remaining'], 2)
        d['linked_count'] = conn.execute(
            "SELECT COUNT(*) FROM installment_transaction_links WHERE installment_id=? AND user_id=? AND status IN ('confirmed','auto_matched')",
            (d['id'], uid)).fetchone()[0]
        if not d.get('status'):
            d['status'] = 'active'
        if not d.get('source'):
            d['source'] = 'manual'
        result.append(d)
    conn.close()
    return jsonify(result)


@app.route('/api/installments', methods=['POST'])
@login_required
def installments_add():
    data = request.json
    total_amount = float(data.get('total_amount', 0))
    total_payments = int(data.get('total_payments', 1))
    monthly_payment = round(total_amount / total_payments, 2) if total_payments > 0 else total_amount
    conn = get_db()

    uid = get_uid()
    if data.get('id'):
        conn.execute("""
            UPDATE installments SET description=?, store=?, total_amount=?,
                total_payments=?, payments_made=?, monthly_payment=?,
                start_date=?, card=?, notes=?
            WHERE id=? AND user_id=?
        """, (data['description'], data.get('store', ''), total_amount,
              total_payments, int(data.get('payments_made', 0)), monthly_payment,
              data.get('start_date', ''), data.get('card', ''),
              data.get('notes', ''), data['id'], uid))
    else:
        source = data.get('source', 'manual')
        vendor_norm = data.get('vendor_normalized', '')
        if not vendor_norm and data.get('store'):
            vendor_norm = _normalize_vendor(data['store'])
        cur = conn.execute("""
            INSERT INTO installments (description, store, total_amount, total_payments,
                payments_made, monthly_payment, start_date, card, notes, user_id,
                source, vendor_normalized, status)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'active')
        """, (data['description'], data.get('store', ''), total_amount,
              total_payments, int(data.get('payments_made', 0)), monthly_payment,
              data.get('start_date', ''), data.get('card', ''),
              data.get('notes', ''), uid, source, vendor_norm))
        new_id = cur.lastrowid

    conn.commit()
    conn.close()
    return jsonify({'status': 'ok', 'id': new_id if not data.get('id') else int(data['id'])})


@app.route('/api/installments/<int:inst_id>', methods=['DELETE'])
@login_required
def installments_delete(inst_id):
    uid = get_uid()
    conn = get_db()
    conn.execute("DELETE FROM installment_transaction_links WHERE installment_id=? AND user_id=?", (inst_id, uid))
    conn.execute("DELETE FROM installments WHERE id=? AND user_id=?", (inst_id, uid))
    conn.commit()
    conn.close()
    return jsonify({'status': 'ok'})


@app.route('/api/installments/<int:inst_id>/linked-transactions', methods=['GET'])
@login_required
def installment_linked_transactions(inst_id):
    uid = get_uid()
    conn = get_db()
    links = conn.execute("""
        SELECT itl.id as link_id, itl.expense_id, itl.status, itl.confidence,
               itl.match_reasons_json, itl.payment_number,
               e.date, e.description, e.amount, e.card
        FROM installment_transaction_links itl
        JOIN expenses e ON itl.expense_id = e.id
        WHERE itl.installment_id=? AND itl.user_id=?
        ORDER BY e.date ASC
    """, (inst_id, uid)).fetchall()
    conn.close()
    return jsonify([dict(r) for r in links])


@app.route('/api/installments/<int:inst_id>/unlink-transaction', methods=['POST'])
@login_required
def installment_unlink_transaction(inst_id):
    uid = get_uid()
    data = request.json or {}
    expense_id = data.get('expense_id')
    if not expense_id:
        return jsonify({'error': 'Missing expense_id'}), 400
    conn = get_db()
    deleted = conn.execute(
        "DELETE FROM installment_transaction_links WHERE installment_id=? AND expense_id=? AND user_id=?",
        (inst_id, expense_id, uid)).rowcount
    if deleted:
        conn.execute("""
            UPDATE installments SET payments_made = MAX(payments_made - 1, 0),
            updated_at=CURRENT_TIMESTAMP WHERE id=? AND user_id=?
        """, (inst_id, uid))
    conn.commit()
    conn.close()
    return jsonify({'status': 'ok', 'unlinked': deleted})


@app.route('/api/installments/<int:inst_id>/complete-early', methods=['PUT'])
@login_required
def installment_complete_early(inst_id):
    conn = get_db()
    conn.execute(
        "UPDATE installments SET status='completed', updated_at=CURRENT_TIMESTAMP WHERE id=? AND user_id=?",
        (inst_id, get_uid()))
    conn.commit()
    conn.close()
    return jsonify({'status': 'ok'})


# ── Installment Suggestion Endpoints ────────────────────────────

@app.route('/api/installment-suggestions/scan', methods=['POST'])
@login_required
def installment_scan():
    result = scan_installment_suggestions(get_uid())
    return jsonify(result)


@app.route('/api/installment-suggestions', methods=['GET'])
@login_required
def installment_suggestions_list():
    uid = get_uid()
    status = request.args.get('status', 'suggested')
    conn = get_db()
    rows = conn.execute(
        "SELECT * FROM installment_suggestions WHERE user_id=? AND status=? ORDER BY confidence_score DESC",
        (uid, status)).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route('/api/installment-suggestions/<int:sug_id>', methods=['PUT'])
@login_required
def installment_suggestion_resolve(sug_id):
    uid = get_uid()
    data = request.json or {}
    action = data.get('action', '')
    if action not in ('confirm', 'reject', 'ignore'):
        return jsonify({'error': 'Invalid action'}), 400

    conn = get_db()
    sug = conn.execute(
        "SELECT * FROM installment_suggestions WHERE id=? AND user_id=?",
        (sug_id, uid)).fetchone()
    if not sug:
        conn.close()
        return jsonify({'error': 'Not found'}), 404

    sug = dict(sug)
    if sug['status'] != 'suggested':
        conn.close()
        return jsonify({'error': 'Already resolved'}), 409

    if action == 'confirm':
        overrides = data.get('overrides', {})
        description = overrides.get('description', sug['vendor_display'])
        store = overrides.get('store', sug['vendor_display'])
        total_payments = int(overrides.get('total_payments', sug['estimated_total_payments']) or sug['estimated_total_payments'])
        payments_made = int(overrides.get('payments_made', sug['estimated_payments_made']) or sug['estimated_payments_made'])
        start_date = overrides.get('start_date', sug['estimated_start_date'])
        card = overrides.get('card', sug['card'])
        notes = overrides.get('notes', '')
        monthly_payment = sug['monthly_amount']
        total_amount = round(monthly_payment * total_payments, 2)
        vendor_norm = sug['vendor_normalized']

        # Allow user to deselect certain expense IDs
        selected_ids = data.get('selected_expense_ids')
        all_ids = json.loads(sug['expense_ids'])
        if selected_ids is not None:
            link_ids = [eid for eid in selected_ids if eid in all_ids]
        else:
            link_ids = all_ids

        if not link_ids:
            conn.close()
            return jsonify({'error': 'No expenses selected'}), 400

        # Create installment plan
        cur = conn.execute("""
            INSERT INTO installments (description, store, total_amount, total_payments,
                payments_made, monthly_payment, start_date, card, notes, user_id,
                source, vendor_normalized, status)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (description, store, total_amount, total_payments,
              len(link_ids), monthly_payment, start_date, card, notes, uid,
              'detected', vendor_norm, 'active'))
        inst_id = cur.lastrowid

        # Backlink selected expenses
        link_expenses = conn.execute(
            f"SELECT id, date FROM expenses WHERE id IN ({','.join('?' * len(link_ids))}) AND user_id=? ORDER BY date ASC",
            link_ids + [uid]).fetchall()
        for i, le in enumerate(link_expenses, 1):
            conn.execute("""
                INSERT OR IGNORE INTO installment_transaction_links
                (user_id, installment_id, expense_id, status, confidence, match_reasons_json, payment_number)
                VALUES (?,?,?,?,?,?,?)
            """, (uid, inst_id, le['id'], 'confirmed', 1.0, '["user_confirmed"]', i))

        # Update suggestion
        conn.execute("""
            UPDATE installment_suggestions SET status='confirmed', linked_installment_id=?,
            updated_at=CURRENT_TIMESTAMP WHERE id=?
        """, (inst_id, sug_id))

        conn.commit()
        conn.close()
        return jsonify({'status': 'ok', 'installment_id': inst_id})

    elif action == 'reject':
        conn.execute(
            "UPDATE installment_suggestions SET status='rejected', updated_at=CURRENT_TIMESTAMP WHERE id=?",
            (sug_id,))
        conn.commit()
        conn.close()
        return jsonify({'status': 'ok'})

    else:  # ignore
        conn.execute(
            "UPDATE installment_suggestions SET status='ignored', updated_at=CURRENT_TIMESTAMP WHERE id=?",
            (sug_id,))
        # Optionally create ignore rule
        if data.get('add_rule'):
            vendor = sug['vendor_normalized']
            reason = data.get('reason', '')
            try:
                conn.execute(
                    "INSERT OR IGNORE INTO installment_ignore_rules (user_id, rule_type, rule_value, reason) VALUES (?,?,?,?)",
                    (uid, 'never_suggest', vendor, reason))
            except Exception:
                pass
            # Bulk-ignore other suggestions for same vendor
            conn.execute("""
                UPDATE installment_suggestions SET status='ignored', updated_at=CURRENT_TIMESTAMP
                WHERE user_id=? AND vendor_normalized=? AND status='suggested' AND id!=?
            """, (uid, sug['vendor_normalized'], sug_id))
        conn.commit()
        conn.close()
        return jsonify({'status': 'ok'})


# ── Installment Ignore Rules Endpoints ──────────────────────────

@app.route('/api/installment-ignore-rules', methods=['GET'])
@login_required
def installment_ignore_rules_list():
    conn = get_db()
    rows = conn.execute(
        "SELECT * FROM installment_ignore_rules WHERE user_id=? ORDER BY created_at DESC",
        (get_uid(),)).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route('/api/installment-ignore-rules', methods=['POST'])
@login_required
def installment_ignore_rules_add():
    data = request.json or {}
    rule_type = data.get('rule_type', 'never_suggest')
    rule_value = data.get('rule_value', '')
    if not rule_value or rule_type not in ('never_suggest', 'always_installment'):
        return jsonify({'error': 'Invalid rule'}), 400
    conn = get_db()
    try:
        conn.execute(
            "INSERT OR IGNORE INTO installment_ignore_rules (user_id, rule_type, rule_value, reason) VALUES (?,?,?,?)",
            (get_uid(), rule_type, _normalize_text(rule_value), data.get('reason', '')))
        conn.commit()
    except Exception:
        pass
    conn.close()
    return jsonify({'status': 'ok'})


@app.route('/api/installment-ignore-rules/<int:rule_id>', methods=['DELETE'])
@login_required
def installment_ignore_rules_delete(rule_id):
    conn = get_db()
    conn.execute("DELETE FROM installment_ignore_rules WHERE id=? AND user_id=?", (rule_id, get_uid()))
    conn.commit()
    conn.close()
    return jsonify({'status': 'ok'})


# ============================================================
# Admin Panel APIs
# ============================================================

@app.route('/api/admin/users', methods=['GET'])
@admin_required
def admin_get_users():
    conn = get_db()
    rows = conn.execute("SELECT id, username, email, phone, verified, is_admin, created_at FROM users ORDER BY id").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route('/api/admin/users/<int:user_id>', methods=['DELETE'])
@admin_required
def admin_delete_user(user_id):
    conn = get_db()
    user = conn.execute("SELECT is_admin FROM users WHERE id=?", (user_id,)).fetchone()
    if user and user['is_admin']:
        conn.close()
        return jsonify({'error': 'Cannot delete admin user'}), 400
    # Cascade delete all user data
    for tbl in ['transaction_links', 'link_rules', 'installment_transaction_links', 'installment_suggestions', 'installment_ignore_rules', 'expenses', 'income', 'budget', 'budget_plans', 'installments', 'reminders', 'financial_products', 'bank_balances', 'excluded_months', 'insurance_suggestions', 'insurance_blacklist', 'insurance_ignore_rules', 'insurance_overlap_alerts', 'assets', 'liabilities', 'net_worth_snapshots', 'tip_events']:
        conn.execute(f"DELETE FROM {tbl} WHERE user_id=?", (user_id,))
    conn.execute("DELETE FROM users WHERE id=?", (user_id,))
    conn.commit()
    conn.close()
    return jsonify({'status': 'ok'})


@app.route('/api/admin/categories', methods=['PUT'])
@admin_required
def admin_update_category():
    data = request.json
    conn = get_db()
    conn.execute("UPDATE categories SET name_he=?, color=? WHERE id=?",
                 (data['name_he'], data['color'], data['id']))
    conn.commit()
    conn.close()
    return jsonify({'status': 'ok'})


@app.route('/api/admin/categories/<cat_id>', methods=['DELETE'])
@admin_required
def admin_delete_category(cat_id):
    conn = get_db()
    # Check if category has expenses
    count = conn.execute("SELECT COUNT(*) FROM expenses WHERE category_id=?", (cat_id,)).fetchone()[0]
    if count > 0:
        conn.close()
        return jsonify({'error': f'Category has {count} expenses, cannot delete'}), 400
    conn.execute("DELETE FROM categories WHERE id=?", (cat_id,))
    conn.execute("DELETE FROM budget WHERE category_id=?", (cat_id,))
    conn.commit()
    conn.close()
    return jsonify({'status': 'ok'})


@app.route('/api/admin/stats', methods=['GET'])
@admin_required
def admin_stats():
    conn = get_db()
    stats = {
        'total_users': conn.execute("SELECT COUNT(*) FROM users").fetchone()[0],
        'total_expenses': conn.execute("SELECT COUNT(*) FROM expenses").fetchone()[0],
        'total_income': conn.execute("SELECT COUNT(*) FROM income").fetchone()[0],
        'total_categories': conn.execute("SELECT COUNT(*) FROM categories").fetchone()[0],
        'total_budget_plans': conn.execute("SELECT COUNT(*) FROM budget_plans").fetchone()[0],
        'db_size': os.path.getsize(DB_PATH) if os.path.exists(DB_PATH) else 0,
    }
    conn.close()
    return jsonify(stats)


import re as _re

# --- AI Chat Agent ---
AI_CONFIG_PATH = os.path.join(DATA_DIR, 'ai_config.json')

def _get_ai_key():
    if os.path.exists(AI_CONFIG_PATH):
        with open(AI_CONFIG_PATH, 'r') as f:
            return json.load(f).get('anthropic_api_key', '')
    return ''

@app.route('/api/settings/ai', methods=['GET'])
@login_required
def get_ai_settings():
    key = _get_ai_key()
    return jsonify({'has_key': bool(key), 'key_preview': key[:8] + '...' if len(key) > 8 else ''})

@app.route('/api/settings/ai', methods=['POST'])
@login_required
def save_ai_settings():
    data = request.json
    key = data.get('api_key', '').strip()
    with open(AI_CONFIG_PATH, 'w') as f:
        json.dump({'anthropic_api_key': key}, f)
    return jsonify({'status': 'ok'})

@app.route('/api/version')
def get_version():
    return jsonify({'version': APP_VERSION})

def _ai_chat(query, lang, conn, uid=None):
    """Use Claude API to understand the query and generate a structured SQL response."""
    api_key = _get_ai_key()
    if not api_key:
        return None

    try:
        import anthropic
        client = anthropic.Anthropic(api_key=api_key)
    except Exception:
        return None

    today_str = date.today().strftime('%Y-%m-%d')
    cats = [dict(r) for r in conn.execute("SELECT id, name_he FROM categories ORDER BY sort_order").fetchall()]
    cats_desc = ', '.join(f'{c["id"]}={c["name_he"]}' for c in cats)
    cards = [r[0] for r in conn.execute("SELECT DISTINCT card FROM expenses WHERE user_id=? AND card IS NOT NULL", (uid,)).fetchall()]

    system_prompt = f"""You are a budget assistant for a Hebrew/English family expense tracker app.
Today's date: {today_str}

DATABASE SCHEMA:
- expenses: id, date (TEXT YYYY-MM-DD), category_id (TEXT), subcategory (TEXT), description (TEXT), amount (REAL), source, frequency, card (TEXT), user_id (INTEGER), created_at
- income: id, date (TEXT YYYY-MM-DD), person (TEXT), source (TEXT), amount (REAL), description (TEXT), is_recurring, user_id (INTEGER), created_at
- categories: id (TEXT PK), name_he (TEXT)

CATEGORIES: {cats_desc}
CARD/SOURCE VALUES: {', '.join(cards)}

YOUR JOB: Parse the user's natural language query and return a JSON object with:
1. "sql" - a SELECT SQL query to answer the question. Use parameterized ? placeholders. For expenses always JOIN categories: SELECT e.*, c.name_he as category_name, c.color as category_color FROM expenses e JOIN categories c ON e.category_id = c.id WHERE ...
2. "params" - array of parameter values for the SQL
3. "text" - a human-readable response in {"Hebrew" if lang == "he" else "English"}. Include the total amount (₪), count, and date range. If showing top items, list them.
4. "action" - navigation action object or null:
   - For expense queries: {{"tab": "expensesTab", "filter": {{"category_id": "...", "subcategory": "...", "from_date": "...", "to_date": "..."}}}}
   - For income queries: {{"tab": "incomeTab"}}
   - For navigation: {{"tab": "dashboardTab|expensesTab|incomeTab|budgetTab|financialTab|insightsTab|analysisTab|importTab"}}
5. "type" - "expenses", "income", or "navigate"

IMPORTANT RULES:
- CRITICAL: ALWAYS add "AND e.user_id = ?" to expense queries and "AND user_id = ?" to income queries. The user_id value ({uid}) must be the FIRST element in the params array.
- The description field contains business names exactly as they appear (e.g. 'מובילנד בתי קולנוע בעם', 'קצביית שור הבר בע"מ'). Search with LIKE '%term%' using the EXACT business name the user provides.
- For date ranges: "חצי שנה" = 6 months, "שנה" = 12 months, "חודש" = 1 month, "3 חודשים" = 3 months. Calculate from today.
- Always ORDER BY date DESC and LIMIT 50 for item queries.
- For totals, use a separate COUNT(*) and SUM(amount) query approach — just put them in the text.
- Keep the SQL simple and correct. No CTEs or subqueries unless necessary.

Respond with ONLY valid JSON, no markdown, no explanation."""

    try:
        msg = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=1024,
            system=system_prompt,
            messages=[{"role": "user", "content": query}]
        )
        raw = msg.content[0].text.strip()
        # Strip markdown code fences if present
        if raw.startswith('```'):
            raw = raw.split('\n', 1)[1] if '\n' in raw else raw[3:]
            if raw.endswith('```'):
                raw = raw[:-3]
        result = json.loads(raw)
    except Exception as e:
        app.logger.error(f"AI chat error: {e}")
        return None

    # Execute the AI-generated SQL safely
    try:
        sql = result.get('sql', '')
        params = result.get('params', [])
        if not sql:
            return result  # Navigation or text-only response

        # Safety: ensure user_id filtering is present
        if ('expenses' in sql.lower() or 'income' in sql.lower()) and 'user_id' not in sql:
            sql = sql.replace('WHERE', f'WHERE user_id = {uid} AND', 1) if 'WHERE' in sql else sql

        rows = conn.execute(sql, params).fetchall()
        items = [dict(r) for r in rows]

        # Get totals
        total = sum(it.get('amount', 0) for it in items)
        count = len(items)

        # If AI returned a count/sum query, use those results differently
        if items and 'amount' not in items[0] and len(items) == 1:
            # Aggregate query result
            pass
        else:
            result_type = result.get('type', 'expenses')
            response = {
                'text': result.get('text', ''),
                'data': {'items': items[:20], 'total': total, 'count': count, 'type': result_type},
                'action': result.get('action')
            }
            return response

    except Exception as e:
        app.logger.error(f"AI SQL execution error: {e}")
        # Return the text response even if SQL fails
        return {
            'text': result.get('text', 'שגיאה בביצוע השאילתה' if lang == 'he' else 'Error executing query'),
            'data': None,
            'action': result.get('action')
        }

    return result


@app.route('/api/chat', methods=['POST'])
@login_required
def chat_assistant():
    """Smart budget assistant — uses AI when available, falls back to keyword matching."""
    data = request.json
    query = data.get('query', '').strip()
    lang = data.get('lang', 'he')

    conn = get_db()

    uid = get_uid()

    # Try AI agent first
    ai_result = _ai_chat(query, lang, conn, uid=uid)
    if ai_result:
        conn.close()
        return jsonify(ai_result)

    # --- Fallback: keyword-based parsing ---
    query_lower = query.lower()
    cats = {r['id']: dict(r) for r in conn.execute("SELECT * FROM categories").fetchall()}
    response = {'text': '', 'data': None, 'action': None}

    today = date.today()
    from_date = None
    to_date = today.strftime('%Y-%m-%d')

    m = _re.search(r'(\d+)\s*(חודש|חדש|month)', query_lower)
    if m:
        n = int(m.group(1))
        from_dt = date(today.year, today.month, 1)
        for _ in range(n - 1):
            from_dt = date(from_dt.year - (1 if from_dt.month == 1 else 0),
                           12 if from_dt.month == 1 else from_dt.month - 1, 1)
        from_date = from_dt.strftime('%Y-%m-%d')

    if any(w in query_lower for w in ['this month', 'החודש', 'החודש הזה']):
        from_date = today.strftime('%Y-%m-01')
    if any(w in query_lower for w in ['this year', 'השנה']):
        from_date = f'{today.year}-01-01'
    if any(w in query_lower for w in ['חצי שנה', 'half year', 'half a year']):
        from_dt = today - timedelta(days=180)
        from_date = from_dt.strftime('%Y-%m-%d')

    if not from_date:
        from_dt = date(today.year, today.month, 1)
        for _ in range(2):
            from_dt = date(from_dt.year - (1 if from_dt.month == 1 else 0),
                           12 if from_dt.month == 1 else from_dt.month - 1, 1)
        from_date = from_dt.strftime('%Y-%m-%d')

    cat_id = None
    cat_name = ''
    for cid, c in cats.items():
        if c['name_he'] and c['name_he'].lower() in query_lower:
            cat_id = cid
            cat_name = c['name_he']
            break

    # Simple search: remove common words, keep the rest
    noise = {'כמה', 'הוצאתי', 'הוצאות', 'ההוצאות', 'תראה', 'תמצא', 'חפש', 'מצא',
             'לי', 'של', 'את', 'על', 'אני', 'כל', 'עם', 'גם', 'רק', 'יש', 'אין',
             'חודש', 'חודשים', 'אחרונים', 'אחרונות', 'האחרונה', 'האחרונים', 'האחרון',
             'השנה', 'החודש', 'שנה', 'חצי', 'בחצי',
             'expenses', 'show', 'find', 'me', 'the', 'last', 'my', 'how', 'much', 'spent', 'all',
             'הכי', 'גדולה', 'יקרה', 'הכנסות', 'הכנסה',
             cat_name.lower()} if cat_name else {'כמה', 'הוצאתי', 'הוצאות', 'ההוצאות', 'תראה', 'תמצא', 'חפש', 'מצא',
             'לי', 'של', 'את', 'על', 'אני', 'כל', 'עם', 'גם', 'רק', 'יש', 'אין',
             'חודש', 'חודשים', 'אחרונים', 'אחרונות', 'האחרונה', 'האחרונים', 'האחרון',
             'השנה', 'החודש', 'שנה', 'חצי', 'בחצי',
             'expenses', 'show', 'find', 'me', 'the', 'last', 'my', 'how', 'much', 'spent', 'all',
             'הכי', 'גדולה', 'יקרה', 'הכנסות', 'הכנסה'}
    words = [w for w in _re.sub(r'[\d\-/]+', ' ', query_lower).split() if len(w) >= 2 and w not in noise]

    search_term = ' '.join(words).strip()

    sql_where = "WHERE e.user_id = ? AND e.date >= ? AND e.date <= ?"
    params = [uid, from_date, to_date]
    if cat_id:
        sql_where += " AND e.category_id = ?"
        params.append(cat_id)
    if search_term:
        kw_list = [w for w in search_term.split() if len(w) >= 2]
        if kw_list:
            kw_clauses = []
            for kw in kw_list:
                # Try both the word as-is AND with Hebrew prefix stripped (מ,ב,ה,ל,ש,כ)
                variants = [f'%{kw}%']
                if len(kw) >= 4 and kw[0] in 'מבהלשכ':
                    variants.append(f'%{kw[1:]}%')
                or_parts = []
                for v in variants:
                    or_parts.append("e.subcategory LIKE ?")
                    or_parts.append("e.description LIKE ?")
                    params += [v, v]
                kw_clauses.append("(" + " OR ".join(or_parts) + ")")
            sql_where += " AND " + " AND ".join(kw_clauses)

    is_income = any(w in query_lower for w in ['income', 'הכנסה', 'הכנסות', 'salary', 'משכורת'])
    is_navigate = (not cat_id and not search_term and
                   any(w in query_lower for w in ['where', 'איפה', 'go to', 'take me', 'קח אותי']))

    if is_income:
        rows = conn.execute("SELECT * FROM income WHERE user_id=? AND date >= ? AND date <= ? ORDER BY date DESC LIMIT 20",
                            [uid, from_date, to_date]).fetchall()
        total = conn.execute("SELECT COALESCE(SUM(amount),0) FROM income WHERE user_id=? AND date >= ? AND date <= ?",
                             [uid, from_date, to_date]).fetchone()[0]
        items = [dict(r) for r in rows]
        response['text'] = (f"סה״כ הכנסות מ-{from_date} עד {to_date}: ₪{total:,.0f} ({len(items)} רשומות)"
                            if lang == 'he' else f"Total income {from_date} to {to_date}: ₪{total:,.0f} ({len(items)} entries)")
        response['data'] = {'items': items, 'total': total, 'type': 'income'}
        response['action'] = {'tab': 'incomeTab'}
    elif is_navigate:
        tab_map = {'דשבורד': 'dashboardTab', 'הוצאות': 'expensesTab', 'הכנסות': 'incomeTab',
                   'תקציב': 'budgetTab', 'ביטוח': 'financialTab', 'תובנות': 'insightsTab',
                   'dashboard': 'dashboardTab', 'expense': 'expensesTab', 'budget': 'budgetTab'}
        found_tab = next((tab for kw, tab in tab_map.items() if kw in query_lower), None)
        if found_tab:
            response['text'] = 'מעביר אותך...' if lang == 'he' else 'Opening...'
            response['action'] = {'tab': found_tab}
        else:
            response['text'] = 'הדפים: דשבורד, הוצאות, הכנסות, תקציב, ביטוח' if lang == 'he' else 'Pages: Dashboard, Expenses, Income, Budget, Insurance'
    else:
        rows = conn.execute(f"""
            SELECT e.*, c.name_he as category_name, c.color as category_color
            FROM expenses e JOIN categories c ON e.category_id = c.id
            {sql_where} ORDER BY e.date DESC LIMIT 50
        """, params).fetchall()
        total = conn.execute(f"SELECT COALESCE(SUM(e.amount),0) FROM expenses e {sql_where}", params).fetchone()[0]
        count = conn.execute(f"SELECT COUNT(*) FROM expenses e {sql_where}", params).fetchone()[0]
        items = [dict(r) for r in rows]

        # "Did you mean?" — if 0 results, retry with OR and suggest closest matches
        suggestions = []
        if count == 0 and search_term:
            kw_list = [w for w in search_term.split() if len(w) >= 2]
            if kw_list:
                or_clauses = []
                or_params = [uid, from_date, to_date]
                if cat_id:
                    or_params.append(cat_id)
                for kw in kw_list:
                    variants = [f'%{kw}%']
                    if len(kw) >= 4 and kw[0] in 'מבהלשכ':
                        variants.append(f'%{kw[1:]}%')
                    for v in variants:
                        or_clauses.append("e.subcategory LIKE ?")
                        or_clauses.append("e.description LIKE ?")
                        or_params += [v, v]
                or_where = "WHERE e.user_id = ? AND e.date >= ? AND e.date <= ?"
                if cat_id:
                    or_where += " AND e.category_id = ?"
                or_where += " AND (" + " OR ".join(or_clauses) + ")"

                or_rows = conn.execute(f"""
                    SELECT e.*, c.name_he as category_name, c.color as category_color
                    FROM expenses e JOIN categories c ON e.category_id = c.id
                    {or_where} ORDER BY e.date DESC LIMIT 50
                """, or_params).fetchall()
                if or_rows:
                    items = [dict(r) for r in or_rows]
                    total = sum(it['amount'] for it in items)
                    count = len(items)
                    # Collect unique descriptions for suggestions
                    suggestions = list(set(
                        it.get('description', '') for it in items if it.get('description')
                    ))[:5]

        filter_desc = f' {cat_name}' if cat_name else ''
        if search_term:
            filter_desc += f' "{search_term}"'

        if suggestions:
            suggest_str = ', '.join(suggestions[:3])
            if lang == 'he':
                response['text'] = f"לא מצאתי התאמה מדויקת.\nאולי התכוונת ל: {suggest_str}?\n\nנמצאו {count} תוצאות דומות מ-{from_date} עד {to_date}\nסה״כ: ₪{total:,.0f}"
            else:
                response['text'] = f"No exact match found.\nDid you mean: {suggest_str}?\n\nFound {count} similar results from {from_date} to {to_date}\nTotal: ₪{total:,.0f}"
        else:
            response['text'] = (f"נמצאו {count} הוצאות{filter_desc} מ-{from_date} עד {to_date}\nסה״כ: ₪{total:,.0f}"
                                if lang == 'he' else f"Found {count}{filter_desc} expenses {from_date} to {to_date}\nTotal: ₪{total:,.0f}")
        response['data'] = {'items': items[:20], 'total': total, 'count': count, 'type': 'expenses'}
        if suggestions:
            response['data']['suggestions'] = suggestions
        response['action'] = {'tab': 'expensesTab', 'filter': {
            'category_id': cat_id or '', 'subcategory': search_term, 'from_date': from_date, 'to_date': to_date
        }}

    conn.close()
    return jsonify(response)


@app.route('/api/chat/confirm', methods=['POST'])
@login_required
def chat_confirm_alias():
    """Learn from user confirmation: save search alias so fuzzy results become instant next time."""
    data = request.json
    user_typed = (data.get('user_typed') or '').strip()
    actual_match = (data.get('actual_match') or '').strip()
    confirmed = data.get('confirmed', False)

    if not user_typed:
        return jsonify({'ok': False, 'error': 'missing user_typed'})

    conn = get_db()
    if confirmed and actual_match:
        # Save alias: next time user types this, go straight to the match
        existing = conn.execute(
            "SELECT id FROM chat_aliases WHERE user_typed = ? AND actual_match = ?",
            [user_typed, actual_match]).fetchone()
        if existing:
            conn.execute("UPDATE chat_aliases SET times_used = times_used + 1 WHERE id = ?", [existing['id']])
        else:
            conn.execute("INSERT INTO chat_aliases (user_typed, actual_match) VALUES (?, ?)",
                         [user_typed, actual_match])
        conn.commit()
        conn.close()
        return jsonify({'ok': True, 'saved': True})
    else:
        # User said "no" — don't save, just acknowledge
        conn.close()
        return jsonify({'ok': True, 'saved': False})


@app.route('/api/chat/feedback', methods=['POST'])
@login_required
def chat_feedback():
    """Save user satisfaction rating for a chat response."""
    data = request.json
    rating = data.get('rating')
    query_text = data.get('query', '')

    if not rating or rating not in [1, 2, 3, 4, 5]:
        return jsonify({'ok': False, 'error': 'rating must be 1-5'})

    user_id = session.get('user_id')
    conn = get_db()
    conn.execute("INSERT INTO chat_feedback (user_id, query, rating) VALUES (?, ?, ?)",
                 [user_id, query_text, rating])
    conn.commit()
    conn.close()
    return jsonify({'ok': True})


@app.route('/api/admin/chat-satisfaction', methods=['GET'])
@login_required
def admin_chat_satisfaction():
    """Get chat satisfaction survey results for admin dashboard."""
    conn = get_db()
    # Check admin
    user = conn.execute("SELECT is_admin FROM users WHERE id = ?", [session.get('user_id')]).fetchone()
    if not user or not user['is_admin']:
        conn.close()
        return jsonify({'error': 'unauthorized'}), 403

    stats = conn.execute("""
        SELECT COUNT(*) as total_ratings,
               ROUND(AVG(rating), 1) as avg_rating,
               SUM(CASE WHEN rating >= 4 THEN 1 ELSE 0 END) as positive,
               SUM(CASE WHEN rating <= 2 THEN 1 ELSE 0 END) as negative
        FROM chat_feedback
    """).fetchone()

    recent = conn.execute("""
        SELECT cf.rating, cf.query, cf.created_at, u.username
        FROM chat_feedback cf LEFT JOIN users u ON cf.user_id = u.id
        ORDER BY cf.created_at DESC LIMIT 20
    """).fetchall()

    distribution = conn.execute("""
        SELECT rating, COUNT(*) as cnt FROM chat_feedback GROUP BY rating ORDER BY rating
    """).fetchall()

    conn.close()
    return jsonify({
        'total': stats['total_ratings'],
        'avg_rating': stats['avg_rating'] or 0,
        'positive': stats['positive'] or 0,
        'negative': stats['negative'] or 0,
        'recent': [dict(r) for r in recent],
        'distribution': [dict(r) for r in distribution]
    })


if __name__ == '__main__':
    if getattr(sys, 'frozen', False):
        import webbrowser
        import threading
        threading.Timer(1.5, lambda: webbrowser.open('http://127.0.0.1:5000')).start()
        print('=== מעקב הוצאות משפחתי ===')
        print('האפליקציה רצה בכתובת: http://127.0.0.1:5000')
        print('לסגירה: סגרו חלון זה או לחצו Ctrl+C')
        app.run(debug=False, port=5000)
    else:
        app.run(debug=True, port=5000)
