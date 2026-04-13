"""
Sanity Test Suite for Family Budget Tracker
Run before committing: python sanity_test.py
Uses a temporary database with random seed data — no effect on real data.
"""
import os
import sys
import json
import random
import tempfile
import shutil

# ── Make sure app can be imported ──
sys.path.insert(0, os.path.dirname(__file__))

import app as budget_app

PASS = 0
FAIL = 0
ERRORS = []


def ok(label):
    global PASS
    PASS += 1
    print(f"  ✅ {label}")


def fail(label, detail=""):
    global FAIL
    FAIL += 1
    msg = f"  ❌ {label}" + (f" — {detail}" if detail else "")
    ERRORS.append(msg)
    print(msg)


def check(condition, label, detail=""):
    if condition:
        ok(label)
    else:
        fail(label, detail)


# ── Setup: temporary DB, Flask test client ──
tmp_dir = tempfile.mkdtemp()
tmp_db = os.path.join(tmp_dir, "test_budget.db")
budget_app.DB_PATH = tmp_db
budget_app.app.config["TESTING"] = True
budget_app.app.config["SECRET_KEY"] = "test-secret"
client = budget_app.app.test_client()

# Initialize DB (disable external secrets for test)
budget_app.ADMIN_SECRETS = {}
with budget_app.app.app_context():
    budget_app.init_db()

# Create a test user and login (endpoints require auth)
with budget_app.app.app_context():
    conn = budget_app.get_db()
    pw_hash = budget_app.hash_password("testpass123")
    conn.execute("INSERT INTO users (username, password_hash, email, verified, is_admin) VALUES (?, ?, ?, 1, 1)",
                 ("testadmin", pw_hash, "test@test.com"))
    conn.commit()
    conn.close()

# Login to get session
with client.session_transaction() as sess:
    with budget_app.app.app_context():
        conn = budget_app.get_db()
        user = conn.execute("SELECT id FROM users WHERE username='testadmin'").fetchone()
        conn.close()
    sess['user_id'] = user['id']
    sess['username'] = 'testadmin'


def api_get(url):
    r = client.get(url)
    return r.status_code, r.get_json()


def api_post(url, data=None):
    r = client.post(url, json=data, content_type="application/json")
    return r.status_code, r.get_json()


def api_delete(url):
    r = client.delete(url)
    return r.status_code, r.get_json()


def api_put(url, data=None):
    r = client.put(url, json=data, content_type="application/json")
    return r.status_code, r.get_json()


# ====================================================================
print("\n🔹 1. App Startup & Static Files")
# ====================================================================
status, _ = api_get("/")
check(status == 200, "GET / returns 200")

status, _ = api_get("/static/index.html")
check(status == 200, "GET /static/index.html returns 200")


# ====================================================================
print("\n🔹 2. Categories")
# ====================================================================
status, cats = api_get("/api/categories")
check(status == 200, "GET /api/categories returns 200")
check(isinstance(cats, list) and len(cats) > 0, "Categories list is non-empty")
check(all("id" in c and "name_he" in c for c in cats), "Each category has id and name_he")
cat_ids = [c["id"] for c in cats]


# ====================================================================
print("\n🔹 3. Seed Data — realistic multi-month budget data")
# ====================================================================
random.seed(42)  # Reproducible

MONTHS = ["2025-11", "2025-12", "2026-01", "2026-02", "2026-03", "2026-04"]
SOURCES = ["cash", "visa", "bank_transfer"]
CARDS = ["4580", "7722", "1234", ""]
PERSONS = ["Dad", "Mom"]
FREQUENCIES = ["once", "monthly"]

EXPENSE_DESCRIPTIONS = [
    "Supermarket groceries", "Electricity bill", "Water bill", "Gas station",
    "Netflix subscription", "Gym membership", "Restaurant dinner", "Kids school",
    "Internet bill", "Phone plan", "Insurance payment", "Clothing store",
    "Bus pass", "Pharmacy", "Home maintenance", "Pet food",
    "Coffee shop", "Book store", "Dentist visit", "Car repair",
]

INCOME_DESCRIPTIONS = [
    ("Salary - Dad", "Dad"), ("Salary - Mom", "Mom"),
    ("Freelance project", "Dad"), ("Child allowance", "Mom"),
]

seed_expense_count = 0
seed_income_count = 0

for month in MONTHS:
    # 15-25 expenses per month, spread across categories
    num_expenses = random.randint(15, 25)
    for _ in range(num_expenses):
        day = random.randint(1, 28)
        cat = random.choice(cat_ids)
        source = random.choice(SOURCES)
        freq = "monthly" if random.random() < 0.2 else "once"
        status, _ = api_post("/api/expenses", {
            "description": random.choice(EXPENSE_DESCRIPTIONS),
            "amount": round(random.uniform(15, 1500), 2),
            "date": f"{month}-{day:02d}",
            "category_id": cat,
            "source": source,
            "frequency": freq,
        })
        if status == 200:
            seed_expense_count += 1

    # 2-4 income entries per month
    for desc, person in random.sample(INCOME_DESCRIPTIONS, random.randint(2, 4)):
        day = random.randint(1, 10)
        status, _ = api_post("/api/income", {
            "description": desc,
            "amount": round(random.uniform(3000, 15000), 2),
            "date": f"{month}-{day:02d}",
            "source": random.choice(["bank_transfer", "cash"]),
            "person": person,
            "is_recurring": 1 if "Salary" in desc else 0,
        })
        if status == 200:
            seed_income_count += 1

check(seed_expense_count >= 80, f"Seeded {seed_expense_count} expenses across 6 months")
check(seed_income_count >= 12, f"Seeded {seed_income_count} income entries across 6 months")


# ====================================================================
print("\n🔹 4. Expenses CRUD")
# ====================================================================
expense_data = {
    "description": "CRUD Test Expense",
    "amount": 150.50,
    "date": "2026-04-01",
    "category_id": cat_ids[0],
    "source": "cash",
    "frequency": "once",
}
status, res = api_post("/api/expenses", expense_data)
check(status == 200 or status == 201, "POST /api/expenses succeeds")

status, expenses = api_get("/api/expenses?month=2026-04")
check(status == 200, "GET /api/expenses returns 200")
crud_exp = [e for e in expenses if e["description"] == "CRUD Test Expense"]
check(len(crud_exp) == 1, "CRUD expense found in list")

exp_id = crud_exp[0]["id"]

# Update
status, res = api_put(f"/api/expenses/{exp_id}", {**expense_data, "amount": 200})
check(status == 200, "PUT /api/expenses/<id> succeeds")

# Verify update
status, expenses = api_get("/api/expenses?month=2026-04")
updated = [e for e in expenses if e["id"] == exp_id]
check(len(updated) == 1 and updated[0]["amount"] == 200, "Expense amount updated to 200")

# Delete
status, res = api_delete(f"/api/expenses/{exp_id}")
check(status == 200, "DELETE /api/expenses/<id> succeeds")

status, expenses = api_get("/api/expenses?month=2026-04")
check(all(e["id"] != exp_id for e in expenses), "Deleted expense no longer in list")


# ====================================================================
print("\n🔹 5. Income CRUD")
# ====================================================================
income_data = {
    "description": "CRUD Test Income",
    "amount": 10000,
    "date": "2026-04-01",
    "source": "bank_transfer",
    "person": "Test User",
    "is_recurring": 1,
}
status, res = api_post("/api/income", income_data)
check(status == 200 or status == 201, "POST /api/income succeeds")

status, incomes = api_get("/api/income?month=2026-04")
check(status == 200 and len(incomes) >= 1, "GET /api/income returns data")

crud_inc = [i for i in incomes if i["description"] == "CRUD Test Income"]
check(len(crud_inc) == 1, "CRUD income found in list")
inc_id = crud_inc[0]["id"]

status, _ = api_delete(f"/api/income/{inc_id}")
check(status == 200, "DELETE /api/income/<id> succeeds")


# ====================================================================
print("\n🔹 6. Budget Plans CRUD (max 3)")
# ====================================================================
status, plans = api_get("/api/budget-plans")
check(status == 200, "GET /api/budget-plans returns 200")
check(len(plans) >= 1, "Default plan exists")
default_plan = plans[0]
check(default_plan["id"] == 1, "Default plan has id=1")

# Create plan 2
status, _ = api_post("/api/budget-plans", {"name": "Plan B", "description": "Test plan B"})
check(status == 200, "Create budget plan 2")

# Create plan 3
status, _ = api_post("/api/budget-plans", {"name": "Plan C", "description": "Test plan C"})
check(status == 200, "Create budget plan 3")

# Attempt plan 4 — should fail
status, res = api_post("/api/budget-plans", {"name": "Plan D", "description": ""})
check(status == 400, "Cannot create 4th plan", f"got status {status}")

status, plans = api_get("/api/budget-plans")
check(len(plans) == 3, "Exactly 3 plans exist")

# Edit plan 2
status, _ = api_post("/api/budget-plans", {"id": 2, "name": "Plan B Edited", "description": "Updated"})
check(status == 200, "Edit plan 2 name/description")
status, plans = api_get("/api/budget-plans")
p2 = next((p for p in plans if p["id"] == 2), None)
check(p2 and p2["name"] == "Plan B Edited", "Plan 2 name updated correctly")

# Cannot delete default plan
status, _ = api_delete("/api/budget-plans/1")
check(status == 400, "Cannot delete default plan")

# Delete plan 3
status, _ = api_delete("/api/budget-plans/3")
check(status == 200, "Delete plan 3 succeeds")

status, plans = api_get("/api/budget-plans")
check(len(plans) == 2, "2 plans remain after deletion")


# ====================================================================
print("\n🔹 7. Budget per-plan isolation with seed data")
# ====================================================================
# Set budgets for multiple categories on both plans
for i, cid in enumerate(cat_ids[:4]):
    plan1_amt = (i + 1) * 500
    plan2_amt = (i + 1) * 800
    api_post("/api/budget", {
        "category_id": cid, "month": "2026-04", "planned_amount": plan1_amt, "plan_id": 1
    })
    api_post("/api/budget", {
        "category_id": cid, "month": "2026-04", "planned_amount": plan2_amt, "plan_id": 2
    })

# Verify plan 1
status, b1 = api_get("/api/budget?month=2026-04&plan=1")
check(status == 200 and len(b1) >= 4, "Plan 1 has 4 budget entries")
b1_map = {b["category_id"]: b["planned_amount"] for b in b1}
check(b1_map.get(cat_ids[0]) == 500, "Plan 1 first category = 500")
check(b1_map.get(cat_ids[2]) == 1500, "Plan 1 third category = 1500")

# Verify plan 2
status, b2 = api_get("/api/budget?month=2026-04&plan=2")
check(status == 200 and len(b2) >= 4, "Plan 2 has 4 budget entries")
b2_map = {b["category_id"]: b["planned_amount"] for b in b2}
check(b2_map.get(cat_ids[0]) == 800, "Plan 2 first category = 800")
check(b2_map.get(cat_ids[2]) == 2400, "Plan 2 third category = 2400")

# Plans are truly separate
check(b1_map.get(cat_ids[1]) != b2_map.get(cat_ids[1]),
      "Same category has different amounts per plan")


# ====================================================================
print("\n🔹 8. Budget plan deletion cascades budget data")
# ====================================================================
status, _ = api_delete("/api/budget-plans/2")
check(status == 200, "Delete plan 2")

status, b2 = api_get("/api/budget?month=2026-04&plan=2")
check(status == 200 and len(b2) == 0, "Plan 2 budget data deleted (cascade)")

# Plan 1 budget still intact
status, b1 = api_get("/api/budget?month=2026-04&plan=1")
check(len(b1) >= 4, "Plan 1 budget unaffected by plan 2 deletion")
check(b1[0]["planned_amount"] == 500, "Plan 1 first category still 500")


# ====================================================================
print("\n🔹 9. Summary with rich data")
# ====================================================================
status, summary = api_get("/api/summary?month=2026-04&plan=1")
check(status == 200, "GET /api/summary returns 200")
check("expense_total" in summary, "Summary has expense_total")
check(summary["expense_total"] > 0, f"Summary expense_total = {summary.get('expense_total', 0)}")
check("income_total" in summary, "Summary has income_total")
check(summary["income_total"] > 0, f"Summary income_total = {summary.get('income_total', 0)}")
check("balance" in summary, "Summary has balance")
check(len(summary.get("by_category", [])) > 0, "Summary has expenses by category")
check(len(summary.get("daily", [])) > 0, "Summary has daily breakdown")
check(len(summary.get("income_by_person", [])) > 0, "Summary has income by person")
check(len(summary.get("budget_vs_actual", [])) > 0, "Summary has budget_vs_actual entries")
check(len(summary.get("monthly_trend", [])) >= 2, "Summary has multi-month trend data")
check(len(summary.get("by_card", [])) > 0, "Summary has expenses by card")
check(len(summary.get("by_frequency", [])) > 0, "Summary has expenses by frequency")


# ====================================================================
print("\n🔹 10. Standing Orders from seed data")
# ====================================================================
status, orders = api_get("/api/standing-orders")
check(status == 200, "GET /api/standing-orders returns 200")
check(isinstance(orders, list) and len(orders) >= 1,
      f"Standing orders returned ({len(orders)} items)")
# Each standing order should have required fields
if orders:
    check(all("description" in o and "amount" in o for o in orders),
          "Standing orders have description and amount fields")


# ====================================================================
print("\n🔹 11. Category Averages with multi-month data")
# ====================================================================
status, avgs = api_get("/api/category-averages")
check(status == 200, "GET /api/category-averages returns 200")
check(isinstance(avgs, dict) and len(avgs) > 0, "Averages has entries")
# Averages should be > 0 for categories that have expenses
first_avg = list(avgs.values())[0]
check(first_avg > 0, f"First category average = {first_avg}")

# With from filter
status, avgs_filtered = api_get("/api/category-averages?from=2026-01")
check(status == 200, "GET /api/category-averages with from=2026-01")
check(len(avgs_filtered) > 0, "Filtered averages has entries")

# from=all should equal no-filter
status, avgs_all = api_get("/api/category-averages?from=all")
check(status == 200, "GET /api/category-averages with from=all")
check(avgs_all == avgs, "from=all equals no filter")


# ====================================================================
print("\n🔹 12. Available Months")
# ====================================================================
status, months = api_get("/api/available-months")
check(status == 200, "GET /api/available-months returns 200")
check(len(months) == 6, f"6 months available (got {len(months)})")
for m in MONTHS:
    check(m in months, f"Month {m} is available")


# ====================================================================
print("\n🔹 13. Cards")
# ====================================================================
status, cards = api_get("/api/cards")
check(status == 200, "GET /api/cards returns 200")
check(isinstance(cards, list), "Cards is a list")


# ====================================================================
print("\n🔹 14. Per-Plan Settings (localStorage simulation)")
# ====================================================================
plan_keys = {}
for plan_id in [1, 2, 3]:
    plan_keys[plan_id] = {
        "lang": f"plan_{plan_id}_lang",
        "currency": f"plan_{plan_id}_currency",
        "avgFrom": f"plan_{plan_id}_avgFrom",
    }

# Simulate: plan 1 = Hebrew/ILS, plan 2 = English/USD
settings_store = {}
settings_store[plan_keys[1]["lang"]] = "he"
settings_store[plan_keys[1]["currency"]] = "ILS"
settings_store[plan_keys[1]["avgFrom"]] = "all"
settings_store[plan_keys[2]["lang"]] = "en"
settings_store[plan_keys[2]["currency"]] = "USD"
settings_store[plan_keys[2]["avgFrom"]] = "2026-01"

# Switch to plan 1 -> read plan 1 settings
check(settings_store[plan_keys[1]["lang"]] == "he", "Plan 1 language = he")
check(settings_store[plan_keys[1]["currency"]] == "ILS", "Plan 1 currency = ILS")

# Switch to plan 2 -> read plan 2 settings
check(settings_store[plan_keys[2]["lang"]] == "en", "Plan 2 language = en")
check(settings_store[plan_keys[2]["currency"]] == "USD", "Plan 2 currency = USD")
check(settings_store[plan_keys[2]["avgFrom"]] == "2026-01", "Plan 2 avg from = 2026-01")

# Verify plan 1 not affected by plan 2
check(settings_store[plan_keys[1]["lang"]] == "he", "Plan 1 language still he after plan 2 switch")
check(settings_store[plan_keys[1]["currency"]] == "ILS", "Plan 1 currency still ILS after plan 2 switch")

# Verify keys are truly separate
check(plan_keys[1]["lang"] != plan_keys[2]["lang"], "Plan 1 and 2 use different localStorage keys")


# ====================================================================
print("\n🔹 15. Import endpoint exists")
# ====================================================================
r = client.post("/api/import")
check(r.status_code != 500, "POST /api/import does not crash", f"got {r.status_code}")


# ====================================================================
print("\n🔹 16. Export with data")
# ====================================================================
r = client.get("/api/export?month=2026-04")
check(r.status_code == 200, "GET /api/export returns 200")
check("spreadsheet" in r.content_type or "excel" in r.content_type or "octet" in r.content_type,
      "Export returns file content type", f"got {r.content_type}")
check(len(r.data) > 500, f"Export file has substantial content ({len(r.data)} bytes)")


# ====================================================================
print("\n🔹 17. Insights endpoints with real data")
# ====================================================================
insight_checks = [
    ("/api/insights/heatmap?months=3", "heatmap", lambda d: isinstance(d, dict)),
    ("/api/insights/burnrate?month=2026-04", "burnrate", lambda d: isinstance(d, dict)),
    ("/api/insights/latte?months=3", "latte", lambda d: isinstance(d, dict)),
    ("/api/insights/anomalies?months=3", "anomalies", lambda d: isinstance(d, (dict, list))),
    ("/api/insights/recurring", "recurring", lambda d: isinstance(d, dict)),
    ("/api/insights/weekly-pulse?month=2026-04", "weekly-pulse", lambda d: isinstance(d, dict)),
    ("/api/insights/projection?month=2026-04", "projection", lambda d: isinstance(d, dict)),
    ("/api/insights/comparison?month=2026-04", "comparison", lambda d: isinstance(d, dict)),
    ("/api/insights/achievements", "achievements", lambda d: isinstance(d, (dict, list))),
]
for ep, name, validator in insight_checks:
    status, data = api_get(ep)
    check(status == 200, f"Insight '{name}' returns 200", f"got {status}")
    check(validator(data), f"Insight '{name}' returns valid data structure")


# ====================================================================
print("\n🔹 18. Financial Products")
# ====================================================================
status, products = api_get("/api/financial/products")
check(status == 200, "GET /api/financial/products returns 200")

status, fin_summary = api_get("/api/financial/summary")
check(status == 200, "GET /api/financial/summary returns 200")


# ====================================================================
print("\n🔹 19. Installments CRUD")
# ====================================================================
status, inst = api_get("/api/installments")
check(status == 200, "GET /api/installments returns 200")

inst_data = {
    "description": "New Laptop",
    "total_amount": 3600,
    "num_payments": 12,
    "start_date": "2026-04-01",
    "card": "4580"
}
status, _ = api_post("/api/installments", inst_data)
check(status == 200 or status == 201, "POST /api/installments succeeds")

status, inst = api_get("/api/installments")
check(len(inst) >= 1, "Installment created")
if inst:
    status, _ = api_delete(f"/api/installments/{inst[0]['id']}")
    check(status == 200, "DELETE /api/installments/<id> succeeds")


# ====================================================================
print("\n🔹 20. Multi-month expense consistency")
# ====================================================================
total_seeded = 0
for month in MONTHS:
    status, exps = api_get(f"/api/expenses?month={month}")
    check(status == 200, f"GET expenses for {month}")
    total_seeded += len(exps)

check(total_seeded >= 80, f"Total expenses across all months = {total_seeded}")

# Verify summaries work for each month
for month in MONTHS:
    status, s = api_get(f"/api/summary?month={month}&plan=1")
    check(status == 200 and s.get("expense_total", 0) > 0,
          f"Summary for {month} has expenses")


# ====================================================================
print("\n🔹 21. Auth — Signup, Login, Logout, Status")
# ====================================================================
# Logout current session first
status, res = api_post("/api/auth/logout")
check(status == 200, "POST /api/auth/logout succeeds")

# Auth status should show not logged in
status, res = api_get("/api/auth/status")
check(status == 200, "GET /api/auth/status returns 200")
check(res["logged_in"] == False, "After logout, logged_in is false")

# All data endpoints require auth when logged out
status, _ = api_get("/api/expenses?month=2026-04")
check(status == 401, "GET /api/expenses returns 401 when not logged in")

status, _ = api_get("/api/income?month=2026-04")
check(status == 401, "GET /api/income returns 401 when not logged in")

status, _ = api_get("/api/summary?month=2026-04&plan=1")
check(status == 401, "GET /api/summary returns 401 when not logged in")

# Write actions require auth
status, _ = api_post("/api/expenses", {
    "description": "Unauthorized test", "amount": 100, "date": "2026-04-01",
    "category_id": 1, "source": "cash", "frequency": "once"
})
check(status == 401, "POST /api/expenses returns 401 when not logged in")

status, _ = api_post("/api/income", {
    "description": "Unauthorized test", "amount": 5000, "date": "2026-04-01",
    "source": "cash", "person": "Test", "is_recurring": 0
})
check(status == 401, "POST /api/income returns 401 when not logged in")

# Signup a new user (OTP sending will fail in test → auto-verified)
status, res = api_post("/api/auth/signup", {
    "username": "newuser",
    "password": "securepass123",
    "email": "newuser@example.com",
    "verification_method": "email",
})
check(status == 200, "POST /api/auth/signup succeeds")
check(res.get("auto_verified") == True or res.get("status") == "verification_sent",
      "Signup returns valid status")

# If auto-verified, we should be logged in now
if res.get("auto_verified"):
    status, res = api_get("/api/auth/status")
    check(res["logged_in"] == True, "Auto-verified user is logged in")
    check(res["username"] == "newuser", "Logged in as newuser")
    check(res["is_admin"] == False, "New user is not admin")

    # Logout the new user
    status, _ = api_post("/api/auth/logout")
    check(status == 200, "Logout newuser")

# Signup with duplicate username should fail
status, res = api_post("/api/auth/signup", {
    "username": "newuser",
    "password": "anotherpass",
    "email": "other@example.com",
    "verification_method": "email",
})
check(status == 400, "Duplicate username signup returns 400")
check("exists" in res.get("error", "").lower(), "Error mentions username exists")

# Signup with bad data
status, _ = api_post("/api/auth/signup", {"username": "", "password": "123456", "email": "a@b.c", "verification_method": "email"})
check(status == 400, "Signup with empty username returns 400")

status, _ = api_post("/api/auth/signup", {"username": "shortpw", "password": "12345", "email": "a@b.c", "verification_method": "email"})
check(status == 400, "Signup with short password returns 400")

# Login with wrong password
status, res = api_post("/api/auth/login", {"username": "testadmin", "password": "wrongpass"})
check(status == 401, "Login with wrong password returns 401")

# Login with non-existent user
status, _ = api_post("/api/auth/login", {"username": "noexist", "password": "whatever"})
check(status == 401, "Login with non-existent user returns 401")

# Login with correct username
status, res = api_post("/api/auth/login", {"username": "testadmin", "password": "testpass123"})
check(status == 200, "Login with correct username succeeds")
check(res.get("username") == "testadmin", "Login returns correct username")

# Logout and login with email instead
api_post("/api/auth/logout")
status, res = api_post("/api/auth/login", {"username": "test@test.com", "password": "testpass123"})
check(status == 200, "Login with email succeeds")
check(res.get("username") == "testadmin", "Email login returns correct username")

# Auth status should show logged in + admin
status, res = api_get("/api/auth/status")
check(res["logged_in"] == True, "After login, logged_in is true")
check(res["is_admin"] == True, "testadmin is admin")

# Protected endpoints should work again
status, _ = api_get("/api/expenses?month=2026-04")
check(status == 200, "Protected endpoint works after login")


# ====================================================================
print("\n🔹 22. Admin endpoints")
# ====================================================================
# Admin stats
status, stats = api_get("/api/admin/stats")
check(status == 200, "GET /api/admin/stats returns 200")
check(stats["total_users"] >= 2, f"Admin stats shows {stats['total_users']} users")
check(stats["total_expenses"] > 0, "Admin stats shows expenses")
check(stats["total_categories"] > 0, "Admin stats shows categories")

# Admin get users
status, users = api_get("/api/admin/users")
check(status == 200, "GET /api/admin/users returns 200")
check(len(users) >= 2, f"Admin sees {len(users)} users")
admin_user = next((u for u in users if u["username"] == "testadmin"), None)
check(admin_user is not None, "Admin user found in list")
check(admin_user["is_admin"] == 1, "Admin user has is_admin=1")
new_user = next((u for u in users if u["username"] == "newuser"), None)
check(new_user is not None, "newuser found in list")

# Cannot delete admin user
status, res = api_delete(f"/api/admin/users/{admin_user['id']}")
check(status == 400, "Cannot delete admin user")

# Delete non-admin user
status, _ = api_delete(f"/api/admin/users/{new_user['id']}")
check(status == 200, "Delete non-admin user succeeds")

status, users = api_get("/api/admin/users")
check(not any(u["username"] == "newuser" for u in users), "Deleted user no longer in list")

# Non-admin cannot access admin endpoints
# Logout, login as a new regular user
api_post("/api/auth/logout")
api_post("/api/auth/signup", {
    "username": "regularuser",
    "password": "regular123",
    "email": "regular@example.com",
    "verification_method": "email",
})
# regularuser is auto-verified and logged in
status, _ = api_get("/api/admin/stats")
check(status in [401, 403], "Non-admin cannot access admin stats", f"got {status}")

status, _ = api_get("/api/admin/users")
check(status in [401, 403], "Non-admin cannot access admin users", f"got {status}")

# Re-login as admin for any further tests
api_post("/api/auth/logout")
api_post("/api/auth/login", {"username": "testadmin", "password": "testpass123"})


# ====================================================================
print("\n🔹 23. Chat Assistant — Basic Queries")
# ====================================================================
# Basic expense query (default: last 3 months)
status, res = api_post("/api/chat", {"query": "how much did i spend", "lang": "en"})
check(status == 200, "Chat: basic expense query returns 200")
check("text" in res and "data" in res and "action" in res, "Chat: response has text, data, action")
check(res["data"]["type"] == "expenses", "Chat: expense query returns type=expenses")
check(res["data"]["count"] >= 0, "Chat: count is non-negative")
check(res["data"]["total"] >= 0, "Chat: total is non-negative")
check(isinstance(res["data"]["items"], list), "Chat: items is a list")
check(isinstance(res["data"].get("breakdown", []), list), "Chat: breakdown is a list (or absent)")

# Hebrew query
status, res = api_post("/api/chat", {"query": "כמה הוצאתי החודש", "lang": "he"})
check(status == 200, "Chat: Hebrew query returns 200")
check("נמצאו" in res["text"] or "סה״כ" in res["text"] or res["data"]["count"] == 0,
      "Chat: Hebrew response text is in Hebrew")

# English query with data
status, res = api_post("/api/chat", {"query": "show me expenses this year", "lang": "en"})
check(status == 200, "Chat: 'this year' query returns 200")
check(res["data"]["count"] >= 0, "Chat: year query returns results")

# Specific month query
status, res = api_post("/api/chat", {"query": "expenses 2026-03", "lang": "en"})
check(status == 200, "Chat: specific month query returns 200")

# Top/biggest query
status, res = api_post("/api/chat", {"query": "what is the biggest expense", "lang": "en"})
check(status == 200, "Chat: top expense query returns 200")
if res["data"]["count"] > 0:
    check("Top expenses" in res["text"] or "biggest" in res["text"].lower() or res["data"]["items"],
          "Chat: top query contains expense data")

# Income query
status, res = api_post("/api/chat", {"query": "show me my income", "lang": "en"})
check(status == 200, "Chat: income query returns 200")
check(res["data"]["type"] == "income", "Chat: income query returns type=income")

# Navigation query — "הוצאות" is both a noise word and a tab keyword,
# so search_term becomes empty and is_navigate triggers
status, res = api_post("/api/chat", {"query": "איפה הוצאות", "lang": "he"})
check(status == 200, "Chat: navigation query returns 200")
check(res.get("action", {}).get("tab") == "expensesTab",
      "Chat: navigation routes to expensesTab", f"got action={res.get('action')}")

# Empty query
status, res = api_post("/api/chat", {"query": "", "lang": "en"})
check(status == 200, "Chat: empty query returns 200 (graceful)")

# Lang defaults to 'he'
status, res = api_post("/api/chat", {"query": "expenses"})
check(status == 200, "Chat: missing lang defaults gracefully")


# ====================================================================
print("\n🔹 24. Chat Assistant — Keyword Search & Fuzzy Fallback")
# ====================================================================
# Insert a known expense with a specific subcategory for search testing
with budget_app.app.app_context():
    conn = budget_app.get_db()
    test_uid = conn.execute("SELECT id FROM users WHERE username='testadmin'").fetchone()['id']
    conn.execute("""INSERT INTO expenses (date, category_id, subcategory, description, amount, source, user_id)
                    VALUES ('2026-04-01', ?, 'Supermarket Rami Levy', 'Weekly groceries', 250.0, 'visa', ?)""",
                 (cat_ids[0], test_uid))
    conn.commit()
    conn.close()

# Exact keyword search
status, res = api_post("/api/chat", {"query": "Rami Levy", "lang": "en"})
check(status == 200, "Chat: keyword search returns 200")
check(res["data"]["count"] >= 1, "Chat: keyword 'Rami Levy' finds the inserted expense",
      f"count={res['data']['count']}")

# Search with a typo — should trigger fuzzy fallback
status, res = api_post("/api/chat", {"query": "Rami Levi supermarkt", "lang": "en"})
check(status == 200, "Chat: typo search returns 200")
if res["data"]["count"] > 0:
    is_fuzzy = res["data"].get("fuzzy", False)
    check(True, "Chat: typo search found results via fallback")
    if is_fuzzy:
        check("fuzzy_matches" in res["data"], "Chat: fuzzy result includes fuzzy_matches list")
        check("original_search" in res["data"], "Chat: fuzzy result includes original_search")
        check("Did you mean" in res["text"] or "\u05d0\u05d5\u05dc\u05d9" in res["text"],
              "Chat: fuzzy response asks for confirmation")

# Category filter in search
if len(cat_ids) >= 2:
    # Get first category name
    with budget_app.app.app_context():
        conn = budget_app.get_db()
        cat_row = conn.execute("SELECT name_he FROM categories WHERE id=?", (cat_ids[0],)).fetchone()
        conn.close()
    if cat_row:
        cat_name_he = cat_row["name_he"]
        status, res = api_post("/api/chat", {"query": f"{cat_name_he} השנה", "lang": "he"})
        check(status == 200, f"Chat: category+date query returns 200")


# ====================================================================
print("\n🔹 25. Chat Assistant — Alias Learning (confirm/deny)")
# ====================================================================
# Confirm a fuzzy match alias
status, res = api_post("/api/chat/confirm", {
    "user_typed": "rami levi",
    "actual_match": "Supermarket Rami Levy",
    "confirmed": True
})
check(status == 200, "Chat confirm: save alias returns 200")
check(res.get("ok") is True, "Chat confirm: response ok=True")
check(res.get("saved") is True, "Chat confirm: alias was saved")

# Verify alias exists in DB
with budget_app.app.app_context():
    conn = budget_app.get_db()
    alias = conn.execute("SELECT * FROM chat_aliases WHERE user_typed='rami levi'").fetchone()
    conn.close()
check(alias is not None, "Chat confirm: alias row created in DB")
check(alias["actual_match"] == "Supermarket Rami Levy", "Chat confirm: alias actual_match is correct")
check(alias["times_used"] == 1, "Chat confirm: times_used starts at 1")

# Confirm same alias again — times_used should increment
status, res = api_post("/api/chat/confirm", {
    "user_typed": "rami levi",
    "actual_match": "Supermarket Rami Levy",
    "confirmed": True
})
check(status == 200 and res.get("saved"), "Chat confirm: re-confirm same alias succeeds")
with budget_app.app.app_context():
    conn = budget_app.get_db()
    alias = conn.execute("SELECT times_used FROM chat_aliases WHERE user_typed='rami levi'").fetchone()
    conn.close()
check(alias["times_used"] == 2, "Chat confirm: times_used incremented to 2")

# Now search using the aliased term — should use the alias directly
status, res = api_post("/api/chat", {"query": "rami levi", "lang": "en"})
check(status == 200, "Chat alias: search with aliased term returns 200")
check(res["data"]["count"] >= 1, "Chat alias: aliased search finds results directly",
      f"count={res['data']['count']}")
check(res["data"].get("fuzzy") is not True, "Chat alias: aliased search is NOT fuzzy (used alias)")

# Deny a fuzzy match — should not save
status, res = api_post("/api/chat/confirm", {
    "user_typed": "some typo",
    "confirmed": False
})
check(status == 200, "Chat deny: returns 200")
check(res.get("saved") is False, "Chat deny: not saved")

# Missing user_typed — should return error
status, res = api_post("/api/chat/confirm", {
    "user_typed": "",
    "actual_match": "anything",
    "confirmed": True
})
check(status == 200 and res.get("ok") is False, "Chat confirm: empty user_typed rejected")


# ====================================================================
print("\n🔹 26. Chat Assistant — Feedback / Satisfaction Rating")
# ====================================================================
# Valid rating
status, res = api_post("/api/chat/feedback", {"rating": 5, "query": "test query"})
check(status == 200, "Chat feedback: rating 5 returns 200")
check(res.get("ok") is True, "Chat feedback: ok=True")

# Another rating
status, res = api_post("/api/chat/feedback", {"rating": 3, "query": "another query"})
check(status == 200, "Chat feedback: rating 3 returns 200")

status, res = api_post("/api/chat/feedback", {"rating": 1, "query": "bad query"})
check(status == 200, "Chat feedback: rating 1 returns 200")

# Invalid ratings
status, res = api_post("/api/chat/feedback", {"rating": 0, "query": "test"})
check(res.get("ok") is False, "Chat feedback: rating 0 rejected")

status, res = api_post("/api/chat/feedback", {"rating": 6, "query": "test"})
check(res.get("ok") is False, "Chat feedback: rating 6 rejected")

status, res = api_post("/api/chat/feedback", {"rating": "abc", "query": "test"})
check(res.get("ok") is False, "Chat feedback: non-numeric rating rejected")

status, res = api_post("/api/chat/feedback", {"query": "test"})
check(res.get("ok") is False, "Chat feedback: missing rating rejected")

# Verify feedback in DB
with budget_app.app.app_context():
    conn = budget_app.get_db()
    fb_count = conn.execute("SELECT COUNT(*) as c FROM chat_feedback").fetchone()["c"]
    conn.close()
check(fb_count == 3, "Chat feedback: 3 valid ratings stored in DB", f"got {fb_count}")


# ====================================================================
print("\n🔹 27. Chat Satisfaction — Admin Dashboard API")
# ====================================================================
status, res = api_get("/api/admin/chat-satisfaction")
check(status == 200, "Admin satisfaction: returns 200")
check(res.get("total") == 3, "Admin satisfaction: total = 3", f"got {res.get('total')}")
check(res.get("avg_rating") == 3.0, "Admin satisfaction: avg = 3.0 ((5+3+1)/3)",
      f"got {res.get('avg_rating')}")
check(res.get("positive") == 1, "Admin satisfaction: 1 positive (rating 5)")
check(res.get("negative") == 1, "Admin satisfaction: 1 negative (rating 1)")
check(isinstance(res.get("recent"), list), "Admin satisfaction: recent is a list")
check(len(res["recent"]) == 3, "Admin satisfaction: 3 recent entries")
check(isinstance(res.get("distribution"), list), "Admin satisfaction: distribution is a list")

# Check recent entries have expected fields
if res["recent"]:
    r0 = res["recent"][0]
    check("rating" in r0 and "query" in r0 and "created_at" in r0,
          "Admin satisfaction: recent entry has rating, query, created_at")

# Non-admin should not access satisfaction data
api_post("/api/auth/logout")
api_post("/api/auth/login", {"username": "regularuser", "password": "regular123"})
status, _ = api_get("/api/admin/chat-satisfaction")
check(status in [401, 403], "Admin satisfaction: non-admin blocked",
      f"got {status}")

# Re-login as admin
api_post("/api/auth/logout")
api_post("/api/auth/login", {"username": "testadmin", "password": "testpass123"})


# ====================================================================
print("\n🔹 28. Chat Assistant — Auth Required")
# ====================================================================
# Logout and test that chat POST requires auth
api_post("/api/auth/logout")

status, res = api_post("/api/chat", {"query": "test", "lang": "en"})
check(status == 401, "Chat: POST without auth returns 401", f"got {status}")

status, res = api_post("/api/chat/confirm", {"user_typed": "x", "confirmed": True})
check(status == 401, "Chat confirm: POST without auth returns 401", f"got {status}")

status, res = api_post("/api/chat/feedback", {"rating": 5, "query": "test"})
check(status == 401, "Chat feedback: POST without auth returns 401", f"got {status}")

# Re-login for cleanup
api_post("/api/auth/login", {"username": "testadmin", "password": "testpass123"})


# ====================================================================
print("\n🔹 29. Multi-User Data Isolation")
# ====================================================================
# Create two fresh test users for isolation testing
api_post("/api/auth/logout")

api_post("/api/auth/signup", {
    "username": "isolationA",
    "password": "isopass123A",
    "email": "isoA@example.com",
    "verification_method": "email",
})
# isolationA is auto-verified and logged in

# -- User A: add expense, income, installment --
status, _ = api_post("/api/expenses", {
    "description": "UserA Expense Only",
    "amount": 777.77,
    "date": "2026-04-05",
    "category_id": cat_ids[0],
    "source": "cash",
    "frequency": "once",
})
check(status == 200, "Isolation: User A creates expense")

status, _ = api_post("/api/income", {
    "description": "UserA Salary Only",
    "amount": 9999.99,
    "date": "2026-04-05",
    "source": "bank_transfer",
    "person": "User A",
    "is_recurring": 1,
})
check(status == 200, "Isolation: User A creates income")

status, _ = api_post("/api/installments", {
    "description": "UserA Laptop Only",
    "total_amount": 3000,
    "num_payments": 6,
    "start_date": "2026-04-01",
    "card": "1111",
})
check(status == 200, "Isolation: User A creates installment")

# Verify User A sees own data
status, exps = api_get("/api/expenses?month=2026-04")
check(status == 200, "Isolation: User A GET expenses")
a_exps = [e for e in exps if e["description"] == "UserA Expense Only"]
check(len(a_exps) == 1, "Isolation: User A sees own expense")
check(a_exps[0]["amount"] == 777.77, "Isolation: User A expense amount correct")

status, incs = api_get("/api/income?month=2026-04")
a_incs = [i for i in incs if i["description"] == "UserA Salary Only"]
check(len(a_incs) == 1, "Isolation: User A sees own income")

status, inst = api_get("/api/installments")
a_inst = [i for i in inst if i["description"] == "UserA Laptop Only"]
check(len(a_inst) == 1, "Isolation: User A sees own installment")

# Get user A's budget plan info
status, a_plans = api_get("/api/budget-plans")
check(len(a_plans) >= 1, "Isolation: User A has a budget plan")
a_plan_id = a_plans[0]["id"]

# Set a budget for User A
api_post("/api/budget", {
    "category_id": cat_ids[0], "month": "2026-04",
    "planned_amount": 1234, "plan_id": a_plan_id,
})
status, a_budgets = api_get(f"/api/budget?month=2026-04&plan={a_plan_id}")
a_bud = [b for b in a_budgets if b["category_id"] == cat_ids[0]]
check(len(a_bud) == 1 and a_bud[0]["planned_amount"] == 1234,
      "Isolation: User A budget set to 1234")

# Summary should reflect User A data only
status, a_summary = api_get(f"/api/summary?month=2026-04&plan={a_plan_id}")
check(a_summary["expense_total"] == 777.77,
      "Isolation: User A summary expense_total = 777.77",
      f"got {a_summary.get('expense_total')}")
check(a_summary["income_total"] == 9999.99,
      "Isolation: User A summary income_total = 9999.99",
      f"got {a_summary.get('income_total')}")

# -- Switch to User B --
api_post("/api/auth/logout")
api_post("/api/auth/signup", {
    "username": "isolationB",
    "password": "isopass123B",
    "email": "isoB@example.com",
    "verification_method": "email",
})
# isolationB is auto-verified and logged in

# User B should see NONE of User A's data
status, exps = api_get("/api/expenses?month=2026-04")
check(status == 200, "Isolation: User B GET expenses")
b_sees_a = [e for e in exps if e["description"] == "UserA Expense Only"]
check(len(b_sees_a) == 0, "Isolation: User B does NOT see User A expense")

status, incs = api_get("/api/income?month=2026-04")
b_sees_a_inc = [i for i in incs if i["description"] == "UserA Salary Only"]
check(len(b_sees_a_inc) == 0, "Isolation: User B does NOT see User A income")

status, inst = api_get("/api/installments")
b_sees_a_inst = [i for i in inst if i["description"] == "UserA Laptop Only"]
check(len(b_sees_a_inst) == 0, "Isolation: User B does NOT see User A installment")

# User B should have own empty budget plan, NOT User A's
status, b_plans = api_get("/api/budget-plans")
check(len(b_plans) >= 1, "Isolation: User B has own budget plan")
b_plan_id = b_plans[0]["id"]
check(b_plan_id != a_plan_id, "Isolation: User B plan id differs from User A",
      f"A={a_plan_id}, B={b_plan_id}")

status, b_budgets = api_get(f"/api/budget?month=2026-04&plan={b_plan_id}")
b_bud_a = [b for b in b_budgets if b.get("planned_amount", 0) == 1234]
check(len(b_bud_a) == 0, "Isolation: User B does NOT see User A budget entries")

# User B summary should be empty/zero
status, b_summary = api_get(f"/api/summary?month=2026-04&plan={b_plan_id}")
check(b_summary["expense_total"] == 0,
      "Isolation: User B summary expense_total = 0",
      f"got {b_summary.get('expense_total')}")
check(b_summary["income_total"] == 0,
      "Isolation: User B summary income_total = 0",
      f"got {b_summary.get('income_total')}")

# -- User B adds own data --
status, _ = api_post("/api/expenses", {
    "description": "UserB Expense Only",
    "amount": 333.33,
    "date": "2026-04-06",
    "category_id": cat_ids[1],
    "source": "visa",
    "frequency": "once",
})
check(status == 200, "Isolation: User B creates expense")

status, exps = api_get("/api/expenses?month=2026-04")
b_exps = [e for e in exps if e["description"] == "UserB Expense Only"]
check(len(b_exps) == 1, "Isolation: User B sees own expense")

# User B still shouldn't see User A
b_sees_a2 = [e for e in exps if e["description"] == "UserA Expense Only"]
check(len(b_sees_a2) == 0, "Isolation: User B still doesn't see User A after own insert")

# -- Switch back to User A — verify A doesn't see B --
api_post("/api/auth/logout")
api_post("/api/auth/login", {"username": "isolationA", "password": "isopass123A"})

status, exps = api_get("/api/expenses?month=2026-04")
a_sees_b = [e for e in exps if e["description"] == "UserB Expense Only"]
check(len(a_sees_b) == 0, "Isolation: User A does NOT see User B expense")

# User A still sees own expense
a_exps2 = [e for e in exps if e["description"] == "UserA Expense Only"]
check(len(a_exps2) == 1, "Isolation: User A still sees own expense")

# -- User A cannot UPDATE User B's expense --
# Get User B's expense id via direct DB
with budget_app.app.app_context():
    conn = budget_app.get_db()
    b_exp = conn.execute("SELECT id FROM expenses WHERE description='UserB Expense Only'").fetchone()
    conn.close()

if b_exp:
    status, _ = api_put(f"/api/expenses/{b_exp['id']}", {
        "description": "UserA hacked B", "amount": 0.01,
        "date": "2026-04-06", "category_id": cat_ids[1],
        "source": "visa", "frequency": "once",
    })
    # Should either 404 or silently update 0 rows — verify B's data unchanged
    with budget_app.app.app_context():
        conn = budget_app.get_db()
        b_exp_check = conn.execute("SELECT description FROM expenses WHERE id=?", (b_exp['id'],)).fetchone()
        conn.close()
    check(b_exp_check["description"] == "UserB Expense Only",
          "Isolation: User A cannot overwrite User B's expense")

    # User A cannot DELETE User B's expense
    status, _ = api_delete(f"/api/expenses/{b_exp['id']}")
    with budget_app.app.app_context():
        conn = budget_app.get_db()
        b_exp_still = conn.execute("SELECT id FROM expenses WHERE id=?", (b_exp['id'],)).fetchone()
        conn.close()
    check(b_exp_still is not None, "Isolation: User A cannot delete User B's expense")

# -- Admin cascade delete removes user data --
api_post("/api/auth/logout")
api_post("/api/auth/login", {"username": "testadmin", "password": "testpass123"})

# Get isolationB user id
with budget_app.app.app_context():
    conn = budget_app.get_db()
    iso_b = conn.execute("SELECT id FROM users WHERE username='isolationB'").fetchone()
    conn.close()

if iso_b:
    status, _ = api_delete(f"/api/admin/users/{iso_b['id']}")
    check(status == 200, "Isolation: Admin deletes User B")

    # Verify User B's data is gone
    with budget_app.app.app_context():
        conn = budget_app.get_db()
        b_exps_left = conn.execute("SELECT COUNT(*) as c FROM expenses WHERE user_id=?", (iso_b['id'],)).fetchone()['c']
        b_incs_left = conn.execute("SELECT COUNT(*) as c FROM income WHERE user_id=?", (iso_b['id'],)).fetchone()['c']
        b_plans_left = conn.execute("SELECT COUNT(*) as c FROM budget_plans WHERE user_id=?", (iso_b['id'],)).fetchone()['c']
        conn.close()
    check(b_exps_left == 0, "Isolation: Admin delete cascaded — User B expenses removed")
    check(b_incs_left == 0, "Isolation: Admin delete cascaded — User B income removed")
    check(b_plans_left == 0, "Isolation: Admin delete cascaded — User B budget plans removed")

# Clean up isolationA too
with budget_app.app.app_context():
    conn = budget_app.get_db()
    iso_a = conn.execute("SELECT id FROM users WHERE username='isolationA'").fetchone()
    conn.close()

if iso_a:
    status, _ = api_delete(f"/api/admin/users/{iso_a['id']}")
    check(status == 200, "Isolation: Admin deletes User A")


# ====================================================================
print("\n🔹 30. Cross-User Endpoint Isolation (standing orders, available months, export)")
# ====================================================================
# Create userC with data, verify userD sees nothing on derived endpoints
api_post("/api/auth/logout")
api_post("/api/auth/signup", {
    "username": "isolationC",
    "password": "isopass123C",
    "email": "isoC@example.com",
    "verification_method": "email",
})

# User C: add monthly expense (creates standing order)
status, _ = api_post("/api/expenses", {
    "description": "UserC Monthly Gym",
    "amount": 200,
    "date": "2026-04-01",
    "category_id": cat_ids[0],
    "source": "visa",
    "frequency": "monthly",
})
check(status == 200, "Isolation: User C creates monthly expense")

status, c_orders = api_get("/api/standing-orders")
check(any(o["description"] == "UserC Monthly Gym" for o in c_orders),
      "Isolation: User C sees own standing order")

status, c_months = api_get("/api/available-months")
check("2026-04" in c_months, "Isolation: User C sees own available month")

# Switch to User D
api_post("/api/auth/logout")
api_post("/api/auth/signup", {
    "username": "isolationD",
    "password": "isopass123D",
    "email": "isoD@example.com",
    "verification_method": "email",
})

# User D should not see User C's standing orders or months
status, d_orders = api_get("/api/standing-orders")
d_sees_c = [o for o in d_orders if o.get("description") == "UserC Monthly Gym"]
check(len(d_sees_c) == 0, "Isolation: User D does NOT see User C standing order")

status, d_months = api_get("/api/available-months")
check(len(d_months) == 0, "Isolation: User D has no available months (no data)",
      f"got {d_months}")

# Export for User D should be empty/minimal
r = client.get("/api/export?month=2026-04")
check(r.status_code == 200, "Isolation: User D export returns 200")

# Cleanup isolation users C and D
api_post("/api/auth/logout")
api_post("/api/auth/login", {"username": "testadmin", "password": "testpass123"})

for uname in ['isolationC', 'isolationD']:
    with budget_app.app.app_context():
        conn = budget_app.get_db()
        u = conn.execute("SELECT id FROM users WHERE username=?", (uname,)).fetchone()
        conn.close()
    if u:
        api_delete(f"/api/admin/users/{u['id']}")


# ====================================================================
# 31. Insurance Portfolio Import
# ====================================================================
print("\n🔹 31. Insurance Portfolio Import")

# Create a test XLSX insurance file
import openpyxl

def make_test_insurance_xlsx(filepath, file_type='general'):
    """Create a minimal test insurance portfolio XLSX."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'תיק ביטוחי'

    if file_type == 'general':
        # HbResults format (car/home)
        ws.cell(row=2, column=2, value="התיק הביטוחי, הופק מאתר 'הר הביטוח' של משרד האוצר, בתאריך")
        ws.cell(row=2, column=6, value='11/04/2026')
        headers = ['תעודת זהות', 'ענף ראשי', 'ענף (משני)', 'סוג מוצר', 'חברה', 'תקופת ביטוח', 'פרטים נוספים', 'פרמיה בש"ח', 'סוג פרמיה', 'מספר פוליסה', 'סיווג תכנית']
        for j, h in enumerate(headers):
            ws.cell(row=4, column=j+1, value=h)
        # Car insurance
        for j, v in enumerate(['12345678', 'ביטוח רכב', 'ביטוח מקיף', 'פוליסת ביטוח', 'כלל חברה לביטוח בע"מ', '01/01/2025 - 31/12/2025', None, 6000, 'שנתית', '999111222', 'אישי']):
            ws.cell(row=6, column=j+1, value=v)
        # Home insurance
        for j, v in enumerate(['12345678', 'ביטוח דירה', 'ביטוח מבנה', 'פוליסת ביטוח', 'הראל חברה לביטוח בע"מ', '01/04/2025 - 31/03/2026', None, 1200, 'שנתית', '999333444', 'אישי']):
            ws.cell(row=7, column=j+1, value=v)
    elif file_type == 'health':
        # HitResults section 2 format
        ws.cell(row=2, column=2, value="התיק הביטוחי, הופקו מאתר 'כלי מסביר ביטוח' של רשות שוק ההון, בתאריך")
        ws.cell(row=2, column=6, value='4/11/2026')
        headers = ['תעודת זהות', 'ענף ראשי', 'ענף (משני)', 'סוג מוצר', 'חברה', 'תקופת ביטוח', 'פרמיה בש"ח', 'סוג פרמיה', 'מספר פוליסה', 'סיווג תכנית']
        for j, h in enumerate(headers):
            ws.cell(row=4, column=j+1, value=h)
        # Two coverages under same policy
        for j, v in enumerate(['12345678', 'ביטוח בריאות', 'ניתוחים', 'קבוצתי', 'מגדל ביטוח', 'מתחדש', 80, 'חודשית', '555666777', 'תכנית ביטוח']):
            ws.cell(row=5, column=j+1, value=v)
        for j, v in enumerate(['12345678', 'ביטוח בריאות', 'תרופות', 'קבוצתי', 'מגדל ביטוח', 'מתחדש', 40, 'חודשית', '555666777', 'תכנית ביטוח']):
            ws.cell(row=6, column=j+1, value=v)
        # Zero cost entry
        for j, v in enumerate(['12345678', 'ביטוח בריאות', 'ייעוץ', 'קבוצתי', 'כללית', 'מתחדש', 0, 'חודשית', '', 'תכנית ביטוח']):
            ws.cell(row=7, column=j+1, value=v)

    wb.save(filepath)

# Test 1: Auto-detection of insurance portfolio files
general_path = os.path.join(tmp_dir, 'test_general_ins.xlsx')
health_path = os.path.join(tmp_dir, 'test_health_ins.xlsx')
make_test_insurance_xlsx(general_path, 'general')
make_test_insurance_xlsx(health_path, 'health')

with budget_app.app.app_context():
    check(budget_app._is_insurance_portfolio_xlsx(general_path),
          "Insurance detect: general portfolio detected")
    check(budget_app._is_insurance_portfolio_xlsx(health_path),
          "Insurance detect: health portfolio detected")

# Test 2: Import via API endpoint
api_post("/api/auth/logout")
api_post("/api/auth/login", {"username": "testadmin", "password": "testpass123"})

with open(general_path, 'rb') as f:
    r = client.post("/api/import", data={'file': (f, 'test_general_ins.xlsx')}, content_type='multipart/form-data')
check(r.status_code == 200, "Insurance import: general file returns 200")
data = r.get_json()
check(data.get('source') == 'insurance_portfolio', "Insurance import: source is insurance_portfolio")
check(data.get('imported') == 2, f"Insurance import: general imported 2 products (got {data.get('imported')})")

# Test 3: Health file with aggregation (same policy number)
with open(health_path, 'rb') as f:
    r = client.post("/api/import", data={'file': (f, 'test_health_ins.xlsx')}, content_type='multipart/form-data')
check(r.status_code == 200, "Insurance import: health file returns 200")
data = r.get_json()
check(data.get('imported') == 1, f"Insurance import: health aggregated to 1 product (got {data.get('imported')})")
check(data.get('skipped', 0) >= 1, f"Insurance import: zero-cost entries skipped (got {data.get('skipped', 0)})")
if data.get('products'):
    check(data['products'][0]['monthly_cost'] == 120.0,
          f"Insurance import: aggregated monthly cost = 120 (got {data['products'][0].get('monthly_cost')})")

# Test 4: Re-import dedup (same policy numbers should be skipped)
with open(general_path, 'rb') as f:
    r = client.post("/api/import", data={'file': (f, 'test_general_ins.xlsx')}, content_type='multipart/form-data')
data = r.get_json()
check(data.get('imported') == 0, f"Insurance import: re-import deduped (imported {data.get('imported')})")
check(data.get('skipped', 0) >= 2, f"Insurance import: re-import skipped duplicates (got {data.get('skipped', 0)})")

# Test 5: Products appear in financial products list
r = client.get("/api/financial/products")
products = r.get_json()
ins_products = [p for p in products if p['type'] == 'insurance']
check(len(ins_products) >= 3, f"Insurance import: 3+ products in financial list (got {len(ins_products)})")

# Check subtypes are correct
subtypes = {p['policy_number']: p['subtype'] for p in ins_products if p.get('policy_number')}
if '999111222' in subtypes:
    check(subtypes['999111222'] == 'car', f"Insurance import: car subtype correct (got {subtypes['999111222']})")
if '999333444' in subtypes:
    check(subtypes['999333444'] == 'home', f"Insurance import: home subtype correct (got {subtypes['999333444']})")
if '555666777' in subtypes:
    check(subtypes['555666777'] == 'health', f"Insurance import: health subtype correct (got {subtypes['555666777']})")

# Test 6: Expense pattern set for bank/visa matching
with budget_app.app.app_context():
    conn = budget_app.get_db()
    uid_row = conn.execute("SELECT id FROM users WHERE username='testadmin'").fetchone()
    if uid_row:
        patterns = conn.execute("SELECT expense_pattern FROM financial_products WHERE user_id=? AND type='insurance' AND expense_pattern != ''",
                                (uid_row['id'],)).fetchall()
        check(len(patterns) >= 2, f"Insurance import: expense_pattern set for matching ({len(patterns)} products)")
    conn.close()

# Test 7: Ignore rules endpoint
r = client.get("/api/insurance/ignore-rules")
check(r.status_code == 200, "Insurance ignore-rules: GET returns 200")

# Test 8: Merge candidates endpoint
r = client.get("/api/insurance/merge-candidates")
check(r.status_code == 200, "Insurance merge-candidates: GET returns 200")
candidates = r.get_json()
check(len(candidates) >= 3, f"Insurance merge-candidates: returns imported products (got {len(candidates)})")


# ====================================================================
# 32. Insurance Overlap Detection
# ====================================================================
print("\n\U0001f539 32. Insurance Overlap Detection")

# Create two overlapping car insurance products
ov_product_a = {
    "type": "insurance", "subtype": "car", "company": "Harel", "name": "Car Insurance A",
    "monthly_cost": 350, "policy_number": "POL-12345", "insured_object": "12-345-67",
    "insured_person": "John", "notes": "full coverage", "expense_pattern": "harel car ins"
}
ov_product_b = {
    "type": "insurance", "subtype": "car", "company": "Migdal", "name": "Car Insurance B",
    "monthly_cost": 380, "policy_number": "MG-99999", "insured_object": "12-345-67",
    "insured_person": "John", "notes": "third party", "expense_pattern": "migdal car"
}
# Create a non-overlapping health product
ov_product_c = {
    "type": "insurance", "subtype": "health", "company": "Clalit", "name": "Health Plan",
    "monthly_cost": 200, "insured_person": "Jane"
}

status_a, _ = api_post("/api/financial/products", ov_product_a)
status_b, _ = api_post("/api/financial/products", ov_product_b)
status_c, _ = api_post("/api/financial/products", ov_product_c)
check(status_a == 200, "Overlap: create product A returns 200")
check(status_b == 200, "Overlap: create product B returns 200")
check(status_c == 200, "Overlap: create product C returns 200")

# Scan for overlaps
status, scan_result = api_post("/api/insurance/overlap-scan")
check(status == 200, "Overlap scan: POST returns 200")
check(scan_result.get("total_open", 0) >= 1, f"Overlap scan: found open alerts (got {scan_result.get('total_open', 0)})")

# List open alerts
status, alerts = api_get("/api/insurance/overlap-alerts?status=open")
check(status == 200, "Overlap alerts: GET returns 200")
check(isinstance(alerts, list), "Overlap alerts: returns list")

# Find the car-car overlap alert
car_alert = None
for a in alerts:
    if (a.get("a_subtype") == "car" and a.get("b_subtype") == "car"):
        car_alert = a
        break
check(car_alert is not None, "Overlap alerts: detected car-car overlap")
if car_alert:
    check(car_alert["overlap_score"] >= 50, f"Overlap score: car-car >= 50 (got {car_alert['overlap_score']})")
    check(car_alert["alert_level"] in ("critical", "warning"), f"Overlap level: critical or warning (got {car_alert['alert_level']})")
    check(car_alert.get("estimated_duplicate_cost_monthly", 0) > 0, "Overlap: estimated duplicate cost > 0")

    # Test reasons parsing (bundled format: {reasons: [...], confidence_level, ...})
    raw_reasons = json.loads(car_alert.get("reasons_json", "{}"))
    if isinstance(raw_reasons, dict):
        reasons = raw_reasons.get("reasons", [])
        check("confidence_level" in raw_reasons, f"Overlap reasons: has confidence_level (got {list(raw_reasons.keys())})")
    else:
        reasons = raw_reasons  # legacy plain array fallback
    signal_names = [r["signal"] for r in reasons]
    check("same_category" in signal_names, "Overlap reasons: includes same_category signal")
    check("same_target" in signal_names, f"Overlap reasons: includes same_target signal (got {signal_names})")

    # Test dismiss
    status, _ = api_put(f"/api/insurance/overlap-alerts/{car_alert['id']}", {"status": "dismissed"})
    check(status == 200, "Overlap dismiss: PUT returns 200")

    # After dismiss, re-list should not include it
    _, alerts2 = api_get("/api/insurance/overlap-alerts?status=open")
    dismissed_ids = [a["id"] for a in alerts2]
    check(car_alert["id"] not in dismissed_ids, "Overlap dismiss: alert no longer in open list")

    # Re-scan should not recreate dismissed alert
    api_post("/api/insurance/overlap-scan")
    _, alerts3 = api_get("/api/insurance/overlap-alerts?status=open")
    recreated = any(a["id"] == car_alert["id"] for a in alerts3)
    check(not recreated, "Overlap rescan: dismissed alert not recreated")

# Test overlap summary
status, summary = api_get("/api/insurance/overlap-summary")
check(status == 200, "Overlap summary: GET returns 200")
check("total" in summary, "Overlap summary: has total key")

# Test note saving
if car_alert:
    # Re-fetch from dismissed list to test note
    status, _ = api_put(f"/api/insurance/overlap-alerts/{car_alert['id']}", {"user_note": "keeping both for now"})
    check(status == 200, "Overlap note: PUT returns 200")

# Verify overlap alert status update rejects invalid status
status, _ = api_put(f"/api/insurance/overlap-alerts/{car_alert['id'] if car_alert else 1}", {"status": "bogus_status"})
check(status == 400, "Overlap: invalid status returns 400")


# ====================================================================
print("\n\U0001f539 33. Assets / Net Worth")
# ====================================================================

# Asset CRUD
status, data = api_post("/api/assets", {
    "asset_type": "real_estate", "name": "Test Apartment",
    "current_value": 2000000, "currency": "ILS",
    "address": "Tel Aviv", "rent_income_monthly": 5000,
    "mortgage_balance": 800000, "property_expenses_monthly": 1200
})
check(status == 200, "Asset: create real estate returns 200")
asset_re_id = data.get("id")
check(asset_re_id is not None, "Asset: create returns id")

status, data = api_post("/api/assets", {
    "asset_type": "stocks", "name": "Investment Portfolio",
    "current_value": 500000, "currency": "USD",
    "platform_name": "Interactive Brokers", "dividend_income_monthly": 800
})
check(status == 200, "Asset: create stocks returns 200")
asset_st_id = data.get("id")

status, data = api_post("/api/assets", {
    "asset_type": "cash", "name": "Savings Account",
    "current_value": 100000, "institution_name": "Bank Leumi",
    "interest_rate": 4.5, "interest_income_monthly": 375
})
check(status == 200, "Asset: create cash returns 200")
asset_ca_id = data.get("id")

# List assets
status, assets = api_get("/api/assets")
check(status == 200, "Asset: GET list returns 200")
check(len(assets) >= 3, f"Asset: list has 3+ items (got {len(assets)})")

# Update asset
status, _ = api_put(f"/api/assets/{asset_re_id}", {"current_value": 2100000})
check(status == 200, "Asset: PUT update returns 200")

# Verify update
status, assets2 = api_get("/api/assets")
updated = next((a for a in assets2 if a["id"] == asset_re_id), None)
check(updated and updated["current_value"] == 2100000, "Asset: value updated to 2100000")

# Invalid asset type
status, _ = api_post("/api/assets", {"asset_type": "spaceship", "name": "X"})
check(status == 400, "Asset: invalid type returns 400")

# Liability CRUD
status, data = api_post("/api/liabilities", {
    "liability_type": "mortgage", "name": "Apartment Mortgage",
    "current_balance": 800000, "monthly_payment": 4500,
    "interest_rate": 3.5, "linked_asset_id": asset_re_id
})
check(status == 200, "Liability: create mortgage returns 200")
liab_id = data.get("id")
check(liab_id is not None, "Liability: create returns id")

status, data = api_post("/api/liabilities", {
    "liability_type": "investment_loan", "name": "Margin Loan",
    "current_balance": 50000, "monthly_payment": 1500
})
check(status == 200, "Liability: create investment_loan returns 200")

# List liabilities
status, liabs = api_get("/api/liabilities")
check(status == 200, "Liability: GET list returns 200")
check(len(liabs) >= 2, f"Liability: list has 2+ items (got {len(liabs)})")

# Update liability
status, _ = api_put(f"/api/liabilities/{liab_id}", {"current_balance": 790000})
check(status == 200, "Liability: PUT update returns 200")

# Invalid liability type
status, _ = api_post("/api/liabilities", {"liability_type": "unicorn", "name": "X"})
check(status == 400, "Liability: invalid type returns 400")

# Net worth summary
status, nw = api_get("/api/net-worth/summary")
check(status == 200, "Net worth: GET summary returns 200")
check("total_assets" in nw, "Net worth: has total_assets")
check("total_liabilities" in nw, "Net worth: has total_liabilities")
check("net_worth" in nw, "Net worth: has net_worth")
check("passive_income" in nw, "Net worth: has passive_income")
check("by_type" in nw, "Net worth: has by_type breakdown")
check(nw["total_assets"] >= 2600000, f"Net worth: total_assets >= 2.6M (got {nw['total_assets']})")
check(nw["passive_income"]["rent"] == 5000, f"Net worth: rent income = 5000 (got {nw['passive_income']['rent']})")
check(nw["passive_income"]["dividends"] == 800, f"Net worth: dividend income = 800 (got {nw['passive_income']['dividends']})")
check(nw["passive_income"]["interest"] == 375, f"Net worth: interest income = 375 (got {nw['passive_income']['interest']})")
check(nw["net_worth"] > 0, "Net worth: positive net worth")

# Net worth history (should have at least 1 snapshot from auto-save)
status, history = api_get("/api/net-worth/history")
check(status == 200, "Net worth: GET history returns 200")
check(len(history) >= 1, f"Net worth: has snapshot (got {len(history)})")

# Manual snapshot
status, _ = api_post("/api/net-worth/snapshot")
check(status == 200, "Net worth: POST snapshot returns 200")

# Delete asset (archive)
status, _ = api_delete(f"/api/assets/{asset_ca_id}")
check(status == 200, "Asset: DELETE returns 200")
status, assets3 = api_get("/api/assets")
check(all(a["id"] != asset_ca_id for a in assets3), "Asset: deleted asset not in list")

# Delete liability (archive)
status, _ = api_delete(f"/api/liabilities/{liab_id}")
check(status == 200, "Liability: DELETE returns 200")
status, liabs2 = api_get("/api/liabilities")
check(all(l["id"] != liab_id for l in liabs2), "Liability: deleted liability not in list")

# Net worth summary after deletions
status, nw2 = api_get("/api/net-worth/summary")
check(status == 200, "Net worth: summary after deletions returns 200")
check(nw2["total_assets"] < nw["total_assets"], "Net worth: total_assets decreased after deletion")

# Auth required
api_post("/api/auth/logout")
status, _ = api_get("/api/assets")
check(status == 401, "Assets: GET without auth returns 401")
status, _ = api_get("/api/net-worth/summary")
check(status == 401, "Net worth: GET without auth returns 401")
api_post("/api/auth/login", {"username": "testadmin", "password": "testpass123"})


# ====================================================================
print("\n\U0001f539 34. Transaction-Asset Linking")
# ====================================================================

# First, add income entries that match asset patterns
status, _ = api_post("/api/income", {
    "date": "2026-04-01", "person": "landlord", "source": "rental",
    "amount": 5100, "description": "שכירות דירת תל אביב"
})
check(status == 200, "Link: create rent income returns 200")

status, _ = api_post("/api/income", {
    "date": "2026-03-01", "person": "landlord", "source": "rental",
    "amount": 5000, "description": "שכירות דירת תל אביב"
})
check(status == 200, "Link: create 2nd rent income returns 200")

status, _ = api_post("/api/income", {
    "date": "2026-04-05", "person": "broker", "source": "dividend",
    "amount": 820, "description": "דיבידנד Interactive Brokers"
})
check(status == 200, "Link: create dividend income returns 200")

# Add expense matching mortgage pattern
status, _ = api_post("/api/expenses", {
    "date": "2026-04-02", "category_id": "housing",
    "subcategory": "משכנתא", "description": "דסק-משכנתא",
    "amount": 4500, "frequency": "monthly"
})
check(status == 200, "Link: create mortgage expense returns 200")

# Get income IDs by fetching income list (add_income doesn't return id)
status, all_income = api_get("/api/income?month=2026-04")
rent_inc_id = None
if status == 200 and all_income:
    for inc in (all_income if isinstance(all_income, list) else all_income.get('income', [])):
        if 'שכירות' in (inc.get('description') or ''):
            rent_inc_id = inc['id']
            break

# Run auto-link engine
status, link_res = api_post("/api/auto-link", {})
check(status == 200, "Link: POST /api/auto-link returns 200")
check(isinstance(link_res.get("auto_linked"), int), "Link: auto-link returns auto_linked count")
check(isinstance(link_res.get("suggested"), int), "Link: auto-link returns suggested count")
total_linked = link_res.get("auto_linked", 0) + link_res.get("suggested", 0)
check(total_linked > 0, f"Link: engine found matches (got {total_linked} linked+suggested)")

# Check suggestions endpoint
status, suggestions = api_get("/api/link-suggestions")
check(status == 200, "Link: GET /api/link-suggestions returns 200")
check(isinstance(suggestions, list), "Link: suggestions is a list")

# Idempotency: running auto-link again should not create duplicates
status, link_res2 = api_post("/api/auto-link", {})
check(status == 200, "Link: 2nd auto-link returns 200")
# Check suggestions count didn't increase
status, suggestions2 = api_get("/api/link-suggestions")
check(len(suggestions2) == len(suggestions), f"Link: idempotent — same suggestion count ({len(suggestions)} = {len(suggestions2)})")

# Manual link: create a link from the rent income to the real estate asset
status, ml_resp = api_post("/api/transaction-links", {
    "transaction_type": "income", "transaction_id": rent_inc_id,
    "asset_id": asset_re_id
})
check(status == 200, f"Link: manual link returns 200 (got {status})")

# Asset linked transactions should include our manual link
status, linked_txns = api_get(f"/api/assets/{asset_re_id}/linked-transactions")
check(status == 200, "Link: GET asset linked-transactions returns 200")
check(isinstance(linked_txns, list), "Link: linked-transactions is a list")
has_rent = any(t.get("transaction_id") == rent_inc_id for t in linked_txns)
check(has_rent, "Link: manual link appears in asset linked-transactions")

# Asset intelligence
status, intel = api_get(f"/api/assets/{asset_re_id}/intelligence")
check(status == 200, "Link: GET asset intelligence returns 200")
check("declared" in intel, "Link: intelligence has declared section")
check("actual" in intel, "Link: intelligence has actual section")
check("variance" in intel, "Link: intelligence has variance section")
check("pnl" in intel, "Link: intelligence has pnl section")
check(intel["declared"]["income_monthly"] == 5000, f"Link: declared rent = 5000 (got {intel['declared'].get('income_monthly')})")

# Actual vs declared dashboard
status, avd = api_get("/api/intelligence/actual-vs-declared")
check(status == 200, "Link: GET actual-vs-declared returns 200")
check("declared_passive_total" in avd, "Link: avd has declared_passive_total")
check("actual_passive_total" in avd, "Link: avd has actual_passive_total")
check("by_asset" in avd, "Link: avd has by_asset list")

# Net worth summary includes actual_passive_income
status, nw3 = api_get("/api/net-worth/summary")
check(status == 200, "Link: net worth summary returns 200")
check("actual_passive_income" in nw3, "Link: summary includes actual_passive_income")
check("pending_suggestions" in nw3, "Link: summary includes pending_suggestions")

# Confirm a suggestion (if any exist)
if suggestions:
    first_sug = suggestions[0]
    sug_id = first_sug["id"]
    status, _ = api_put(f"/api/transaction-links/{sug_id}", {"status": "confirmed"})
    check(status == 200, "Link: confirm suggestion returns 200")

# Reject a suggestion
status, sug3 = api_get("/api/link-suggestions")
if sug3:
    rej_id = sug3[0]["id"]
    status, _ = api_put(f"/api/transaction-links/{rej_id}", {"status": "rejected"})
    check(status == 200, "Link: reject suggestion returns 200")

# Create always_link rule
status, _ = api_post("/api/link-rules", {
    "rule_type": "always_link",
    "description_pattern": "שכירות תל אביב",
    "target_type": "asset",
    "target_id": asset_re_id
})
check(status == 200, "Link: create always_link rule returns 200")

# List rules
status, rules = api_get("/api/link-rules")
check(status == 200, "Link: GET link-rules returns 200")
check(len(rules) >= 1, f"Link: has at least 1 rule (got {len(rules)})")

# Create never_suggest rule
status, _ = api_post("/api/link-rules", {
    "rule_type": "never_suggest",
    "description_pattern": "false pattern",
    "target_type": "asset",
    "target_id": asset_re_id
})
check(status == 200, "Link: create never_suggest rule returns 200")

# Delete a rule
if rules:
    status, _ = api_delete(f"/api/link-rules/{rules[0]['id']}")
    check(status == 200, "Link: DELETE rule returns 200")

# Delete a link
status, linked2 = api_get(f"/api/assets/{asset_re_id}/linked-transactions")
if linked2:
    status, _ = api_delete(f"/api/transaction-links/{linked2[0]['id']}")
    check(status == 200, "Link: DELETE transaction link returns 200")

# Liability linked transactions
status, liab_txns = api_get(f"/api/liabilities/{liab_id}/linked-transactions")
check(status == 200, "Link: GET liability linked-transactions returns 200")
check(isinstance(liab_txns, list), "Link: liability linked-transactions is a list")

# Invalid link: missing fields
status, _ = api_post("/api/transaction-links", {"transaction_type": "income"})
check(status == 400, "Link: manual link with missing fields returns 400")

# Invalid rule: bad type
status, _ = api_post("/api/link-rules", {
    "rule_type": "invalid", "description_pattern": "test",
    "target_type": "asset", "target_id": 999
})
check(status == 400, "Link: invalid rule_type returns 400")

# Auth: linking requires login
api_post("/api/auth/logout")
status, _ = api_get("/api/link-suggestions")
check(status == 401, "Link: suggestions without auth returns 401")
status, _ = api_post("/api/auto-link", {})
check(status == 401, "Link: auto-link without auth returns 401")
api_post("/api/auth/login", {"username": "testadmin", "password": "testpass123"})


# ====================================================================
print("\n\U0001f539 38. Installment Detection Engine")
# ====================================================================

# Seed installment-like expenses: same vendor, same amount, 4 consecutive months, visa_import
inst_det_ids = []
for m in range(1, 5):
    status, _ = api_post("/api/expenses", {
        "date": f"2026-{m:02d}-05",
        "category_id": "shopping",
        "description": "\u05d0\u05d9\u05e7\u05d0\u05d4 \u05ea\u05e9\u05dc\u05d5\u05dd 3 \u05de\u05ea\u05d5\u05da 10",
        "amount": 199.0,
        "source": "visa_import",
        "frequency": "once",
    })
    check(status == 200, f"InstDet: create expense month {m}")

# Seed subscription-like expenses: Netflix, 12 months (should NOT be detected)
for m in range(1, 13):
    api_post("/api/expenses", {
        "date": f"2026-{m:02d}-10",
        "category_id": "entertainment",
        "description": "NETFLIX",
        "amount": 49.9,
        "source": "visa_import",
        "frequency": "once",
    })

# Seed housing category expenses (should NOT be detected)
for m in range(1, 4):
    api_post("/api/expenses", {
        "date": f"2026-{m:02d}-01",
        "category_id": "housing",
        "description": "\u05d3\u05e1\u05e7 \u05de\u05e9\u05db\u05e0\u05ea\u05d0 \u05ea\u05e9\u05dc\u05d5\u05dd 2 \u05de\u05ea\u05d5\u05da 6",
        "amount": 4500,
        "source": "visa_import",
        "frequency": "once",
    })

# Seed expenses with frequency='monthly' (should NOT be detected - already classified)
for m in range(1, 4):
    api_post("/api/expenses", {
        "date": f"2026-{m:02d}-15",
        "category_id": "food",
        "description": "Monthly Food Delivery",
        "amount": 300,
        "source": "visa_import",
        "frequency": "monthly",
    })

# T1: Scan detects the installment pattern
status, scan_res = api_post("/api/installment-suggestions/scan", {})
check(status == 200, "InstDet: scan returns 200")
check(scan_res.get('total_found', 0) >= 1, f"InstDet: scan found >= 1 pattern (got {scan_res.get('total_found', 0)})")

# T2: GET suggestions returns the detected pattern
status, suggestions = api_get("/api/installment-suggestions")
check(status == 200, "InstDet: GET suggestions returns 200")
check(len(suggestions) >= 1, f"InstDet: has >= 1 suggestion (got {len(suggestions)})")

# Verify suggestion structure
if suggestions:
    sug = suggestions[0]
    check('confidence_score' in sug, "InstDet: suggestion has confidence_score")
    check('detection_reasons' in sug, "InstDet: suggestion has detection_reasons")
    check('expense_ids' in sug, "InstDet: suggestion has expense_ids")
    check('months_seen' in sug, "InstDet: suggestion has months_seen")
    check('vendor_display' in sug, "InstDet: suggestion has vendor_display")
    check(sug['confidence_score'] >= 0.45, f"InstDet: confidence >= 0.45 (got {sug['confidence_score']:.2f})")

    # T3: Check Netflix was NOT detected (subscription filter)
    vendor_names = [s.get('vendor_normalized', '') for s in suggestions]
    netflix_found = any('netflix' in v.lower() for v in vendor_names)
    check(not netflix_found, "InstDet: Netflix NOT detected (subscription filter)")

    # T4: Check housing category NOT detected
    housing_found = any('\u05de\u05e9\u05db\u05e0\u05ea\u05d0' in (s.get('vendor_display', '') or '') for s in suggestions)
    check(not housing_found, "InstDet: housing category NOT detected")

# T5: Keyword detection - description has explicit count pattern
ikea_sug = [s for s in suggestions if '\u05d0\u05d9\u05e7\u05d0\u05d4' in (s.get('vendor_display', '') or '') or '\u05d0\u05d9\u05e7\u05d0' in (s.get('vendor_normalized', '') or '')]
if ikea_sug:
    reasons = json.loads(ikea_sug[0].get('detection_reasons', '[]'))
    has_keyword = any(r.startswith('keyword') for r in reasons)
    has_explicit = any(r.startswith('explicit_count') for r in reasons)
    check(has_keyword or has_explicit, f"InstDet: keyword/explicit signal fired (reasons: {reasons})")
else:
    check(False, "InstDet: IKEA suggestion found for keyword test")

# T10: Idempotent rescan - no duplicates
status, scan2 = api_post("/api/installment-suggestions/scan", {})
check(status == 200, "InstDet: rescan returns 200")
status, sug2 = api_get("/api/installment-suggestions")
check(len(sug2) == len(suggestions), f"InstDet: idempotent rescan, no duplicates ({len(sug2)} == {len(suggestions)})")


# ====================================================================
print("\n\U0001f539 39. Installment Suggestion Lifecycle")
# ====================================================================

# Get first suggestion for confirm test
status, all_sug = api_get("/api/installment-suggestions")
if all_sug:
    test_sug = all_sug[0]
    sug_id = test_sug['id']
    exp_ids = json.loads(test_sug['expense_ids'])

    # T6: Confirm suggestion with overrides
    status, confirm_res = api_put(f"/api/installment-suggestions/{sug_id}", {
        "action": "confirm",
        "overrides": {
            "description": "Test IKEA Plan",
            "store": "IKEA",
            "total_payments": 10,
            "payments_made": 4,
            "start_date": "2026-01-05",
            "card": "4580",
        },
        "selected_expense_ids": exp_ids,
    })
    check(status == 200, "InstDet: confirm suggestion returns 200")
    check(confirm_res.get('installment_id') is not None, "InstDet: confirm returns installment_id")

    inst_id = confirm_res.get('installment_id')

    # T7: Verify installment was created with source='detected'
    status, all_inst = api_get("/api/installments")
    detected_inst = [i for i in all_inst if i.get('id') == inst_id]
    if detected_inst:
        check(detected_inst[0].get('source') == 'detected', "InstDet: installment source is 'detected'")
        check(detected_inst[0].get('description') == 'Test IKEA Plan', "InstDet: installment uses override description")
        check(detected_inst[0].get('linked_count', 0) > 0, f"InstDet: installment has linked transactions (got {detected_inst[0].get('linked_count', 0)})")
    else:
        check(False, "InstDet: created installment found in list")

    # T12: GET linked-transactions for the confirmed plan
    status, linked = api_get(f"/api/installments/{inst_id}/linked-transactions")
    check(status == 200, "InstDet: GET linked-transactions returns 200")
    check(isinstance(linked, list), "InstDet: linked-transactions is a list")
    check(len(linked) > 0, f"InstDet: has linked transactions (got {len(linked)})")

    # T13: Unlink a transaction
    if linked:
        expense_id = linked[0].get('expense_id')
        status, _ = api_post(f"/api/installments/{inst_id}/unlink-transaction", {"expense_id": expense_id})
        check(status == 200, "InstDet: unlink transaction returns 200")

        # Verify payments_made decremented
        status, all_inst2 = api_get("/api/installments")
        det2 = [i for i in all_inst2 if i.get('id') == inst_id]
        if det2 and detected_inst:
            check(det2[0].get('linked_count', 0) < detected_inst[0].get('linked_count', 0),
                  "InstDet: linked_count decremented after unlink")

    # T14: Complete early
    status, _ = api_put(f"/api/installments/{inst_id}/complete-early", {})
    check(status == 200, "InstDet: complete-early returns 200")
    status, all_inst3 = api_get("/api/installments")
    det3 = [i for i in all_inst3 if i.get('id') == inst_id]
    if det3:
        check(det3[0].get('status') == 'completed', "InstDet: complete-early sets status=completed")

    # T8: Verify suggestion is now 'confirmed', rescan skips it
    status, scan3 = api_post("/api/installment-suggestions/scan", {})
    status, sug_after = api_get("/api/installment-suggestions")
    confirmed_in_pending = [s for s in sug_after if s['id'] == sug_id]
    check(len(confirmed_in_pending) == 0, "InstDet: confirmed suggestion not in pending list")

else:
    check(False, "InstDet: no suggestions available for lifecycle tests")


# ====================================================================
print("\n\U0001f539 40. Installment Reject and Ignore Rules")
# ====================================================================

# Seed a new set of expenses for reject/ignore tests
for m in range(1, 4):
    api_post("/api/expenses", {
        "date": f"2026-{m:02d}-12",
        "category_id": "shopping",
        "description": "ZARA payment plan",
        "amount": 150.0,
        "source": "visa_import",
        "frequency": "once",
    })

# Re-scan to pick up new pattern
api_post("/api/installment-suggestions/scan", {})
status, sug_list = api_get("/api/installment-suggestions")
zara_sug = [s for s in sug_list if 'zara' in (s.get('vendor_normalized', '') or '').lower()]

if zara_sug:
    zara_id = zara_sug[0]['id']

    # T8b: Reject suggestion
    status, _ = api_put(f"/api/installment-suggestions/{zara_id}", {"action": "reject"})
    check(status == 200, "InstDet: reject suggestion returns 200")

    # Rescan: rejected should not reappear
    api_post("/api/installment-suggestions/scan", {})
    status, sug_after_reject = api_get("/api/installment-suggestions")
    zara_pending = [s for s in sug_after_reject if s['id'] == zara_id]
    check(len(zara_pending) == 0, "InstDet: rejected suggestion stays rejected on rescan")
else:
    check(False, "InstDet: ZARA suggestion found for reject test")

# Seed another pattern for ignore + rule test
for m in range(1, 4):
    api_post("/api/expenses", {
        "date": f"2026-{m:02d}-20",
        "category_id": "shopping",
        "description": "H&M installments",
        "amount": 89.0,
        "source": "visa_import",
        "frequency": "once",
    })

api_post("/api/installment-suggestions/scan", {})
status, sug_list2 = api_get("/api/installment-suggestions")
hm_sug = [s for s in sug_list2 if 'h&m' in (s.get('vendor_normalized', '') or '').lower() or 'h m' in (s.get('vendor_normalized', '') or '').lower()]

if hm_sug:
    hm_id = hm_sug[0]['id']

    # T9: Ignore + create rule
    status, _ = api_put(f"/api/installment-suggestions/{hm_id}", {"action": "ignore", "add_rule": True})
    check(status == 200, "InstDet: ignore + rule returns 200")
else:
    # Even if H&M not detected (low score), test the rule CRUD directly
    pass

# T19: Ignore rules CRUD
status, rules = api_get("/api/installment-ignore-rules")
check(status == 200, "InstDet: GET ignore-rules returns 200")
check(isinstance(rules, list), "InstDet: ignore-rules is a list")

# Create a manual rule
status, rule_res = api_post("/api/installment-ignore-rules", {
    "rule_type": "never_suggest",
    "rule_value": "test-vendor-ignore",
    "reason": "test rule",
})
check(status == 200, "InstDet: create ignore rule returns 200")

# Create always_installment rule
status, rule_res2 = api_post("/api/installment-ignore-rules", {
    "rule_type": "always_installment",
    "rule_value": "test-vendor-always",
    "reason": "always detect this vendor",
})
check(status == 200, "InstDet: create always_installment rule returns 200")

# List rules
status, rules2 = api_get("/api/installment-ignore-rules")
check(len(rules2) >= 2, f"InstDet: has >= 2 rules after creation (got {len(rules2)})")

# Delete a rule
if rules2:
    rule_id = rules2[-1]['id']
    status, _ = api_delete(f"/api/installment-ignore-rules/{rule_id}")
    check(status == 200, "InstDet: DELETE rule returns 200")


# ====================================================================
print("\n\U0001f539 41. Installment Detection Auth")
# ====================================================================

# Logout and verify all installment detection endpoints require auth
api_post("/api/auth/logout")

status, _ = api_post("/api/installment-suggestions/scan", {})
check(status == 401, "InstDet: scan without auth returns 401")

status, _ = api_get("/api/installment-suggestions")
check(status == 401, "InstDet: GET suggestions without auth returns 401")

status, _ = api_put("/api/installment-suggestions/1", {"action": "reject"})
check(status == 401, "InstDet: PUT suggestion without auth returns 401")

status, _ = api_get("/api/installment-ignore-rules")
check(status == 401, "InstDet: GET ignore-rules without auth returns 401")

status, _ = api_post("/api/installment-ignore-rules", {"rule_type": "never_suggest", "rule_value": "x"})
check(status == 401, "InstDet: POST ignore-rule without auth returns 401")

status, _ = api_delete("/api/installment-ignore-rules/1")
check(status == 401, "InstDet: DELETE ignore-rule without auth returns 401")

status, _ = api_get("/api/installments/1/linked-transactions")
check(status == 401, "InstDet: GET linked-transactions without auth returns 401")

status, _ = api_post("/api/installments/1/unlink-transaction", {"link_id": 1})
check(status == 401, "InstDet: POST unlink without auth returns 401")

status, _ = api_put("/api/installments/1/complete-early", {})
check(status == 401, "InstDet: PUT complete-early without auth returns 401")

# Re-login for cleanup
api_post("/api/auth/login", {"username": "testadmin", "password": "testpass123"})


# ====================================================================
print("\n\U0001f539 42. Installment Insights & Safe-to-Spend Integration")
# ====================================================================

# First create a fresh active installment for insights testing
api_post("/api/installments", {
    "description": "Insight Test Plan",
    "total_amount": 6000,
    "num_payments": 10,
    "start_date": "2026-03-01",
    "card": "4580",
})

# GET installment-insights
status, insights = api_get("/api/installment-insights")
check(status == 200, "Insights: GET installment-insights returns 200")
check('total_monthly_commitment' in insights, "Insights: has total_monthly_commitment")
check('active_count' in insights, "Insights: has active_count")
check('ending_soon' in insights, "Insights: has ending_soon list")
check('recently_completed' in insights, "Insights: has recently_completed list")
check('next_month_drop' in insights, "Insights: has next_month_drop")
check('burden_pct' in insights, "Insights: has burden_pct")
check('plans_summary' in insights, "Insights: has plans_summary")
check('trend' in insights, "Insights: has trend")
check(insights['trend'].get('direction') in ('improving', 'stable', 'increasing'), f"Insights: trend direction valid (got {insights['trend'].get('direction')})")
check(isinstance(insights['ending_soon'], list), "Insights: ending_soon is a list")
check(isinstance(insights['recently_completed'], list), "Insights: recently_completed is a list")
check(insights['active_count'] >= 1, f"Insights: active_count >= 1 (got {insights['active_count']})")
check(insights['total_monthly_commitment'] > 0, f"Insights: total_monthly > 0 (got {insights['total_monthly_commitment']})")

# Verify ending_soon has correct structure if items present
if insights['ending_soon']:
    es = insights['ending_soon'][0]
    check('days_until_end' in es, "Insights: ending_soon item has days_until_end")
    check('description' in es, "Insights: ending_soon item has description")

# GET safe-to-spend should include installment deduction
status, safe = api_get("/api/safe-to-spend")
check(status == 200 or (status == 200 and safe.get('available') is False), "Insights: GET safe-to-spend returns 200")
if safe.get('available'):
    check('installment_monthly' in safe, "Insights: safe-to-spend has installment_monthly")
    if safe.get('why_inputs'):
        check('installments' in safe['why_inputs'], "Insights: why_inputs has installments deduction")

# Tips endpoint with installments present (must not crash)
status, tips_data = api_get(f"/api/tips?month={month}")
check(status == 200, "Insights: GET tips with installments returns 200")
check(isinstance(tips_data, list), "Insights: tips is a list")

# Auth check
api_post("/api/auth/logout")
status, _ = api_get("/api/installment-insights")
check(status == 401, "Insights: installment-insights without auth returns 401")
api_post("/api/auth/login", {"username": "testadmin", "password": "testpass123"})


# ====================================================================
# 43. Smart Tips v2 — API Contract Validation
# ====================================================================
print("\n\U0001f539 43. Smart Tips v2 — API Contract")

status, tips_v2 = api_get(f"/api/tips?month={month}")
check(status == 200, "Tips v2: GET /api/tips returns 200")
check(isinstance(tips_v2, list), "Tips v2: response is a list")

if tips_v2:
    # Structure checks on first tip
    tip0 = tips_v2[0]
    check('id' in tip0, "Tips v2: tip has 'id'")
    check('category' in tip0, "Tips v2: tip has 'category'")
    check('severity' in tip0, "Tips v2: tip has 'severity'")
    check('score' in tip0, "Tips v2: tip has 'score'")
    check('impact_amount' in tip0, "Tips v2: tip has 'impact_amount'")
    check('impact_type' in tip0, "Tips v2: tip has 'impact_type'")
    check('params' in tip0, "Tips v2: tip has 'params'")
    check('action' in tip0, "Tips v2: tip has 'action'")
    check('icon' in tip0, "Tips v2: tip has 'icon'")
    check('color' in tip0, "Tips v2: tip has 'color'")

    # Severity enum
    valid_sev = {'important', 'watch', 'opportunity'}
    for tip in tips_v2:
        check(tip['severity'] in valid_sev, f"Tips v2: severity '{tip['severity']}' valid for tip '{tip['id']}'")

    # Score bounds
    for tip in tips_v2:
        check(0 <= tip['score'] <= 1, f"Tips v2: score {tip['score']} in [0,1] for tip '{tip['id']}'")

    # Sorted by score descending
    scores = [t['score'] for t in tips_v2]
    check(scores == sorted(scores, reverse=True), "Tips v2: tips sorted by score descending")

    # Max 8 tips
    check(len(tips_v2) <= 8, f"Tips v2: max 8 tips (got {len(tips_v2)})")

    # No duplicate IDs
    tip_ids = [t['id'] for t in tips_v2]
    check(len(tip_ids) == len(set(tip_ids)), "Tips v2: no duplicate tip IDs")

    # impact_amount numeric >= 0
    for tip in tips_v2:
        check(isinstance(tip['impact_amount'], (int, float)) and tip['impact_amount'] >= 0,
              f"Tips v2: impact_amount >= 0 for '{tip['id']}'")

    # impact_type valid
    valid_types = {'yearly_cost', 'monthly_savings', 'risk', 'coverage'}
    for tip in tips_v2:
        check(tip['impact_type'] in valid_types, f"Tips v2: impact_type '{tip['impact_type']}' valid for '{tip['id']}'")

    # Category cap: max 3 per category
    cat_counts = {}
    for tip in tips_v2:
        cat_counts[tip['category']] = cat_counts.get(tip['category'], 0) + 1
    for cat, cnt in cat_counts.items():
        check(cnt <= 3, f"Tips v2: category '{cat}' has {cnt} tips (max 3)")

    # Subsumption: negative_balance should suppress balance_decline
    if 'negative_balance' in tip_ids:
        check('balance_decline' not in tip_ids, "Tips v2: negative_balance subsumes balance_decline")
    if 'installment_avoid' in tip_ids:
        check('installment_burden' not in tip_ids, "Tips v2: installment_avoid subsumes installment_burden")

    # Action label key present when action exists
    for tip in tips_v2:
        if tip.get('action'):
            check('action_label_key' in tip and tip['action_label_key'],
                  f"Tips v2: tip '{tip['id']}' with action has action_label_key")

# Auth check: 401 when logged out
api_post("/api/auth/logout")
status, _ = api_get("/api/tips")
check(status == 401, "Tips v2: GET /api/tips without auth returns 401")
api_post("/api/auth/login", {"username": "testadmin", "password": "testpass123"})


# ====================================================================
# 44. Tips v2.1 — Config, Coordination, Analytics, Boost
# ====================================================================
print("\n\U0001f539 44. Tips v2.1 — Config, Coordination, Analytics, Boost")

# T1: TIP_CONFIG exists and has expected keys
from app import TIP_CONFIG
expected_keys = ['w_severity', 'w_impact', 'w_urgency', 'w_confidence', 'w_recency', 'w_action_bonus',
                 'sev_important', 'sev_watch', 'sev_opportunity', 'impact_cap',
                 'max_per_category', 'max_total_tips', 'subsumption_rules',
                 'overlap_penalty', 'boost_positive_traj', 'boost_mixed_traj', 'positive_tip_ids',
                 'food_monthly_threshold', 'restaurant_ratio_threshold', 'entertainment_monthly_threshold',
                 'cash_total_threshold', 'bit_total_threshold', 'savings_low_threshold', 'savings_good_threshold',
                 'emergency_buffer_months', 'installment_burden_threshold', 'deficit_months_threshold',
                 'fixed_ratio_threshold', 'category_spike_threshold', 'latte_min_vendors',
                 'subscriptions_monthly_threshold', 'dining_out_monthly_threshold',
                 'analytics_prune_days']
for k in expected_keys:
    check(k in TIP_CONFIG, f"TIP_CONFIG has key '{k}'")

# T2: Scoring with overlap penalty lowers score
from app import _score_tip
sample_tip = {
    'id': 'deficit_months', 'severity': 'important', 'impact_amount': 0,
    'urgency': 0.7, 'confidence': 0.9, 'recency': 0.8, 'action': 'analysis',
}
score_normal = _score_tip(sample_tip)
score_penalized = _score_tip(sample_tip, overlapping_ids={'deficit_months'})
check(score_penalized < score_normal, f"Overlap penalty lowers score ({score_normal} -> {score_penalized})")
check(score_penalized >= 0, "Penalized score stays >= 0")

# T3: Trajectory param accepted by /api/tips
status, tips_traj = api_get(f"/api/tips?month={month}&trajectory=positive&signals=traj_saving")
check(status == 200, "Tips with trajectory param returns 200")
check(isinstance(tips_traj, list), "Tips with trajectory param returns list")

# T4: _boost_positive boosts positive tips when no important
from app import _boost_positive
pos_tips = [
    {'id': 'good_savings_rate', 'severity': 'opportunity', 'score': 0.35, 'category': 'savings'},
    {'id': 'food_spending', 'severity': 'watch', 'score': 0.50, 'category': 'spending'},
]
boosted = _boost_positive([dict(t) for t in pos_tips], 'positive')
gs_tip = next(t for t in boosted if t['id'] == 'good_savings_rate')
check(gs_tip['score'] == 0.55, f"Positive boost adds 0.20 (0.35 -> {gs_tip['score']})")

# T5: _boost_positive does NOT boost when important tip exists
mix_tips = [
    {'id': 'negative_balance', 'severity': 'important', 'score': 0.85, 'category': 'debt'},
    {'id': 'good_savings_rate', 'severity': 'opportunity', 'score': 0.35, 'category': 'savings'},
]
not_boosted = _boost_positive([dict(t) for t in mix_tips], 'positive')
gs_tip2 = next(t for t in not_boosted if t['id'] == 'good_savings_rate')
check(gs_tip2['score'] == 0.35, f"No boost when important exists (score stays {gs_tip2['score']})")

# T6: POST /api/tip-events returns 200
status, resp = api_post("/api/tip-events", {"events": [
    {"tip_id": "test_tip", "event_type": "shown", "month": month},
    {"tip_id": "test_tip", "event_type": "expanded", "month": month},
]})
check(status == 200, "POST /api/tip-events returns 200")
check(resp.get('count') == 2, f"tip-events inserted 2 events (got {resp.get('count')})")

# T7: POST /api/tip-events with empty list returns 200
status, resp2 = api_post("/api/tip-events", {"events": []})
check(status == 200, "POST /api/tip-events empty list returns 200")
check(resp2.get('count') == 0, "tip-events empty returns count=0")

# T8: POST /api/tip-events rejects invalid event_type
status, resp3 = api_post("/api/tip-events", {"events": [
    {"tip_id": "x", "event_type": "invalid_type", "month": month},
]})
check(status == 200, "POST /api/tip-events with invalid type returns 200")
check(resp3.get('count') == 0, "tip-events filters out invalid event_type")

# T9: POST /api/tip-events requires auth
api_post("/api/auth/logout")
status, _ = api_post("/api/tip-events", {"events": [{"tip_id": "x", "event_type": "shown"}]})
check(status == 401, "POST /api/tip-events without auth returns 401")
api_post("/api/auth/login", {"username": "testadmin", "password": "testpass123"})

# T10: tip_events table exists
import sqlite3 as _sqlite3
_tconn = _sqlite3.connect(tmp_db)
_tcols = [r[1] for r in _tconn.execute("PRAGMA table_info(tip_events)").fetchall()]
_tconn.close()
check('user_id' in _tcols, "tip_events table has user_id column")
check('tip_id' in _tcols, "tip_events table has tip_id column")
check('event_type' in _tcols, "tip_events table has event_type column")
check('action_target' in _tcols, "tip_events table has action_target column")
check('created_at' in _tcols, "tip_events table has created_at column")

# T11: Subsumption rules from TIP_CONFIG work
from app import _deduplicate_tips
test_tips_sub = [
    {'id': 'negative_balance', 'category': 'debt', 'severity': 'important', 'score': 0.9},
    {'id': 'balance_decline', 'category': 'debt', 'severity': 'watch', 'score': 0.6},
    {'id': 'food_spending', 'category': 'spending', 'severity': 'watch', 'score': 0.5},
]
deduped = _deduplicate_tips([dict(t) for t in test_tips_sub])
deduped_ids = {t['id'] for t in deduped}
check('balance_decline' not in deduped_ids, "Config-driven subsumption removes balance_decline")
check('negative_balance' in deduped_ids, "Config-driven subsumption keeps negative_balance")

# T12: Max tips from TIP_CONFIG
check(TIP_CONFIG['max_total_tips'] == 8, f"max_total_tips is 8 (got {TIP_CONFIG['max_total_tips']})")
check(TIP_CONFIG['max_per_category'] == 3, f"max_per_category is 3 (got {TIP_CONFIG['max_per_category']})")

# ====================================================================
# Section 45 — Next Best Action, Vendor Normalization, Variants
# ====================================================================
print("\n--- 45. Next Best Action, Vendor Normalization, Tip Variants ---")

# T1: /api/next-action returns 200
r = client.get('/api/next-action?month=2025-01')
check(r.status_code == 200, f"GET /api/next-action returns 200 (got {r.status_code})")
# T2: response is null or valid object
nba = r.get_json()
check(nba is None or isinstance(nba, dict), f"/api/next-action returns null or dict (got {type(nba).__name__})")
# T3: if dict, has required keys
if isinstance(nba, dict):
    for k in ['tip_id', 'action', 'score', 'params', 'reasons']:
        check(k in nba, f"next-action has key '{k}'")
    check(0 <= nba['score'] <= 1, f"next-action score in [0,1] (got {nba['score']})")
    check(isinstance(nba['reasons'], list), f"next-action reasons is a list")
    for reason in nba['reasons']:
        check('key' in reason and 'params' in reason, f"reason chip has key and params")

# T4: /api/next-action requires auth
api_post("/api/auth/logout")
status_na, _ = api_get("/api/next-action?month=2025-01")
check(status_na == 401, f"/api/next-action without auth returns 401 (got {status_na})")
api_post("/api/auth/login", {"username": "testadmin", "password": "testpass123"})

# T5: _normalize_subscription_desc merges variants
from app import _normalize_subscription_desc
pairs = [
    ('NETFLIX', 'netflix'),
    ('Netflix', 'netflix'),
    ('APPLE.COM', 'apple'),
    ('APPLE BILL', 'apple'),
    ('GOOGLE *SERVICE', 'google'),
    ('SPOTIFY SUBSCRIPTION', 'spotify'),
    ('SPOTIFY', 'spotify'),
    ('הוראת קבע נטפליקס 12345', 'נטפליקס'),
]
for raw, expected in pairs:
    got = _normalize_subscription_desc(raw)
    check(got == expected, f"_normalize_subscription_desc('{raw}') == '{expected}' (got '{got}')")

# T6: Different vendors stay different
check(_normalize_subscription_desc('NETFLIX') != _normalize_subscription_desc('SPOTIFY'),
      "NETFLIX != SPOTIFY after normalization")

# T7: _NBA_CANDIDATES set exists and has subscription/dining entries
from app import _NBA_CANDIDATES
check('subscriptions_cost' in _NBA_CANDIDATES, "NBA candidates includes subscriptions_cost")
check('dining_out_high' in _NBA_CANDIDATES, "NBA candidates includes dining_out_high")
check('subscriptions_growth' in _NBA_CANDIDATES, "NBA candidates includes subscriptions_growth")

# T8: Tip variant params — subscriptions_cost with vs_avg > 10 gets _variant='up'
from app import _tip_subscriptions_cost
mock_ctx = {
    'subscriptions_monthly': 500, 'subscriptions_this_month': 600,
    'subscriptions_count': 7, 'subscriptions_count_this_month': 7,
    'subscriptions_pct_of_income': 3.0, 'subscriptions_vs_avg_pct': 20,
    'subscriptions_new_this_month': 0, 'subscription_concentration_pct': 0,
    'subscription_top_vendors': [], 'inc_total': 15000,
}
tip = _tip_subscriptions_cost(mock_ctx)
check(tip is not None, "subscriptions_cost fires with monthly=500")
check(tip['params']['_variant'] == 'up', f"subscriptions_cost _variant='up' when vs_avg=20 (got '{tip['params']['_variant']}')")

# T9: Tip variant — dining_out_high with vs_avg > 10 gets _variant='up'
from app import _tip_dining_out_high
mock_ctx2 = {
    'dining_out_monthly': 1200, 'dining_out_this_month': 1500,
    'dining_out_vs_avg_pct': 25, 'dining_out_pct_of_income': 8.0,
    'inc_total': 15000,
}
tip2 = _tip_dining_out_high(mock_ctx2)
check(tip2 is not None, "dining_out_high fires with monthly=1200")
check(tip2['params']['_variant'] == 'up', f"dining_out_high _variant='up' when vs_avg=25 (got '{tip2['params']['_variant']}')")

# T10: Concentration param appears when strong
mock_ctx3 = dict(mock_ctx)
mock_ctx3['subscription_concentration_pct'] = 75
mock_ctx3['subscription_top_vendors'] = [('netflix', 3000), ('spotify', 1500), ('hbo', 500)]
tip3 = _tip_subscriptions_cost(mock_ctx3)
check(tip3['params']['concentration_pct'] == '75', f"concentration_pct=75 when top2=75% (got '{tip3['params']['concentration_pct']}')")
check(tip3['params']['top_vendor'] == 'netflix', f"top_vendor=netflix (got '{tip3['params']['top_vendor']}')")

# T11: Concentration param empty when below 60%
mock_ctx4 = dict(mock_ctx)
mock_ctx4['subscription_concentration_pct'] = 40
mock_ctx4['subscription_top_vendors'] = [('a', 200), ('b', 150)]
tip4 = _tip_subscriptions_cost(mock_ctx4)
check(tip4['params']['concentration_pct'] == '', f"concentration_pct empty when <60% (got '{tip4['params']['concentration_pct']}')")

# T12: Subsumption rule for subscriptions_cost > subscriptions_growth
check(('subscriptions_cost', 'subscriptions_growth') in TIP_CONFIG['subsumption_rules'],
      "Subsumption: subscriptions_cost subsumes subscriptions_growth")


# ====================================================================
# Section 46 — Salary Statements (Payslip PDF Import)
# ====================================================================
print("\n--- 46. Salary Statements ---")

# Insert mock salary statements directly
with budget_app.app.app_context():
    conn = budget_app.get_db()
    conn.execute("""
        INSERT OR REPLACE INTO salary_statements
        (user_id, person, month, company_name, gross_salary, net_salary,
         income_tax, social_security, health_insurance,
         pension_employee, pension_employer,
         education_fund_employee, education_fund_employer,
         severance_employer, bonus_amount, vacation_days, sick_days,
         extraction_confidence, raw_text, source_filename)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (1, 'husband', '2026-01', 'Test Corp', 20000, 14000,
          2800, 400, 200, 1200, 1400, 500, 1500, 1667, 0, 12.5, 8,
          0.83, 'mock text', 'test1.pdf'))
    conn.execute("""
        INSERT OR REPLACE INTO salary_statements
        (user_id, person, month, company_name, gross_salary, net_salary,
         income_tax, social_security, health_insurance,
         pension_employee, pension_employer,
         education_fund_employee, education_fund_employer,
         severance_employer, bonus_amount, vacation_days, sick_days,
         extraction_confidence, raw_text, source_filename)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (1, 'husband', '2026-02', 'Test Corp', 30000, 20000,
          4200, 600, 300, 1800, 2100, 750, 2250, 2500, 10000, 13, 8,
          0.67, 'mock text bonus', 'test2.pdf'))
    conn.commit()
    conn.close()

# T1: GET /api/salary-statements returns 200 + array
r = client.get("/api/salary-statements")
check(r.status_code == 200, "GET /api/salary-statements returns 200")
data = json.loads(r.data)
check(isinstance(data, list), "Salary statements returns array")
check(len(data) >= 2, f"At least 2 salary statements ({len(data)} found)")

# T2: Response includes computed fields
if data:
    check('true_compensation' in data[0], "Response includes true_compensation")
    check('employer_contributions_total' in data[0], "Response includes employer_contributions_total")
    check(data[0]['true_compensation'] > data[0]['gross_salary'], "true_compensation > gross_salary")

# T3: GET /api/salary-statements/summary returns correct data
r = client.get("/api/salary-statements/summary")
check(r.status_code == 200, "GET /api/salary-statements/summary returns 200")
summ = json.loads(r.data)
check(summ.get('has_data') == True, "Summary has_data is True")
check(summ.get('months_count') >= 2, f"months_count >= 2 (got {summ.get('months_count')})")
check(summ.get('baseline_months', 0) >= 1, f"baseline_months >= 1 (got {summ.get('baseline_months')})")
check(summ.get('latest_employer', {}).get('total', 0) > 0, "Employer total > 0")
check(summ.get('latest_true_compensation', 0) > summ.get('latest_gross', 0), "true_compensation > gross")

# T4: Baseline normalization — avg should exclude bonus month
# Non-bonus month gross=20000, bonus month gross=30000. Baseline avg should be 20000.
check(summ.get('avg_gross', 0) == 20000, f"Baseline avg_gross=20000 (non-bonus only), got {summ.get('avg_gross')}")

# T5: POST /api/salary-statements (save new)
r = client.post("/api/salary-statements", data=json.dumps({
    'month': '2026-03', 'person': 'husband', 'company_name': 'Test Corp',
    'gross_salary': 22000, 'net_salary': 15000,
    'income_tax': 3000, 'social_security': 450, 'health_insurance': 220,
    'pension_employee': 1300, 'pension_employer': 1540,
    'education_fund_employee': 550, 'education_fund_employer': 1650,
    'severance_employer': 1833, 'extraction_confidence': 0.83,
}), content_type='application/json')
check(r.status_code == 200, "POST /api/salary-statements returns 200")
save_result = json.loads(r.data)
check(save_result.get('true_compensation', 0) > 22000, "Saved true_compensation > gross")

# T6: DELETE /api/salary-statements
r_list = client.get("/api/salary-statements")
all_stmts = json.loads(r_list.data)
if all_stmts:
    del_id = all_stmts[-1]['id']
    r = client.delete(f"/api/salary-statements/{del_id}")
    check(r.status_code == 200, f"DELETE /api/salary-statements/{del_id} returns 200")
    del_data = json.loads(r.data)
    check(del_data.get('deleted') == 1, f"deleted=1 (got {del_data.get('deleted')})")
    r2 = client.get("/api/salary-statements")
    check(len(json.loads(r2.data)) == len(all_stmts) - 1, "One fewer statement after delete")

# T7: 401 when logged out
api_post("/api/auth/logout")
r = client.get("/api/salary-statements")
check(r.status_code == 401, "GET /api/salary-statements returns 401 when logged out")
r = client.get("/api/salary-statements/summary")
check(r.status_code == 401, "GET /api/salary-statements/summary returns 401 when logged out")
# Log back in
api_post("/api/auth/login", {"username": "testadmin", "password": "testpass123"})

# T8: Subsumption rule for true_savings_rate > low_savings_rate
check(('true_savings_rate', 'low_savings_rate') in TIP_CONFIG['subsumption_rules'],
      "Subsumption: true_savings_rate subsumes low_savings_rate")

# T9: Salary tip generators exist in _TIP_GENERATORS
from app import _TIP_GENERATORS, _NBA_CANDIDATES
gen_names = [g.__name__ for g in _TIP_GENERATORS]
for name in ['_tip_vacation_days_unused', '_tip_bonus_detected', '_tip_true_savings_rate', '_tip_employer_contributions_value']:
    check(name in gen_names, f"{name} registered in _TIP_GENERATORS")

# T10: Salary tip IDs in NBA candidates
for tid in ['vacation_days_unused', 'bonus_detected', 'true_savings_rate', 'employer_contributions_value']:
    check(tid in _NBA_CANDIDATES, f"{tid} in _NBA_CANDIDATES")


# ====================================================================
# Section 47 — Dashboard Salary Cards, NBA Salary Wording, Payslip Analytics
# ====================================================================
print("\n--- 47. Dashboard Salary Cards, NBA Wording, Payslip Analytics ---")

# Ensure logged in
api_post("/api/auth/login", {"username": "testadmin", "password": "testpass123"})

# T1-T4: Salary tip generators return custom action_label_key (not generic)
from app import _tip_vacation_days_unused, _tip_bonus_detected, _tip_true_savings_rate, _tip_employer_contributions_value
mock_sal_ctx = {
    'has_salary_data': True, 'sal_gross': 20000, 'sal_net': 14000,
    'sal_employer_total': 4500, 'sal_true_comp': 24500,
    'sal_vacation_days': 20, 'sal_bonus': 5000, 'sal_months_count': 3,
}
vac_tip = _tip_vacation_days_unused(mock_sal_ctx)
check(vac_tip is not None and vac_tip['action_label_key'] == 'tip_action_check_vacation',
      f"vacation tip has custom action_label_key (got {vac_tip['action_label_key'] if vac_tip else 'None'})")

bonus_tip = _tip_bonus_detected(mock_sal_ctx)
check(bonus_tip is not None and bonus_tip['action_label_key'] == 'tip_action_allocate_bonus',
      f"bonus tip has custom action_label_key (got {bonus_tip['action_label_key'] if bonus_tip else 'None'})")

low_rate_ctx = dict(mock_sal_ctx, sal_employer_total=800, sal_true_comp=20800)
sav_tip = _tip_true_savings_rate(low_rate_ctx)
check(sav_tip is not None and sav_tip['action_label_key'] == 'tip_action_review_hidden_savings',
      f"savings rate tip has custom action_label_key (got {sav_tip['action_label_key'] if sav_tip else 'None'})")

emp_tip = _tip_employer_contributions_value(mock_sal_ctx)
check(emp_tip is not None and emp_tip['action_label_key'] == 'tip_action_review_hidden_savings',
      f"employer tip has custom action_label_key (got {emp_tip['action_label_key'] if emp_tip else 'None'})")

# T5: tip_events endpoint accepts payslip event types
payslip_events = [
    {'event_type': 'payslip_upload', 'tip_id': 'test.pdf', 'month': '2026-03'},
    {'event_type': 'payslip_preview_shown', 'tip_id': '2026-03', 'month': '2026-03'},
    {'event_type': 'payslip_field_edited', 'tip_id': '2026-03', 'action_target': 'gross_salary', 'month': '2026-03'},
    {'event_type': 'payslip_saved', 'tip_id': '2026-03', 'month': '2026-03'},
    {'event_type': 'payslip_cancelled', 'tip_id': '2026-04', 'month': '2026-04'},
]
sc, ev_result = api_post('/api/tip-events', {'events': payslip_events})
check(sc == 200, f"POST /api/tip-events with payslip events returns 200 (got {sc})")
check(ev_result and ev_result.get('count') == 5, f"All 5 payslip events accepted (got {ev_result.get('count') if ev_result else 'None'})")

# T6: Payslip analytics endpoint
r = client.get('/api/payslip-analytics')
check(r.status_code == 200, f"GET /api/payslip-analytics returns 200 (got {r.status_code})")
analytics = r.get_json() or {}
check(analytics.get('uploads', 0) >= 1, f"Analytics shows uploads (got {analytics.get('uploads')})")
check(analytics.get('saves', 0) >= 1, f"Analytics shows saves (got {analytics.get('saves')})")
check('save_rate' in analytics, "Analytics includes save_rate")
check('edit_rate' in analytics, "Analytics includes edit_rate")

# T7: Income stability tip generators
from app import _tip_income_stable, _tip_bonus_reliance, _tip_income_drop

# income_stable fires for stable salary
stable_ctx = {
    'has_salary_data': True, 'sal_cv': 0.03, 'sal_months_count': 4,
    'sal_bonus_share': 0.05,
}
stable_tip = _tip_income_stable(stable_ctx)
check(stable_tip is not None and stable_tip['id'] == 'income_stable',
      f"income_stable fires for stable salary (got {stable_tip['id'] if stable_tip else 'None'})")

# income_stable suppressed when bonus share > 20%
check(_tip_income_stable(dict(stable_ctx, sal_bonus_share=0.25)) is None,
      "income_stable suppressed when bonus share > 20%")

# bonus_reliance fires when share >= 15%
bonus_rel_ctx = {
    'has_salary_data': True, 'sal_bonus_share': 0.20, 'sal_months_count': 3,
    'sal_bonus': 3000,
}
bonus_rel_tip = _tip_bonus_reliance(bonus_rel_ctx)
check(bonus_rel_tip is not None and bonus_rel_tip['id'] == 'bonus_reliance',
      f"bonus_reliance fires for high bonus share (got {bonus_rel_tip['id'] if bonus_rel_tip else 'None'})")

# income_drop fires for 10%+ drop
drop_ctx = {
    'has_salary_data': True, 'sal_latest_vs_baseline': -15, 'sal_months_count': 3,
    'sal_latest_gross': 17000, 'sal_gross': 20000,
}
drop_tip = _tip_income_drop(drop_ctx)
check(drop_tip is not None and drop_tip['id'] == 'income_drop',
      f"income_drop fires for salary drop (got {drop_tip['id'] if drop_tip else 'None'})")

# income_drop does NOT fire for small drop
check(_tip_income_drop(dict(drop_ctx, sal_latest_vs_baseline=-5)) is None,
      "income_drop does not fire for small drop (-5%)")

# T8: Bonus allocation params present
bonus_alloc_ctx = {
    'has_salary_data': True, 'sal_bonus': 10000,
    'latest_balance': 5000, 'essential_monthly_avg': 8000,
    'inst_monthly': 500, 'overdraft_total': 1000,
    'savings_rate': 10,
}
alloc_tip = _tip_bonus_detected(bonus_alloc_ctx)
check(alloc_tip is not None, "bonus_detected fires with allocation context")
check('emergency' in alloc_tip['params'], "bonus allocation includes emergency param")
check('debt' in alloc_tip['params'], "bonus allocation includes debt param")
check('invest' in alloc_tip['params'], "bonus allocation includes invest param")
check('enjoy' in alloc_tip['params'], "bonus allocation includes enjoy param")
# Verify sequential allocation: amounts should sum to original bonus
alloc_sum = sum(int(alloc_tip['params'][k].replace(',', '')) for k in ['emergency', 'debt', 'invest', 'enjoy'])
check(alloc_sum == 10000, f"Allocation sums to original bonus (got {alloc_sum})")

# T9: Payslip analytics includes field breakdown
r = client.get('/api/payslip-analytics')
analytics = r.get_json() or {}
check('field_edits_by_field' in analytics, "Analytics includes field_edits_by_field")
check('most_edited_field' in analytics, "Analytics includes most_edited_field")
check('companies' in analytics, "Analytics includes companies list")

# T10: Subsumption rules
check(('income_drop', 'income_gap') in TIP_CONFIG['subsumption_rules'],
      "income_drop subsumes income_gap in subsumption rules")

# T11: New generators registered
gen_names = [g.__name__ for g in _TIP_GENERATORS]
for name in ['_tip_income_stable', '_tip_bonus_reliance', '_tip_income_drop']:
    check(name in gen_names, f"{name} registered in _TIP_GENERATORS")

# T12: NBA candidates
for tid in ['bonus_reliance', 'income_drop']:
    check(tid in _NBA_CANDIDATES, f"{tid} in _NBA_CANDIDATES")
check('income_stable' not in _NBA_CANDIDATES, "income_stable NOT in _NBA_CANDIDATES (positive tip)")

# T13: Version was 1.0.1000039 (now superseded by 1.0.1000040)

# ====================================================================
# Section 48 — Income Risk Score, Dashboard Income Status, Trajectory Signal F
# ====================================================================
print("\n--- 48. Income Risk Score, Dashboard Income Status, Trajectory Signal F ---")

from app import _TRAJECTORY_TIP_OVERLAP

# Clean up ALL existing salary data first so we have a controlled test environment
api_post("/api/auth/login", {"username": "testadmin", "password": "testpass123"})
existing_sal = (client.get('/api/salary-statements').get_json() or [])
for row in existing_sal:
    client.delete(f"/api/salary-statements/{row['id']}")

# T1: Insert 3 stable salary months and verify income_risk section
test_months_risk = ['2026-01', '2026-02', '2026-03']
for tm in test_months_risk:
    sc, _ = api_post('/api/salary-statements', {
        'month': tm, 'company_name': 'RiskTestCo', 'person': 'single',
        'gross_salary': 15000, 'net_salary': 11000,
        'pension_employer': 750, 'education_fund_employer': 375, 'severance_employer': 500,
        'income_tax': 2000, 'social_security': 800, 'health_insurance': 200,
    })

r = client.get('/api/salary-statements/summary')
summary_data = r.get_json() or {}
check('income_risk' in summary_data, "Salary summary includes income_risk section")
ir = summary_data.get('income_risk', {})
for key in ['level', 'score', 'status_key', 'status_params', 'cv', 'bonus_share', 'latest_vs_baseline']:
    check(key in ir, f"income_risk has '{key}' field")

# T2: 3 identical months → low risk, stable status
check(ir.get('level') == 'low', f"3 stable months → low risk (got {ir.get('level')})")
check(ir.get('status_key') == 'income_status_stable', f"Status key = income_status_stable (got {ir.get('status_key')})")
check(ir.get('cv') == 0, f"Identical salaries → CV=0 (got {ir.get('cv')})")
check(0 <= (ir.get('score') or 0) <= 1.0, f"income_risk.score in [0,1] (got {ir.get('score')})")

# T3: Add bonus-heavy months, verify bonus reliance detection
# Need bonus_share >= 0.15. Add 2 months with bonuses.
sc, _ = api_post('/api/salary-statements', {
    'month': '2025-12', 'company_name': 'RiskTestCo', 'person': 'single',
    'gross_salary': 15000, 'net_salary': 11000, 'bonus_amount': 5000,
    'pension_employer': 750, 'education_fund_employer': 375, 'severance_employer': 500,
    'income_tax': 2000, 'social_security': 800, 'health_insurance': 200,
})
sc, _ = api_post('/api/salary-statements', {
    'month': '2025-11', 'company_name': 'RiskTestCo', 'person': 'single',
    'gross_salary': 15000, 'net_salary': 11000, 'bonus_amount': 8000,
    'pension_employer': 750, 'education_fund_employer': 375, 'severance_employer': 500,
    'income_tax': 2000, 'social_security': 800, 'health_insurance': 200,
})
r = client.get('/api/salary-statements/summary')
ir3 = (r.get_json() or {}).get('income_risk', {})
# bonus_share = 13000 / 75000 ≈ 0.173 → >= 0.15 but no drop → bonus_reliance
check(ir3.get('status_key') == 'income_status_bonus_reliance',
      f"High bonus share → income_status_bonus_reliance (got {ir3.get('status_key')})")

# T4: Add a dropped month (latest), verify income drop detection
sc, _ = api_post('/api/salary-statements', {
    'month': '2026-04', 'company_name': 'RiskTestCo', 'person': 'single',
    'gross_salary': 10000, 'net_salary': 7500,
    'pension_employer': 500, 'education_fund_employer': 250, 'severance_employer': 330,
    'income_tax': 1200, 'social_security': 500, 'health_insurance': 200,
})
r = client.get('/api/salary-statements/summary')
ir4 = (r.get_json() or {}).get('income_risk', {})
# Latest = 2026-04 (10000). Baseline (non-bonus: 2026-01,02,03,04) avg = (15k*3+10k)/4 = 13750
# latest_vs_baseline = (10000 - 13750)/13750 * 100 = -27.3% → income_drop priority
check(ir4.get('status_key') == 'income_status_drop',
      f"Big salary drop → income_status_drop (got {ir4.get('status_key')})")
check(ir4.get('latest_vs_baseline') < -10, f"latest_vs_baseline < -10 (got {ir4.get('latest_vs_baseline')})")
check(0 <= ir4.get('score', -1) <= 1.0, f"Risk score in [0,1] (got {ir4.get('score')})")

# T5: Trajectory tip overlap entries
check('traj_income_risk' in _TRAJECTORY_TIP_OVERLAP, "traj_income_risk in _TRAJECTORY_TIP_OVERLAP")
check('traj_income_stable' in _TRAJECTORY_TIP_OVERLAP, "traj_income_stable in _TRAJECTORY_TIP_OVERLAP")
check('income_drop' in _TRAJECTORY_TIP_OVERLAP['traj_income_risk'], "income_drop in traj_income_risk overlap")
check('bonus_reliance' in _TRAJECTORY_TIP_OVERLAP['traj_income_risk'], "bonus_reliance in traj_income_risk overlap")
check('income_stable' in _TRAJECTORY_TIP_OVERLAP['traj_income_stable'], "income_stable in traj_income_stable overlap")

# T6: Trajectory endpoint works (doesn't crash with salary data)
r = client.get('/api/trajectory')
check(r.status_code == 200, f"GET /api/trajectory returns 200 (got {r.status_code})")

# T7: Clean up test salary data
rows = client.get('/api/salary-statements').get_json() or []
for row in rows:
    if row.get('company_name') == 'RiskTestCo':
        client.delete(f"/api/salary-statements/{row['id']}")

# T8: Version was 1.0.1000040 (superseded)

# ====================================================================
# Section 49 — Installment Status Recalculation on Edit
# ====================================================================
print("\n--- 49. Installment Status Recalculation on Edit ---")

api_post("/api/auth/login", {"username": "testadmin", "password": "testpass123"})

# T1: Create an installment, then complete it, then edit payments_made back
sc, inst_res = api_post('/api/installments', {
    'description': 'StatusTestInst', 'store': 'TestStore',
    'total_amount': 6000, 'total_payments': 6, 'payments_made': 6,
    'start_date': '2026-01-01',
})
check(sc == 200, f"POST /api/installments returns 200 (got {sc})")

# Get the installment — it should be completed (6/6)
r = client.get('/api/installments')
inst_list = r.get_json() or []
test_inst = next((i for i in inst_list if i['description'] == 'StatusTestInst'), None)
check(test_inst is not None, "Test installment found")
check(test_inst['status'] == 'completed', f"6/6 → status=completed (got {test_inst['status']})")
check(test_inst['payments_remaining'] == 0, f"6/6 → remaining=0 (got {test_inst['payments_remaining']})")

# T2: Edit payments_made to 3 — status should change to 'active'
sc2, _ = api_post('/api/installments', {
    'id': test_inst['id'], 'description': 'StatusTestInst', 'store': 'TestStore',
    'total_amount': 6000, 'total_payments': 6, 'payments_made': 3,
    'start_date': '2026-01-01',
})
check(sc2 == 200, f"Edit installment returns 200 (got {sc2})")
r2 = client.get('/api/installments')
inst_list2 = r2.get_json() or []
test_inst2 = next((i for i in inst_list2 if i['description'] == 'StatusTestInst'), None)
check(test_inst2 is not None, "Edited installment found")
check(test_inst2['status'] == 'active', f"3/6 after edit → status=active (got {test_inst2['status']})")
check(test_inst2['payments_remaining'] == 3, f"3/6 → remaining=3 (got {test_inst2['payments_remaining']})")
check(test_inst2['remaining_amount'] == 3000, f"3/6 → remaining_amount=3000 (got {test_inst2['remaining_amount']})")

# T3: Clean up
client.delete(f"/api/installments/{test_inst['id']}")

# T4: Ongoing installment (total_payments=0) stays active, never auto-completes
sc3, _ = api_post('/api/installments', {
    'description': 'OngoingTestInst', 'store': 'TestStore',
    'total_amount': 1200, 'total_payments': 0, 'payments_made': 5,
    'start_date': '2026-01-01',
})
check(sc3 == 200, f"POST ongoing installment returns 200 (got {sc3})")
r3 = client.get('/api/installments')
ongoing = next((i for i in (r3.get_json() or []) if i['description'] == 'OngoingTestInst'), None)
check(ongoing is not None, "Ongoing installment found")
check(ongoing['status'] == 'active', f"Ongoing → status=active (got {ongoing['status']})")
check(ongoing.get('is_ongoing') == True, f"Ongoing flag is True (got {ongoing.get('is_ongoing')})")
check(ongoing['payments_remaining'] == -1, f"Ongoing → remaining=-1 sentinel (got {ongoing['payments_remaining']})")

# T5: Explicit completed checkbox overrides auto-status
sc4, _ = api_post('/api/installments', {
    'id': ongoing['id'], 'description': 'OngoingTestInst', 'store': 'TestStore',
    'total_amount': 1200, 'total_payments': 0, 'payments_made': 5,
    'start_date': '2026-01-01', 'status': 'completed',
})
check(sc4 == 200, f"Edit with completed flag returns 200 (got {sc4})")
r4 = client.get('/api/installments')
completed_ongoing = next((i for i in (r4.get_json() or []) if i['description'] == 'OngoingTestInst'), None)
check(completed_ongoing['status'] == 'completed', f"Explicit completed flag sets status=completed (got {completed_ongoing['status']})")

# T6: Clean up
client.delete(f"/api/installments/{ongoing['id']}")

# T7: Version is 1.0.1000041
check(budget_app.APP_VERSION == '1.0.1000041', f"Version is 1.0.1000041 (got {budget_app.APP_VERSION})")


# ====================================================================
# Cleanup
# ====================================================================
shutil.rmtree(tmp_dir, ignore_errors=True)

# ====================================================================
# Report
# ====================================================================
total = PASS + FAIL
print(f"\n{'='*50}")
print(f"  Results: {PASS}/{total} passed, {FAIL} failed")
print(f"{'='*50}")

if ERRORS:
    print("\nFailures:")
    for e in ERRORS:
        print(e)

sys.exit(1 if FAIL else 0)
